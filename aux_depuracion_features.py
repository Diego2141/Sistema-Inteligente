# -*- coding: utf-8 -*-
"""
aux_depuracion_features.py
Genera un Excel para depuración de features con dos hojas:
  1. "Depuración Features" — catálogo completo con columnas de decisión para el equipo
  2. "Recomendación"       — análisis técnico con sugerencia Mantener/Eliminar/Evaluar
  3. "Leyenda"             — guía de criterios
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RUTA_OUTPUT = (r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente"
               r"\2. Output\depuracion_features.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# Definición completa de features
# (var, categoría, fuente, descripción, riesgo_multicolinealidad, notas)
# ─────────────────────────────────────────────────────────────────────────────
FEATURES = [
    # ── Identidad ─────────────────────────────────────────────────────────────
    ("fecha_t",  "Identidad", "Sistema",
     "Fecha de origen de la predicción", "No aplica", ""),
    ("banco",    "Identidad", "Sistema",
     "Identificador del banco",           "No aplica", ""),

    # ── Horizonte ─────────────────────────────────────────────────────────────
    ("h",      "Horizonte", "Sistema",
     "Horizonte de predicción en días hábiles (2 a 90)",
     "Baja", ""),
    ("log_h",  "Horizonte", "Sistema",
     "Logaritmo natural del horizonte h — captura relación no lineal con el horizonte",
     "Alta — log_h = log(h), relación determinista con h",
     "XGBoost no requiere transformación log; evaluar si aporta frente a h solo"),

    # ── Bancarios: niveles ────────────────────────────────────────────────────
    ("R_t0", "Bancarios - Nivel", "Datos bancarios",
     "Retiro realizado en t (conocido hoy con 2d de anticipación)",
     "Media — correlacionado con R_t-1, ma_R_5d", "Feature clave"),
    ("D_t0", "Bancarios - Nivel", "Datos bancarios",
     "Depósito realizado en t (conocido hoy con 1d de anticipación)",
     "Media — correlacionado con D_t-1, ma_D_5d", "Feature clave"),

    # ── Bancarios: rezagos ────────────────────────────────────────────────────
    ("R_t-1",  "Bancarios - Rezagos", "Datos bancarios",
     "Retiro del banco en t-1",
     "Alta — correlacionado con R_t0, R_t-2, ma_R_5d", ""),
    ("D_t-1",  "Bancarios - Rezagos", "Datos bancarios",
     "Depósito del banco en t-1",
     "Alta — correlacionado con D_t0, D_t-2, ma_D_5d", ""),
    ("R_t-2",  "Bancarios - Rezagos", "Datos bancarios",
     "Retiro del banco en t-2",
     "Alta — correlacionado con R_t-1, R_t-3", ""),
    ("D_t-2",  "Bancarios - Rezagos", "Datos bancarios",
     "Depósito del banco en t-2",
     "Alta — correlacionado con D_t-1, D_t-3", ""),
    ("R_t-3",  "Bancarios - Rezagos", "Datos bancarios",
     "Retiro del banco en t-3",
     "Alta — correlacionado con R_t-2, R_t-5", ""),
    ("D_t-3",  "Bancarios - Rezagos", "Datos bancarios",
     "Depósito del banco en t-3",
     "Alta — correlacionado con D_t-2, D_t-5", ""),
    ("R_t-5",  "Bancarios - Rezagos", "Datos bancarios",
     "Retiro del banco en t-5 (mismo día de la semana anterior)",
     "Media — captura patrón semanal", ""),
    ("D_t-5",  "Bancarios - Rezagos", "Datos bancarios",
     "Depósito del banco en t-5 (mismo día de la semana anterior)",
     "Media — captura patrón semanal", ""),
    ("R_t-22", "Bancarios - Rezagos", "Datos bancarios",
     "Retiro del banco en t-22 (mismo día del mes anterior)",
     "Baja — captura patrón mensual", ""),
    ("D_t-22", "Bancarios - Rezagos", "Datos bancarios",
     "Depósito del banco en t-22 (mismo día del mes anterior)",
     "Baja — captura patrón mensual", ""),

    # ── Bancarios: estadísticos ───────────────────────────────────────────────
    ("sigma_R_5d",  "Bancarios - Estadísticos", "Datos bancarios",
     "Desv. estándar rolling 5d de retiros",
     "Alta — correlacionado con sigma_R_22d, garch_vol", ""),
    ("sigma_D_5d",  "Bancarios - Estadísticos", "Datos bancarios",
     "Desv. estándar rolling 5d de depósitos",
     "Alta — correlacionado con sigma_D_22d", ""),
    ("ma_R_5d",     "Bancarios - Estadísticos", "Datos bancarios",
     "Media móvil 5d de retiros",
     "Alta — combinación lineal de R_t0 a R_t-4",
     "Redundante si ya se tienen R_t0, R_t-1, R_t-2, R_t-3, R_t-5"),
    ("ma_D_5d",     "Bancarios - Estadísticos", "Datos bancarios",
     "Media móvil 5d de depósitos",
     "Alta — combinación lineal de D_t0 a D_t-4",
     "Redundante si ya se tienen D_t0, D_t-1, D_t-2, D_t-3, D_t-5"),
    ("sigma_R_22d", "Bancarios - Estadísticos", "Datos bancarios",
     "Desv. estándar rolling 22d de retiros",
     "Alta — correlacionado con sigma_R_5d", ""),
    ("sigma_D_22d", "Bancarios - Estadísticos", "Datos bancarios",
     "Desv. estándar rolling 22d de depósitos",
     "Alta — correlacionado con sigma_D_5d", ""),
    ("ma_R_22d",    "Bancarios - Estadísticos", "Datos bancarios",
     "Media móvil 22d de retiros",
     "Alta — correlacionado con ma_R_5d y rezagos", ""),
    ("ma_D_22d",    "Bancarios - Estadísticos", "Datos bancarios",
     "Media móvil 22d de depósitos",
     "Alta — correlacionado con ma_D_5d y rezagos", ""),
    ("delta_R",     "Bancarios - Estadísticos", "Datos bancarios",
     "Variación diaria de retiros: R_t0 - R(t-1)",
     "Alta — combinación lineal de R_t0 y R_t-1",
     "Redundante si ya están R_t0 y R_t-1"),
    ("delta_D",     "Bancarios - Estadísticos", "Datos bancarios",
     "Variación diaria de depósitos: D_t0 - D(t-1)",
     "Alta — combinación lineal de D_t0 y D_t-1",
     "Redundante si ya están D_t0 y D_t-1"),

    # ── Flujo neto ────────────────────────────────────────────────────────────
    ("sigma_flujo_5d",       "Flujo neto", "Datos bancarios",
     "Desv. estándar rolling 5d del flujo neto D−R",
     "Alta — correlacionado con sigma_flujo_20d y garch_vol",
     "Captura régimen de volatilidad del flujo"),
    ("ma_flujo_5d",          "Flujo neto", "Datos bancarios",
     "Media móvil 5d del flujo neto D−R",
     "Alta — correlacionado con ma_flujo_20d y rezagos", ""),
    ("sigma_flujo_20d",      "Flujo neto", "Datos bancarios",
     "Desv. estándar rolling 20d del flujo neto D−R",
     "Alta — correlacionado con sigma_flujo_5d y garch_vol", ""),
    ("ma_flujo_20d",         "Flujo neto", "Datos bancarios",
     "Media móvil 20d del flujo neto D−R",
     "Alta — correlacionado con ma_flujo_5d", ""),
    ("garch_vol",            "Flujo neto", "Datos bancarios",
     "Volatilidad condicional GARCH(1,1) del flujo neto D−R",
     "Alta — correlacionado con sigma_flujo_5d y sigma_flujo_20d",
     "Captura clusters de volatilidad; más sofisticado que rolling std"),
    ("flujo_neto_acum_mes",  "Flujo neto", "Datos bancarios",
     "Acumulado del flujo neto D−R desde el primer día hábil del mes hasta t. "
     "Captura la lógica de reversión intramensual: acumulación de depósitos netos "
     "anticipa mayores retiros en cierre de mes, y viceversa.",
     "Baja — no correlacionado directamente con los otros features de flujo",
     "Feature nuevo sugerido por el equipo; evaluar importancia en step005"),

    # ── Confirmados ───────────────────────────────────────────────────────────
    ("R_conf_t1", "Confirmados", "Confirmados operativos",
     "Retiro confirmado para t+1 (aviso real o proxy histórico)",
     "Media — correlacionado con R_t0",
     "Feature de alto valor si se tienen los avisos reales"),
    ("R_conf_t2", "Confirmados", "Confirmados operativos",
     "Retiro confirmado para t+2 (aviso real o proxy histórico)",
     "Media — correlacionado con R_conf_t1", ""),
    ("D_conf_t1", "Confirmados", "Confirmados operativos",
     "Depósito confirmado para t+1 (aviso real o proxy histórico)",
     "Media — correlacionado con D_t0", ""),

    # ── Mercado financiero ────────────────────────────────────────────────────
    ("VIX",       "Mercado financiero", "Yahoo Finance",
     "Índice de volatilidad implícita S&P 500 — mide estrés global",
     "Alta — correlacionado con VIX_ma22 y delta_VIX", ""),
    ("delta_VIX", "Mercado financiero", "Yahoo Finance",
     "Variación diaria del VIX",
     "Alta — derivada de VIX",
     "Captura shocks puntuales; diferente información que el nivel"),
    ("VIX_ma22",  "Mercado financiero", "Yahoo Finance",
     "Media móvil 22d del VIX — nivel de estrés de fondo",
     "Alta — combinación lineal de VIX de los últimos 22 días",
     "Redundante si VIX ya está; XGBoost puede aprender la tendencia"),
    ("T10Y",      "Mercado financiero", "Yahoo Finance",
     "Rendimiento bono del Tesoro EE.UU. a 10 años",
     "Media — correlacionado con FED_FUNDS y diferencial_tasas",
     "Proxy de expectativas de inflación y apetito por riesgo global"),

    # ── Tipo de cambio ────────────────────────────────────────────────────────
    ("TC_PEN_USD",   "Tipo de cambio", "BCRP Add-In",
     "Tipo de cambio PEN/USD (promedio compra/venta BCRP)",
     "Media — correlacionado con delta_TC (su primera diferencia)", ""),
    ("delta_TC",     "Tipo de cambio", "BCRP Add-In",
     "Variación diaria del tipo de cambio",
     "Alta — derivada de TC_PEN_USD", ""),
    ("tc_vol_5d",    "Tipo de cambio", "BCRP Add-In",
     "Volatilidad rolling 5d de retornos del TC",
     "Alta — correlacionado con tc_vol_22d y garch_vol_tc", ""),
    ("tc_vol_22d",   "Tipo de cambio", "BCRP Add-In",
     "Volatilidad rolling 22d de retornos del TC",
     "Alta — correlacionado con tc_vol_5d y garch_vol_tc", ""),
    ("garch_vol_tc", "Tipo de cambio", "BCRP Add-In",
     "Volatilidad condicional GARCH(1,1) de retornos del TC",
     "Alta — correlacionado con tc_vol_5d y tc_vol_22d",
     "Más sofisticado que rolling vol; captura estrés cambiario"),

    # ── Riesgo país ───────────────────────────────────────────────────────────
    ("EMBI_PERU",      "Riesgo país", "BCRP Add-In",
     "EMBI Perú — spread soberano en puntos básicos",
     "Media — correlacionado con delta_EMBI y garch_vol_embi", ""),
    ("delta_EMBI",     "Riesgo país", "BCRP Add-In",
     "Variación diaria del EMBI Perú",
     "Alta — derivada de EMBI_PERU", ""),
    ("garch_vol_embi", "Riesgo país", "BCRP Add-In",
     "Volatilidad condicional GARCH(1,1) del EMBI",
     "Alta — correlacionado con EMBI_PERU y delta_EMBI",
     "Detecta estrés político/soberano"),

    # ── Política monetaria ────────────────────────────────────────────────────
    ("TASA_REF_BCRP",    "Política monetaria", "BCRP Add-In",
     "Tasa de referencia del BCRP",
     "Alta — diferencial_tasas = TASA_REF_BCRP − FED_FUNDS", ""),
    ("FED_FUNDS",        "Política monetaria", "FRED API",
     "Tasa de política monetaria de la Reserva Federal (EE.UU.)",
     "Alta — diferencial_tasas = TASA_REF_BCRP − FED_FUNDS", ""),
    ("diferencial_tasas","Política monetaria", "Calculado",
     "TASA_REF_BCRP − FED_FUNDS — diferencial de tasas PEN/USD",
     "Muy alta — combinación lineal exacta de TASA_REF_BCRP y FED_FUNDS",
     "Incluir solo uno de los tres para evitar multicolinealidad exacta"),

    # ── Calendario: mes ───────────────────────────────────────────────────────
    ("dias_al_cierre_mes",    "Calendario - Mes", "Calendario",
     "Días hábiles restantes hasta fin de mes en t+h",
     "Alta — correlacionado con dias_desde_cierre_mes y pos_en_mes", ""),
    ("dias_desde_cierre_mes", "Calendario - Mes", "Calendario",
     "Días hábiles transcurridos desde inicio de mes en t+h",
     "Alta — dias_al_cierre + dias_desde_cierre ≈ total_bdays_mes",
     "Casi redundante dado dias_al_cierre_mes y total_bdays_mes"),
    ("pos_en_mes",            "Calendario - Mes", "Calendario",
     "Posición del día hábil dentro del mes (1=primero)",
     "Alta — correlacionado con dias_al_cierre_mes",
     "Información similar a dias_al_cierre_mes"),
    ("total_bdays_mes",       "Calendario - Mes", "Calendario",
     "Total de días hábiles del mes de t+h",
     "Media — varía poco entre meses", ""),
    ("is_quincena",           "Calendario - Mes", "Calendario",
     "1 si t+h es día 15 o último hábil del mes",
     "Baja — evento puntual", "Captura pagos de quincena"),
    ("is_cierre_encaje",      "Calendario - Mes", "Calendario",
     "1 si t+h está en los últimos 2 días hábiles del mes",
     "Baja — evento puntual", "Captura presión de encaje bancario"),

    # ── Calendario: trimestral ────────────────────────────────────────────────
    ("is_penult_bday_trim", "Calendario - Trimestre", "Calendario",
     "1 si t+h es penúltimo día hábil del trimestre",
     "Baja — mutuamente excluyentes entre sí", ""),
    ("is_ultimo_bday_trim", "Calendario - Trimestre", "Calendario",
     "1 si t+h es último día hábil del trimestre",
     "Baja — mutuamente excluyentes entre sí", ""),
    ("is_1er_bday_trim",    "Calendario - Trimestre", "Calendario",
     "1 si t+h es primer día hábil del trimestre",
     "Baja — mutuamente excluyentes entre sí", ""),
    ("is_2do_bday_trim",    "Calendario - Trimestre", "Calendario",
     "1 si t+h es segundo día hábil del trimestre",
     "Baja — mutuamente excluyentes entre sí",
     "Evaluar si el segundo día tiene patrón diferenciado"),
    ("is_3er_bday_trim",    "Calendario - Trimestre", "Calendario",
     "1 si t+h es tercer día hábil del trimestre",
     "Baja — mutuamente excluyentes entre sí",
     "Evaluar si el tercer día tiene patrón diferenciado"),

    # ── Calendario: general ───────────────────────────────────────────────────
    ("dia_semana",      "Calendario - General", "Calendario",
     "Día de la semana (0=lunes, 4=viernes)",
     "Baja", "Captura patrones de inicio/fin de semana"),
    ("mes",             "Calendario - General", "Calendario",
     "Mes del año (1–12)",
     "Baja", "Captura estacionalidad anual"),
    ("is_fin_anio",     "Calendario - General", "Calendario",
     "1 si t+h es 28–31 de diciembre",
     "Baja — evento puntual anual", ""),
    ("is_pre_feriado",  "Calendario - General", "Calendario / holidays",
     "1 si el día siguiente a t+h es feriado PE o US",
     "Baja", ""),
    ("is_post_feriado", "Calendario - General", "Calendario / holidays",
     "1 si el día anterior a t+h es feriado PE o US",
     "Baja", ""),
    ("is_pre_eleccion", "Calendario - General", "Calendario",
     "1 si t+h está dentro de los 7 días previos a elecciones presidenciales",
     "Baja — evento muy esporádico (~5 observaciones en 10 años)", ""),
    ("is_post_eleccion","Calendario - General", "Calendario",
     "1 si t+h está dentro de los 7 días posteriores a elecciones presidenciales",
     "Baja — evento muy esporádico (~5 observaciones en 10 años)", ""),

    # ── Target ────────────────────────────────────────────────────────────────
    ("target", "Target", "Datos bancarios",
     "Flujo neto = D(b, t+h) - R(b, t+h). Variable a predecir.",
     "No aplica", ""),
]

# ─────────────────────────────────────────────────────────────────────────────
# Recomendaciones técnicas por feature
# (var, recomendación, razón_principal, acción_sugerida)
# Criterios: multicolinealidad, parsimonia, lógica económica, evidencia empírica
# ─────────────────────────────────────────────────────────────────────────────
RECOMENDACIONES = [
    # Identidad / sistema
    ("fecha_t",             "—",        "Columna de índice, no es feature del modelo", ""),
    ("banco",               "—",        "Columna de índice, no es feature del modelo", ""),
    ("target",              "—",        "Variable dependiente, no es feature",         ""),

    # Horizonte
    ("h",                   "Mantener", "Feature estructural del modelo; interactúa con todas las variables de calendario",
     ""),
    ("log_h",               "Eliminar", "Relación determinista con h (log_h = ln(h)). XGBoost aprende relaciones no lineales sin transformación previa.",
     "Si se quiere capturar decaimiento logarítmico, es preferible incluir h y dejar que el árbol lo aprenda"),

    # Bancarios nivel
    ("R_t0",                "Mantener", "Retiro observado hoy; señal más fuerte y reciente del proceso",
     ""),
    ("D_t0",                "Mantener", "Depósito observado hoy; señal más fuerte y reciente del proceso",
     ""),

    # Rezagos — cortos
    ("R_t-1",               "Mantener", "Rezago inmediato; captura inercia de retiros de un día al siguiente",
     ""),
    ("D_t-1",               "Mantener", "Rezago inmediato; captura inercia de depósitos",
     ""),
    ("R_t-2",               "Evaluar",  "Correlacionado con R_t-1. Aporta si hay patrón de 2 días; verificar importancia en heatmap",
     "Si importancia baja y estable en todos los folds → eliminar"),
    ("D_t-2",               "Evaluar",  "Correlacionado con D_t-1. Mismo razonamiento que R_t-2",
     ""),
    ("R_t-3",               "Evaluar",  "Cadena de correlación R_t-1 → R_t-2 → R_t-3; señal marginal decreciente",
     "Mantener solo si importancia media o alta en step005"),
    ("D_t-3",               "Evaluar",  "Mismo razonamiento que R_t-3",
     ""),

    # Rezagos — estructurales
    ("R_t-5",               "Mantener", "Mismo día de la semana anterior; captura patrón semanal estructural",
     ""),
    ("D_t-5",               "Mantener", "Mismo día de la semana anterior; captura patrón semanal",
     ""),
    ("R_t-22",              "Mantener", "Mismo día del mes anterior; captura estacionalidad mensual",
     ""),
    ("D_t-22",              "Mantener", "Mismo día del mes anterior; captura estacionalidad mensual",
     ""),

    # Estadísticos
    ("sigma_R_5d",          "Mantener", "Régimen de volatilidad de corto plazo; información distinta a los rezagos",
     ""),
    ("sigma_D_5d",          "Mantener", "Régimen de volatilidad de corto plazo de depósitos",
     ""),
    ("ma_R_5d",             "Eliminar", "Combinación lineal de R_t0…R_t-4; si estos rezagos están en el modelo, ma_R_5d es redundante",
     "XGBoost aprende el promedio por sí solo con los rezagos disponibles"),
    ("ma_D_5d",             "Eliminar", "Mismo razonamiento que ma_R_5d",
     ""),
    ("sigma_R_22d",         "Mantener", "Régimen de volatilidad de largo plazo; distinto de sigma_R_5d por la escala temporal",
     ""),
    ("sigma_D_22d",         "Mantener", "Régimen de volatilidad de largo plazo de depósitos",
     ""),
    ("ma_R_22d",            "Evaluar",  "Correlacionado con ma_R_5d y rezagos; la señal mensual ya la captura R_t-22",
     "Verificar si aporta sobre R_t-22; si no → eliminar"),
    ("ma_D_22d",            "Evaluar",  "Mismo razonamiento que ma_R_22d",
     ""),
    ("delta_R",             "Eliminar", "R_t0 − R_t-1: combinación lineal exacta de dos features ya presentes. No añade información",
     ""),
    ("delta_D",             "Eliminar", "D_t0 − D_t-1: mismo razonamiento que delta_R",
     ""),

    # Flujo neto
    ("sigma_flujo_5d",      "Mantener", "Régimen de volatilidad del flujo neto; garch_vol también captura esto pero con mayor sofisticación",
     "Evaluar si eliminar uno de los dos (sigma_flujo_5d vs garch_vol) una vez se vea importancia"),
    ("ma_flujo_5d",         "Evaluar",  "Correlacionado con ma_flujo_20d y derivable de los rezagos de D y R. Aportar marginal",
     "Si importancia baja → eliminar; si alta → mantener como señal de nivel reciente del flujo"),
    ("sigma_flujo_20d",     "Mantener", "Régimen de volatilidad de largo plazo del flujo; escala temporal distinta a sigma_flujo_5d",
     ""),
    ("ma_flujo_20d",        "Evaluar",  "Correlacionado con ma_flujo_5d. Verificar cuál tiene mayor importancia y mantener solo uno",
     "Mantener el que tenga mayor importancia estable entre folds"),
    ("garch_vol",           "Mantener", "Volatilidad condicional GARCH(1,1): captura clusters de volatilidad mejor que rolling std",
     "Feature más sofisticado del grupo; priorizar sobre sigma_flujo si hay que reducir"),
    ("flujo_neto_acum_mes", "Mantener", "Feature nuevo con lógica económica sólida: acumulación intramensual anticipa reversión en cierre de mes",
     "Verificar importancia en primer run de step005; si es consistente → mantener definitivamente"),

    # Confirmados
    ("R_conf_t1",           "Mantener", "Aviso real o proxy del retiro de mañana; información forward-looking de alto valor",
     ""),
    ("R_conf_t2",           "Mantener", "Aviso del retiro de pasado mañana; señal directa del flujo futuro",
     ""),
    ("D_conf_t1",           "Mantener", "Aviso del depósito de mañana; información forward-looking",
     ""),

    # Mercado financiero
    ("VIX",                 "Mantener", "Indicador de estrés global; captura episodios de aversión al riesgo que afectan retiros",
     ""),
    ("delta_VIX",           "Evaluar",  "Shocks puntuales del VIX; información distinta al nivel. Verificar importancia",
     "Si importancia estable y alta → mantener; si baja → eliminar"),
    ("VIX_ma22",            "Eliminar", "Combinación lineal de los últimos 22 días del VIX. Con VIX en el modelo, XGBoost puede aprender el promedio",
     ""),
    ("T10Y",                "Mantener", "Rendimiento bono a 10 años: proxy de expectativas de tasa y condiciones financieras globales",
     ""),

    # Tipo de cambio
    ("TC_PEN_USD",          "Mantener", "Nivel del TC: captura depreciación/apreciación sostenida que afecta comportamiento bancario",
     ""),
    ("delta_TC",            "Mantener", "Variación diaria: captura shocks cambiarios; señal distinta al nivel",
     ""),
    ("tc_vol_5d",           "Evaluar",  "Correlacionado con garch_vol_tc. Si garch_vol_tc se mantiene, evaluar si tc_vol_5d añade algo",
     "Preferir garch_vol_tc por mayor sofisticación; verificar importancia relativa"),
    ("tc_vol_22d",          "Evaluar",  "Correlacionado con tc_vol_5d y garch_vol_tc. Señal marginal si los otros dos están presentes",
     ""),
    ("garch_vol_tc",        "Mantener", "Volatilidad condicional GARCH del TC: detecta episodios de estrés cambiario con memoria",
     ""),

    # Riesgo país
    ("EMBI_PERU",           "Mantener", "Riesgo soberano: refleja condiciones de liquidez y confianza en el sistema financiero peruano",
     ""),
    ("delta_EMBI",          "Mantener", "Shocks de riesgo soberano; señal distinta al nivel del EMBI",
     ""),
    ("garch_vol_embi",      "Evaluar",  "Volatilidad del EMBI: puede capturar inestabilidad política; verificar importancia vs. delta_EMBI",
     "Si importancia baja → eliminar; delta_EMBI ya captura los shocks"),

    # Política monetaria
    ("TASA_REF_BCRP",       "Mantener", "Tasa de política monetaria local: afecta costo de liquidez y comportamiento bancario",
     ""),
    ("FED_FUNDS",           "Mantener", "Tasa de la Fed: condiciones financieras globales; explicativo del diferencial de tasas",
     ""),
    ("diferencial_tasas",   "Eliminar", "Combinación lineal exacta de TASA_REF_BCRP y FED_FUNDS. Introduce multicolinealidad exacta",
     "Eliminar este; TASA_REF_BCRP y FED_FUNDS aportan más información por separado"),

    # Calendario mes
    ("dias_al_cierre_mes",    "Mantener", "Posición en el mes desde el fin: captura presión de cierre; el feature más informativo del grupo",
     ""),
    ("dias_desde_cierre_mes", "Eliminar", "dias_al_cierre + dias_desde_cierre ≈ total_bdays_mes: casi redundante dado que los otros dos están presentes",
     ""),
    ("pos_en_mes",            "Eliminar", "Información similar a dias_al_cierre_mes pero medida desde el inicio; redundante",
     ""),
    ("total_bdays_mes",       "Mantener", "Controla por la longitud variable del mes (meses de 20 vs 23 días hábiles)",
     ""),
    ("is_quincena",           "Mantener", "Captura pagos de quincena: evento económico real con impacto en flujos",
     ""),
    ("is_cierre_encaje",      "Mantener", "Presión de encaje bancario en últimos 2 días del mes: mecanismo regulatorio real",
     ""),

    # Calendario trimestral
    ("is_penult_bday_trim",   "Mantener", "Penúltimo día del trimestre: ajustes de balance y provisiones",
     ""),
    ("is_ultimo_bday_trim",   "Mantener", "Último día del trimestre: cierre de balance, alto impacto en flujos",
     ""),
    ("is_1er_bday_trim",      "Mantener", "Primer día del trimestre: reversión de flujos de fin de trimestre",
     ""),
    ("is_2do_bday_trim",      "Evaluar",  "El segundo día hábil del trimestre; el patrón de reversión puede persistir 2-3 días",
     "Verificar importancia; si baja → eliminar"),
    ("is_3er_bday_trim",      "Evaluar",  "El tercer día hábil del trimestre; patrón de reversión posiblemente ya absorbido",
     ""),

    # Calendario general
    ("dia_semana",            "Mantener", "Patrón semanal estructural; lunes y viernes tienen comportamientos distintos",
     ""),
    ("mes",                   "Mantener", "Estacionalidad anual; diciembre, marzo y junio tienen patrones de flujo distintos",
     ""),
    ("is_fin_anio",           "Mantener", "Cierre de año: retiros extraordinarios, ajustes de portafolio; evento de alto impacto",
     ""),
    ("is_pre_feriado",        "Mantener", "Anticipación de feriado: los agentes retiran antes del feriado",
     ""),
    ("is_post_feriado",       "Mantener", "Normalización post-feriado: flujo de retorno",
     ""),
    ("is_pre_eleccion",       "Evaluar",  "Solo ~5 eventos en 10 años de datos: el modelo no tiene suficientes observaciones para aprender este patrón de forma confiable",
     "Evaluar con expertos; si el equipo cree que es importante → mantener como señal cualitativa"),
    ("is_post_eleccion",      "Evaluar",  "Mismo problema de escasez que is_pre_eleccion",
     ""),
]

# ─────────────────────────────────────────────────────────────────────────────
# Colores
# ─────────────────────────────────────────────────────────────────────────────
COLORES_CAT = {
    "Identidad":                "D9D9D9",
    "Horizonte":                "F2F2F2",
    "Bancarios - Nivel":        "DEEAF1",
    "Bancarios - Rezagos":      "BDD7EE",
    "Bancarios - Estadísticos": "9DC3E6",
    "Flujo neto":               "2E75B6",
    "Confirmados":              "FCE4D6",
    "Mercado financiero":       "E2EFDA",
    "Tipo de cambio":           "C6EFCE",
    "Riesgo país":              "FFEB9C",
    "Política monetaria":       "FFC7CE",
    "Calendario - Mes":         "EAD1DC",
    "Calendario - Trimestre":   "D5B8E0",
    "Calendario - General":     "C9B8E8",
    "Target":                   "404040",
}

COLORES_RIESGO = {
    "No aplica": "FFFFFF",
    "Baja":      "C6EFCE",
    "Media":     "FFEB9C",
    "Alta":      "FFCC99",
    "Muy alta":  "FFC7CE",
}

COLORES_REC = {
    "Mantener": "C6EFCE",   # verde
    "Eliminar": "FFC7CE",   # rojo
    "Evaluar":  "FFEB9C",   # amarillo
    "—":        "F2F2F2",   # gris
}


# ─────────────────────────────────────────────────────────────────────────────
# Builders
# ─────────────────────────────────────────────────────────────────────────────

def build_df():
    rows = []
    for (var, cat, fuente, desc, riesgo, notas) in FEATURES:
        nivel = riesgo.split(" —")[0].strip()
        rows.append({
            "Variable":                 var,
            "Categoría":                cat,
            "Fuente":                   fuente,
            "Descripción":              desc,
            "Riesgo multicolinealidad": nivel,
            "Detalle multicolinealidad": riesgo,
            "Notas técnicas":           notas,
            "Decisión":                 "",
            "Justificación equipo":     "",
        })
    return pd.DataFrame(rows)


def build_rec_df():
    rec_map = {var: (rec, razon, accion)
               for var, rec, razon, accion in RECOMENDACIONES}
    rows = []
    for (var, cat, fuente, desc, riesgo, _) in FEATURES:
        nivel = riesgo.split(" —")[0].strip()
        rec, razon, accion = rec_map.get(var, ("Evaluar", "", ""))
        rows.append({
            "Variable":          var,
            "Categoría":         cat,
            "Recomendación":     rec,
            "Razón principal":   razon,
            "Acción sugerida":   accion,
            "Riesgo multicol.":  nivel,
        })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Formato hoja 1 — Depuración Features
# ─────────────────────────────────────────────────────────────────────────────

def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def format_depuracion(ws, df):
    ws.title = "Depuración Features"
    anchos = {"A": 22, "B": 26, "C": 24, "D": 60,
              "E": 14, "F": 55, "G": 55, "H": 14, "I": 45}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    hf = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.fill = hf
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[1].height = 30

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        cat    = df.iloc[row_idx - 2]["Categoría"]
        riesgo = df.iloc[row_idx - 2]["Riesgo multicolinealidad"]
        cc     = COLORES_CAT.get(cat, "FFFFFF")
        rc     = COLORES_RIESGO.get(riesgo, "FFFFFF")
        fc     = "FFFFFF" if cat in ("Target", "Flujo neto") else "000000"

        for cell in row:
            col_l = get_column_letter(cell.column)
            cell.border = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(color=fc, size=10)

            if col_l == "E":
                cell.fill = PatternFill("solid", fgColor=rc)
                cell.font = Font(color="000000", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_l == "H":
                cell.fill = PatternFill("solid", fgColor="FFFACD")
                cell.font = Font(color="000000", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_l in ("A", "B"):
                cell.fill = PatternFill("solid", fgColor=cc)
            else:
                cell.fill = PatternFill("solid",
                                        fgColor=cc if col_l not in ("F", "G", "I")
                                        else "FAFAFA")
        ws.row_dimensions[row_idx].height = 40

    dv = DataValidation(type="list",
                        formula1='"Mantener,Eliminar,Evaluar,Agregar variante"',
                        allow_blank=True, showDropDown=False)
    dv.sqref = f"H2:H{ws.max_row}"
    ws.add_data_validation(dv)
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
# Formato hoja 2 — Recomendación técnica
# ─────────────────────────────────────────────────────────────────────────────

def format_recomendacion(ws, df_rec):
    ws.title = "Recomendación"

    anchos = {"A": 22, "B": 26, "C": 14, "D": 75, "E": 65, "F": 14}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    # ── Resumen en las primeras filas ─────────────────────────────────────────
    conteo = df_rec[df_rec["Recomendación"] != "—"]["Recomendación"].value_counts()
    total  = conteo.sum()

    resumen = [
        ("RESUMEN DE RECOMENDACIONES TÉCNICAS", "", "", "", "", ""),
        ("", "", "", "", "", ""),
        ("Recomendación", "N°", "% del total", "Criterio base", "", ""),
        ("Mantener",  str(conteo.get("Mantener", 0)),
         f"{conteo.get('Mantener',0)/total:.0%}",
         "Signal confirmada: baja multicolinealidad O lógica económica fuerte + no redundante", "", ""),
        ("Evaluar",   str(conteo.get("Evaluar",  0)),
         f"{conteo.get('Evaluar', 0)/total:.0%}",
         "Señal posible pero multicolinealidad o baja frecuencia; verificar con heatmap de step005", "", ""),
        ("Eliminar",  str(conteo.get("Eliminar", 0)),
         f"{conteo.get('Eliminar',0)/total:.0%}",
         "Redundante matemático (combinación lineal) o superable por otro feature del mismo grupo", "", ""),
        ("", "", "", "", "", ""),
    ]

    hf = PatternFill("solid", fgColor="1F4E79")
    row_offset = len(resumen) + 1

    for i, fila in enumerate(resumen, 1):
        for j, val in enumerate(fila, 1):
            c = ws.cell(i, j, val)
            if i == 1:
                c.fill = hf
                c.font = Font(bold=True, color="FFFFFF", size=12)
            elif i == 3:
                c.fill = PatternFill("solid", fgColor="D9D9D9")
                c.font = Font(bold=True, size=10)
            elif i in (4, 5, 6) and j == 1:
                rec_val = fila[0]
                c.fill = PatternFill("solid", fgColor=COLORES_REC.get(rec_val, "FFFFFF"))
                c.font = Font(bold=True, size=10)
            ws.row_dimensions[i].height = 20

    # ── Encabezado tabla ──────────────────────────────────────────────────────
    headers = ["Variable", "Categoría", "Recomendación",
               "Razón principal", "Acción sugerida", "Riesgo multicol."]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row_offset, j, h)
        c.fill = hf
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row_offset].height = 28

    # ── Filas de datos ────────────────────────────────────────────────────────
    for i, (_, row) in enumerate(df_rec.iterrows(), 1):
        r = row_offset + i
        vals = [row["Variable"], row["Categoría"], row["Recomendación"],
                row["Razón principal"], row["Acción sugerida"], row["Riesgo multicol."]]

        rec   = row["Recomendación"]
        cat   = row["Categoría"]
        rc    = COLORES_REC.get(rec, "FFFFFF")
        cc    = COLORES_CAT.get(cat, "FFFFFF")
        risco = COLORES_RIESGO.get(row["Riesgo multicol."], "FFFFFF")

        for j, val in enumerate(vals, 1):
            c = ws.cell(r, j, val)
            c.border = _border()
            c.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
            c.font = Font(size=10)

            if j == 3:   # Recomendación
                c.fill = PatternFill("solid", fgColor=rc)
                c.font = Font(bold=True, size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif j == 6:  # Riesgo multicol.
                c.fill = PatternFill("solid", fgColor=risco)
                c.font = Font(bold=True, size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif j in (1, 2):
                c.fill = PatternFill("solid", fgColor=cc)
                if cat in ("Target", "Flujo neto"):
                    c.font = Font(color="FFFFFF", size=10)
            else:
                c.fill = PatternFill("solid", fgColor="FAFAFA")

        ws.row_dimensions[r].height = 45

    ws.freeze_panes = f"A{row_offset + 1}"


# ─────────────────────────────────────────────────────────────────────────────
# Formato hoja 3 — Leyenda
# ─────────────────────────────────────────────────────────────────────────────

def format_leyenda(ws):
    ws.title = "Leyenda"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 65

    leyenda = [
        ("LEYENDA — RIESGO MULTICOLINEALIDAD", ""),
        ("Baja",      "Feature independiente; bajo riesgo de redundancia"),
        ("Media",     "Cierta correlación; vigilar importancia en step005"),
        ("Alta",      "Correlacionado con varios features del grupo; evaluar si aporta información nueva"),
        ("Muy alta",  "Relación matemática casi exacta; considerar eliminar"),
        ("", ""),
        ("RECOMENDACIONES TÉCNICAS", ""),
        ("Mantener",  "Feature validado: señal propia + lógica económica + bajo riesgo de redundancia"),
        ("Evaluar",   "Señal posible pero requiere verificación con heatmap de importancia (step005)"),
        ("Eliminar",  "Redundante matemático O superable por otro feature del mismo grupo"),
        ("", ""),
        ("CRITERIOS DE DEPURACIÓN", ""),
        ("Criterio 1 — Redundancia matemática",
         "Si el feature es combinación lineal de otros ya presentes → Eliminar "
         "(ej: delta_R = R_t0 − R_t-1, diferencial_tasas = TASA_REF − FED_FUNDS)"),
        ("Criterio 2 — Multicolinealidad",
         "Si el feature aporta información similar a otro del mismo grupo → "
         "Elegir el más informativo económicamente y evaluar el otro"),
        ("Criterio 3 — Parsimonia + colsample_bytree",
         "Cada feature inútil compite en el sorteo de columnas (colsample_bytree). "
         "Menos features → Optuna converge mejor con el mismo presupuesto de trials"),
        ("Criterio 4 — Frecuencia del evento",
         "Features binarios con < 20 eventos en 10 años (ej: elecciones) proveen "
         "señal insuficiente para que el modelo aprenda un patrón robusto"),
        ("", ""),
        ("PROCESO SUGERIDO", ""),
        ("Paso 1", "Aplicar las recomendaciones 'Eliminar' → lista reducida a ~55 features"),
        ("Paso 2", "Correr step005 con lista reducida; comparar heatmap de importancia por fold"),
        ("Paso 3", "Para features 'Evaluar': si importancia < mediana del fold en > 50% de folds → eliminar"),
        ("Paso 4", "Verificar coverage_90 y pinball antes/después de eliminar cada grupo"),
        ("Paso 5", "Documentar decisión final en hoja 'Depuración Features' con justificación"),
    ]

    hf = PatternFill("solid", fgColor="1F4E79")
    for i, (a, b) in enumerate(leyenda, 1):
        wa, wb = ws.cell(i, 1, a), ws.cell(i, 2, b)
        is_header = b == "" and a
        if is_header:
            for c in (wa, wb):
                c.fill = hf
                c.font = Font(bold=True, color="FFFFFF", size=11)
        else:
            wa.font = Font(bold=True, size=10)
            wb.font = Font(size=10)
            wb.alignment = Alignment(wrap_text=True)
            # Color por nivel de riesgo en la leyenda de multicolinealidad
            if a in COLORES_RIESGO:
                wa.fill = PatternFill("solid", fgColor=COLORES_RIESGO[a])
            elif a in COLORES_REC:
                wa.fill = PatternFill("solid", fgColor=COLORES_REC[a])
        ws.row_dimensions[i].height = 22 if not is_header else 24


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    Path(RUTA_OUTPUT).parent.mkdir(parents=True, exist_ok=True)

    df     = build_df()
    df_rec = build_rec_df()

    # Exportar hoja principal
    df.to_excel(RUTA_OUTPUT, index=False, sheet_name="Depuración Features")

    wb = load_workbook(RUTA_OUTPUT)
    ws_dep = wb.active
    format_depuracion(ws_dep, df)

    # Hoja de recomendaciones
    ws_rec = wb.create_sheet("Recomendación")
    # Escribir df_rec en la hoja (a partir de la fila 9, después del resumen)
    n_resumen = 8
    headers = list(df_rec.columns)
    for j, h in enumerate(headers, 1):
        ws_rec.cell(n_resumen, j, h)
    for i, (_, row) in enumerate(df_rec.iterrows(), 1):
        for j, val in enumerate(row.values, 1):
            ws_rec.cell(n_resumen + i, j, val)
    format_recomendacion(ws_rec, df_rec)

    # Hoja de leyenda
    ws_ley = wb.create_sheet("Leyenda")
    format_leyenda(ws_ley)

    wb.save(RUTA_OUTPUT)
    print(f"Archivo guardado en: {RUTA_OUTPUT}")

    # ── Resumen en consola ────────────────────────────────────────────────────
    features_modelo = df_rec[df_rec["Recomendación"] != "—"]
    conteo = features_modelo["Recomendación"].value_counts()

    print(f"\nTotal features (incluyendo identidad/target): {len(df)}")
    print(f"Features del modelo (excluye fecha_t, banco, target): {len(features_modelo)}")
    print("\nRecomendación técnica:")
    for rec in ["Mantener", "Evaluar", "Eliminar"]:
        n = conteo.get(rec, 0)
        print(f"  {rec:10s}: {n:2d} features")

    print("\nFeatures recomendados a ELIMINAR:")
    for _, r in df_rec[df_rec["Recomendación"] == "Eliminar"].iterrows():
        print(f"  - {r['Variable']:25s}  [{r['Categoría']}]  →  {r['Razón principal'][:60]}...")

    print("\nFeatures a EVALUAR (verificar con heatmap step005):")
    for _, r in df_rec[df_rec["Recomendación"] == "Evaluar"].iterrows():
        print(f"  - {r['Variable']:25s}  [{r['Categoría']}]")

    print(f"\nLista depurada mínima (solo Mantener): {conteo.get('Mantener', 0)} features")
    print(f"Lista conservadora (Mantener + Evaluar): "
          f"{conteo.get('Mantener', 0) + conteo.get('Evaluar', 0)} features")
