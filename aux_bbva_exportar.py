# -*- coding: utf-8 -*-
"""
aux_bbva_exportar.py
====================
Lee bbva_encaje.xlsx y exporta bbva_encaje_enriquecido.xlsx con tres
columnas adicionales:
  - Encaje         = Cta Cte BCR + Caja
  - Faltante       = Exigible mensual total - Encaje acumulado del mes
  - Avance         = Encaje acumulado / Exigible mensual total  (%)
"""

import pathlib
import numpy as np
import pandas as pd

RUTA_IN  = pathlib.Path(
    r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Raw\bbva_encaje.xlsx"
)
RUTA_OUT = RUTA_IN.parent / "bbva_encaje_enriquecido.xlsx"

# ── Carga ─────────────────────────────────────────────────────────────────────
print("Leyendo bbva_encaje.xlsx...")
df = pd.read_excel(RUTA_IN)

# Asignar nombres por posición (9 columnas)
df = df.iloc[:, :9].copy()
df.columns = [
    "Fecha", "Entidad", "Codigo",
    "Overnight_BCR", "Cta_Cte_BCR", "Caja",
    "TOSE_Total_Deducido", "Exigible_Diario", "Retiro_Neto"
]

df["Fecha"] = pd.to_datetime(df["Fecha"])
df = df.sort_values("Fecha").reset_index(drop=True)

df["Ano"] = df["Fecha"].dt.year
df["Mes"] = df["Fecha"].dt.month

# ── Columnas nuevas ───────────────────────────────────────────────────────────

# 1. Encaje = Cta Cte BCR + Caja  (overnight NO computa)
df["Encaje"] = df["Cta_Cte_BCR"] + df["Caja"]

# 2. Exigible total del mes = suma de Exigible_Diario de todos los días del mes
exigible_mes = (
    df.groupby(["Ano", "Mes"])["Exigible_Diario"]
    .transform("sum")
)

# Encaje acumulado dentro del mes (suma corrida día a día)
encaje_acum = (
    df.groupby(["Ano", "Mes"])["Encaje"]
    .cumsum()
)

# 3. Faltante = exigible total del mes - encaje acumulado hasta hoy
df["Faltante"] = exigible_mes - encaje_acum

# 4. Avance = encaje acumulado / exigible total del mes  (en %)
df["Avance_pct"] = (encaje_acum / exigible_mes.replace(0, np.nan) * 100).round(2)

# Quitar columnas auxiliares
df.drop(columns=["Ano", "Mes"], inplace=True)

# ── Export ────────────────────────────────────────────────────────────────────
print(f"Exportando → {RUTA_OUT.name} ...")
with pd.ExcelWriter(RUTA_OUT, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Datos", index=False)

    # Formato de la hoja
    ws = writer.sheets["Datos"]
    from openpyxl.styles import PatternFill, Font
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    new_fill    = PatternFill("solid", fgColor="E2EFDA")  # verde claro para nuevas cols

    new_cols = {"Encaje", "Faltante", "Avance_pct"}
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        if cell.value in new_cols:
            cell.fill = PatternFill("solid", fgColor="375623")

    # Ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(14, min(max_len + 2, 30))

print(f"\nListo: {RUTA_OUT}")
print(f"  Filas    : {len(df):,}")
print(f"  Columnas : {list(df.columns)}")
print(f"\n  Avance al final de cada mes (muestra últimos 5):")
fin_mes = df.groupby([df['Fecha'].dt.year, df['Fecha'].dt.month]).last()
print(fin_mes[["Fecha","Encaje","Faltante","Avance_pct"]].tail(5).to_string(index=False))
