# -*- coding: utf-8 -*-
"""
build_feature_matrix.py  ·  v2
Construye la matriz de features para el sistema de predicción de liquidez
en moneda extranjera del banco central peruano.

Modelo: Gradient Boosting con quantile regression (LightGBM).
h (horizonte) es un feature más → UN SOLO modelo por banco.
Los bancos pequeños se agrupan en "Otros_bancos" antes de construir features.

QUÉ AGREGA v2 SOBRE v1
----------------------
Particiones del sistema. SISTEMA está dominado por un puñado de bancos (BBVA
concentra ~94% del neto en la ventana de cierre), así que el modelo agregado
termina ajustando ese banco más ruido, y la dinámica del resto de la banca queda
ahogada. v2 permite partir el sistema en dos grupos complementarios y modelar
cada uno por separado:

    PARTICION_ACTIVA = "bbva"      -> FOCO_BBVA + RESTO_BBVA
    PARTICION_ACTIVA = "globales"  -> FOCO_GLOBALES + RESTO_GLOBALES
    PARTICION_ACTIVA = None        -> sin partición, idéntico a v1

Los grupos entran a lista_bancos_full como dos "bancos" más, así que reusan toda
la maquinaria de features existente sin tocarla: la familia *_pos, las
volatilidades y las medias móviles se calculan sobre la serie del grupo.

DÓNDE SE CALCULA Y POR QUÉ AHÍ
------------------------------
La partición se aplica sobre el pivot CRUDO por banco, antes de agrupar_bancos().
Es obligatorio: agrupar_bancos() suma los bancos chicos en Otros_bancos y ELIMINA
sus columnas, así que después de esa llamada un banco global pequeño (Deutsche,
ICBC, Bank of China, BCI) ya no se puede separar del bucket. Aplicada antes, la
partición ve el universo completo.

SUMAR CUANTILES NO ES CUANTIL DE LA SUMA
----------------------------------------
Advertencia para quien consuma estas matrices: el objetivo del producto sigue
siendo el TOTAL del sistema. Reconstruirlo a partir de los dos grupos NO se hace
sumando sus cuantiles predichos —eso supone que ambos salen siempre en el mismo
percentil a la vez, y acá la dependencia es probablemente NEGATIVA (cuando BBVA
retira, otros bancos aportan)—. Hay que simular trayectorias conjuntas que
preserven esa dependencia, sumarlas trayectoria por trayectoria, y recién ahí
tomar cuantiles.
"""

import os
import unicodedata
import time
import logging
import warnings
import getpass
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# Constantes para codificación cíclica de variables de calendario
# ─────────────────────────────────────────────────────────────────────────────
_P_ELEC = 1260  # 5 años × ~252 días hábiles

# Primeras vueltas de elecciones generales peruanas (cada 5 años, sin extraordinarias)
_ELECCIONES_GENERALES_NP = np.array([
    "2001-04-08", "2006-04-09", "2011-04-10",
    "2016-04-10", "2021-04-11", "2026-04-12",
    "2031-04-13", "2036-04-12", "2041-04-14",
], dtype="datetime64[D]")

# ---------------------------------------------------------------------------
# Proxy corporativo: lee credenciales de variable de entorno o las pide
# ---------------------------------------------------------------------------
if not os.environ.get("BCRP_PROXY"):
    print("=" * 60)
    print("  Configuración de proxy corporativo BCRP")
    print("=" * 60)
    _usuario = input("  Usuario de red (ej: 2579): ").strip()
    _password = getpass.getpass("  Contraseña: ")
    os.environ["BCRP_PROXY"] = f"http://{_usuario}:{_password}@bcrproxy:8080"
    print("  Proxy configurado correctamente.\n")

# ---------------------------------------------------------------------------
# Logging con timestamps
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


###############################################################################
# PARTE 0 — Parámetros globales
###############################################################################

# Fecha de corte dinámica: ayer (último día hábil confirmado).
# Lógica: D(t) se conoce con 1 día hábil de anticipación y R(t) con 2.
# Usando t-1 día hábil garantizamos que tanto R como D del último día
# estén completamente confirmados en Transacciones_BancaLocal.xlsx.
_hoy = pd.Timestamp.today().normalize()
_fin_historico = (_hoy - pd.offsets.BDay(1)).strftime("%Y-%m-%d")

# ─────────────────────────────────────────────────────────────────────────────
# FEATURES A EXCLUIR
# Columnas que se descartan del parquet final antes de guardarlo.
# Se calculan internamente (pueden ser necesarias para derivar otras) y luego
# se eliminan. Para agregar más tras revisar el heatmap de step005, añadir
# el nombre aquí. Las columnas identidad (fecha_t, banco, h, target) están
# protegidas y no pueden eliminarse desde aquí.
#
# Fuente: diccionario_variables_filtro.xlsx — columna "Color":
#   Rojo  → eliminar sin reemplazo (leakage, redundancia o escala bruta).
#   Azul  → reemplazadas por transformación sin/cos o ratio (ver abajo).
#   flujo_neto_acum_mes: re-activado (sin leakage; acumula desde inicio de mes
#                        usando D(t) conocido en t-1 y R(t) conocido en t-2).
#
# NOTA: sigma_flujo_5d/20d y tc_vol_5d/22d se calculan internamente para los
# ratios antes de que este listado los elimine del parquet final.
# ─────────────────────────────────────────────────────────────────────────────
FEATURES_EXCLUIR = [
    # ── Rojo: eliminar sin reemplazo ─────────────────────────────────────────
    "log_h",
    # R_t0/D_t0/R_t-1/D_t-1: eran la excepción dentro de esta familia —el resto
    # de los rezagos (t-2..t-22) ya estaba excluido— con la idea de que el
    # nivel más reciente fuera la señal autocorrelativa más directa para h
    # cortos. El heatmap de SHAP |mean| (participación q99, por fold) mostró
    # los cuatro con participación despreciable frente a dias_al_cierre_mes,
    # dias_desde_cierre_mes y la familia *_pos — se excluyen también.
    "R_t0", "D_t0", "R_t-1", "D_t-1",
    "R_t-2", "D_t-2", "R_t-3", "D_t-3",
    "R_t-5", "D_t-5", "R_t-22", "D_t-22",
    "sigma_R_5d", "sigma_D_5d", "ma_R_5d", "ma_D_5d",
    "sigma_R_22d", "sigma_D_22d", "ma_R_22d", "ma_D_22d",
    "delta_VIX",
    "delta_R", "delta_D",
    "VIX_ma22", "TC_PEN_USD", "delta_TC",
    "EMBI_PERU", "delta_EMBI", "garch_vol_embi",
    "diferencial_tasas",
    "EMBI_PERU_frac", "T10Y_frac", "VIX_frac",
    # dias_desde_cierre_mes e is_cierre_encaje se ACTIVAN: ver la nota en la
    # sección "reemplazadas por transformación sin/cos" más abajo.
    "pos_en_mes",          # colineal: pos_en_mes = dias_desde_cierre_mes + 1
    "total_bdays_mes",
    "is_quincena",         # mezcla el día 15 con el cierre en una sola bandera
    # Duplicado de es_post_feriado (int32, calendario PE+USA completo)
    "is_post_feriado",
    # ── SADF: sin señal en episodios de stress (hit rate P95 < 5%) ───────────
    "sadf_vol_60d", "sadf_vol_120d", "sadf_vol_252d",
    # ── Encaje: variables intermedias / con leakage ───────────────────────────
    "encaje_lag1", "exceso_lag1", "faltante_lag1", "techo_10h",
    "techo_restante_lag1", "proporcion_usada", "encaje_urgencia_lag1",
    # ── HMM: estado de régimen (inestable en ventana corta) ──────────────────
    "hmm_estado",
    # ── Azul: reemplazadas por transformación sin/cos o ratio ────────────────
    "sigma_flujo_5d", "sigma_flujo_20d",              # → sigma_flujo_ratio
    "tc_vol_5d", "tc_vol_22d",                        # → tc_vol_ratio
    # dias_al_cierre_mes y dias_desde_cierre_mes NO se excluyen: son el reloj
    # hábil crudo, con el split entero (dcm<=1) que el diagnóstico de
    # posición en el mes (aux_diagnostico_ventana_mala.py, sección F) mostró
    # necesario — los 5 días de borde concentraban el 116% del déficit de
    # cobertura. is_cierre_encaje es la binaria equivalente: agrupa los dos
    # últimos días hábiles, el bloque más chico que los modelos de cola
    # pueden aislar sin violar min_child_weight.
    # Su forma sin/cos (dias_al_cierre_mes_sin/cos) y dias_cal_al_cierre_mes
    # SÍ se excluyen (sección 8 abajo): la convivencia con la forma cruda se
    # dio por deliberada en su momento, pero gain/perm/SHAP de las corridas
    # recientes muestran que le quitan aporte al crudo en vez de sumar señal
    # propia — ver el detalle en la sección 8.
    "is_penult_bday_trim", "is_ultimo_bday_trim",     # → dias_al_cierre_trim_sin/cos
    "is_1er_bday_trim", "is_2do_bday_trim", "is_3er_bday_trim",
    "dia_semana",                                     # → dias_sem_sin/cos
    "mes",                                            # → mes_sin/cos
    "is_fin_anio",                                    # → dias_al_cierre_anio_sin/cos
    # OJO con la coma: si se comenta ANTES de ella, Python concatena este literal
    # con el de la línea siguiente y produce una entrada basura
    # ("is_post_elecciones_post_feriado") en vez de dos exclusiones. Ambas
    # aparecen igual más abajo, así que no cambia el resultado, pero es una
    # trampa si algún día se borran esas otras entradas.
    "is_pre_eleccion", "is_post_eleccion",            # → elec_sin/cos
    "es_post_feriado", "is_pre_feriado",


# 1. Información endógena del flujo: modelo base
    # R_t0/D_t0/R_t-1/D_t-1 se excluyen más arriba (línea ~90), junto a sus
    # hermanos R_t-2..t-22: R_t-1/D_t-1 SÍ existen como columna —se generan en
    # el bucle "for l in todos_lags" (línea 1389), vía f"R_t-{l}"— y por eso
    # aparecían en el heatmap de SHAP aunque una búsqueda por el texto literal
    # "R_t-1" no los encontrara en el código.
    "R_conf_t1",
    # R_conf_t2 NO se excluye — a propósito, aunque su SHAP en el heatmap sea
    # bajo. Es insumo ESTRUCTURAL de DESCOMPONER_H2 en step005 (línea ~2306):
    # el target de h=2 se redefine a D(t+2) y R_conf_t2 se resta después de
    # predecir para reconstruir el target real. step005 lo lee de
    # X_train["R_conf_t2"] — si esta columna faltara en la matriz, la
    # descomposición se desactivaría en silencio (log.warning, no error) y
    # h=2 perdería la ventaja de usar el retiro confirmado por la banca 2
    # días antes. Su baja SHAP es la firma de que el mecanismo funciona bien
    # -no necesita que el árbol la use como split, la resta es exacta y
    # ocurre fuera del modelo-, no evidencia de que sobre.
    #"R_conf_t2",
    "D_conf_t1",
    #"ma_flujo_5d",
    #"ma_flujo_20d",
    "flujo_neto_acum_mes",
    "flujo_neto_sum_5d",
    "flujo_neto_sum_22d",
    "flujo_neto_sum_66d",


# 2. Estado de volatilidad
    "sigma_22d",
    #"sigma_flujo_ratio",
    "tc_vol_ratio",
    "garch_vol_tc",

# 3. Condiciones macrofinancieras externas y locales
    "VIX",
    "FED_FUNDS",
    "T10Y",
    # TASA_REF_BCRP se excluye esta sesión: plana en gain, perm y SHAP en
    # las 5 corridas de convergencia (q01/q05/q50/q95/q99) — sin señal en
    # ninguna de las tres métricas, en ningún cuantil. A diferencia de la
    # frec_flujo_pos: excluida en revision de features. No estaba en la lista
    # —ni comentada—, asi que seguia entrando a la matriz. Ojo con la convencion
    # de este archivo: una entrada COMENTADA significa feature ACTIVA (no se la
    # excluye), o sea que comentarla para "sacarla" hace lo contrario.
    "frec_flujo_pos",

    # familia *_pos (SHAP alto, perm bajo por redundancia entre features
    # correlacionadas), acá no hay ni SHAP que la respalde.
    "TASA_REF_BCRP",
    #"CDS_PERU_5Y_frac",
    "COPPER_frac",
    

# 4. Cumplimiento y posición dentro del período de encaje
    #"avance_mes_lag1",
    #"exceso_abs_lag1",
    "exceso_dia_lag1",
    "encaje_ovn_lag1",
    "ratio_ovn_total_lag1",
    

# 5. Stocks y comportamiento de liquidez del sistema y de la entidad modelada
#    v2: ccovn_bbva_lag1/var_ccovn_bbva_lag1 se renombraron a
#    ccovn_propio_lag1/var_ccovn_propio_lag1 (dejan de ser "el saldo de BBVA
#    para todos" y pasan a ser "el saldo propio de la entidad que se modela",
#    resuelto por resolver_ccovn_lados). bbva_share_lag1 y
#    var_ccovn_bbva_exceso_lag1 se renombran igual pero SIGUEN EXCLUIDOS —
#    solo cambia el nombre para que coincida con la nueva semántica, el estado
#    activo/excluido de cada feature no cambió con esta revisión.
    #"ccovn_sistema_lag1",
    #"ccovn_propio_lag1",
    #"var_ccovn_sistema_lag1",
    #"var_ccovn_propio_lag1",
    #"ccovn_contraparte_lag1",       # nuevo en v2, activo — ver el hallazgo
    #"var_ccovn_contraparte_lag1",   # correspondiente en el diccionario
    "share_propio_lag1",
    "var_ccovn_propio_exceso_lag1",
    "ccovn_vs_dia_mes_lag1",
    "residuo_ccovn_lag1",


# 6. Discontinuidades del calendario operativo
    "dias_desde_ultimo_habil",
    "es_post_feriado","is_pre_feriado",

# 7. Estacionalidad general
    # mes_sin/cos y dias_sem_sin/cos se excluyen esta sesión: sin gain, perm
    # ni SHAP consistentes en las corridas recientes — a diferencia de los
    # ejes de proximidad a cierre (mes/trim/año), no tenían ya evidencia
    # fuerte a favor, así que se sacan directo sin necesitar el paso
    # intermedio de "coexiste con la cruda" que sí se probó para esos otros.
    "mes_sin", "mes_cos",
    "dias_sem_sin", "dias_sem_cos",

# 8. Proximidad a cierres institucionales
    # dias_al_cierre_mes_sin/cos y dias_cal_al_cierre_mes se excluyen esta
    # sesión: revierte la decisión de la sesión anterior, que las había
    # dejado activas por la evidencia de aux_diagnostico_ventana_mala.py
    # (sección F). Esa evidencia comparaba sin/cos contra NO tener nada —
    # no contra dias_al_cierre_mes (el crudo hábil, que sigue activo y sin
    # tocar). Con ambas formas conviviendo, gain/perm/SHAP muestran que
    # dias_al_cierre_mes_sin/cos y dias_cal_al_cierre_mes le quitan aporte a
    # dias_al_cierre_mes en vez de sumar señal propia — la redundancia que
    # se daba por deliberada no se sostuvo para el eje de mes con datos.
    "dias_al_cierre_mes_sin", "dias_al_cierre_mes_cos",
    "dias_cal_al_cierre_mes",
    # dias_al_cierre_trim_sin/cos y dias_al_cierre_anio_sin/cos se excluyen
    # esta sesión: el heatmap de importancia por permutación (Δloss relativo,
    # folds 1-4) muestra su forma cruda (dias_al_cierre_trim, dias_al_cierre_anio)
    # con bandas de alto impacto consistentes en todo el horizonte, mientras
    # que sus pares sin/cos apenas se distinguen del ruido de fondo.
    "dias_al_cierre_trim_sin", "dias_al_cierre_trim_cos",
    "dias_al_cierre_anio_sin", "dias_al_cierre_anio_cos",
    # is_cierre_encaje y es_mes_cierre_trim se excluyen esta sesión: mismo
    # motivo que TASA_REF_BCRP arriba — planas en gain, perm y SHAP en las 5
    # corridas de convergencia, sin señal propia en ningún cuantil. Ambas
    # eran binarias derivadas de dias_al_cierre_mes/mes, que siguen activos
    # en su forma entera y ya cubren la misma información sin necesitar el
    # split adicional que la binaria pretendía facilitar.
    "is_cierre_encaje", "es_mes_cierre_trim",

# 9. Episodios extraordinarios
    "is_post_eleccion",
    "elec_sin", "elec_cos",
# 10. Features atados al horizonte
    # presion_deadline_th/_t se eliminaron esta sesión (reemplazadas por
    # capacidad_retiro_th, sección 8c de build_feature_matrix): medían la
    # obligación de depositar, no la capacidad de retirar, y su nivel era
    # casi constante dentro del mes por construcción algebraica — ver el
    # diccionario de features para el análisis completo.
    #
    # capacidad_retiro_th excluida (no borrada): sin poder predictivo medido,
    # cobertura <3% en horizontes largos por diseño (solo proyecta dentro del
    # mes calendario de fecha_t), y agregó tiempo de corrida en step005 sin
    # mejorar cobertura — la empeoró levemente. El código y el resguardo de
    # frescura quedan intactos por si vale retomarla junto con
    # capacidad_retiro_pos (mes/trimestre) u otro ángulo.
    "capacidad_retiro_th",
    # esc_deposito_pos (anclada al cierre) queda superada por
    # esc_deposito_pos_ap (anclada a la apertura): en importancia por
    # permutación, esc_deposito_pos_ap superó notablemente a esta versión.
    # Tiene sentido mecánico — los depósitos fuertes ocurren a inicios de
    # mes, así que anclar al cierre desalinea la posición buscada en los
    # meses previos con el día donde realmente ocurre el depósito grande.
    # NO se replica el mismo argumento para esc_neto_max_pos: su espejo de
    # apertura (esc_neto_max_pos_ap) ya se probó y salió débil, así que el
    # extremo del NETO al cierre captura algo distinto del depósito puro
    # (ver hallazgo "asimetría cierre/apertura" en el diccionario de
    # features) y se mantiene activo.
    "esc_deposito_pos",
]

# ─────────────────────────────────────────────────────────────────────────────
# DIFERENCIACIÓN FRACCIONAL (FFD) — López de Prado Cap. 5
# Calibrar d mínimo solo con datos hasta esta fecha → sin leakage futuro.
# Usar la mayor fecha disponible ANTES del período de test de producción.
# ─────────────────────────────────────────────────────────────────────────────
_FD_TRAIN_CUTOFF = "2010-12-31"   # calibrar d en ventana histórica inicial

# ─────────────────────────────────────────────────────────────────────────────
# HMM — Gaussian Hidden Markov Model (3 estados: calma / moderado / severo)
# Se pre-computa UNA SOLA VEZ por banco con ventana expanding (sin leakage).
# Para etiquetar el año Y se entrena en [HMM_INICIO, Y-1].
# Requiere hmmlearn y sklearn. Si no están instalados, la columna queda en NaN.
# Si no aporta señal, agregar "hmm_estado" a FEATURES_EXCLUIR.
# ─────────────────────────────────────────────────────────────────────────────
HMM_N_ESTADOS = 3
HMM_INICIO    = "2010-01-01"   # primer año de historia disponible
HMM_MIN_AÑOS  = 2              # mínimo de años antes de etiquetar
                               # → primeras etiquetas desde HMM_INICIO + 2 años

try:
    from hmmlearn import hmm as _hmmlearn
    from sklearn.preprocessing import StandardScaler as _StandardScaler
    _HMM_DISPONIBLE = True
except ImportError:
    _HMM_DISPONIBLE = False
    logger.warning("hmmlearn/sklearn no instalados — hmm_estado quedará en NaN. "
                   "Instalar: pip install hmmlearn scikit-learn")

PARAMS = {
    # Fechas
    "fecha_inicio_historico": "2010-01-01",
    "fecha_fin_historico": _fin_historico,   # dinámico: ayer (1 día hábil de rezago)

    # Modelo
    "h_min": 2,
    "h_max": 75,
    "quantiles": [0.01, 0.05, 0.50, 0.95, 0.99],

    # Modo demo (True mientras no lleguen los datos reales)
    "usar_datos_demo": False,

    # Alias de bancos: cambios de nombre históricos que deben unificarse
    # clave = nombre en los datos, valor = nombre canónico
    "alias_bancos": {
        "CONTINEN": "BBVA",   # Banco Continental → BBVA desde 2020
    },

    # Agrupación de bancos
    "umbral_banco_pequeño_pct": 0.01,   # bancos con < 1% del volumen → Otros
    "bancos_otros": [],                  # lista fija (tiene prioridad si no está vacía)
    "nombre_otros": "Otros_bancos",

    # Partición del sistema en dos grupos complementarios. Ver PARTICIONES más
    # abajo para las opciones. None = sin partición (comportamiento de v1).
    "particion_activa": "bbva",

    # Rezagos bancarios
    "lags_cortos": [1, 2, 3],
    "lag_semana": 5,
    "lag_mes": 22,
    "ventanas_vol": [5, 22],

    # Calendario
    "años_calendario": list(range(2009, 2042)),

    # Días no hábiles adicionales: gobierno peruano los declaró inhábiles mediante
    # decreto, pero no pertenecen a ningún calendario estándar (PE ni US).
    # Se excluyen de fecha_t y se anulan en bancarios para no corromper lags/rolling.
    # Fuente: DS 354-2016-EF (APEC Lima 2016), DS 286-2019-EF (puente Santa Rosa).
    "dias_no_habiles_adicionales": [
        "2016-11-17",   # APEC Lima — DS 354-2016-EF (días 17-20 no hábiles)
        "2016-11-18",   # APEC Lima — DS 354-2016-EF
        "2019-08-29",   # Puente Santa Rosa de Lima — DS 286-2019-EF
    ],

    # Rutas archivos manuales
    "ruta_datos_bancarios": r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Raw\Transacciones_BancaLocal.xlsx",
    "ruta_encaje": r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Raw\EncajeD.xlsx",
    "ruta_bbva_encaje_features": r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\2. Output\encaje_bbva\bbva_encaje_features_modelo.xlsx",
    "banco_encaje": "BBVA",   # banco al que aplican los datos de encaje
    "ruta_ccovn":  r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Raw\Saldos_CCOVN.xlsx",
    "bbva_keyword": "bbva",   # subcadena (case-insensitive) que identifica la columna BBVA en Saldos_CCOVN
    "ruta_confirmados": r"RUTA\confirmados.xlsx",
    "ruta_intervencion": r"RUTA\intervencion.xlsx",
    # "ruta_igv" eliminado: pagos IGV en soles, no relevante para liquidez ME
    # ruta_elecciones: eliminado — fechas presidenciales hardcodeadas en build_peru_calendar()
    "ruta_aux_xgboost": r"RUTA\Aux_XGBoost.py",
    "ruta_output":      r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Clean\matriz_features.parquet",
    "ruta_diccionario": r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Clean\diccionario_variables.xlsx",

    # Series BCRP descargadas manualmente con Add-In BCRPData
    # Un solo archivo Excel, cada serie en su propia hoja
    "ruta_bcrp": r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Raw\series_bcrp.xlsx",
    "hoja_bcrp_embi":     "EMBIG",      # hoja con PD04709XD (Spread EMBIG Perú)
    "hoja_bcrp_tasa_ref": "TasaPM",     # hoja con PD12301MD (Tasa de Referencia)
    "hoja_bcrp_tc_compra": "TC Compra", # hoja con TC compra BCRP
    "hoja_bcrp_tc_venta":  "TC Venta",  # hoja con TC venta BCRP

    # Series Bloomberg — DataBBG.xlsx con valores estáticos (hojas: BVL, CDS, Cobre)
    "ruta_bloomberg": r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Raw\DataBBG.xlsx",

    # APIs externas
    "fred_api_key": "96fa168ee9a9a4c1fcf323983db5ba64",
    "proxy": os.environ.get("BCRP_PROXY", "http://bcrproxy:8080"),

    # Códigos BCRP (solo como referencia; la lectura es desde archivo)
    "bcrp_embi": "PD04709XD",      # Spread EMBIG Perú (pbs) — diaria
    "bcrp_tasa_ref": "PD12301MD",  # Tasa de Referencia Política Monetaria — diaria
    # TC se lee de BCRP (promedio compra/venta) — Yahoo Finance tenía outliers

    # Tickers Yahoo Finance
    "ticker_vix": "^VIX",
    "ticker_tc": "PEN=X",          # Tipo de cambio USD/PEN (venta)
    "ticker_t10y": "^TNX",

    # Código FRED
    "fred_fedfunds": "DFF",            # Daily Federal Funds Effective Rate (diaria)
}


###############################################################################
# PARTE 1 — Calendario hábil peruano
###############################################################################
def build_peru_calendar(años, ruta_elecciones=None):
    """
    Retorna (peru_bday, peru_holidays, fechas_elecciones).

    Calendario hábil conjunto PE + US: excluye feriados de ambos países.
    Elecciones presidenciales hardcodeadas 2000-2041 (no requieren archivo externo).
    """
    logger.info("PARTE 1: Construyendo calendario hábil peruano...")

    # ── Base garantizada: AbstractHolidayCalendar con Easter hardcoded ──────────
    # No depende de la librería 'holidays' para Jueves/Viernes Santo.
    # Incluye todos los feriados fijos PE + US más los variables de Semana Santa.
    from pandas.tseries.holiday import (
        AbstractHolidayCalendar, Holiday, GoodFriday, USFederalHolidayCalendar, Easter,
    )
    from pandas.tseries.offsets import CustomBusinessDay, Day as _Day

    _f0 = str(min(años)) + "-01-01"
    _f1 = str(max(años)) + "-12-31"

    class _PeruCalendar(AbstractHolidayCalendar):
        rules = [
            Holiday("AnioNuevo",   month=1,  day=1),
            Holiday("JuevesSanto", month=1,  day=1, offset=[Easter(), _Day(-3)]),
            GoodFriday,                              # Viernes Santo
            Holiday("Trabajo",     month=5,  day=1),
            Holiday("SanPedro",    month=6,  day=29),
            Holiday("FiestasP1",   month=7,  day=28),
            Holiday("FiestasP2",   month=7,  day=29),
            Holiday("SantaRosa",   month=8,  day=30),
            Holiday("Angamos",     month=10, day=8),
            Holiday("TodosSantos", month=11, day=1),
            Holiday("Inmaculada",  month=12, day=8),
            Holiday("Nochebuena",  month=12, day=24),
            Holiday("Navidad",     month=12, day=25),
        ]

    _hols_pe = set(_PeruCalendar().holidays(_f0, _f1).normalize())
    _hols_us = set(USFederalHolidayCalendar().holidays(_f0, _f1).normalize())
    _hols_base = _hols_pe | _hols_us

    # ── Suplemento opcional: librería 'holidays' añade fechas adicionales ────────
    # (p.e. feriados municipales, reemplazos por lunes, etc.)
    # Si no está instalada se usa la base garantizada — Semana Santa siempre incluida.
    try:
        import holidays as _hlib
        _sup_pe = set(pd.to_datetime(list(_hlib.Peru(years=años).keys())).normalize())
        _sup_us = set(pd.to_datetime(list(_hlib.UnitedStates(years=años).keys())).normalize())
        _hols_all = _hols_base | _sup_pe | _sup_us
        logger.info(
            f"  Feriados: {len(_hols_base)} base garantizada + "
            f"{len(_hols_all) - len(_hols_base)} adicionales (librería 'holidays') "
            f"= {len(_hols_all)} únicos en {min(años)}-{max(años)}"
        )
    except ImportError:
        _hols_all = _hols_base
        logger.warning(
            f"  Librería 'holidays' no instalada — usando {len(_hols_base)} feriados "
            f"base (PE+US fijos + Semana Santa). Instalar con: pip install holidays"
        )

    peru_holidays = pd.to_datetime(sorted(_hols_all))
    peru_bday     = CustomBusinessDay(holidays=peru_holidays)

    # Elecciones presidenciales peruanas hardcodeadas 2000-2041
    # Fechas desde 2031 son estimadas (2do domingo de abril / 1er domingo de junio)
    ULTIMA_FECHA_CONFIRMADA = pd.Timestamp("2026-06-07")
    ULTIMO_ANIO_HORIZONTE   = 2041

    _elecciones_raw = [
        # ── Históricas confirmadas ──────────────────────────────
        ("2000-04-09", "1ra vuelta"),  ("2000-05-28", "2da vuelta"),
        ("2001-04-08", "1ra vuelta"),  ("2001-06-03", "2da vuelta"),
        ("2006-04-09", "1ra vuelta"),  ("2006-06-04", "2da vuelta"),
        ("2011-04-10", "1ra vuelta"),  ("2011-06-05", "2da vuelta"),
        ("2016-04-10", "1ra vuelta"),  ("2016-06-05", "2da vuelta"),
        ("2021-04-11", "1ra vuelta"),  ("2021-06-06", "2da vuelta"),
        # ── Programadas / confirmadas ───────────────────────────
        ("2026-04-12", "1ra vuelta"),  ("2026-06-07", "2da vuelta"),
        # ── Estimadas (cada 5 años, 2do domingo abril / ~7 sem después) ──
        ("2031-04-13", "1ra vuelta"),  ("2031-06-08", "2da vuelta"),
        ("2036-04-12", "1ra vuelta"),  ("2036-06-07", "2da vuelta"),
        ("2041-04-14", "1ra vuelta"),  ("2041-06-08", "2da vuelta"),
    ]

    fechas_elecciones = [pd.Timestamp(f) for f, _ in _elecciones_raw]
    logger.info(
        f"  Elecciones presidenciales cargadas: {len(fechas_elecciones)} fechas "
        f"(2000-{ULTIMO_ANIO_HORIZONTE}, estimadas desde 2031)"
    )

    # Avisar cuando el horizonte de predicción se acerca al límite estimado
    fecha_hoy = pd.Timestamp.today().normalize()
    dias_al_limite = (pd.Timestamp(f"{ULTIMO_ANIO_HORIZONTE}-12-31") - fecha_hoy).days
    if dias_al_limite < 365 * 3:
        logger.warning(
            f"  ⚠ HORIZONTE ELECTORAL: quedan ~{dias_al_limite // 365} años hasta el límite "
            f"estimado ({ULTIMO_ANIO_HORIZONTE}). Considera extender las fechas en el código."
        )
    primera_estimada = pd.Timestamp("2031-04-13")
    dias_a_estimada = (primera_estimada - fecha_hoy).days
    if dias_a_estimada < 365 * 2:
        logger.warning(
            f"  ⚠ ELECCIONES ESTIMADAS: en ~{dias_a_estimada // 30} meses se entra a fechas "
            f"estimadas (desde {primera_estimada.date()}). Verificar con JNE."
        )

    logger.info("  Calendario construido correctamente.")
    return peru_bday, peru_holidays, fechas_elecciones


def get_future_business_dates(t, h_max, peru_bday):
    """Retorna los h_max días hábiles siguientes desde t (excluye t)."""
    fechas = pd.date_range(start=t + peru_bday, periods=h_max, freq=peru_bday)
    return fechas


###############################################################################
# PARTE 2 — Carga de datos manuales
###############################################################################
def load_manual_data(params):
    """
    Carga datos manuales desde archivos Excel definidos en PARAMS.

    Si un archivo no está disponible, carga un DataFrame vacío y muestra advertencia.
    Retorna dict con claves: 'bancarios', 'confirmados', 'intervencion'.
    """
    logger.info("PARTE 2: Cargando datos manuales...")

    resultado = {
        "bancarios": pd.DataFrame(),
        "confirmados": pd.DataFrame(columns=["banco", "R_conf_t1", "R_conf_t2", "D_conf_t1"]),
        "intervencion": pd.Series(dtype=float),
    }

    # Datos bancarios — formato Transacciones_BancaLocal.xlsx
    # Columnas: Broker (banco), Fecha Valor (fecha), Delivery Principal Usd (monto)
    # Monto < 0 → Retiro (R) | Monto > 0 → Depósito (D)
    try:
        df_raw = pd.read_excel(params["ruta_datos_bancarios"])
        if df_raw.empty:
            logger.warning(f"  Archivo bancarios vacío: {params['ruta_datos_bancarios']}")
        else:
            df_raw["Fecha Valor"] = pd.to_datetime(df_raw["Fecha Valor"])
            df_raw["monto"] = pd.to_numeric(df_raw["Delivery Principal Usd"], errors="coerce")

            # Separar retiros y depósitos
            df_raw["R"] = df_raw["monto"].clip(upper=0).abs()  # negativos → positivos
            df_raw["D"] = df_raw["monto"].clip(lower=0)        # positivos

            # Agregar por banco y fecha
            df_banco = (
                df_raw.groupby(["Broker", "Fecha Valor"])[["R", "D"]]
                .sum()
                .reset_index()
                .rename(columns={"Broker": "banco", "Fecha Valor": "fecha"})
            )
            df_banco["fecha"] = pd.to_datetime(df_banco["fecha"])

            # Unificar alias históricos (ej. CONTINEN → BBVA)
            # Comparación case-insensitive para tolerar variaciones de capitalización
            alias = params.get("alias_bancos", {})
            if alias:
                alias_lower = {k.lower(): v for k, v in alias.items()}
                def _aplicar_alias(nombre):
                    return alias_lower.get(nombre.lower(), nombre)
                antes = set(df_banco["banco"].unique())
                df_banco["banco"] = df_banco["banco"].apply(_aplicar_alias)
                despues = set(df_banco["banco"].unique())
                aplicados = antes - despues
                for viejo in aplicados:
                    nuevo = _aplicar_alias(viejo)
                    logger.info(f"  Alias aplicado: '{viejo}' → '{nuevo}'")
                # Re-agregar por si el mismo banco tiene filas con ambos nombres en la misma fecha
                df_banco = (
                    df_banco.groupby(["banco", "fecha"])[["R", "D"]]
                    .sum()
                    .reset_index()
                )

            df_banco = df_banco.sort_values(["banco", "fecha"])
            n_bancos = df_banco["banco"].nunique()
            f_min = df_banco["fecha"].min().date()
            f_max = df_banco["fecha"].max().date()
            logger.info(
                f"  Datos bancarios cargados: {len(df_banco):,} filas | "
                f"{n_bancos} bancos | {f_min} → {f_max}"
            )

            # Convertir formato LARGO → ANCHO: índice=fecha, columnas={banco}_R / {banco}_D
            # Necesario para compatibilidad con agrupar_bancos y build_feature_matrix
            df_wide = df_banco.pivot_table(
                index="fecha", columns="banco", values=["R", "D"], aggfunc="sum"
            )
            # MultiIndex ('R', 'BBVA') → 'BBVA_R'
            df_wide.columns = [f"{col[1]}_{col[0]}" for col in df_wide.columns]
            df_wide = df_wide.sort_index()
            df_wide.index.name = "fecha"

            # Rellenar calendario completo con 0: días sin transacción = flujo cero,
            # no dato desconocido. Evita NaN masivos en lags, sigmas y medias móviles.
            # reindex() añade fechas faltantes (fill_value=0).
            # fillna(0.0) cubre además los NaN internos del pivot (banco sin transacción
            # en una fecha donde otro banco sí la tuvo — pivot_table los deja como NaN).
            idx_completo = pd.bdate_range(start=df_wide.index.min(), end=df_wide.index.max())
            df_wide = df_wide.reindex(idx_completo, fill_value=0.0).fillna(0.0)
            df_wide.index.name = "fecha"

            resultado["bancarios"] = df_wide
            logger.info(
                f"  Matriz bancaria pivotada: {df_wide.shape[0]:,} fechas × "
                f"{df_wide.shape[1]} columnas ({n_bancos} bancos × 2) "
                f"[calendario completo, ceros en días sin transacción]"
            )
    except Exception as e:
        logger.warning(f"  No se pudo cargar datos bancarios: {params['ruta_datos_bancarios']} | {e}")

    # Confirmados (proxy histórico: valores realizados en t+1/t+2)
    ruta_conf = params.get("ruta_confirmados", "")
    if ruta_conf and ruta_conf != r"RUTA\confirmados.xlsx":
        try:
            df = pd.read_excel(ruta_conf)
            if df.empty:
                logger.warning(f"  Archivo confirmados vacío: {ruta_conf}")
            else:
                resultado["confirmados"] = df
                logger.info(f"  Datos confirmados cargados: {df.shape}")
        except Exception as e:
            logger.warning(f"  No se pudo cargar confirmados: {ruta_conf} | {e}")
    else:
        logger.info("  Confirmados: usando proxy histórico (D_t+1, R_t+2 realizados).")

    logger.info("  Carga de datos manuales completada.")
    return resultado


def load_encaje_data(params):
    """
    Carga EncajeD.xlsx con datos de encaje de BBVA.

    Columnas esperadas (nombres normalizados a minúsculas y sin tildes):
      fecha, caja, cta_cte_bcr, overnight_bcr, activos, encaje_exigible, retiro_neto

    Retorna DataFrame indexado por fecha con las columnas normalizadas,
    o DataFrame vacío si el archivo no está disponible.
    """
    ruta = params.get("ruta_encaje", "")
    if not ruta or ruta == r"RUTA\EncajeD.xlsx":
        logger.info("  Encaje: ruta no configurada — features de encaje omitidas.")
        return pd.DataFrame()

    try:
        df = pd.read_excel(ruta)
        if df.empty:
            logger.warning(f"  EncajeD.xlsx vacío: {ruta}")
            return pd.DataFrame()

        # Normalizar nombres de columna: minúsculas, sin tildes, sin espacios
        import unicodedata
        def _norm(s):
            s = str(s).lower().strip()
            s = "".join(c for c in unicodedata.normalize("NFD", s)
                        if unicodedata.category(c) != "Mn")
            return s.replace(" ", "_").replace(".", "_")

        df.columns = [_norm(c) for c in df.columns]

        # Mapear variantes de nombres al estándar interno
        _alias = {
            "cta__cte__bcr": "cta_cte_bcr",
            "cta_cte": "cta_cte_bcr",
            "overnight": "overnight_bcr",
            "exigible": "encaje_exigible",
            "retiro": "retiro_neto",
        }
        df = df.rename(columns=_alias)

        # Detectar columna de fecha
        fecha_col = next((c for c in df.columns if "fecha" in c), None)
        if fecha_col is None:
            logger.warning("  EncajeD.xlsx: no se encontró columna de fecha.")
            return pd.DataFrame()

        df["fecha"] = pd.to_datetime(df[fecha_col]).dt.normalize()
        df = df.set_index("fecha").sort_index()

        # Conservar solo columnas numéricas relevantes
        cols_requeridas = ["caja", "cta_cte_bcr", "overnight_bcr", "activos",
                           "encaje_exigible", "retiro_neto"]
        cols_presentes = [c for c in cols_requeridas if c in df.columns]
        df = df[cols_presentes].apply(pd.to_numeric, errors="coerce")

        logger.info(
            f"  EncajeD cargado: {len(df):,} filas | "
            f"{df.index.min().date()} → {df.index.max().date()} | "
            f"cols: {cols_presentes}"
        )
        return df

    except Exception as e:
        logger.warning(f"  No se pudo cargar EncajeD.xlsx: {ruta} | {e}")
        return pd.DataFrame()


def load_ccovn_data(params):
    """
    Carga Saldos_CCOVN.xlsx: saldos de cierre CC + OVN por banco en el BCR.

    Formato esperado:
      - Col A: Fecha (d-m-yyyy)
      - Cols B…: un banco por columna (valores numéricos en USD)

    v2: a diferencia de v1, NO colapsa a solo sistema/bbva. Retorna todas las
    columnas por banco tal cual están en el Excel (numéricas, sin normalizar el
    nombre todavía), más 'sistema' = suma de todas. La normalización y el
    emparejamiento contra los bancos canónicos de Transacciones_BancaLocal.xlsx
    ocurre después, en armar_ccovn_ancho() — separar los dos pasos permite que
    esta función siga sin conocer bancos_canonicos, que en build_full_matrix
    recién se conoce tras agrupar_bancos().

    bbva_keyword queda sin uso (la extracción de "la columna BBVA" ahora es un
    caso particular del emparejamiento genérico por partición), pero el
    parámetro se conserva en PARAMS para no romper configuraciones existentes.

    Indexado por fecha, o DataFrame vacío si el archivo no está disponible.
    """
    ruta = params.get("ruta_ccovn", "")
    if not ruta:
        logger.info("  CCOVN: ruta no configurada — features ccovn omitidas.")
        return pd.DataFrame()

    try:
        raw = pd.read_excel(ruta, header=0)
        if raw.empty:
            logger.warning(f"  Saldos_CCOVN.xlsx vacío: {ruta}")
            return pd.DataFrame()

        col_fecha   = raw.columns[0]
        cols_bancos = raw.columns[1:].tolist()

        raw[col_fecha] = pd.to_datetime(raw[col_fecha], dayfirst=True, errors="coerce")
        raw = raw.dropna(subset=[col_fecha]).sort_values(col_fecha)

        for c in cols_bancos:
            raw[c] = pd.to_numeric(raw[c], errors="coerce")

        df_out = raw[cols_bancos].copy()
        df_out["sistema"] = raw[cols_bancos].sum(axis=1, skipna=True)
        df_out.index = pd.DatetimeIndex(raw[col_fecha].values)
        df_out.index.name = "fecha"
        df_out = df_out.sort_index()

        logger.info(
            f"  Saldos_CCOVN cargado: {len(df_out):,} filas | "
            f"{df_out.index.min().date()} → {df_out.index.max().date()} | "
            f"{len(cols_bancos)} columnas de banco: {cols_bancos}"
        )
        return df_out

    except Exception as e:
        logger.warning(f"  No se pudo cargar Saldos_CCOVN.xlsx: {ruta} | {e}")
        return pd.DataFrame()


###############################################################################
# PARTE 3 — Agrupación de bancos
###############################################################################
# ─────────────────────────────────────────────────────────────────────────────
# Particiones del sistema (v2)
# ─────────────────────────────────────────────────────────────────────────────
# Claves de coincidencia por SUBCADENA sobre el nombre normalizado, no por nombre
# exacto. Es el mismo criterio que usa aux_graficos_hallazgos.py, y el motivo es
# que un mismo banco aparece con varias grafías en los datos fuente
# ("BANCO DE CREDITO", "B. CREDITO", "BCP") y con cambios de razón social a lo
# largo de 2010-2026. PARAMS["alias_bancos"] ya unifica los casos conocidos, pero
# la subcadena cubre los que no están mapeados sin que la partición falle en
# silencio dejando un banco del lado equivocado.
CLAVES_GLOBALES = [
    "BBVA", "CONTINEN",        # BBVA / ex Banco Continental
    "CITI",                    # Citibank
    "SCOTIABANK",              # Scotiabank (matriz canadiense)
    "SANTANDER",
    "HSBC",
    "DEUTSCHE",
    "JPMORGAN",
    "BANKOFCHINA",
    "ICBC",
    "BCI",                     # Banco de Crédito e Inversiones (Chile)
]

# Cada partición corta el sistema en DOS grupos complementarios y exhaustivos:
# todo banco cae en foco o en resto, nunca en ambos ni en ninguno. Esa propiedad
# se verifica en aplicar_particion(), no se asume.
PARTICIONES = {
    "bbva": {
        "claves": ["BBVA", "CONTINEN"],
        "foco":  "FOCO_BBVA",
        "resto": "RESTO_BBVA",
        "descripcion": "BBVA contra el resto del sistema",
    },
    "globales": {
        "claves": CLAVES_GLOBALES,
        "foco":  "FOCO_GLOBALES",
        "resto": "RESTO_GLOBALES",
        "descripcion": "bancos con matriz en el exterior contra la banca local",
    },
}


def _normalizar_banco(nombre):
    """
    Mayúsculas, sin tildes, sin espacios ni puntuación, para comparar por
    subcadena.

    Quitar tildes es necesario y no cosmético: "CREDITO" (nombre canónico de
    Transacciones_BancaLocal) no es subcadena de "CRÉDITO" con tilde (headers
    de Saldos_CCOVN.xlsx u otras fuentes pueden traer el nombre completo con
    acentos, ej. "Banco de Crédito del Perú") porque .upper() preserva la Í/É/Ó
    acentuada como carácter distinto de la letra sin acento — ch.isalnum() la
    deja pasar igual, así que el filtro anterior no la sacaba. Se usa
    normalización NFKD para separar cada letra de su diacrítico y luego se
    descartan los diacríticos (categoría Unicode Mn), antes de filtrar alnum.
    """
    sin_tildes = "".join(
        ch for ch in unicodedata.normalize("NFKD", str(nombre))
        if unicodedata.category(ch) != "Mn"
    )
    return "".join(ch for ch in sin_tildes.upper() if ch.isalnum())


def _es_del_foco(banco, claves):
    n = _normalizar_banco(banco)
    return any(k in n for k in claves)


def aplicar_particion(df_bancarios, particion, nombre_otros=None):
    """
    Agrega las series agregadas de los dos grupos de una partición.

    TIENE QUE LLAMARSE ANTES DE agrupar_bancos(). Esa función suma los bancos
    chicos en Otros_bancos y ELIMINA sus columnas individuales, así que después
    un banco del foco que sea pequeño (Deutsche, ICBC, Bank of China o BCI están
    típicamente por debajo del umbral del 1%) ya quedó dentro del bucket y no hay
    forma de separarlo. Aplicada antes, la partición ve el universo completo.

    Devuelve (df con las 4 columnas nuevas, reporte). Si particion es None
    devuelve el df intacto y un reporte vacío, que es el comportamiento de v1.
    """
    if not particion:
        return df_bancarios, {"activa": None}

    cfg = PARTICIONES.get(particion)
    if cfg is None:
        raise ValueError(
            f"particion_activa={particion!r} no existe. "
            f"Opciones: {sorted(PARTICIONES)} o None.")

    bancos = sorted({c[:-2] for c in df_bancarios.columns
                     if c.endswith(("_R", "_D"))})
    # Otros_bancos no debería existir todavía —esta función corre antes de
    # agrupar_bancos()— pero si alguien invierte el orden, sumarlo al resto
    # mezclaría bancos del foco que ya cayeron adentro. Mejor abortar.
    if nombre_otros and nombre_otros in bancos:
        raise RuntimeError(
            f"aplicar_particion() encontró la columna agregada {nombre_otros!r}: "
            f"se está llamando DESPUÉS de agrupar_bancos(), y en ese punto los "
            f"bancos chicos del foco ya son inseparables del bucket.")

    foco  = [b for b in bancos if _es_del_foco(b, cfg["claves"])]
    resto = [b for b in bancos if b not in foco]

    if not foco:
        raise ValueError(
            f"La partición {particion!r} no encontró ningún banco. "
            f"Claves: {cfg['claves']} | bancos disponibles: {bancos}")

    df = df_bancarios.copy()
    for sufijo in ("_R", "_D"):
        cols_f = [f"{b}{sufijo}" for b in foco  if f"{b}{sufijo}" in df.columns]
        cols_r = [f"{b}{sufijo}" for b in resto if f"{b}{sufijo}" in df.columns]
        df[f"{cfg['foco']}{sufijo}"]  = df[cols_f].sum(axis=1) if cols_f else 0.0
        df[f"{cfg['resto']}{sufijo}"] = df[cols_r].sum(axis=1) if cols_r else 0.0

    # Los dos grupos tienen que reconstruir el total exactamente. Si no cierra,
    # algún banco quedó de los dos lados o de ninguno, y el error se propagaría
    # en silencio a todas las features de ambos grupos.
    for sufijo in ("_R", "_D"):
        crudas = [f"{b}{sufijo}" for b in bancos if f"{b}{sufijo}" in df.columns]
        total = df[crudas].sum(axis=1)
        suma  = df[f"{cfg['foco']}{sufijo}"] + df[f"{cfg['resto']}{sufijo}"]
        peor = float((total - suma).abs().max())
        if peor > 1e-6:
            raise AssertionError(
                f"partición {particion!r}: foco + resto != total en {sufijo} "
                f"(peor desvío {peor:.6g}). Revisar solapamiento de claves.")

    reporte = {
        "activa": particion,
        "descripcion": cfg["descripcion"],
        "nombre_foco": cfg["foco"],
        "nombre_resto": cfg["resto"],
        "bancos_foco": foco,
        "bancos_resto": resto,
    }
    logger.info(f"  Partición '{particion}' ({cfg['descripcion']})")
    logger.info(f"    {cfg['foco']}  = {len(foco)} banco(s): {', '.join(foco)}")
    logger.info(f"    {cfg['resto']} = {len(resto)} banco(s): "
                f"{', '.join(resto) if len(resto) <= 8 else str(len(resto)) + ' bancos'}")
    return df, reporte


def columnas_derivadas(reporte_particion, nombre_sistema="SISTEMA"):
    """
    Nombres de banco que NO son bancos reales sino agregados calculados.

    Existe para que ni agrupar_bancos() ni la suma de SISTEMA los traten como una
    entidad más: si FOCO y RESTO entraran a esa suma, el sistema quedaría al
    doble, y si entraran al conteo de volumen de agrupar_bancos() competirían con
    los bancos reales por el umbral del 1%. Ninguno de los dos errores rompe nada
    visible, solo produce números mal.
    """
    derivadas = {nombre_sistema}
    if reporte_particion and reporte_particion.get("activa"):
        derivadas.add(reporte_particion["nombre_foco"])
        derivadas.add(reporte_particion["nombre_resto"])
    return derivadas


# ─────────────────────────────────────────────────────────────────────────────
# CCOVN por partición (v2)
# ─────────────────────────────────────────────────────────────────────────────
def _mapear_bancos_ccovn(cols_ccovn, bancos_canonicos):
    """
    Empareja cada banco CANÓNICO (nombres que salen de Transacciones_BancaLocal,
    ej. "BBVA", "CREDITO") con la columna de Saldos_CCOVN.xlsx cuyo nombre lo
    contiene o lo contienen, normalizado.

    Los dos archivos no tienen por qué compartir la misma grafía de encabezado
    ("CREDITO" contra "Banco de Crédito", por ejemplo), así que no se busca
    igualdad exacta — igual que la clasificación de la partición, que ya usa
    coincidencia por subcadena para el mismo problema.

    Retorna {banco_canonico: columna_ccovn_o_None}. Un banco sin match queda en
    None y su columna derivada será NaN — degradación explícita, no un error
    silencioso, porque el llamador reporta los faltantes.
    """
    normal_cols = {c: _normalizar_banco(c) for c in cols_ccovn}
    mapeo = {}
    for b in bancos_canonicos:
        nb = _normalizar_banco(b)
        candidatos = [c for c, nc in normal_cols.items() if nb in nc or nc in nb]
        if len(candidatos) > 1:
            # Ambigüedad (varias columnas contienen la clave): la más corta es
            # la menos probable candidata a coincidencia espuria por casualidad.
            candidatos.sort(key=len)
        mapeo[b] = candidatos[0] if candidatos else None
    return mapeo


def armar_ccovn_ancho(df_ccovn_raw, bancos_canonicos, reporte_particion):
    """
    Convierte el CCOVN crudo (una columna por header de Excel) en un DataFrame
    ancho con columnas 'sistema', 'banco_<X>' por cada banco canónico emparejado,
    y 'foco'/'resto' si hay partición activa.

    'sistema' es la suma de TODAS las columnas del Excel, emparejadas o no — no
    depende del matching, es la misma cifra que v1 ya reportaba.

    'foco'/'resto' SÍ dependen del matching: son la suma de las columnas
    emparejadas a los bancos que la partición clasificó de cada lado, según
    aplicar_particion() sobre Transacciones_BancaLocal. Si un banco del foco no
    tiene columna en Saldos_CCOVN.xlsx, ccovn_foco queda subestimado para esa
    fecha en la parte de ese banco — se reporta la cobertura para que quede
    visible cuánto del foco realmente sostiene la cifra.

    Retorna (df_ancho, reporte_matching).
    """
    if df_ccovn_raw.empty:
        return pd.DataFrame(), {"cols_ccovn": [], "mapeo": {}, "sin_match": list(bancos_canonicos)}

    cols_raw = [c for c in df_ccovn_raw.columns if c != "sistema"]
    mapeo = _mapear_bancos_ccovn(cols_raw, bancos_canonicos)
    sin_match = sorted(b for b, c in mapeo.items() if c is None)

    ancho = pd.DataFrame(index=df_ccovn_raw.index)
    ancho["sistema"] = df_ccovn_raw["sistema"]
    for b, col in mapeo.items():
        ancho[f"banco_{b}"] = df_ccovn_raw[col] if col is not None else np.nan

    reporte = {"cols_ccovn": cols_raw, "mapeo": mapeo, "sin_match": sin_match}
    if sin_match:
        logger.warning(f"  CCOVN: sin columna emparejada para {sin_match} — "
                       f"su 'propio' quedará en NaN, y si son parte del foco o "
                       f"resto de la partición activa, subestiman esa suma.")

    if reporte_particion and reporte_particion.get("activa"):
        f_bancos = reporte_particion["bancos_foco"]
        r_bancos = reporte_particion["bancos_resto"]
        cols_f = [mapeo[b] for b in f_bancos if mapeo.get(b) is not None]
        cols_r = [mapeo[b] for b in r_bancos if mapeo.get(b) is not None]
        ancho["foco"]  = df_ccovn_raw[cols_f].sum(axis=1) if cols_f else np.nan
        ancho["resto"] = df_ccovn_raw[cols_r].sum(axis=1) if cols_r else np.nan
        cob_f = len(cols_f) / max(len(f_bancos), 1)
        cob_r = len(cols_r) / max(len(r_bancos), 1)
        reporte["cobertura_foco"], reporte["cobertura_resto"] = cob_f, cob_r
        logger.info(f"  CCOVN partición '{reporte_particion['activa']}': "
                    f"cobertura foco {cob_f:.0%} ({len(cols_f)}/{len(f_bancos)}), "
                    f"resto {cob_r:.0%} ({len(cols_r)}/{len(r_bancos)})")

    return ancho, reporte


def resolver_ccovn_lados(banco, nombre_sistema, reporte_particion):
    """
    Claves (dentro del df ancho de armar_ccovn_ancho / de las columnas
    ccovn_<clave>_lag1 que produce build_ccovn_features) que corresponden al
    saldo PROPIO de esta entidad y, si aplica, al de su CONTRAPARTE.

    Misma lógica de composición exacta que destinos_encaje_bbva(): un banco
    individual solo hereda el rol de un lado de la partición si su composición
    coincide EXACTO con ese lado (no "contiene"), para no atribuirle a un banco
    suelto el saldo de un grupo de varios.

    Devuelve (clave_propio, clave_contraparte). clave_contraparte es None cuando
    no hay una contraparte natural: SISTEMA es el total, y un banco individual
    fuera de la partición activa no tiene "el otro lado" definido.
    """
    if banco == nombre_sistema:
        # Con partición activa, la contraparte natural de SISTEMA es el FOCO.
        # No es cosmético: v1 le daba al modelo de SISTEMA la columna
        # bbva_share_lag1, que es la señal de concentración del hallazgo 3. Sin
        # esta línea, SISTEMA se queda con share_propio == 1 por construcción y
        # pierde esa señal en el camino al renombre relativo.
        if reporte_particion and reporte_particion.get("activa"):
            return "sistema", "foco"
        return "sistema", None
    if reporte_particion and reporte_particion.get("activa"):
        f_nom = reporte_particion["nombre_foco"]
        r_nom = reporte_particion["nombre_resto"]
        if banco == f_nom:
            return "foco", "resto"
        if banco == r_nom:
            return "resto", "foco"
        if set(reporte_particion.get("bancos_foco", [])) == {banco}:
            return f"banco_{banco}", "resto"
        if set(reporte_particion.get("bancos_resto", [])) == {banco}:
            return f"banco_{banco}", "foco"
    return f"banco_{banco}", None


CCOVN_COMUNES = ("ccovn_sistema_lag1", "var_ccovn_sistema_lag1",
                 "ccovn_vs_dia_mes_lag1", "residuo_ccovn_lag1")


def armar_sub_ccovn(ccovn_features, clave_propio, clave_contraparte):
    """
    Sub-DataFrame de CCOVN con los roles relativos ya resueltos.

    Se asigna columna por columna en vez de hacer un select + rename porque
    cuando la entidad ES el sistema, clave_propio vale "sistema" y la fuente del
    rol "propio" coincide con una de las comunes. Con select + rename esa
    columna entraba dos veces y el rename convertia AMBAS copias, dejando
    ccovn_propio_lag1 duplicada (df["ccovn_propio_lag1"] devolvia un DataFrame y
    la division posterior reventaba) y ccovn_sistema_lag1 desaparecida, para ser
    rellenada con NaN mas abajo sin aviso. Copiando por separado, una misma
    fuente alimenta los dos roles sin colisionar.
    """
    renombre = {
        f"ccovn_{clave_propio}_lag1":     "ccovn_propio_lag1",
        f"var_ccovn_{clave_propio}_lag1": "var_ccovn_propio_lag1",
    }
    if clave_contraparte is not None:
        renombre[f"ccovn_{clave_contraparte}_lag1"]     = "ccovn_contraparte_lag1"
        renombre[f"var_ccovn_{clave_contraparte}_lag1"] = "var_ccovn_contraparte_lag1"

    sub = pd.DataFrame(index=ccovn_features.index)
    for origen, destino in renombre.items():
        if origen in ccovn_features.columns:
            sub[destino] = ccovn_features[origen]
    for c in CCOVN_COMUNES:
        if c in ccovn_features.columns:
            sub[c] = ccovn_features[c]

    dups = sub.columns[sub.columns.duplicated()].tolist()
    if dups:
        raise AssertionError(f"armar_sub_ccovn produjo columnas duplicadas: {dups}")
    return sub


def agrupar_bancos(df_bancarios, umbral_pct, bancos_otros, nombre_otros,
                   excluir=None):
    """
    Agrupa bancos pequeños en una categoría única antes de modelar.

    Retorna (df_agrupado, lista_bancos, reporte).
    """
    logger.info("PARTE 3: Agrupando bancos pequeños...")

    if df_bancarios.empty:
        logger.warning("  DataFrame bancarios vacío. No hay bancos que agrupar.")
        return df_bancarios, [], {}

    # Detectar bancos disponibles (columnas tipo banco_X_R y banco_X_D)
    cols_R = [c for c in df_bancarios.columns if c.endswith("_R")]
    cols_D = [c for c in df_bancarios.columns if c.endswith("_D")]
    bancos_disponibles = [c[:-2] for c in cols_R if c[:-2] + "_D" in cols_D]

    # v2: los agregados de la partición viajan en el mismo DataFrame pero no son
    # bancos. Si entraran acá competirían con los bancos reales por el umbral de
    # volumen —FOCO_BBVA solo tiene ~la mitad del volumen, RESTO_BBVA la otra— y
    # además podrían terminar sumados dentro de Otros_bancos.
    if excluir:
        _fuera = [b for b in bancos_disponibles if b in excluir]
        bancos_disponibles = [b for b in bancos_disponibles if b not in excluir]
        if _fuera:
            logger.info(f"  Excluidos del agrupamiento por ser agregados: {_fuera}")

    if not bancos_disponibles:
        logger.warning("  No se detectaron columnas con formato {banco}_R / {banco}_D.")
        return df_bancarios, [], {}

    # Determinar bancos pequeños
    if bancos_otros:
        # Opción B: lista fija
        pequeños = [b for b in bancos_otros if b in bancos_disponibles]
        logger.info(f"  Usando lista fija de bancos pequeños: {pequeños}")
    else:
        # Opción A: umbral automático por volumen
        volumen = {}
        for b in bancos_disponibles:
            r = df_bancarios.get(f"{b}_R", pd.Series(dtype=float))
            d = df_bancarios.get(f"{b}_D", pd.Series(dtype=float))
            volumen[b] = (r.abs() + d.abs()).mean()

        total_vol = sum(volumen.values())
        participacion = {b: v / total_vol for b, v in volumen.items()} if total_vol > 0 else {b: 0 for b in bancos_disponibles}
        pequeños = [b for b, p in participacion.items() if p < umbral_pct]
        logger.info(f"  Bancos detectados automáticamente como pequeños (< {umbral_pct:.0%}): {pequeños}")

    grandes = [b for b in bancos_disponibles if b not in pequeños]

    df_agrupado = df_bancarios.copy()

    if pequeños:
        # Sumar R y D de bancos pequeños
        cols_R_pequeños = [f"{b}_R" for b in pequeños if f"{b}_R" in df_agrupado.columns]
        cols_D_pequeños = [f"{b}_D" for b in pequeños if f"{b}_D" in df_agrupado.columns]

        df_agrupado[f"{nombre_otros}_R"] = df_agrupado[cols_R_pequeños].sum(axis=1)
        df_agrupado[f"{nombre_otros}_D"] = df_agrupado[cols_D_pequeños].sum(axis=1)

        cols_a_eliminar = cols_R_pequeños + cols_D_pequeños
        df_agrupado = df_agrupado.drop(columns=cols_a_eliminar)

        lista_bancos = grandes + [nombre_otros]
    else:
        lista_bancos = grandes

    # Reporte
    reporte = {
        "bancos_individuales": grandes,
        "bancos_agrupados": pequeños,
        "nombre_grupo": nombre_otros,
    }

    if not df_bancarios.empty and not bancos_otros:
        reporte["participaciones"] = participacion if "participacion" in dir() else {}

    # Imprimir tabla resumen
    logger.info(f"\n{'='*55}")
    logger.info(f"  RESUMEN DE AGRUPACIÓN DE BANCOS")
    logger.info(f"{'='*55}")
    logger.info(f"  Bancos individuales ({len(grandes)}): {grandes}")
    logger.info(f"  Bancos agrupados en '{nombre_otros}' ({len(pequeños)}): {pequeños}")
    logger.info(f"  Total entidades a modelar: {len(lista_bancos)}")
    logger.info(f"{'='*55}\n")

    return df_agrupado, lista_bancos, reporte


###############################################################################
# PARTE 4 — Descarga automática de series externas
###############################################################################
def _retry_download(func, max_intentos=3, espera=2):
    """Ejecuta func hasta max_intentos veces con espera exponencial."""
    for intento in range(1, max_intentos + 1):
        try:
            resultado = func()
            if resultado is not None:
                return resultado
        except Exception as e:
            logger.warning(f"  Intento {intento}/{max_intentos} fallido: {e}")
            if intento < max_intentos:
                time.sleep(espera * intento)
    return None


def _descargar_yahoo(ticker, inicio, fin, proxies, nombre):
    """
    Descarga una serie desde Yahoo Finance via requests directo con headers de navegador.
    Funciona detrás de proxies corporativos que interceptan SSL.
    """
    import urllib3
    urllib3.disable_warnings()

    ticker_encoded = ticker.replace("^", "%5E")
    inicio_ts = int(pd.Timestamp(inicio).timestamp())
    fin_ts    = int(pd.Timestamp(fin).timestamp())
    url = (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker_encoded}"
        f"?interval=1d&period1={inicio_ts}&period2={fin_ts}"
    )
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://finance.yahoo.com",
    }

    def _descarga():
        import requests as req
        r = req.get(
            url,
            headers=headers,
            proxies=proxies if proxies else None,
            verify=False,
            timeout=30,
        )
        if r.status_code != 200:
            raise ValueError(f"HTTP {r.status_code} para {ticker}")
        data   = r.json()
        result = data["chart"]["result"]
        if not result:
            raise ValueError(f"Sin datos en respuesta para {ticker}")
        timestamps = result[0]["timestamp"]
        closes     = result[0]["indicators"]["quote"][0]["close"]
        s = pd.Series(
            closes,
            index=pd.to_datetime(timestamps, unit="s").tz_localize(None),
            name=nombre,
            dtype=float,
        )
        s.index = s.index.normalize()
        if s.dropna().empty:
            raise ValueError(f"Serie vacía para {ticker}")
        return s

    resultado = _retry_download(_descarga)
    if resultado is None:
        logger.warning(f"  No se pudo descargar {nombre} ({ticker}) desde Yahoo Finance.")
        return pd.Series(dtype=float, name=nombre)
    logger.info(f"  {nombre} descargado: {len(resultado)} observaciones.")
    return resultado


def _descargar_fred(serie_id, inicio, fin, api_key, proxies, nombre):
    """Descarga una serie desde la API REST de FRED (sin fredapi, solo requests)."""
    import urllib3
    urllib3.disable_warnings()

    url = (
        f"https://api.stlouisfed.org/fred/series/observations"
        f"?series_id={serie_id}"
        f"&api_key={api_key}"
        f"&file_type=json"
        f"&observation_start={pd.Timestamp(inicio).strftime('%Y-%m-%d')}"
        f"&observation_end={pd.Timestamp(fin).strftime('%Y-%m-%d')}"
    )

    def _descarga():
        import requests
        r = requests.get(
            url,
            proxies=proxies if proxies else None,
            verify=False,
            timeout=30,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        if r.status_code != 200:
            raise ValueError(f"FRED HTTP {r.status_code}: {r.text[:200]}")
        obs = r.json().get("observations", [])
        if not obs:
            raise ValueError(f"Sin observaciones para {serie_id}")
        registros = []
        for o in obs:
            try:
                val = float(o["value"])
                registros.append({"fecha": pd.Timestamp(o["date"]), nombre: val})
            except (ValueError, TypeError):
                pass  # "." = dato faltante en FRED
        s = pd.DataFrame(registros).set_index("fecha")[nombre]
        return s

    resultado = _retry_download(_descarga)
    if resultado is None:
        logger.warning(f"  No se pudo descargar {nombre} ({serie_id}) desde FRED.")
        return pd.Series(dtype=float, name=nombre)
    logger.info(f"  {nombre} descargado: {len(resultado)} observaciones.")
    return resultado


def _leer_bcrp_excel(ruta, nombre, hoja=0):
    """
    Lee una serie BCRP desde una hoja de Excel descargado con el Add-In BCRPData.

    Formato esperado del Add-In:
        Fila 1: código de serie
        Fila 2: Descripción
        Fila 3: Frecuencia
        Fila 4: Periodo
        Fila 5: vacía
        Fila 6: encabezados  Date | Valores
        Fila 7+: datos       fecha | valor

    hoja: nombre o índice de la hoja dentro del Excel (default=0, primera hoja).
    Si el archivo no existe devuelve Series vacía con advertencia.
    """
    if not ruta or not os.path.exists(ruta):
        logger.warning(
            f"  Archivo BCRP '{nombre}' no encontrado en: {ruta}\n"
            f"  → Descargarlo con Add-In BCRPData y guardarlo en esa ruta."
        )
        return pd.Series(dtype=float, name=nombre)

    try:
        # header=None para leer desde fila 1 sin asumir encabezados
        raw = pd.read_excel(ruta, sheet_name=hoja, header=None)

        # Buscar la fila de encabezado "Date" / "Valores"
        header_row = None
        for i, row in raw.iterrows():
            vals = [str(v).strip().lower() for v in row if pd.notna(v)]
            if any(v in ("date", "fecha") for v in vals):
                header_row = i
                break

        if header_row is None:
            header_row = 5  # fallback: fila 6

        df = pd.read_excel(ruta, sheet_name=hoja, header=header_row)
        df.columns = [str(c).strip() for c in df.columns]

        # Identificar columnas de fecha y valor
        col_fecha = next((c for c in df.columns if c.lower() in ("date", "fecha")), df.columns[0])
        col_valor = next((c for c in df.columns if c.lower() in ("valores", "value", "values")), df.columns[1])

        df = df[[col_fecha, col_valor]].copy()
        df.columns = ["fecha", nombre]

        # El Add-In BCRPData usa meses en español: Set=Sep, Ene=Jan, etc.
        _meses_es = {
            "Ene": "Jan", "Feb": "Feb", "Mar": "Mar", "Abr": "Apr",
            "May": "May", "Jun": "Jun", "Jul": "Jul", "Ago": "Aug",
            "Set": "Sep", "Oct": "Oct", "Nov": "Nov", "Dic": "Dec",
        }
        def _parse_fecha_bcrp(s):
            s = str(s).strip()
            for es, en in _meses_es.items():
                s = s.replace(es, en)
            return pd.to_datetime(s, errors="coerce", dayfirst=True)

        df["fecha"] = df["fecha"].apply(_parse_fecha_bcrp)
        df[nombre] = pd.to_numeric(
            df[nombre].astype(str).str.replace(",", "."), errors="coerce"
        )
        df = df.dropna(subset=["fecha"]).set_index("fecha").sort_index()

        # Advertir si los datos tienen más de 7 días de antigüedad
        ultima = df.index.max()
        dias_atras = (pd.Timestamp.today() - ultima).days
        if dias_atras > 7:
            logger.warning(
                f"  {nombre}: última fecha {ultima.date()} ({dias_atras} días atrás). "
                f"Considera actualizar el archivo."
            )

        logger.info(f"  {nombre} leído desde archivo: {len(df)} obs, hasta {ultima.date()}")
        return df[nombre]

    except Exception as e:
        logger.warning(f"  Error leyendo {nombre} desde {ruta}: {e}")
        return pd.Series(dtype=float, name=nombre)


_BBG_HOJAS     = {"CDS_PERU_5Y": "CDS", "COPPER": "Cobre"}
_BBG_DATE_FMTS = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y%m%d"]


def _leer_bloomberg_excel(ruta, nombre):
    """
    Lee directamente desde DataBBG.xlsx (hojas BVL / CDS / Cobre).
    Maneja el formato Bloomberg estándar: filas de metadatos + datos.
    """
    if not ruta or not os.path.exists(ruta):
        logger.warning(f"  Bloomberg '{nombre}' no encontrado en: {ruta}")
        return pd.Series(dtype=float, name=nombre)
    hoja = _BBG_HOJAS.get(nombre, nombre)
    try:
        raw = pd.read_excel(ruta, sheet_name=hoja, header=None, dtype=str)
        etiquetas = {"security", "start date", "end date", "period", "currency",
                     "pricing source", "date", "dates", "nan", ""}
        fila_inicio = None
        for i, row in raw.iterrows():
            val = str(row.iloc[0]).strip().lower()
            if val in etiquetas:
                continue
            for fmt in _BBG_DATE_FMTS:
                try:
                    ts = pd.to_datetime(val.split(" ")[0], format=fmt)
                    if pd.Timestamp("1990-01-01") <= ts <= pd.Timestamp("2100-01-01"):
                        fila_inicio = i
                        break
                except Exception:
                    continue
            if fila_inicio is not None:
                break
        if fila_inicio is None:
            logger.warning(f"  {nombre}: no se encontraron filas de datos en hoja '{hoja}'")
            return pd.Series(dtype=float, name=nombre)

        datos = raw.iloc[fila_inicio:, :2].copy()
        datos.columns = ["fecha", nombre]
        datos = datos.dropna(subset=["fecha"])
        datos = datos[datos["fecha"].str.strip() != ""]
        for fmt in _BBG_DATE_FMTS:
            try:
                datos["fecha"] = pd.to_datetime(datos["fecha"], format=fmt, errors="raise")
                break
            except Exception:
                continue
        else:
            datos["fecha"] = pd.to_datetime(datos["fecha"], dayfirst=True, errors="coerce")
        datos[nombre] = (datos[nombre].astype(str)
                         .str.replace(",", ".", regex=False)
                         .pipe(pd.to_numeric, errors="coerce"))
        datos = datos.dropna(subset=["fecha", nombre]).set_index("fecha").sort_index()
        datos = datos[~datos.index.duplicated(keep="last")]
        s = datos[nombre]
        logger.info(f"  {nombre} [{hoja}]: {len(s):,} obs, "
                    f"{s.index.min().date()} → {s.index.max().date()}")
        return s
    except Exception as e:
        logger.warning(f"  Error leyendo Bloomberg '{nombre}' (hoja '{hoja}'): {e}")
        return pd.Series(dtype=float, name=nombre)


def download_external_series(params):
    """
    Descarga todas las series externas y las retorna en un DataFrame
    alineado al calendario hábil peruano.

    Forward fill máximo 3 días para series de menor frecuencia.
    """
    logger.info("PARTE 4: Descargando series externas...")

    inicio = params["fecha_inicio_historico"]
    fin = params["fecha_fin_historico"]
    proxy = params.get("proxy", "")
    proxies = {"http": proxy, "https": proxy} if proxy and "USUARIO" not in proxy else None

    series = {}

    # 4a. Yahoo Finance
    series["VIX"]  = _descargar_yahoo(params["ticker_vix"],  inicio, fin, proxies, "VIX")
    series["T10Y"] = _descargar_yahoo(params["ticker_t10y"], inicio, fin, proxies, "T10Y")

    # 4b. FRED
    api_key = params.get("fred_api_key", "")
    if api_key and api_key != "TU_API_KEY":
        series["FED_FUNDS"] = _descargar_fred(
            params["fred_fedfunds"], inicio, fin, api_key, proxies, "FED_FUNDS"
        )
    else:
        logger.warning("  FRED API key no configurada. FED_FUNDS quedará como NaN.")
        series["FED_FUNDS"] = pd.Series(dtype=float, name="FED_FUNDS")

    # 4c. Series BCRP — leídas desde Excel descargado manualmente con Add-In BCRPData
    ruta_bcrp = params["ruta_bcrp"]
    series["EMBI_PERU"]     = _leer_bcrp_excel(ruta_bcrp, "EMBI_PERU",     hoja=params["hoja_bcrp_embi"])
    series["TASA_REF_BCRP"] = _leer_bcrp_excel(ruta_bcrp, "TASA_REF_BCRP", hoja=params["hoja_bcrp_tasa_ref"])

    # TC_PEN_USD = promedio(TC Compra, TC Venta) del BCRP — reemplaza Yahoo Finance (tenía outliers)
    tc_compra = _leer_bcrp_excel(ruta_bcrp, "TC_COMPRA", hoja=params["hoja_bcrp_tc_compra"])
    tc_venta  = _leer_bcrp_excel(ruta_bcrp, "TC_VENTA",  hoja=params["hoja_bcrp_tc_venta"])
    if not tc_compra.empty and not tc_venta.empty:
        tc_df = pd.concat([tc_compra.rename("compra"), tc_venta.rename("venta")], axis=1)
        series["TC_PEN_USD"] = tc_df.mean(axis=1).rename("TC_PEN_USD")
        logger.info("  TC_PEN_USD: promedio BCRP compra/venta cargado.")
    elif not tc_venta.empty:
        series["TC_PEN_USD"] = tc_venta.rename("TC_PEN_USD")
        logger.warning("  TC_PEN_USD: solo TC Venta disponible (sin TC Compra).")
    elif not tc_compra.empty:
        series["TC_PEN_USD"] = tc_compra.rename("TC_PEN_USD")
        logger.warning("  TC_PEN_USD: solo TC Compra disponible (sin TC Venta).")
    else:
        logger.warning("  TC_PEN_USD: no se encontraron hojas TC Compra/Venta en series_bcrp.xlsx.")
        series["TC_PEN_USD"] = pd.Series(dtype=float, name="TC_PEN_USD")

    # 4d. Bloomberg — CDS Perú 5Y y Cobre LME (DataBBG.xlsx, hojas CDS / Cobre)
    ruta_bbg = params.get("ruta_bloomberg", "")
    for _nombre_bbg in ("CDS_PERU_5Y", "COPPER"):
        series[_nombre_bbg] = _leer_bloomberg_excel(ruta_bbg, _nombre_bbg)

    # Alinear al índice de fechas del rango
    idx = pd.bdate_range(start=inicio, end=fin)
    df = pd.DataFrame(index=idx)
    for nombre, s in series.items():
        if not s.empty:
            s_alineada = s.reindex(idx)
            s_alineada = s_alineada.ffill()   # sin límite: series mensuales/trimestrales se propagan correctamente
            df[nombre] = s_alineada
        else:
            df[nombre] = np.nan

    cols_esperadas = ["VIX", "TC_PEN_USD", "T10Y", "FED_FUNDS", "EMBI_PERU", "TASA_REF_BCRP",
                      "CDS_PERU_5Y", "COPPER"]
    for c in cols_esperadas:
        if c not in df.columns:
            df[c] = np.nan

    logger.info(f"  DataFrame externo construido: {df.shape}")
    return df


###############################################################################
# PARTE 5 — Construcción de features
###############################################################################

# 5a. Features bancarias
def _garch_vol(flujo: pd.Series) -> pd.Series:
    """
    GARCH(1,1) implementado en NumPy + scipy.optimize — sin dependencia del
    paquete 'arch' ni de statsmodels, evitando incompatibilidades con scipy 1.9+.

    Modelo:
        σ²_t = ω + α·ε²_{t-1} + β·σ²_{t-1}

    donde ε_t = x_t (flujo neto centrado). Los parámetros ω, α, β se estiman
    por Máxima Verosimilitud Gaussiana usando L-BFGS-B.

    σ_t usa solo información hasta t-1 → sin look-ahead bias.
    Fallback a rolling std 20d si la serie es constante o muy corta.
    """
    if flujo.dropna().std() < 1e-9 or len(flujo.dropna()) < 60:
        return flujo.rolling(20).std()

    from scipy.optimize import minimize

    # Escalar a varianza unitaria para estabilidad numérica del optimizador
    escala  = flujo.std()
    x       = (flujo / escala).fillna(0.0).values.astype(float)
    n       = len(x)
    var_unc = float(np.var(x))   # varianza incondicional → σ²_0 inicial

    def _sigma2(params):
        """Calcula la varianza condicional para un vector de parámetros."""
        omega, alpha, beta = params
        s2 = np.empty(n)
        s2[0] = var_unc
        for t in range(1, n):
            s2[t] = omega + alpha * x[t-1]**2 + beta * s2[t-1]
        return s2

    def _neg_loglik(params):
        """Log-verosimilitud gaussiana negativa (a minimizar)."""
        omega, alpha, beta = params
        # Restricciones de estacionariedad: ω>0, α>0, β>0, α+β<1
        if omega <= 0 or alpha <= 0 or beta <= 0 or alpha + beta >= 0.9999:
            return 1e10
        s2 = _sigma2(params)
        if np.any(s2 <= 0):
            return 1e10
        return 0.5 * float(np.sum(np.log(s2) + x**2 / s2))

    # Punto de partida: parámetros típicos de series financieras diarias
    p0     = [0.01, 0.08, 0.88]
    bounds = [(1e-7, 0.5), (1e-7, 0.5), (1e-7, 0.9999)]

    try:
        res = minimize(_neg_loglik, p0, method="L-BFGS-B", bounds=bounds,
                       options={"maxiter": 500, "ftol": 1e-10, "gtol": 1e-7})
        if not res.success and res.fun > 1e9:
            raise ValueError(res.message)

        sigma2     = _sigma2(res.x)
        sigma_cond = pd.Series(np.sqrt(np.maximum(sigma2, 0)) * escala,
                               index=flujo.index)
        logger.info(
            f"    GARCH(1,1) ajustado: ω={res.x[0]:.5f}  "
            f"α={res.x[1]:.4f}  β={res.x[2]:.4f}  "
            f"persistencia={res.x[1]+res.x[2]:.4f}"
        )
        return sigma_cond

    except Exception as e:
        logger.warning(f"  GARCH no convergió ({e}) — usando rolling std 20d")
        return flujo.rolling(20).std()


###############################################################################
# FFD — Diferenciación Fraccional de Ancho Fijo (López de Prado, Cap. 5)
###############################################################################

def _ffd_weights(d: float, thresh: float = 1e-5) -> np.ndarray:
    """Calcula los pesos de FFD por expansión binomial: w_k = Π_{j=1}^{k} (d-j+1)/j * (-1)^k."""
    w, k = [1.0], 1
    while True:
        w_k = -w[-1] * (d - k + 1) / k
        if abs(w_k) < thresh:
            break
        w.append(w_k)
        k += 1
    return np.array(w[::-1])   # más reciente al final → conveniente para dot()


def _fracdiff_fixed_width(series: pd.Series, d: float, thresh: float = 1e-5) -> pd.Series:
    """Aplica FFD con ventana de ancho fijo a una Serie de pandas."""
    w     = _ffd_weights(d, thresh)
    width = len(w)
    vals  = series.values.astype(float)
    n     = len(vals)
    out   = np.full(n, np.nan)
    for i in range(width - 1, n):
        chunk = vals[i - width + 1: i + 1]
        if not np.any(np.isnan(chunk)):
            out[i] = float(np.dot(w, chunk))
    return pd.Series(out, index=series.index, name=series.name)


def _find_min_d(
    series: pd.Series,
    thresh: float = 1e-5,
    max_d: float  = 1.0,
    target_pval: float = 0.05,
    n_steps: int  = 20,
) -> float:
    """
    Encuentra el d mínimo en (0, max_d] tal que la serie diferenciada
    sea estacionaria (ADF p-value ≤ target_pval).
    Requiere statsmodels; si no está disponible retorna 0.4 como fallback.
    """
    try:
        from statsmodels.tsa.stattools import adfuller
    except ImportError:
        logger.warning("  statsmodels no disponible → d_opt fallback = 0.4")
        return 0.4

    for d in np.linspace(0.05, max_d, n_steps):
        fd = _fracdiff_fixed_width(series.dropna(), round(float(d), 4), thresh)
        fd_clean = fd.dropna()
        if len(fd_clean) < 30:
            continue
        try:
            pval = adfuller(fd_clean, maxlag=1, regression="c", autolag=None)[1]
            if pval <= target_pval:
                return round(float(d), 4)
        except Exception:
            continue
    return round(max_d, 4)


def _calcular_hmm_expanding(flujo: pd.Series, sigma_22d: pd.Series,
                            primer_ventana_años: int = 6) -> pd.Series:
    """
    Pre-computa hmm_estado[t] para todo t con expanding window, sin leakage.

    Alineado con la estructura walk-forward (igual que GARCH):
      1. Primer modelo: entrena en [HMM_INICIO, HMM_INICIO + primer_ventana_años - 1]
         → etiqueta ESE mismo periodo (in-sample, como GARCH) + el siguiente año
      2. Años siguientes: expanding, entrena en [HMM_INICIO, Y-1] → etiqueta Y

    Con primer_ventana_años=6 e HMM_INICIO=2010:
      - Modelo 1: 2010-2015 → etiqueta 2010-2015 (in-sample) + 2016
      - Modelo 2: 2010-2016 → etiqueta 2017
      - Modelo 3: 2010-2017 → etiqueta 2018
      - ...

    Estados ordenados por det(Σ_i) ascendente: 0=calma, 1=moderado, 2=severo.
    Devuelve pd.Series de Int8 sin NaN (toda la historia desde HMM_INICIO etiquetada).
    """
    if not _HMM_DISPONIBLE:
        return pd.Series(pd.NA, index=flujo.index, dtype="Int8", name="hmm_estado")

    df_base = pd.concat([flujo.rename("flujo"), sigma_22d.rename("sigma_22d")],
                        axis=1).dropna()
    df_base = df_base[df_base.index >= HMM_INICIO]
    if df_base.empty:
        return pd.Series(pd.NA, index=flujo.index, dtype="Int8", name="hmm_estado")

    año_inicio    = df_base.index.year.min()
    año_fin_prim  = año_inicio + primer_ventana_años - 1   # último año del primer bloque
    años          = sorted(df_base.index.year.unique())
    resultado     = pd.Series(pd.NA, index=df_base.index, dtype="Int8", name="hmm_estado")

    def _fit_and_label(X_train, df_target):
        """Ajusta HMM en X_train y predice estados para df_target."""
        scaler = _StandardScaler()
        Xs     = scaler.fit_transform(X_train)
        modelo = _hmmlearn.GaussianHMM(
            n_components=HMM_N_ESTADOS, covariance_type="full",
            n_iter=1000, random_state=42,
        )
        modelo.fit(Xs)
        varianzas     = [np.linalg.det(modelo.covars_[s]) for s in range(HMM_N_ESTADOS)]
        sorted_states = np.argsort(varianzas)
        mapa          = {sorted_states[i]: i for i in range(HMM_N_ESTADOS)}
        Xs_target     = scaler.transform(df_target.values)
        estados_raw   = modelo.predict(Xs_target)
        return np.array([mapa[e] for e in estados_raw], dtype="int8")

    # ── Paso 1: primer bloque (in-sample, alineado con primer fold de train) ──
    df_prim = df_base[df_base.index.year <= año_fin_prim]
    if len(df_prim) >= 50:
        try:
            estados_prim = _fit_and_label(df_prim.values, df_prim)
            resultado.loc[df_prim.index] = pd.array(estados_prim, dtype="Int8")
        except Exception:
            pass

    # ── Paso 2: expanding año a año desde año_fin_prim + 1 ───────────────────
    for año in años:
        if año <= año_fin_prim:
            continue
        X_train = df_base[df_base.index.year < año].values
        df_año  = df_base[df_base.index.year == año]
        if df_año.empty or len(X_train) < 50:
            continue
        try:
            estados_año = _fit_and_label(X_train, df_año)
            resultado.loc[df_año.index] = pd.array(estados_año, dtype="Int8")
        except Exception:
            continue

    return resultado


def build_bank_features(df_banco, lags_cortos, lag_semana, lag_mes, ventanas_vol):
    """
    Recibe serie temporal de un banco con columnas R y D.
    Genera features de rezagos, volatilidades, medias móviles y cambios diarios.

    Incluye features de régimen de volatilidad del flujo neto (D−R):
      sigma_flujo_5d / 20d : std rolling del flujo neto realizado.
      ma_flujo_5d / 20d    : media rolling del flujo neto (nivel reciente).
      sigma_22d            : desviación estándar rolling 22d del flujo neto.
                             Más reactiva que GARCH al salir de periodos de estrés.
    """
    VENTANAS_FLUJO = [5, 20]

    if df_banco.empty or "R" not in df_banco.columns or "D" not in df_banco.columns:
        cols = (
            ["R_t0", "D_t0"]
            + [f"R_t-{l}" for l in lags_cortos + [lag_semana, lag_mes]]
            + [f"D_t-{l}" for l in lags_cortos + [lag_semana, lag_mes]]
            + [f"sigma_R_{v}d" for v in ventanas_vol]
            + [f"sigma_D_{v}d" for v in ventanas_vol]
            + [f"ma_R_{v}d" for v in ventanas_vol]
            + [f"ma_D_{v}d" for v in ventanas_vol]
            + ["delta_R", "delta_D"]
            + ["R_conf_t1", "R_conf_t2", "D_conf_t1"]
            + [f"sigma_flujo_{v}d" for v in VENTANAS_FLUJO]
            + [f"ma_flujo_{v}d"    for v in VENTANAS_FLUJO]
            + ["sigma_22d", "flujo_neto_acum_mes"]
            + ["flujo_neto_sum_5d", "flujo_neto_sum_22d", "flujo_neto_sum_66d"]
        )
        return pd.DataFrame(columns=cols)

    df = df_banco[["R", "D"]].copy()
    resultado = pd.DataFrame(index=df.index)

    # Valores de hoy: disponibles en t por aviso anticipado
    resultado["R_t0"] = df["R"]
    resultado["D_t0"] = df["D"]

    todos_lags = lags_cortos + [lag_semana, lag_mes]
    for l in todos_lags:
        resultado[f"R_t-{l}"] = df["R"].shift(l)
        resultado[f"D_t-{l}"] = df["D"].shift(l)

    for v in ventanas_vol:
        resultado[f"sigma_R_{v}d"] = df["R"].rolling(v).std()
        resultado[f"sigma_D_{v}d"] = df["D"].rolling(v).std()
        resultado[f"ma_R_{v}d"]    = df["R"].rolling(v).mean()
        resultado[f"ma_D_{v}d"]    = df["D"].rolling(v).mean()

    resultado["delta_R"] = df["R"].diff(1)
    resultado["delta_D"] = df["D"].diff(1)

    resultado["R_conf_t1"] = df["R"].shift(-1)
    resultado["R_conf_t2"] = df["R"].shift(-2)
    resultado["D_conf_t1"] = df["D"].shift(-1)

    # Features de régimen: rolling std/media + GARCH(1,1)
    flujo = df["D"] - df["R"]
    for v in VENTANAS_FLUJO:
        resultado[f"sigma_flujo_{v}d"] = flujo.rolling(v).std()
        resultado[f"ma_flujo_{v}d"]    = flujo.rolling(v).mean()

    resultado["sigma_22d"] = flujo.rolling(22, min_periods=5).std()

    # Ratio de volatilidad corta/larga del flujo neto: > 1 → régimen de alta vol reciente
    resultado["sigma_flujo_ratio"] = (
        resultado["sigma_flujo_5d"] / resultado["sigma_flujo_20d"].replace(0, np.nan)
    )

    # Acumulado del flujo neto (D−R) desde el primer día hábil del mes hasta t.
    # Lógica: una acumulación de depósitos netos a lo largo del mes anticipa
    # mayores retiros en los últimos días, y viceversa.
    resultado["flujo_neto_acum_mes"] = (
        flujo.groupby(flujo.index.to_period("M")).cumsum()
    )

    # Agregados rolling del flujo neto — ventana siempre relativa a t
    # min_periods=N garantiza NaN para filas con historia insuficiente (→ imputadas por mediana del fold)
    resultado["flujo_neto_sum_5d"]  = flujo.rolling( 5, min_periods= 5).sum()
    resultado["flujo_neto_sum_22d"] = flujo.rolling(22, min_periods=22).sum()
    resultado["flujo_neto_sum_66d"] = flujo.rolling(66, min_periods=66).sum()

    return resultado


# 5b. Features macroeconómicas
def build_macro_features(macro_df):
    """
    Genera features macroeconómicas a partir del DataFrame de series externas.
    """
    if macro_df.empty:
        cols = [
            "VIX", "delta_VIX", "VIX_ma22",
            "TC_PEN_USD", "delta_TC", "tc_vol_5d", "tc_vol_22d", "garch_vol_tc",
            "EMBI_PERU", "delta_EMBI", "garch_vol_embi",
            "TASA_REF_BCRP", "FED_FUNDS", "diferencial_tasas", "T10Y",
            "EMBI_PERU_frac", "T10Y_frac", "CDS_PERU_5Y_frac", "COPPER_frac", "VIX_frac",
        ]
        return pd.DataFrame(columns=cols)

    df = macro_df.copy()
    resultado = pd.DataFrame(index=df.index)

    resultado["VIX"] = df.get("VIX", np.nan)
    resultado["delta_VIX"] = df.get("VIX", pd.Series(dtype=float)).diff(1)
    resultado["VIX_ma22"] = df.get("VIX", pd.Series(dtype=float)).rolling(22).mean()

    resultado["TC_PEN_USD"] = df.get("TC_PEN_USD", np.nan)
    tc = df.get("TC_PEN_USD", pd.Series(dtype=float))
    resultado["delta_TC"] = tc.diff(1)
    retornos_tc = tc.pct_change()
    resultado["tc_vol_5d"] = retornos_tc.rolling(5).std()
    resultado["tc_vol_22d"] = retornos_tc.rolling(22).std()

    # Ratio de volatilidad cambiaria corta/larga: > 1 → estrés cambiario reciente
    resultado["tc_vol_ratio"] = (
        resultado["tc_vol_5d"] / resultado["tc_vol_22d"].replace(0, np.nan)
    )

    retornos_log_tc = np.log(tc / tc.shift(1))
    resultado["garch_vol_tc"] = _garch_vol(retornos_log_tc)

    resultado["EMBI_PERU"] = df.get("EMBI_PERU", np.nan)
    delta_embi = df.get("EMBI_PERU", pd.Series(dtype=float)).diff(1)
    resultado["delta_EMBI"] = delta_embi
    resultado["garch_vol_embi"] = _garch_vol(delta_embi)

    resultado["TASA_REF_BCRP"] = df.get("TASA_REF_BCRP", np.nan)
    resultado["FED_FUNDS"] = df.get("FED_FUNDS", np.nan)

    tasa_ref = df.get("TASA_REF_BCRP", pd.Series(dtype=float))
    fed_funds = df.get("FED_FUNDS", pd.Series(dtype=float))
    if not tasa_ref.empty and not fed_funds.empty:
        resultado["diferencial_tasas"] = tasa_ref.values - fed_funds.reindex(tasa_ref.index).values
    else:
        resultado["diferencial_tasas"] = np.nan

    resultado["T10Y"] = df.get("T10Y", np.nan)

    # ── FFD: diferenciación fraccional para series I(1) ────────────────────────
    # d_opt calibrado en datos hasta _FD_TRAIN_CUTOFF (sin leakage).
    # Si statsmodels no está disponible usa d_opt=0.4 como fallback conservador.
    _fd_cutoff = pd.Timestamp(_FD_TRAIN_CUTOFF)
    _ffd_targets = [
        ("EMBI_PERU",   "EMBI_PERU_frac"),
        ("T10Y",        "T10Y_frac"),
        ("CDS_PERU_5Y", "CDS_PERU_5Y_frac"),
        ("COPPER",      "COPPER_frac"),
        ("VIX",         "VIX_frac"),
    ]
    for col_raw, col_frac in _ffd_targets:
        serie = df.get(col_raw, pd.Series(dtype=float))
        if isinstance(serie, pd.Series) and serie.dropna().shape[0] >= 60:
            s_train = serie[serie.index <= _fd_cutoff].dropna()
            d_opt   = _find_min_d(s_train) if len(s_train) >= 60 else 0.4
            resultado[col_frac] = _fracdiff_fixed_width(serie, d_opt)
            logger.info(f"  FFD {col_raw} → {col_frac}  d_opt={d_opt:.4f}")
        else:
            resultado[col_frac] = np.nan

    return resultado


# 5c. Features de encaje (BBVA-específicas, con rezago 1 día)
def build_encaje_features(df_encaje, peru_bday):
    """
    Deriva features de encaje a partir de EncajeD.xlsx.
    Todas las variables usan información de t-1 (sin leakage).

    Features generadas:
      encaje_lag1         : Caja + Cta Cte BCR del día t-1 (M USD)
      exceso_lag1         : encaje_lag1 - encaje_exigible_lag1
      faltante_lag1       : exigible_total_mes - encaje_acum(t-1)
                            = lo que aún necesita acumular para cerrar el mes
      techo_10h           : CC + ON al día hábil que queda a 10 días hábiles
                            del cierre del mes. Disponible solo cuando t está
                            dentro de los últimos 10 días hábiles del mes.
      techo_restante_lag1 : techo_10h - retiro_acum_mes(t-1)
                            = presupuesto de retiro aún disponible
      proporcion_usada    : retiro_acum_mes(t-1) / techo_10h (0 si techo=0)
      encaje_urgencia_lag1: faltante(t-1) / días_hábiles_restantes_en_mes(t-1)
                            = tasa diaria de acumulación requerida para cerrar el mes.
                            Colapsa la interacción faltante × tiempo en una sola variable.

    Retorna DataFrame indexado por fecha, alineado con los datos de encaje.
    """
    if df_encaje.empty:
        return pd.DataFrame()

    df = df_encaje.copy()
    resultado = pd.DataFrame(index=df.index)

    # Encaje disponible (solo Caja + Cta Cte BCR; Overnight NO cuenta)
    if "caja" in df.columns and "cta_cte_bcr" in df.columns:
        encaje = df["caja"] + df["cta_cte_bcr"]
    else:
        logger.warning("  build_encaje_features: faltan columnas caja / cta_cte_bcr")
        return pd.DataFrame()

    overnight = df.get("overnight_bcr", pd.Series(0.0, index=df.index))
    exigible  = df.get("encaje_exigible", pd.Series(np.nan, index=df.index))
    retiro    = df.get("retiro_neto", pd.Series(np.nan, index=df.index))

    # ── 1. Encaje y exceso con rezago 1 día ─────────────────────────────────
    resultado["encaje_lag1"]   = encaje.shift(1)
    resultado["exceso_lag1"]   = (encaje - exigible).shift(1)

    # ── 2. Faltante = lo que el banco aún necesita acumular este mes ─────────
    # exigible_total_mes = encaje_exigible × días_calendario_en_mes
    # encaje_acum(t-1)   = suma de encaje diario desde inicio del mes hasta t-1
    #
    # Sólo días donde encaje_exigible no es NaN participan en el cálculo.
    exigible_mensual = exigible.groupby(exigible.index.to_period("M")).transform("first")
    dias_mes = pd.Series(
        df.index.to_series().dt.daysinmonth.values, index=df.index
    )
    exigible_total = exigible_mensual * dias_mes

    encaje_acum = encaje.groupby(encaje.index.to_period("M")).cumsum()
    faltante_raw = exigible_total - encaje_acum          # positivo = necesita más
    resultado["faltante_lag1"] = faltante_raw.shift(1)

    # ── 3. Techo (CC + ON a 10 días hábiles del cierre del mes) ─────────────
    # Para cada fecha, identificar el día que queda a 10 días hábiles del
    # cierre del mes en su respectivo mes. El techo se define solo una vez
    # por mes (en ese día puntual) y se propaga hacia adelante dentro del mes.
    cc_on = encaje + overnight   # CC + ON es el instrumento observado antes de ON

    techo_por_fecha = pd.Series(np.nan, index=df.index)
    for periodo, grupo in cc_on.groupby(cc_on.index.to_period("M")):
        inicio = periodo.start_time
        if inicio.month == 12:
            fin = pd.Timestamp(year=inicio.year + 1, month=1, day=1) - pd.Timedelta(days=1)
        else:
            fin = pd.Timestamp(year=inicio.year, month=inicio.month + 1, day=1) - pd.Timedelta(days=1)

        bdays_mes = pd.bdate_range(start=inicio, end=fin, freq=peru_bday)
        n_bd = len(bdays_mes)
        if n_bd < 10:
            continue

        fecha_techo = bdays_mes[n_bd - 10]   # día hábil que está a 10 días del cierre
        if fecha_techo in cc_on.index:
            val_techo = cc_on.loc[fecha_techo]
            fechas_grupo = grupo.index
            fechas_post = fechas_grupo[fechas_grupo >= fecha_techo]
            techo_por_fecha.loc[fechas_post] = val_techo

    resultado["techo_10h"] = techo_por_fecha.shift(1)   # rezago 1 día (conocido en t)

    # ── 4. Retiro acumulado del mes hasta t-1 ────────────────────────────────
    retiro_acum_mes = retiro.groupby(retiro.index.to_period("M")).cumsum().shift(1)

    # ── 5. Techo restante y proporción usada ─────────────────────────────────
    resultado["techo_restante_lag1"] = resultado["techo_10h"] - retiro_acum_mes.abs()
    resultado["proporcion_usada"] = (
        retiro_acum_mes.abs() / resultado["techo_10h"].replace(0, np.nan)
    ).clip(0, None)

    # ── 6. Urgencia de encaje: tasa de acumulación diaria requerida ──────────
    # encaje_urgencia = faltante(t-1) / días_hábiles_restantes_en_mes(t-1)
    # Normaliza el monto faltante por el tiempo disponible: lo que el tesorero
    # calcula como "cuánto debo depositar en BCR por día hábil para cerrar el mes".
    # El árbol no puede derivar esto eficientemente solo desde faltante_lag1 y
    # dias_al_cierre_mes (requiere split en dos dimensiones; este feature lo colapsa).
    # Solo tiene valor para el banco que tiene faltante real (banco_encaje = BBVA);
    # para SISTEMA queda en NaN junto con faltante_lag1.
    dias_habiles_restantes = pd.Series(np.nan, index=df.index)
    for periodo, grupo in encaje.groupby(encaje.index.to_period("M")):
        inicio = periodo.start_time
        if inicio.month == 12:
            fin = pd.Timestamp(year=inicio.year + 1, month=1, day=1) - pd.Timedelta(days=1)
        else:
            fin = pd.Timestamp(year=inicio.year, month=inicio.month + 1, day=1) - pd.Timedelta(days=1)
        bdays_mes = pd.bdate_range(start=inicio, end=fin, freq=peru_bday)
        for fecha in grupo.index:
            dias_restantes = (bdays_mes > fecha).sum()
            if fecha in dias_habiles_restantes.index:
                dias_habiles_restantes.loc[fecha] = max(dias_restantes, 1)

    resultado["encaje_urgencia_lag1"] = (
        faltante_raw.shift(1) / dias_habiles_restantes.shift(1).replace(0, np.nan)
    ).clip(lower=0)

    n_validos = resultado["faltante_lag1"].notna().sum()
    logger.info(
        f"  build_encaje_features: {n_validos:,} filas con faltante_lag1 válido | "
        f"techo_10h disponible: {resultado['techo_10h'].notna().sum():,} filas"
    )
    return resultado


# 5c-bis. Features de encaje BBVA desde Excel generado por aux_encaje_2.py
def load_bbva_encaje_features(params):
    """
    Lee bbva_encaje_features_modelo.xlsx (generado por aux_encaje_2.py).
    Retorna un DataFrame indexado en días CALENDARIO con las features lag1
    ya calculadas correctamente sobre ese índice — incluyendo
    ExigibleTotalMes_est_lag1 y EncajeAcumMes_lag1, los insumos que
    build_feature_matrix usa para proyectar capacidad_retiro_th a fecha_th.

    La adaptación al calendario hábil de step001 se hace con merge_asof
    en build_feature_matrix (busca la fecha calendario más reciente ≤ fecha_t).
    Así se evita el problema de construir un índice hábil pre-computado que
    no coincide exactamente con los fecha_t del dataset real.
    """
    _COLS = ["avance_mes_lag1", "exceso_abs_lag1", "exceso_dia_lag1",
             "encaje_ovn_lag1", "ratio_ovn_total_lag1",
             # encaje_diario_lag1 requiere que aux_encaje_2.py haya sido
             # re-corrido con la línea
             # df["encaje_diario_lag1"] = df["encaje"].shift(1) agregada al
             # bloque de versiones lag1 (mismo patrón que encaje_ovn_lag1).
             # Nombre distinto de "encaje_lag1" (ya usado por la fuente EncajeD
             # en la sección 8, línea ~1540) para no colisionar: son series
             # DISTINTAS —Caja+CtaCte allá, encaje diario de aux_encaje_2 acá—
             # y compartir nombre haría que una sobrescribiera a la otra en
             # silencio al mergear. Si el Excel es de una corrida anterior sin
             # esta columna, capacidad_retiro_th queda en NaN — el resto no se
             # ve afectado (ver cols_missing más abajo).
             "encaje_diario_lag1",
             # ExigibleTotalMes_est y EncajeAcumMes: necesarias para proyectar
             # capacidad_retiro_th a fecha_th. Se toman SIN lag y se rezagan acá
             # (renombrando a *_lag1), porque aux_encaje_2 no exporta su versión
             # lag1: el estado de t no se conoce hasta t+1, así que usarlas
             # crudas sería fuga.
             "ExigibleTotalMes_est", "EncajeAcumMes"]

    ruta = params.get("ruta_bbva_encaje_features", "")
    if not ruta or not Path(ruta).exists():
        logger.info("  bbva_encaje_features: archivo no encontrado — features de avance/exceso omitidas.")
        return pd.DataFrame()

    try:
        df_raw = pd.read_excel(ruta, sheet_name="Datos")
        # Seleccionar solo columnas disponibles — si el Excel no tiene alguna
        # (p.ej. versión anterior sin exceso_dia_lag1) se omite sin fallar.
        cols_ok = ["fecha"] + [c for c in _COLS if c in df_raw.columns]
        cols_missing = [c for c in _COLS if c not in df_raw.columns]
        if cols_missing:
            logger.warning(f"  bbva_encaje_features: columnas ausentes en Excel — {cols_missing}")
        df = df_raw[cols_ok].copy()
        df["fecha"] = pd.to_datetime(df["fecha"]).dt.normalize()
        df = df.sort_values("fecha").reset_index(drop=True)

        # ExigibleTotalMes_est y EncajeAcumMes llegan SIN lag (valor del día
        # mismo). Se rezagan acá, sobre el índice CALENDARIO del archivo —que
        # es como está definido el período de encaje—, y se renombran con el
        # sufijo _lag1 para que build_feature_matrix las trate igual que al
        # resto de columnas de esta función: todo lo que sale de acá ya es
        # seguro de usar en fecha_t sin rezagarlo de nuevo.
        if {"ExigibleTotalMes_est", "EncajeAcumMes"}.issubset(df.columns):
            df["ExigibleTotalMes_est_lag1"] = df["ExigibleTotalMes_est"].shift(1)
            df["EncajeAcumMes_lag1"]        = df["EncajeAcumMes"].shift(1)
            df = df.drop(columns=["ExigibleTotalMes_est", "EncajeAcumMes"])
            cols_ok = [c for c in cols_ok
                       if c not in ("ExigibleTotalMes_est", "EncajeAcumMes")]
            cols_ok += ["ExigibleTotalMes_est_lag1", "EncajeAcumMes_lag1"]

        n_val = df["avance_mes_lag1"].notna().sum() if "avance_mes_lag1" in df.columns else 0
        logger.info(
            f"  bbva_encaje_features: {len(df):,} filas calendario cargadas "
            f"({df['fecha'].min().date()} → {df['fecha'].max().date()}) | "
            f"avance_mes_lag1 válido: {n_val:,} | cols: {cols_ok[1:]}"
        )
        return df

    except Exception as e:
        logger.warning(f"  bbva_encaje_features: error al cargar — {e}")
        return pd.DataFrame()


# 5d. Features CC+OVN en BCR (Saldos_CCOVN.xlsx, rezago 1 día)
def build_ccovn_features(df_ccovn, peru_bday, flujo_sistema=None):
    """
    Deriva features de saldos CC+OVN en el BCR a partir del df ANCHO que produce
    armar_ccovn_ancho() (columnas 'sistema', 'banco_<X>', y 'foco'/'resto' si hay
    partición activa). Todas las variables usan información de t-1 (shift(1)) →
    sin leakage.

    v2: GENÉRICA sobre cualquier columna del df ancho, no solo sistema/bbva. Para
    cada columna 'clave' produce ccovn_<clave>_lag1 y var_ccovn_<clave>_lag1. Esto
    es lo que permite que el mismo cálculo sirva para 'sistema', 'foco', 'resto' y
    cada 'banco_<X>' sin repetir código por caso. La selección de qué clave usar
    como "propio" y cuál como "contraparte" para una entidad dada ocurre DESPUÉS,
    en build_feature_matrix vía resolver_ccovn_lados() — esta función no conoce
    qué banco se está modelando.

    Reindexación a freq="B" (lunes-viernes) antes de diff():
      - Los feriados peruanos (sin datos) se forward-fill desde el día previo.
      - Los feriados de EE.UU. que NO son feriados en Perú (ej. MLK Day,
        Presidents Day) quedan incluidos como días hábiles normales, lo que
        es correcto porque los bancos peruanos operan esos días y el BCRP
        reporta el saldo. No se usa peru_bday (joint PE+US) para este
        reindex porque ese calendario los excluiría, generando NaN en la
        matriz donde sí existen filas de fecha_t.

    Parámetros:
      df_ccovn      : salida de armar_ccovn_ancho() — columnas 'sistema',
                      'banco_<X>' por banco emparejado, 'foco'/'resto' opcional.
      flujo_sistema : pd.Series opcional (D-R del sistema, indexada por fecha).
                      Si se provee, se calcula residuo_ccovn_lag1.

    Features generadas:
      ccovn_<clave>_lag1, var_ccovn_<clave>_lag1   por cada clave del df ancho
      ccovn_vs_dia_mes_lag1  : saldo_sistema(t-1) - media_historica[rango_dia_habil_mes]
                               (desviación respecto al nivel estacional esperado para ese
                               puesto del mes: elimina estacionalidad intramonth). Basada
                               en 'sistema', ajena a la partición.
      residuo_ccovn_lag1     : Δsaldo_sistema(t-1) - flujo_neto_sistema(t-1)
                               (error de la identidad Δsaldo≈flujo; solo si flujo_sistema
                               no es None). Basada en 'sistema'.

    share_propio_lag1 y var_ccovn_propio_exceso_lag1 (el share y el componente
    idiosincrático que v1 llamaba bbva_share_lag1 / var_ccovn_bbva_exceso_lag1)
    YA NO se calculan acá: dependen de "propio", que no existe hasta que se
    resuelve por banco. Se recomponen en build_feature_matrix a partir de
    columnas ya rezagadas — shift(1) distribuye sobre razón y producto, así que
    ccovn_propio_lag1/ccovn_sistema_lag1 y
    var_ccovn_propio_lag1 − share_propio_lag1×var_ccovn_sistema_lag1
    dan exactamente lo mismo que calcularlos antes de rezagar.

    Retorna DataFrame indexado por días lun-vie (freq="B").
    """
    if df_ccovn.empty:
        return pd.DataFrame()

    # Reindexar a lun-vie: cubre feriados PE (ffill) y feriados US (incluidos
    # como hábiles en Perú). Esto alinea con todas las fechas posibles de la
    # matriz aunque el calendario sea PE+US.
    idx_bd = pd.bdate_range(
        start=df_ccovn.index.min(),
        end=df_ccovn.index.max(),
        freq="B",
    )
    df_bd = df_ccovn.reindex(idx_bd).ffill()

    sis = df_bd["sistema"]

    resultado = pd.DataFrame(index=df_bd.index)
    for clave in df_bd.columns:
        serie = df_bd[clave]
        resultado[f"ccovn_{clave}_lag1"]     = serie.shift(1)
        # diff() sobre idx_bd → variación entre días lun-vie consecutivos (correcto)
        resultado[f"var_ccovn_{clave}_lag1"] = serie.diff().shift(1)

    # ── ccovn_vs_dia_mes_lag1 ────────────────────────────────────────────────
    # Desviación del saldo del sistema respecto a su media histórica para ese
    # mismo puesto (rango) dentro del mes hábil:
    #   rango = 1 si es el 1er día hábil del mes, 2 si es el 2do, etc.
    #   media_por_rango = media expandida (solo datos hasta t-1) del saldo para
    #                     cada rango, evitando leakage de información futura.
    #   ccovn_vs_dia_mes = saldo_sistema - media_por_rango
    # Elimina el patrón estacional intramonth (ingresos inicio / retiros fin).
    # XGBoost ya tiene dia_habil_del_mes como feature; esta variable aporta
    # la desviación INESPERADA respecto a ese patrón estacional.
    rango_dia = (
        pd.Series(np.arange(len(df_bd)), index=df_bd.index)
        .groupby(df_bd.index.to_period("M"))
        .transform(lambda x: np.arange(1, len(x) + 1))
    )
    media_por_rango = (
        sis.groupby(rango_dia)
           .transform(lambda x: x.expanding().mean().shift(1))
    )
    resultado["ccovn_vs_dia_mes_lag1"] = (sis - media_por_rango).shift(1)

    # ── residuo_ccovn_lag1 ───────────────────────────────────────────────────
    # Error de la identidad Δsaldo_sistema ≈ flujo_neto_sistema:
    #   residuo = Δsaldo_sistema - flujo_neto_sistema
    # Captura variación en el saldo del BCR no explicada por los flujos D-R
    # reportados (diferencias de timing, componentes no observados, etc.).
    # Para horizontes cortos actúa como señal de reversión; si residuo > 0,
    # el saldo SUBIÓ más de lo que justifican los flujos → corrección esperada.
    # (El comentario anterior decía "cayó", que contradice la fórmula: residuo > 0
    #  equivale a Δsaldo > flujo. El cálculo nunca estuvo mal, solo la nota.)
    if flujo_sistema is not None and not flujo_sistema.empty:
        flujo_bd = flujo_sistema.reindex(idx_bd).ffill()
        resultado["residuo_ccovn_lag1"] = (sis.diff() - flujo_bd).shift(1)
        n_res = resultado["residuo_ccovn_lag1"].notna().sum()
        logger.info(f"  build_ccovn_features: residuo_ccovn_lag1 calculado "
                    f"({n_res:,} filas válidas)")
    else:
        logger.info("  build_ccovn_features: flujo_sistema no disponible — "
                    "residuo_ccovn_lag1 omitido")

    n_val = resultado["ccovn_sistema_lag1"].notna().sum()
    logger.info(f"  build_ccovn_features: {n_val:,} filas válidas en ccovn_sistema_lag1 "
                f"| claves generadas: {list(df_bd.columns)}")
    return resultado


# 5e. Features estacionales en t+h
def _get_bdays_en_mes(fecha, peru_bday):
    """Retorna todos los días hábiles del mes de la fecha dada."""
    inicio_mes = fecha.replace(day=1)
    if fecha.month == 12:
        fin_mes = fecha.replace(year=fecha.year + 1, month=1, day=1) - pd.Timedelta(days=1)
    else:
        fin_mes = fecha.replace(month=fecha.month + 1, day=1) - pd.Timedelta(days=1)
    return pd.bdate_range(start=inicio_mes, end=fin_mes, freq=peru_bday)


def _get_bdays_en_trim(fecha, peru_bday):
    """Retorna todos los días hábiles del trimestre de la fecha dada."""
    trim = (fecha.month - 1) // 3
    inicio_trim = fecha.replace(month=trim * 3 + 1, day=1)
    if trim == 3:
        fin_trim = fecha.replace(year=fecha.year + 1, month=1, day=1) - pd.Timedelta(days=1)
    else:
        fin_trim = fecha.replace(month=(trim + 1) * 3 + 1, day=1) - pd.Timedelta(days=1)
    return pd.bdate_range(start=inicio_trim, end=fin_trim, freq=peru_bday)


def _build_lag_posicion_mes(serie_valores, fechas_th, fechas_t, lags_meses,
                            peru_bday, referencia="cierre", n_objetivo=None):
    """
    Rezagos del flujo alineados por POSICIÓN EN EL MES, y su máximo absoluto.

    Para cada fecha objetivo t+h se toma su distancia en días hábiles al cierre
    de su mes, y se busca el día con esa misma distancia en los meses anteriores.
    Se alinea al CIERRE y no al inicio porque los meses tienen entre 19 y 23 días
    hábiles: anclando al final el último día siempre es 0, y es donde se ata la
    restricción de encaje.

    DISPONIBILIDAD (crítico): el rezago k de t+h cae en t+h-21k días hábiles, que
    solo es observable en t si h <= 21k. Un rezago cuya fecha sea posterior a t se
    descarta. Sin esta máscara el feature usaría el futuro.

    VENTANA DESLIZANTE: con n_objetivo se toman los n_objetivo rezagos MÁS
    RECIENTES que hayan sobrevivido a esa máscara, y los demás candidatos se
    ignoran. lags_meses pasa a ser una lista de candidatos, no de rezagos que
    entran.

        h=5   -> 1,2,3,4      h=43  -> 3,4,5,6
        h=22  -> 2,3,4,5      h=75  -> 4,5,6,7

    El motivo es que el extremo sobre 4 observaciones y el extremo sobre 1 no son
    el mismo estadístico. Sobre el SESGO no había problema: E[máx de 4] > E[máx de
    1], pero con un modelo por horizonte el conteo es constante dentro de cada
    modelo y un corrimiento constante no altera el orden que usa un árbol. El
    problema es la VARIANZA. Sin la ventana, a h largo el feature quedaba siendo
    el valor de UN solo día de hace ~4 meses, y a h corto un extremo sobre cuatro:
    el modelo de h largo recibía un feature mucho más ruidoso, y eso no lo arregla
    tener un modelo por horizonte.

    Fijar el conteo con una lista rígida y lejana ([4,5,6,7] para todo h) también
    lo resolvía, pero pagaba recencia en los horizontes cortos, que era donde no
    había ningún problema. La ventana deslizante da conteo fijo Y recencia máxima
    a cada h.

    Con n_objetivo=None entran todos los disponibles, que es el comportamiento
    previo. Hasta h=21 las dos formas coinciden: los cuatro primeros candidatos
    están disponibles, así que se eligen 1,2,3,4 en ambos casos.

    Se agrega con el MÁXIMO de los |rezagos| elegidos. Medido contra el error
    realizado del modelo, el máximo supera a la mediana entre 45% y 95% según la
    métrica, sobre la misma submuestra.

    Parámetros
    ----------
    serie_valores : Series indexada por fecha con el valor diario (flujo o retiro)
    fechas_th     : Series de fechas objetivo t+h, alineada con fechas_t
    fechas_t      : Series de fechas de origen, para la máscara de disponibilidad
    lags_meses    : candidatos de rezago en meses. Se recorre en orden ASCENDENTE
                    (del más reciente al más viejo); la función lo ordena, no
                    confía en que venga ordenado, porque de ese orden depende
                    cuáles se eligen.
    n_objetivo    : cuántos rezagos entran al extremo. None = todos los disponibles.

    Retorna (max_abs, min_sgn, max_sgn, n_elegidos) como Series alineadas con la
    entrada. n_elegidos queda topeado por n_objetivo cuando este no es None.
    """
    # El orden decide QUÉ rezagos se eligen cuando hay tope, así que se impone
    # acá en vez de documentarlo como precondición del llamador.
    lags_meses = sorted(lags_meses)
    # El CALENDARIO de posiciones y la SERIE DE VALORES son cosas distintas y
    # tienen que construirse por separado.
    #
    # El calendario se genera con peru_bday, EL MISMO que usa _build_seasonal_table
    # para dias_al_cierre_mes y el mismo con el que se derivan las fechas t+h. Es
    # deliberado y no es intercambiable con el índice de la serie: df_bancarios se
    # reindexa con pd.bdate_range (Lun-Vie SIN excluir feriados) para preservar los
    # ceros de días festivos en los lags, de modo que su índice tiene ~291 días de
    # más. Calcular la distancia al cierre sobre él contaría feriados como hábiles
    # y produciría una posición distinta de la que el resto de la matriz entiende
    # por "días al cierre de mes"; además los rezagos podrían caer en feriados,
    # donde R = D = 0 por construcción y el valor no significa nada.
    #
    # Debe cubrir hasta la última fecha objetivo: t+h se extiende más allá del
    # último dato bancario en los orígenes recientes, y SIEMPRE en producción,
    # donde t+h es futuro por definición. Si el calendario terminara con los datos,
    # la posición de esas fechas sería NaN y el cálculo abortaría, aunque los
    # rezagos —que son pasado— estén disponibles.
    #
    # Los valores siguen saliendo solo de la serie: un día del calendario sin dato
    # produce NaN y queda fuera del máximo.
    _th = pd.DatetimeIndex(fechas_th).dropna()
    idx = pd.DatetimeIndex(pd.bdate_range(
        start=min(pd.DatetimeIndex(serie_valores.index).min(), _th.min()),
        end=max(pd.DatetimeIndex(serie_valores.index).max(), _th.max()),
        freq=peru_bday,
    ))

    # Posición dentro del mes, en días hábiles, según la referencia pedida:
    #   "cierre" -> rank descendente: el ÚLTIMO día hábil del mes = 0
    #   "inicio" -> cumcount:         el PRIMER día hábil del mes = 0
    # Con meses de largo variable las dos no son intercambiables; ver la nota de
    # ANCLA_POSICION_MES.
    _tmp = pd.DataFrame({"f": idx, "ym": idx.to_period("M")})
    if referencia == "inicio":
        dcm = _tmp.groupby("ym").cumcount().values.astype(int)
    else:
        dcm = (_tmp.groupby("ym")["f"].rank(ascending=False, method="first")
                                      .values - 1).astype(int)

    # Todo lo que sigue es vectorizado a propósito. La versión con diccionarios y
    # tuplas construía ~1.2M de objetos Python (una fila por horizonte por rezago),
    # lo que en una máquina con memoria ajustada pesa más que los datos mismos.
    #
    # El par (año-mes, distancia_al_cierre) se codifica en un entero: el ordinal
    # del mes por 64 más la distancia, que nunca llega a 64 porque ningún mes tiene
    # tantos días hábiles. Así el lookup es un reindex de pandas en C.
    _K = 64
    ym_ord = idx.to_period("M").astype("int64").values
    clave_a_fecha = pd.Series(idx.values, index=ym_ord * _K + dcm)

    # Distancia máxima disponible en cada mes. Los meses tienen entre 19 y 23 días
    # hábiles, así que una posición lejana al cierre puede no existir en el mes del
    # rezago. Sin recorte esos rezagos se pierden y —contra la intuición— los
    # horizontes CORTOS quedan con menos cobertura que los largos, porque son los
    # que más rezagos piden. Recortar mapea "primer día del mes" contra "primer día
    # del mes", que es la correspondencia correcta.
    dcm_max_mes = pd.Series(dcm).groupby(ym_ord).max()

    th = pd.DatetimeIndex(fechas_th)
    dcm_th = (pd.Series(dcm, index=idx).reindex(th)
                                       .to_numpy(dtype="float64"))
    ym_th_ord = th.to_period("M").astype("int64").values
    ft = pd.DatetimeIndex(fechas_t).values

    n_lags  = np.zeros(len(th), dtype="int64")
    max_abs = np.full(len(th), np.nan)   # mayor excursión, sin signo
    min_sgn = np.full(len(th), np.nan)   # peor excursión NEGATIVA (retiro neto)
    max_sgn = np.full(len(th), np.nan)   # mayor excursión POSITIVA (depósito neto)

    for k in lags_meses:
        ym_lag = ym_th_ord - k
        d_max  = dcm_max_mes.reindex(ym_lag).to_numpy(dtype="float64")
        d_use  = np.minimum(dcm_th, d_max)               # NaN se propaga
        ok     = np.isfinite(d_use)

        clave  = np.where(ok, ym_lag * _K + np.nan_to_num(d_use), -1).astype("int64")
        f_lag  = clave_a_fecha.reindex(clave).to_numpy()
        # Máscara point-in-time: descartar lo que aún no ocurrió en t
        ok &= ~pd.isna(f_lag)
        ok &= (f_lag <= ft)

        # Valor CON SIGNO; el absoluto se toma después, para poder devolver
        # también los extremos direccionales.
        v = serie_valores.reindex(
            pd.DatetimeIndex(np.where(ok, f_lag, np.datetime64("NaT")))
        ).to_numpy(dtype="float64")
        v = np.where(ok, v, np.nan)
        a = np.abs(v)

        # Ventana deslizante: además de estar disponible, el rezago solo entra si
        # todavía falta para llegar al objetivo. Como el bucle va del más reciente
        # al más viejo, los que entran son siempre los n_objetivo más recientes.
        usable = np.isfinite(v)
        if n_objetivo is not None:
            usable &= (n_lags < n_objetivo)
        n_lags += usable
        max_abs = np.where(usable & (~np.isfinite(max_abs) | (a > max_abs)),
                           a, max_abs)
        min_sgn = np.where(usable & (~np.isfinite(min_sgn) | (v < min_sgn)),
                           v, min_sgn)
        max_sgn = np.where(usable & (~np.isfinite(max_sgn) | (v > max_sgn)),
                           v, max_sgn)

    return (pd.Series(max_abs), pd.Series(min_sgn),
            pd.Series(max_sgn), pd.Series(n_lags))


# CANDIDATOS de rezago en meses para los features de posición del mes, ordenados
# del más reciente al más viejo. Ojo con la lectura: no son los rezagos que entran
# al extremo, son los que se ofrecen. De cada fila se toman los
# N_REZAGOS_OBJETIVO primeros que estén disponibles, y el resto se ignora.
#
# Antes la lista era [1,2,3,4] y entraban TODOS los disponibles, así que el
# conteo caía con el horizonte: la máscara h <= 21k deja 4 rezagos hasta h=21 y
# uno solo desde h~64. Un extremo sobre 4 observaciones y uno sobre 1 no son el
# mismo estadístico, y a h largo el feature terminaba siendo el valor de un único
# día de hace ~4 meses.
#
# Con ventana deslizante el conteo queda fijo SIN perder recencia: a cada h entran
# los 4 rezagos más recientes que ya ocurrieron.
#
#     h=5   -> 1,2,3,4      h=43  -> 3,4,5,6
#     h=22  -> 2,3,4,5      h=75  -> 4,5,6,7
#
# Propiedad útil para no invalidar lo ya medido: hasta h=21 los cuatro primeros
# candidatos están disponibles, así que se eligen 1,2,3,4 y el resultado es
# IDÉNTICO al de la lista vieja. El cambio solo toca los horizontes donde antes
# faltaban rezagos, que es donde estaba el problema.
#
# Se ofrecen 8 candidatos y no 7 porque el corte real no es exactamente h <= 21k:
# los meses tienen 19-23 ruedas y el lookup usa aritmética de mes calendario, así
# que el h en que cae cada rezago se corre unas ruedas según el calendario. El
# octavo es margen para que el conteo no baje de 4 en el extremo del rango.
LAGS_POSICION_MES = [1, 2, 3, 4, 5, 6, 7, 8]

# Cuántos rezagos entran al extremo. En None se toman todos los disponibles, que
# es el comportamiento anterior a la ventana deslizante.
N_REZAGOS_OBJETIVO = 4

# n_lags_pos (cuántos rezagos había disponibles) es diagnóstico, no señal: dentro
# de un modelo de h fijo toma casi siempre el mismo valor, y una variable casi
# constante solo agrega candidatos de corte inútiles. Se descarta de la matriz por
# defecto. En True queda para auditar la cobertura sin tener que tocar
# FEATURES_EXCLUIR.
GUARDAR_N_LAGS_POS = False

# Resguardo de frescura para capacidad_retiro_th: por encima de este número de
# días calendario entre fecha_t y la fecha REPORTADA del último dato de
# bbva_encaje.xlsx, merge_asof(backward) estaría proyectando desde un insumo
# obsoleto en vez de "ayer". 10 deja margen sobre un fin de semana largo
# (3-4 días) sin penalizar la operación normal, pero corta antes de que un mes
# sin refrescar el Excel contamine la matriz con valores congelados en
# silencio — ver el bloque 8c de build_feature_matrix().
UMBRAL_STALENESS_ENCAJE_DIAS = 10

# Rezagos y ventana del umbral para el feature de RECURRENCIA. Son 12 y no 4
# porque el estadístico es una proporción y necesita denominador; además con
# k hasta 12 la máscara h <= 21k cubre los 74 horizontes.
# ─────────────────────────────────────────────────────────────────────────────
# Destino de las features de encaje de BBVA (bloques 8 y 8b)
# ─────────────────────────────────────────────────────────────────────────────
# avance_mes_lag1, exceso_abs_lag1 y companhia describen la posicion de encaje de
# UN banco. v1 las repartia a todas las entidades con el argumento de que BBVA es
# el driver sistemico, lo cual es defendible para SISTEMA pero deja de serlo con
# particiones activas:
#
#   FOCO_BBVA       exacto, el grupo ES BBVA
#   SISTEMA         aproximado pero razonable, BBVA es ~94% del neto al cierre
#   RESTO_BBVA      incorrecto, describe un banco que NO esta en el grupo
#   FOCO_GLOBALES   enganhoso, atribuye el encaje de uno a un grupo de cinco
#
#   "exacto"             solo banco_encaje y los grupos cuya composicion es
#                        exactamente {banco_encaje}
#   "exacto_y_sistema"   lo anterior mas SISTEMA (por defecto: cumple lo pedido
#                        para los grupos sin cambiar en silencio el modelo de
#                        SISTEMA, que es el que corre hoy)
#   "todos"              comportamiento de v1
POLITICA_ENCAJE_BBVA = "exacto_y_sistema"


def destinos_encaje_bbva(lista_bancos_full, banco_encaje, reporte_particion,
                         nombre_sistema="SISTEMA", politica=None):
    """Nombres de entidad que reciben las features de encaje de banco_encaje."""
    politica = politica or POLITICA_ENCAJE_BBVA
    if politica == "todos":
        return set(lista_bancos_full)

    destinos = {banco_encaje}
    if politica == "exacto_y_sistema":
        destinos.add(nombre_sistema)

    if reporte_particion and reporte_particion.get("activa"):
        # Solo si el grupo es EXACTAMENTE el banco de encaje. "Contiene" no
        # alcanza: FOCO_GLOBALES contiene a BBVA y aun asi atribuirle su encaje
        # seria inventar un dato para los otros cuatro bancos del grupo.
        for lado in ("foco", "resto"):
            miembros = reporte_particion.get(f"bancos_{lado}", [])
            if set(miembros) == {banco_encaje}:
                destinos.add(reporte_particion[f"nombre_{lado}"])
    return destinos & set(lista_bancos_full)


LAGS_FRECUENCIA_MES = list(range(1, 13))
VENTANA_UMBRAL_FREC = 250

# Referencia de la posición dentro del mes, por serie:
#   "cierre" -> ruedas hasta el ÚLTIMO día hábil  (0 = cierre de mes)
#   "inicio" -> ruedas desde el PRIMER día hábil  (0 = primer hábil)
#
# Si todos los meses tuvieran el mismo largo las dos serían idénticas; difieren
# solo porque tienen entre 19 y 23 ruedas, y entonces la referencia decide dónde
# se absorbe el desajuste. Cada una es EXACTA para posiciones contadas desde su
# propio extremo, así que la elección debe seguir al fenómeno: el encaje vence al
# cierre, mientras que las entradas fuertes se observan al arranque del mes.
# Las dos referencias eligen días distintos en el 74% de las fechas, con ~2 días
# de desfase, así que no es una distinción cosmética.
ANCLA_POSICION_MES = {
    "flujo":    "cierre",
    "retiro":   "cierre",
    "deposito": "cierre",   # "inicio" para probar la hipótesis de inicio de mes
    "acum":     "cierre",
}


def _build_frec_posicion_mes(serie_valores, fechas_th, fechas_t, lags_meses,
                             peru_bday, ventana_umbral, referencia="cierre"):
    """
    RECURRENCIA en la posición del mes: en qué fracción de los últimos meses la
    posición equivalente a t+h superó su umbral.

    Mide con qué frecuencia esa altura del mes viene cargada, en vez de cuánto
    valió el extremo. Para modelar colas la recurrencia puede informar más que la
    magnitud: una posición cargada en 8 de 12 meses es distinta de una cargada en
    1 con un valor enorme.

    El umbral es la mediana móvil de |serie| sobre `ventana_umbral` ruedas hasta t,
    de modo que es point-in-time: nunca usa información posterior a la fecha de
    origen, y se adapta al nivel de actividad de cada época.

    Se usan más rezagos que en _build_lag_posicion_mes (12 contra 4) porque el
    estadístico es una proporción y necesita denominador. Como el rezago k solo es
    observable si h <= 21k, con k hasta 12 la cobertura alcanza h = 252 — o sea los
    74 horizontes, a diferencia del máximo que se queda sin rezagos más allá de 84.

    Retorna (frecuencia en [0,1], nº de rezagos disponibles).
    """
    _th = pd.DatetimeIndex(fechas_th).dropna()
    idx = pd.DatetimeIndex(pd.bdate_range(
        start=min(pd.DatetimeIndex(serie_valores.index).min(), _th.min()),
        end=max(pd.DatetimeIndex(serie_valores.index).max(), _th.max()),
        freq=peru_bday,
    ))

    _tmp = pd.DataFrame({"f": idx, "ym": idx.to_period("M")})
    if referencia == "inicio":
        dcm = _tmp.groupby("ym").cumcount().values.astype(int)
    else:
        dcm = (_tmp.groupby("ym")["f"].rank(ascending=False, method="first")
                                      .values - 1).astype(int)

    _K = 64
    ym_ord = idx.to_period("M").astype("int64").values
    clave_a_fecha = pd.Series(idx.values, index=ym_ord * _K + dcm)
    dcm_max_mes   = pd.Series(dcm).groupby(ym_ord).max()

    th = pd.DatetimeIndex(fechas_th)
    dcm_th    = pd.Series(dcm, index=idx).reindex(th).to_numpy(dtype="float64")
    ym_th_ord = th.to_period("M").astype("int64").values
    ft        = pd.DatetimeIndex(fechas_t).values

    # Umbral móvil evaluado en la FECHA DE ORIGEN (no en t+h): es lo único que se
    # conoce al predecir.
    # El umbral se calcula sobre el MISMO calendario del que salen los rezagos.
    # No es intercambiable con la serie cruda: df_bancarios se reindexa a Lun-Vie
    # incluyendo feriados, con R = D = 0 en ellos, de modo que su |flujo| arrastra
    # ~3.5% de ceros que jamás pueden aparecer entre los valores a clasificar.
    # Medido, esos ceros bajan la mediana móvil un 4% en promedio — y hasta 16% —
    # con una variación en el tiempo que depende de cuántos feriados caen en cada
    # ventana. Eso metería variación temporal espuria en un feature cuyo propósito
    # es justamente medir recurrencia.
    umbral_serie = (serie_valores.reindex(idx).abs()
                    .rolling(ventana_umbral, min_periods=max(20, ventana_umbral // 4))
                    .median())
    umbral = (umbral_serie.reindex(pd.DatetimeIndex(fechas_t))
                          .to_numpy(dtype="float64"))

    n_lags = np.zeros(len(th), dtype="int64")
    n_sup  = np.zeros(len(th), dtype="int64")

    for k in lags_meses:
        ym_lag = ym_th_ord - k
        d_max  = dcm_max_mes.reindex(ym_lag).to_numpy(dtype="float64")
        d_use  = np.minimum(dcm_th, d_max)
        ok     = np.isfinite(d_use)

        clave = np.where(ok, ym_lag * _K + np.nan_to_num(d_use), -1).astype("int64")
        f_lag = clave_a_fecha.reindex(clave).to_numpy()
        ok &= ~pd.isna(f_lag)
        ok &= (f_lag <= ft)                      # máscara point-in-time

        v = np.abs(serie_valores.reindex(
            pd.DatetimeIndex(np.where(ok, f_lag, np.datetime64("NaT")))
        ).to_numpy(dtype="float64"))
        usable = ok & np.isfinite(v) & np.isfinite(umbral)
        n_lags += usable
        n_sup  += usable & (v > umbral)

    with np.errstate(invalid="ignore", divide="ignore"):
        frec = np.where(n_lags > 0, n_sup / np.maximum(n_lags, 1), np.nan)
    return pd.Series(frec), pd.Series(n_lags)


###############################################################################
# PARTE 6 — Construcción de la matriz de features completa
###############################################################################
def _build_seasonal_table(fechas_unicas, peru_holidays, fechas_elecciones, peru_bday):
    """
    Calcula features estacionales para un conjunto de fechas únicas (t+h).
    Retorna DataFrame indexado por fecha.
    """
    # Pre-calcular días hábiles por mes y trimestre una sola vez
    meses_unicos = pd.PeriodIndex(fechas_unicas, freq="M").unique()
    trims_unicos = pd.PeriodIndex(fechas_unicas, freq="Q").unique()

    # Mapa: fecha → días hábiles de su mes
    bdays_por_mes = {}
    for m in meses_unicos:
        inicio = m.start_time
        fin    = m.end_time
        bdays_por_mes[m] = pd.bdate_range(start=inicio, end=fin, freq=peru_bday)

    # Mapa: fecha → días hábiles de su trimestre
    bdays_por_trim = {}
    for q in trims_unicos:
        inicio = q.start_time
        fin    = q.end_time
        bdays_por_trim[q] = pd.bdate_range(start=inicio, end=fin, freq=peru_bday)

    # Mapa: año → días hábiles + dict de posición O(1)
    anios_unicos = pd.PeriodIndex(fechas_unicas, freq="Y").unique()
    bdays_por_anio     = {}
    bdays_pos_por_anio = {}
    for _a in anios_unicos:
        _bd = pd.bdate_range(start=_a.start_time, end=_a.end_time, freq=peru_bday)
        bdays_por_anio[_a]     = _bd
        bdays_pos_por_anio[_a] = {d: i + 1 for i, d in enumerate(_bd)}

    # Sets para lookups rápidos
    hols_set  = set(peru_holidays)
    elec_set  = set(pd.DatetimeIndex(fechas_elecciones)) if fechas_elecciones else set()
    elec_arr  = np.array(sorted(elec_set), dtype="datetime64[D]") if elec_set else np.array([], dtype="datetime64[D]")

    # Feriados en formato numpy para np.busday_count (ciclo electoral)
    hols_np = (peru_holidays.values.astype("datetime64[D]")
               if len(peru_holidays) > 0
               else np.array([], dtype="datetime64[D]"))

    TWO_PI = 2.0 * np.pi

    registros = []
    for fecha in fechas_unicas:
        ts = pd.Timestamp(fecha)
        m  = ts.to_period("M")
        q  = ts.to_period("Q")

        bd_mes  = bdays_por_mes.get(m,  pd.DatetimeIndex([]))
        bd_trim = bdays_por_trim.get(q, pd.DatetimeIndex([]))
        bd_mes_list  = list(bd_mes)
        bd_trim_list = list(bd_trim)

        total_bdays_mes  = len(bd_mes_list)
        total_bdays_trim = len(bd_trim_list)

        try:
            pos_en_mes  = bd_mes_list.index(ts) + 1
        except ValueError:
            pos_en_mes  = 1
        try:
            pos_en_trim = bd_trim_list.index(ts) + 1
        except ValueError:
            pos_en_trim = 1

        dias_al_cierre_mes    = total_bdays_mes - pos_en_mes
        dias_desde_cierre_mes = pos_en_mes - 1
        dias_al_cierre_trim   = total_bdays_trim - pos_en_trim

        # Días CALENDARIO (no hábiles) hasta fin de mes. El período de encaje
        # cierra en día calendario, no hábil: cuando el mes termina en fin de
        # semana el último día hábil es viernes pero el deadline regulatorio
        # es domingo, y la presión de ese viernes no es la de un cierre entre
        # semana. dias_al_cierre_mes mide el reloj operativo (hábil); este mide
        # el reloj regulatorio (calendario). Se complementan, no se solapan.
        dias_cal_al_cierre_mes = int((ts + pd.offsets.MonthEnd(0) - ts).days)

        # Cierre de trimestre calendario: Mar/Jun/Sep/Dic. BBVA concentra ahí
        # su retiro (heatmap de intensidad, sesión 2026-08-13): la activación
        # aparece casi solo en esos cuatro meses, con años recientes saturando
        # cerca de 100% del exceso disponible. dias_al_cierre_mes por sí solo
        # no distingue un cierre de marzo de uno de mayo; esta bandera sí.
        es_mes_cierre_trim = int(ts.month in (3, 6, 9, 12))

        is_penult_bday_trim = int(pos_en_trim == total_bdays_trim - 1)
        is_ultimo_bday_trim = int(pos_en_trim == total_bdays_trim)
        is_1er_bday_trim    = int(pos_en_trim == 1)
        is_2do_bday_trim    = int(pos_en_trim == 2)
        is_3er_bday_trim    = int(pos_en_trim == 3)

        # Feriados (busca en ventana ±2 días calendario)
        is_pre_feriado  = int(any((ts + pd.Timedelta(days=d)) in hols_set for d in [1, 2]))
        is_post_feriado = int(any((ts - pd.Timedelta(days=d)) in hols_set for d in [1, 2]))

        # Elecciones
        if len(elec_arr) > 0:
            ts_d  = np.datetime64(ts, "D")
            diffs = (elec_arr - ts_d).astype(int)
            futuros_e        = diffs[diffs > 0]
            pasados_e        = (-diffs)[diffs < 0]
            is_pre_eleccion  = int(len(futuros_e) > 0 and futuros_e.min() <= 7)
            is_post_eleccion = int(len(pasados_e) > 0 and pasados_e.min() <= 7)
        else:
            is_pre_eleccion = is_post_eleccion = 0

        mes = ts.month

        # ── Posición en el año (días hábiles) ────────────────────────────────
        _a               = ts.to_period("Y")
        total_bdays_anio = len(bdays_por_anio.get(_a, []))
        pos_en_anio      = bdays_pos_por_anio.get(_a, {}).get(ts, 1)
        dias_al_cierre_anio = total_bdays_anio - pos_en_anio

        # ── Días hábiles desde última elección general ────────────────────────
        _ts64     = np.datetime64(ts, "D")
        _pasados  = _ELECCIONES_GENERALES_NP[_ELECCIONES_GENERALES_NP <= _ts64]
        n_bd_elec = (int(np.busday_count(_pasados[-1], _ts64, holidays=hols_np))
                     if len(_pasados) > 0 else 0)

        # ── Codificaciones cíclicas (sin/cos) ─────────────────────────────────
        # mes (P=12, calendario)
        mes_sin = float(np.sin(TWO_PI * mes / 12))
        mes_cos = float(np.cos(TWO_PI * mes / 12))

        # día de semana (P=5 hábiles; dayofweek 0=lun→4=vie, usamos 1-5)
        _dsem        = ts.dayofweek + 1
        dias_sem_sin = float(np.sin(TWO_PI * _dsem / 5))
        dias_sem_cos = float(np.cos(TWO_PI * _dsem / 5))

        # cierre de mes (P=total_bdays_mes hábiles)
        _p_mes = total_bdays_mes or 1
        dias_al_cierre_mes_sin = float(np.sin(TWO_PI * pos_en_mes / _p_mes))
        dias_al_cierre_mes_cos = float(np.cos(TWO_PI * pos_en_mes / _p_mes))

        # cierre de trimestre (P=total_bdays_trim hábiles)
        _p_trim = total_bdays_trim or 1
        dias_al_cierre_trim_sin = float(np.sin(TWO_PI * pos_en_trim / _p_trim))
        dias_al_cierre_trim_cos = float(np.cos(TWO_PI * pos_en_trim / _p_trim))

        # cierre de año (P=total_bdays_anio hábiles)
        _p_anio = total_bdays_anio or 1
        dias_al_cierre_anio_sin = float(np.sin(TWO_PI * pos_en_anio / _p_anio))
        dias_al_cierre_anio_cos = float(np.cos(TWO_PI * pos_en_anio / _p_anio))

        # ciclo electoral (P=1260 hábiles ≈ 5 años, solo elecciones generales)
        elec_sin = float(np.sin(TWO_PI * n_bd_elec / _P_ELEC))
        elec_cos = float(np.cos(TWO_PI * n_bd_elec / _P_ELEC))

        registros.append({
            "fecha_th"             : ts,
            "dias_al_cierre_mes"   : dias_al_cierre_mes,
            "dias_desde_cierre_mes": dias_desde_cierre_mes,
            "dias_al_cierre_trim"  : dias_al_cierre_trim,
            "dias_cal_al_cierre_mes": dias_cal_al_cierre_mes,
            "es_mes_cierre_trim"   : es_mes_cierre_trim,
            "pos_en_mes"           : pos_en_mes,
            "total_bdays_mes"      : total_bdays_mes,
            "is_penult_bday_trim"  : is_penult_bday_trim,
            "is_ultimo_bday_trim"  : is_ultimo_bday_trim,
            "is_1er_bday_trim"     : is_1er_bday_trim,
            "is_2do_bday_trim"     : is_2do_bday_trim,
            "is_3er_bday_trim"     : is_3er_bday_trim,
            "dia_semana"           : ts.dayofweek,
            "mes"                  : mes,
            "is_quincena"          : int(pos_en_mes == 15 or dias_al_cierre_mes == 0),
            "is_cierre_encaje"     : int(dias_al_cierre_mes <= 1),
            "is_fin_anio"          : int(mes == 12 and ts.day >= 28),
            "is_pre_feriado"       : is_pre_feriado,
            "is_post_feriado"      : is_post_feriado,
            "is_pre_eleccion"      : is_pre_eleccion,
            "is_post_eleccion"     : is_post_eleccion,
            # ── Cíclicas (sin/cos) ────────────────────────────────────────────
            "mes_sin"                : mes_sin,
            "mes_cos"                : mes_cos,
            "dias_sem_sin"           : dias_sem_sin,
            "dias_sem_cos"           : dias_sem_cos,
            "dias_al_cierre_mes_sin" : dias_al_cierre_mes_sin,
            "dias_al_cierre_mes_cos" : dias_al_cierre_mes_cos,
            "dias_al_cierre_trim_sin": dias_al_cierre_trim_sin,
            "dias_al_cierre_trim_cos": dias_al_cierre_trim_cos,
            "dias_al_cierre_anio_sin": dias_al_cierre_anio_sin,
            "dias_al_cierre_anio_cos": dias_al_cierre_anio_cos,
            "elec_sin"               : elec_sin,
            "elec_cos"               : elec_cos,
        })

    return pd.DataFrame(registros).set_index("fecha_th")


def build_feature_matrix(
    banco,
    datos_manuales,
    macro_features,
    bank_features,
    peru_bday,
    peru_holidays,
    fechas_elecciones,
    h_min,
    h_max,
    hmm_features=None,           # pd.Series(hmm_estado, index=fecha_t) pre-computada
    encaje_features=None,        # pd.DataFrame(index=fecha_t) con features de encaje (BBVA)
    banco_encaje="BBVA",         # banco al que aplican esas features
    bancos_con_encaje=None,      # v2: nombres que reciben las features de encaje
    recibe_encaje_bbva=True,     # v2: si esta entidad recibe el bloque 8b
    bbva_encaje_feat=None,       # pd.DataFrame(index=fecha_t) avance/exceso desde aux_encaje_2
    ccovn_features=None,         # pd.DataFrame(index=fecha_t), salida de build_ccovn_features (genérica por clave)
    reporte_particion=None,      # v2: salida de aplicar_particion(), para resolver propio/contraparte de ccovn
    nombre_sistema="SISTEMA",    # v2: nombre de la entidad sistema, para resolver_ccovn_lados()
    dias_no_habiles_adicionales=None,  # lista de Timestamps (auto-detectados + manual PARAMS)
):
    """
    Construye el dataset de entrenamiento para un banco de forma vectorizada.
    """
    logger.info(f"  Construyendo matriz para banco: {banco}")

    df_bancarios  = datos_manuales.get("bancarios",  pd.DataFrame())
    df_confirmados = datos_manuales.get("confirmados", pd.DataFrame())

    # Índice temporal de referencia
    if not df_bancarios.empty and f"{banco}_R" in df_bancarios.columns:
        fechas_t = df_bancarios.index
    elif not macro_features.empty:
        fechas_t = macro_features.index
    else:
        logger.warning(f"  Sin fechas de referencia para {banco}.")
        return pd.DataFrame()

    fechas_t = pd.DatetimeIndex(fechas_t)

    # Eliminar feriados de fecha_t.
    # El reindex de bancarios usa pd.bdate_range (Lun-Vie sin exclusión de feriados)
    # para preservar los ceros de días festivos en los lags — correcto, el banco
    # realmente tuvo 0 transacciones ese día. Pero esos días NO deben ser puntos
    # de predicción: (a) target duplica al día hábil anterior porque peru_bday los
    # salta al calcular fecha_th, y (b) todas sus features son cero o forward-fill.
    if len(peru_holidays) > 0:
        _n_antes = len(fechas_t)
        fechas_t = fechas_t[~fechas_t.normalize().isin(peru_holidays.normalize())]
        _n_feriados = _n_antes - len(fechas_t)
        if _n_feriados:
            logger.info(
                f"    {banco}: {_n_feriados} fechas de feriados eliminadas de fecha_t "
                f"({_n_antes} → {len(fechas_t)})"
            )

    # Días no hábiles adicionales (decretos gobierno: APEC 2016, puentes, etc.)
    # Se excluyen de fechas_t → no generan filas en la matriz de entrenamiento.
    # NO se nullifica df_bancarios: esos días tienen R=D=0 reales → target=0 correcto.
    # La contaminación de features (shift/rolling) se evita en build_all_matrices
    # mediante drop explícito sobre _peru_bdays_idx antes de llamar a build_bank_features.
    if dias_no_habiles_adicionales:
        _extra_ts = pd.DatetimeIndex(
            pd.to_datetime(dias_no_habiles_adicionales).normalize()
        )
        _n_antes = len(fechas_t)
        fechas_t = fechas_t[~fechas_t.normalize().isin(_extra_ts)]
        _n_extra = _n_antes - len(fechas_t)
        if _n_extra:
            logger.info(
                f"    {banco}: {_n_extra} días no hábiles adicionales eliminados de fecha_t"
            )

    # ── 1. Grid (fecha_t × h) ────────────────────────────────────────────────
    hs = np.arange(h_min, h_max + 1)
    grid = pd.MultiIndex.from_product([fechas_t, hs], names=["fecha_t", "h"])
    df = pd.DataFrame(index=grid).reset_index()
    df["log_h"] = np.log(df["h"])
    df["banco"] = banco

    # Calcular fecha_th vectorialmente: aplicar el offset h veces
    # Construimos un mapa fecha_t → array de fechas futuras
    logger.info(f"    Calculando fechas futuras ({len(fechas_t)} × {len(hs)})...")
    th_map = {}
    for t in fechas_t:
        th_map[t] = list(get_future_business_dates(t, h_max, peru_bday))

    def get_th(row):
        fechas = th_map.get(row["fecha_t"], [])
        idx = row["h"] - 1
        return fechas[idx] if idx < len(fechas) else pd.NaT

    df["fecha_th"] = pd.to_datetime(df.apply(get_th, axis=1))  # object→datetime64[ns], sin tz
    df = df.dropna(subset=["fecha_th"])

    # ── 1b. Features de gap de calendario ───────────────────────────────────
    # dias_desde_ultimo_habil: días calendarios desde el BDay anterior.
    #   Lun normal = 3 | Lun post-Semana Santa = 5 | cualquier Mar-Vie = 1
    # es_post_feriado: 1 si el gap es mayor a un fin de semana estándar (>3 días).
    #   Señala al árbol que los flujos entrantes pueden incluir acumulación de feriados.
    # Nota: al calcular sobre fechas_t ya filtradas de feriados, shift(1) apunta
    # siempre al día hábil inmediato anterior, con el gap correcto.
    _ft_sorted = fechas_t.sort_values()
    _gap = pd.Series(
        (_ft_sorted.to_series().diff().dt.days).fillna(1).astype(int).values,
        index=_ft_sorted,
        name="dias_desde_ultimo_habil",
    )
    _post_feriado = (_gap > 3).astype(int).rename("es_post_feriado")
    _cal_gap_df = pd.DataFrame({"dias_desde_ultimo_habil": _gap,
                                 "es_post_feriado": _post_feriado})

    # ── 2. Features bancarias (merge por fecha_t) ────────────────────────────
    bf = bank_features.get(banco, pd.DataFrame())
    if not bf.empty:
        df = df.merge(bf.add_prefix(""), left_on="fecha_t", right_index=True, how="left")

    df = df.merge(_cal_gap_df, left_on="fecha_t", right_index=True, how="left")

    # ── 3. Datos bancarios → target ──────────────────────────────────────────
    tiene_bancarios = (
        not df_bancarios.empty
        and f"{banco}_R" in df_bancarios.columns
        and f"{banco}_D" in df_bancarios.columns
    )
    if tiene_bancarios:
        df_banco = df_bancarios[[f"{banco}_R", f"{banco}_D"]].copy()
        df_banco.columns = ["R_th", "D_th"]
        df_banco["target"] = df_banco["D_th"] - df_banco["R_th"]
        df = df.merge(df_banco[["target"]], left_on="fecha_th", right_index=True, how="left")
    else:
        df["target"] = np.nan

    # ── 3b. Recalcular R_conf_t1/R_conf_t2/D_conf_t1 sin feriados ni días extra ─
    # build_bank_features calcula shift(-1/-2) sobre el índice denso (bdate_range,
    # Lun-Vie incluyendo feriados) → en fechas pre-feriado el shift aterriza en el
    # feriado (R=0) en lugar del siguiente día hábil real.
    # Adicionalmente, los días no hábiles adicionales (APEC, puentes) fueron anulados
    # con NaN en df_bancarios → si el siguiente bday es un día extra, shift(-1) da NaN.
    # Solución: filtrar AMBOS conjuntos del índice antes de hacer shift.
    if tiene_bancarios:
        _df_bk = df_bancarios[[f"{banco}_R", f"{banco}_D"]].copy()
        _df_bk.columns = ["R", "D"]
        # Excluir feriados estándar PE+USA
        _all_excl = peru_holidays.normalize() if len(peru_holidays) > 0 \
                    else pd.DatetimeIndex([])
        # Excluir también días no hábiles adicionales (decretos, APEC, puentes)
        if dias_no_habiles_adicionales:
            _extra_norm = pd.DatetimeIndex(
                pd.to_datetime(dias_no_habiles_adicionales).normalize()
            )
            _all_excl = _all_excl.union(_extra_norm)
        _df_nohol = _df_bk[~_df_bk.index.normalize().isin(_all_excl)]
        _r_conf_t1 = _df_nohol["R"].shift(-1)
        _r_conf_t2 = _df_nohol["R"].shift(-2)
        _d_conf_t1 = _df_nohol["D"].shift(-1)
        for _col, _ser in [("R_conf_t1", _r_conf_t1),
                            ("R_conf_t2", _r_conf_t2),
                            ("D_conf_t1", _d_conf_t1)]:
            if _col in df.columns:
                df[_col] = df["fecha_t"].map(_ser)

    # ── 4. Confirmados ───────────────────────────────────────────────────────
    # Base: valores calculados arriba (shift -1/-2 sobre realizados sin feriados)
    # Override: si existen confirmados operativos reales (correos), reemplazan la fila exacta
    if not df_confirmados.empty and "banco" in df_confirmados.columns and "fecha" in df_confirmados.columns:
        conf = df_confirmados[df_confirmados["banco"] == banco].set_index("fecha")
        for col in ["R_conf_t1", "R_conf_t2", "D_conf_t1"]:
            if col in conf.columns:
                override = df["fecha_t"].map(conf[col])
                df[col] = override.combine_first(df[col])

    # ── 5. Macro (merge por fecha_t) ─────────────────────────────────────────
    if not macro_features.empty:
        df = df.merge(macro_features, left_on="fecha_t", right_index=True, how="left")
    else:
        for col in ["VIX","delta_VIX","VIX_ma22","TC_PEN_USD","delta_TC",
                    "tc_vol_5d","tc_vol_22d","garch_vol_tc",
                    "EMBI_PERU","delta_EMBI","garch_vol_embi",
                    "TASA_REF_BCRP","FED_FUNDS","diferencial_tasas","T10Y",
                    "EMBI_PERU_frac","T10Y_frac","CDS_PERU_5Y_frac","COPPER_frac","VIX_frac"]:
            df[col] = np.nan

    # ── 6. Estacionales (calculadas una vez por fecha_th única) ──────────────
    fechas_th_unicas = pd.DatetimeIndex(df["fecha_th"].unique())
    logger.info(f"    Calculando estacionales para {len(fechas_th_unicas):,} fechas únicas...")
    df_seasonal = _build_seasonal_table(
        fechas_th_unicas, peru_holidays, fechas_elecciones, peru_bday
    )

    # Downcast ANTES del merge, no después. Este merge es el pico de memoria de
    # toda la función: reindexa cada columna sobre ~291k filas y con float64 pide
    # el doble de lo necesario. Al final se downcastea igual, así que adelantarlo
    # no cambia ningún resultado — solo evita el pico.
    _f64 = df.select_dtypes(include="float64").columns
    if len(_f64):
        df[_f64] = df[_f64].astype("float32")
    _f64s = df_seasonal.select_dtypes(include="float64").columns
    if len(_f64s):
        df_seasonal[_f64s] = df_seasonal[_f64s].astype("float32")

    df = df.merge(df_seasonal, left_on="fecha_th", right_index=True, how="left")
    del df_seasonal

    # ── 6b. Rezagos alineados por posición del mes ───────────────────────────
    # Escala esperada del flujo en la posición del mes a la que apunta t+h,
    # tomada de los mismos días de los meses previos. Complementa a sigma_22d,
    # que al promediar sobre las 22 ruedas mezcla posiciones y borra esta señal:
    # medido contra el error del modelo, sigma_22d aporta +0.002 y este rezago
    # +0.127 (parcial de Spearman, controlando por la anchura predicha y por h).
    # El flujo NETO se expone con signo, separado en su peor excursión negativa y
    # su mayor excursión positiva, en vez de un solo máximo absoluto. El máximo
    # absoluto es ciego al signo: devuelve la mayor excursión venga de donde
    # venga, así que el modelo de q01 —que necesita el peor retiro— recibía el
    # mismo número que el de q99 aunque proviniera de un depósito enorme. Los dos
    # extremos con signo contienen a max|·| (es el máximo entre max_pos y
    # -min_neg), pero no al revés.
    #
    # R y D se mantienen como componentes: no son redundantes con el neto ni entre
    # sí. R dirige la cola baja (retiros, lo operativo) y D la alta, donde el
    # diagnóstico del oráculo encontró la peor calibración (s_hi hasta 3.1x).
    # Nota: como R y D son no negativos, su |·| es un no-op y su extremo con signo
    # coincide con el máximo — por eso de ellos solo se guarda el máximo.
    if tiene_bancarios:
        _serie_flujo  = (df_bancarios[f"{banco}_D"] - df_bancarios[f"{banco}_R"])
        _serie_retiro = df_bancarios[f"{banco}_R"]
        _serie_dep    = df_bancarios[f"{banco}_D"]
        # Acumulado del flujo DENTRO de su mes, reiniciado en cada período. No es
        # un extremo de un día sino una cantidad de TRAYECTORIA: cuánto flujo neto
        # se había construido a esa altura del ciclo. Anclado en la fecha de
        # origen esta información no resultó relevante —de ahí que
        # flujo_neto_acum_mes esté excluido— pero leído en la posición del mes
        # responde otra pregunta: cuánto se había acumulado en meses previos al
        # llegar a la altura a la que apunta t+h.
        _serie_acum = _serie_flujo.groupby(
            pd.PeriodIndex(_serie_flujo.index, freq="M")).cumsum()
        for _nom, _ser in (("flujo", _serie_flujo), ("retiro", _serie_retiro),
                           ("deposito", _serie_dep), ("acum", _serie_acum)):
            # Sin dropna: el índice completo de df_bancarios es parte del
            # calendario. Los días sin dato dan NaN al buscar el valor y quedan
            # fuera del máximo, pero siguen contando para la posición en el mes.
            _ref = ANCLA_POSICION_MES.get(_nom, "cierre")
            _mx, _mn_s, _mx_s, _n = _build_lag_posicion_mes(
                _ser, df["fecha_th"], df["fecha_t"], LAGS_POSICION_MES, peru_bday,
                referencia=_ref, n_objetivo=N_REZAGOS_OBJETIVO,
            )
            if _nom == "flujo":
                df["esc_neto_min_pos"] = _mn_s.values   # peor retiro neto  -> q01
                df["esc_neto_max_pos"] = _mx_s.values   # mayor depósito neto -> q99
                _n_lags = _n.values
            elif _nom == "acum":
                # Con signo, igual que el flujo: un mes con acumulado muy negativo
                # a esa altura describe un ciclo distinto de uno muy positivo, y
                # un máximo absoluto los confundiría.
                df["acum_neto_min_pos"] = _mn_s.values
                df["acum_neto_max_pos"] = _mx_s.values
            else:
                df[f"esc_{_nom}_pos"] = _mx.values
        _cob = df["esc_neto_min_pos"].notna().mean()
        # Reparto de n_lags al log aunque la columna no se guarde: es como se
        # audita que la máscara point-in-time esté haciendo lo suyo.
        _rep = ", ".join(f"{k}:{(_n_lags == k).mean():.0%}"
                         for k in range(len(LAGS_POSICION_MES) + 1)
                         if (_n_lags == k).any())
        logger.info(f"    Rezagos por posición del mes: cobertura {_cob:.1%} "
                    f"(candidatos {LAGS_POSICION_MES}, objetivo "
                    f"{N_REZAGOS_OBJETIVO}) | anclas {ANCLA_POSICION_MES} "
                    f"| n_lags -> {_rep}")
        # El conteo constante es la premisa de la ventana deslizante, así que se
        # verifica en vez de asumirse. Lo esperable es ~100% en el objetivo: solo
        # el arranque de la muestra, donde no hay suficientes meses hacia atrás,
        # debería quedar por debajo. Si el faltante NO está al inicio, los
        # candidatos se acabaron antes de juntar el objetivo y hay que alargar
        # LAGS_POSICION_MES.
        if N_REZAGOS_OBJETIVO is not None:
            _lleno = float((_n_lags >= N_REZAGOS_OBJETIVO).mean())
            if _lleno < 0.99:
                logger.warning(
                    f"    posición del mes: solo {_lleno:.1%} de las filas llegó "
                    f"a {N_REZAGOS_OBJETIVO} rezagos. Revisar si el faltante está "
                    f"al inicio de la muestra (esperado) o en los horizontes "
                    f"largos (faltan candidatos en LAGS_POSICION_MES).")
        # ── frec_flujo_pos: recurrencia, no magnitud ──────────────────────
        _fr, _ = _build_frec_posicion_mes(
            _serie_flujo, df["fecha_th"], df["fecha_t"], LAGS_FRECUENCIA_MES,
            peru_bday, VENTANA_UMBRAL_FREC,
            referencia=ANCLA_POSICION_MES.get("flujo", "cierre"),
        )
        df["frec_flujo_pos"] = _fr.values
        logger.info(f"    frec_flujo_pos: cobertura "
                    f"{df['frec_flujo_pos'].notna().mean():.1%} "
                    f"({len(LAGS_FRECUENCIA_MES)} rezagos)")

        # ── Rezagos anclados a la APERTURA del mes (referencia="inicio") ────
        # Todo lo de arriba está anclado al CIERRE: encuentra, en los meses
        # previos, el día a la misma distancia del ÚLTIMO día hábil. Bueno
        # para la cola baja —ahí es donde ata la restricción de encaje— pero
        # ciego a la reversión que se observó al abrir el mes (depósito fuerte
        # seguido de retiro fuerte unos días después): esa señal vive a la
        # misma distancia del PRIMER día hábil, no del último, y con meses de
        # 19-23 días esas dos distancias no son la misma posición calendario.
        #
        # Ya hay evidencia de que el desdoblamiento importa: en
        # aux_importancia_calendario.py, dias_al_cierre_mes (crudo) domina
        # q01/q05 mientras dias_desde_cierre_mes (crudo) domina q95/q99 —
        # mismo mecanismo, medido a nivel de los features de calendario. Esto
        # lo extiende a la familia *_pos, que hasta ahora solo usa el ancla de
        # cierre para las 7 columnas activas.
        #
        # Solo 2 columnas, no las 7: la reversión es sobre DEPÓSITOS —D y
        # D-R—, no sobre retiros (que se concentran al cierre por diseño del
        # encaje, no al abrir) ni sobre el acumulado (que al abrir el mes es
        # chico casi por definición, con poca varianza que aprovechar). Se
        # mantiene acotado por el mismo motivo que la corrida de
        # capacidad_retiro_th se sintió lenta: cada columna nueva es más
        # trabajo de importancia por permutación en step005.
        #
        # referencia="inicio" nunca se había ejercido en producción (todas las
        # anclas en ANCLA_POSICION_MES son "cierre"); el recorte por longitud
        # de mes (dcm_max_mes) es agnóstico al ancla, así que la cobertura por
        # h debería ser idéntica a la de esc_deposito_pos/esc_neto_max_pos —
        # se audita en el log de abajo para confirmarlo, no se asume.
        for _nom, _ser in (("flujo", _serie_flujo), ("deposito", _serie_dep)):
            _mx_ap, _, _mx_s_ap, _n_ap = _build_lag_posicion_mes(
                _ser, df["fecha_th"], df["fecha_t"], LAGS_POSICION_MES, peru_bday,
                referencia="inicio", n_objetivo=N_REZAGOS_OBJETIVO,
            )
            if _nom == "flujo":
                # Solo el extremo POSITIVO: el ancla de apertura apunta a la
                # cola alta a propósito. El extremo negativo cerca de la
                # apertura ya lo cubre esc_neto_min_pos (cierre) razonablemente
                # — agregar la contraparte sin evidencia sería el mismo exceso
                # de columnas que se quiere evitar.
                df["esc_neto_max_pos_ap"] = _mx_s_ap.values
                _n_lags_ap = _n_ap.values
            else:
                df["esc_deposito_pos_ap"] = _mx_ap.values
        _cob_ap = df["esc_deposito_pos_ap"].notna().mean()
        logger.info(f"    Rezagos ancla apertura: cobertura {_cob_ap:.1%} "
                    f"(esperado ≈ cobertura de esc_deposito_pos arriba)")

        if GUARDAR_N_LAGS_POS:
            df["n_lags_pos"] = _n_lags
    else:
        df["esc_neto_min_pos"] = np.nan
        df["esc_neto_max_pos"] = np.nan
        df["esc_retiro_pos"]   = np.nan
        df["esc_deposito_pos"] = np.nan
        df["frec_flujo_pos"]   = np.nan
        df["acum_neto_min_pos"] = np.nan
        df["acum_neto_max_pos"] = np.nan
        df["esc_neto_max_pos_ap"]  = np.nan
        df["esc_deposito_pos_ap"]  = np.nan
        if GUARDAR_N_LAGS_POS:
            df["n_lags_pos"] = 0

    # ── 7. HMM estado de régimen (pre-computado, sin leakage) ────────────────
    # hmm_estado es una Serie indexada por fecha_t calculada una sola vez en
    # build_full_matrix con expanding window. El merge por fecha_t es seguro:
    # hmm_estado[Y] fue calculado con datos hasta Y-1, exactamente lo que el
    # fold vería como train. Años de burn-in quedan en NaN.
    if hmm_features is not None and not hmm_features.empty:
        df = df.merge(
            hmm_features.rename("hmm_estado").to_frame(),
            left_on="fecha_t", right_index=True, how="left",
        )
    else:
        df["hmm_estado"] = pd.NA

    # ── 8. Features de encaje (solo para el banco configurado, ej. BBVA) ─────
    # encaje_features está indexado por fecha_t y contiene valores del día t-1.
    # Columnas: encaje_lag1, exceso_lag1, faltante_lag1, techo_10h,
    #           techo_restante_lag1, proporcion_usada.
    # Para bancos distintos al banco_encaje → columnas quedan en NaN.
    #
    # v2: la comparación por nombre exacto deja fuera a los grupos de la
    # partición. FOCO_BBVA *es* BBVA con otro nombre, así que perdería las
    # features de encaje en silencio y su modelo saldría sistemáticamente peor
    # que el de BBVA por una razón que no tiene nada que ver con la partición.
    # bancos_con_encaje trae los nombres que deben recibirlas, resuelto por quien
    # arma la partición. Para un grupo de varios bancos las features son del
    # banco_encaje solamente, o sea una aproximación: se avisa en el log en vez
    # de dejarlo implícito.
    if encaje_features is not None and not encaje_features.empty:
        _con_encaje = set(bancos_con_encaje or {banco_encaje})
        if banco in _con_encaje:
            df = df.merge(encaje_features, left_on="fecha_t", right_index=True, how="left")
            _nota = "" if banco == banco_encaje else \
                f" (aproximación: son de {banco_encaje}, que es parte del grupo)"
            logger.info(f"  {banco}: features de encaje incorporadas "
                        f"({len(encaje_features.columns)} cols){_nota}")
        else:
            for col in encaje_features.columns:
                df[col] = np.nan

    # ── 8b. Features de avance/exceso encaje BBVA ─────────────────────────────
    # v1 las aplicaba a TODAS las entidades. v2 las restringe segun
    # POLITICA_ENCAJE_BBVA: describen la posicion de encaje de un solo banco, y
    # con particiones activas repartirlas a un grupo que no lo contiene (o que lo
    # contiene junto a otros cuatro) es atribuir un dato que no corresponde.
    # recibe_encaje_bbva lo resuelve destinos_encaje_bbva() en build_full_matrix.
    # bbva_encaje_feat está en días calendario con lag1 ya computado en aux_encaje_2.
    # merge_asof alinea al calendario hábil de cada entidad.
    _bbva_feat_cols = ["avance_mes_lag1", "exceso_abs_lag1", "exceso_dia_lag1",
                       "encaje_ovn_lag1", "ratio_ovn_total_lag1",
                       "encaje_diario_lag1", "ExigibleTotalMes_est_lag1",
                       "EncajeAcumMes_lag1"]
    if bbva_encaje_feat is not None and not bbva_encaje_feat.empty \
            and not recibe_encaje_bbva:
        logger.info(f"  {banco}: features de encaje de {banco_encaje} OMITIDAS "
                    f"(política {POLITICA_ENCAJE_BBVA})")
        for _c in _bbva_feat_cols:
            df[_c] = np.nan
    elif bbva_encaje_feat is not None and not bbva_encaje_feat.empty:
        # Construir lookup: fecha_t_única → valor del Excel más reciente ≤ fecha_t.
        # merge_asof sobre fechas únicas evita ambigüedades con múltiples h por fecha.
        _ft_norm = pd.to_datetime(df["fecha_t"]).dt.normalize()
        _dates_unicas = pd.DataFrame(
            {"fecha_t": _ft_norm.unique()}
        ).sort_values("fecha_t")

        _feat_r = bbva_encaje_feat[["fecha"] + _bbva_feat_cols].copy()
        _feat_r["fecha"] = pd.to_datetime(_feat_r["fecha"]).dt.normalize()
        # Se conserva la fecha REPORTADA antes de renombrarla a fecha_t (la
        # clave del join): merge_asof(backward) no distingue "encontré un
        # dato de ayer" de "encontré el único dato disponible, de hace 18
        # días" — sin esta columna, capacidad_retiro_th no puede detectar
        # cuándo está proyectando desde un insumo obsoleto (ver 8c).
        _feat_r["_fecha_reportada"] = _feat_r["fecha"]
        _feat_r = _feat_r.rename(columns={"fecha": "fecha_t"}) \
                          .sort_values("fecha_t")

        _lookup = pd.merge_asof(
            _dates_unicas,
            _feat_r,
            on="fecha_t",
            direction="backward",
        ).set_index("fecha_t")

        for col in _bbva_feat_cols:
            df[col] = _ft_norm.map(_lookup[col]).values
        # Insumo interno para el resguardo de frescura de 8c; no es feature,
        # se descarta junto con el resto de insumos de proyección.
        df["_fecha_reportada_bbva"] = _ft_norm.map(_lookup["_fecha_reportada"]).values

        n_ok = df["avance_mes_lag1"].notna().sum()
        logger.info(f"  {banco}: bbva_encaje_feat incorporadas — "
                    f"{n_ok:,}/{len(df):,} filas con valores")

        # ── 8c. capacidad_retiro_th ────────────────────────────────────────
        # Reemplaza a presion_deadline_th/_t (eliminadas esta sesión). Aquellas
        # medían la OBLIGACIÓN de depositar; esta mide la CAPACIDAD de retirar,
        # que es la magnitud directamente relevante para la cola de retiro del
        # target. Además, al proyectar min_por_dia hacia adelante en vez de
        # usar el valor de hoy, no pierde magnitud por el MAX(0,·): cuando el
        # banco va adelantado el requerimiento se satura en 0 y la capacidad
        # reportada es el saldo completo, en vez de aplanarse a un valor fijo
        # como hacía presion_deadline_th con su propio clip.
        #
        # SUPUESTO ÚNICO Y EXPLÍCITO: el banco mantiene desde t-1 hasta t+h el
        # mismo ritmo de depósito diario (encaje_diario_lag1) y el mismo saldo
        # overnight. Es la misma disciplina que ya usaba presion_deadline_th
        # —proyectar solo lo proyectable sin calendario— llevada a la variable
        # de capacidad en vez de a la de obligación:
        #
        #   dias_h = (fecha_th − fecha_t).días + 1   [EncajeAcumMes_lag1 cubre
        #             hasta fecha_t−1, así que el +1 cierra ese día también]
        #   EncajeAcum_proy(t+h) = EncajeAcumMes_lag1 + encaje_diario_lag1 · dias_h
        #   min_por_dia(t+h)     = MAX(0, (ExigibleTotalMes_est_lag1
        #                                   − EncajeAcum_proy(t+h))
        #                                  / dias_restantes_calendario(t+h))
        #   capacidad_retiro_th  = MAX(0, encaje_diario_lag1 − min_por_dia(t+h))
        #                          + overnight_lag1
        #
        # overnight_lag1 se obtiene por diferencia: encaje_ovn_lag1 −
        # encaje_diario_lag1 (aux_encaje_2 no exporta overnight_lag1 directo;
        # encaje_ovn_lag1 sí existía antes de esta sesión).
        #
        # VALIDEZ: igual que antes, el período se reinicia al cierre, así que
        # ExigibleTotalMes_est_lag1 deja de aplicar si fecha_th cae en un mes
        # calendario distinto al de fecha_t. A diferencia de presion_deadline_th
        # —donde "fuera del período" se resolvía a 0 porque ese deadline ya
        # había pasado—, acá NO corresponde poner 0: desconocer la capacidad
        # futura no es lo mismo que no tener capacidad. Queda en NaN.
        _cols_cap = {"encaje_diario_lag1", "encaje_ovn_lag1",
                    "ExigibleTotalMes_est_lag1", "EncajeAcumMes_lag1"}
        if _cols_cap.issubset(df.columns):
            _th_n = pd.to_datetime(df["fecha_th"])
            # EncajeAcumMes_lag1 cubre hasta fecha_t − 1 (el .shift(1) del
            # archivo calendario), no hasta fecha_t: por eso son
            # (fecha_th − fecha_t) + 1 días de crecimiento a proyectar, no
            # (fecha_th − fecha_t). Verificado con un mes sintético: sin el
            # +1 la capacidad proyectada quedaba sistemáticamente un día de
            # ritmo por debajo de la esperada en todo el mes.
            _dias_h = (_th_n - _ft_norm).dt.days + 1
            _dias_rest_th  = ((_th_n + pd.offsets.MonthEnd(0)) - _th_n).dt.days
            _mismo_periodo = (_th_n.dt.to_period("M") == _ft_norm.dt.to_period("M"))

            _overnight_lag1 = df["encaje_ovn_lag1"] - df["encaje_diario_lag1"]
            _encaje_acum_proy = (
                df["EncajeAcumMes_lag1"] + df["encaje_diario_lag1"] * _dias_h
            )
            _min_por_dia_th = np.maximum(
                0.0,
                (df["ExigibleTotalMes_est_lag1"] - _encaje_acum_proy)
                / np.maximum(_dias_rest_th, 1),
            )
            _capacidad = (
                np.maximum(0.0, df["encaje_diario_lag1"] - _min_por_dia_th)
                + _overnight_lag1
            )
            # RESGUARDO DE FRESCURA: merge_asof(backward) no distingue "el dato
            # más reciente es de ayer" de "el más reciente disponible es de
            # hace 18 días porque el archivo no se actualizó". Sin esto, toda
            # fecha_t posterior al último dato de bbva_encaje.xlsx recibe LOS
            # MISMOS insumos congelados, y como la fórmula tiene dos MAX(0,·)
            # que saturan, produce el mismo número exacto para cientos de
            # filas — confirmado empíricamente: tras el corte del archivo, un
            # solo valor (2,596,363,008) cubrió 1,320 filas en 11 fecha_t
            # consecutivas (aux_diagnostico_capacidad_retiro.py, sesión
            # 2026-08-13). UMBRAL_STALENESS_ENCAJE_DIAS=10 dista un poco de un
            # fin de semana largo (3-4 días) para no penalizar la operación
            # normal, pero corta antes de que un mes sin actualizar el Excel
            # contamine la matriz en silencio.
            _dias_stale = (_ft_norm - pd.to_datetime(df["_fecha_reportada_bbva"])).dt.days
            _fresco = _dias_stale <= UMBRAL_STALENESS_ENCAJE_DIAS
            df["capacidad_retiro_th"] = _capacidad.where(_mismo_periodo & _fresco)

            _cob = df["capacidad_retiro_th"].notna().mean()
            _sin_dato = df["encaje_diario_lag1"].isna().mean()
            _n_stale = int((_mismo_periodo & ~_fresco).sum())
            logger.info(
                f"  {banco}: capacidad_retiro_th — cobertura {_cob:.1%} "
                f"| fuera del período {(~_mismo_periodo).mean():.1%} (NaN, "
                f"no proyectable sin datos del mes siguiente) "
                f"| sin dato de encaje {_sin_dato:.1%} "
                f"| descartadas por dato obsoleto (>{UMBRAL_STALENESS_ENCAJE_DIAS}d): "
                f"{_n_stale:,} filas"
            )
        # Insumos de proyección, no features: se descartan tras usarlos.
        # encaje_diario_lag1 y encaje_ovn_lag1 quedan también fuera —el
        # segundo ya estaba excluido antes de esta sesión— porque lo que
        # aportan a la matriz es capacidad_retiro_th, ya calculada.
        df = df.drop(columns=["ExigibleTotalMes_est_lag1", "EncajeAcumMes_lag1",
                              "encaje_diario_lag1", "_fecha_reportada_bbva"],
                     errors="ignore")

    # Si no hubo archivo de encaje, o le faltaban las columnas necesarias, la
    # columna no se creó. Se materializa en NaN para que el esquema de la
    # matriz sea el mismo en todos los casos y coincida con el registro.
    if "capacidad_retiro_th" not in df.columns:
        df["capacidad_retiro_th"] = np.nan

    # ── 9. Features CC+OVN en BCR (Saldos_CCOVN.xlsx) ────────────────────────
    # ccovn_features está indexado por fecha y contiene valores del día t-1,
    # generados genéricamente por clave (sistema, foco, resto, banco_<X>).
    #
    # v1 le daba a TODAS las entidades la misma columna ccovn_bbva_lag1: era
    # literalmente el saldo de BBVA, sin importar qué banco se estuviera
    # modelando. Con particiones activas eso deja de tener sentido — el saldo
    # "propio" de FOCO_BBVA no es el mismo objeto que el "propio" de RESTO_BBVA.
    #
    # v2 resuelve, por banco, cuál es su propio saldo y —cuando la entidad es un
    # lado de la partición activa— el de su contraparte (resolver_ccovn_lados).
    # Nace ahí la señal nueva que no existía antes: el saldo del OTRO lado, que
    # es donde viviría una compensación entre BBVA y el resto de la banca.
    _ccovn_cols_finales = ["ccovn_sistema_lag1", "var_ccovn_sistema_lag1",
                           "ccovn_propio_lag1", "var_ccovn_propio_lag1",
                           "share_propio_lag1", "var_ccovn_propio_exceso_lag1",
                           "ccovn_contraparte_lag1", "var_ccovn_contraparte_lag1",
                           "share_contraparte_lag1",
                           "ccovn_vs_dia_mes_lag1", "residuo_ccovn_lag1"]
    if ccovn_features is not None and not ccovn_features.empty:
        clave_p, clave_c = resolver_ccovn_lados(banco, nombre_sistema, reporte_particion)
        _sub = armar_sub_ccovn(ccovn_features, clave_p, clave_c)
        df = df.merge(_sub, left_on="fecha_t", right_index=True, how="left")

        for _c in _ccovn_cols_finales:
            # ccovn_vs_dia_mes_lag1 y residuo_ccovn_lag1 también pueden faltar:
            # el segundo solo se calcula si flujo_sistema estaba disponible.
            if _c not in df.columns:
                df[_c] = np.nan

        # share_propio_lag1 y var_ccovn_propio_exceso_lag1 se recomponen acá, no
        # en build_ccovn_features: shift(1) distribuye sobre razón y producto, así
        # que operar sobre columnas YA rezagadas da lo mismo que rezagar el
        # resultado. Ver la nota de build_ccovn_features.
        _sis_nz = df["ccovn_sistema_lag1"].replace(0, np.nan)
        df["share_propio_lag1"]      = df["ccovn_propio_lag1"] / _sis_nz
        df["share_contraparte_lag1"] = df["ccovn_contraparte_lag1"] / _sis_nz
        df["var_ccovn_propio_exceso_lag1"] = (
            df["var_ccovn_propio_lag1"]
            - df["share_propio_lag1"] * df["var_ccovn_sistema_lag1"]
        )

        # Cuando la entidad ES el sistema, "propio" y "sistema" son la misma
        # serie: share_propio da 1 y el exceso da 0, en todas las filas. Un
        # feature constante no aporta un corte y encima ensucia los heatmaps de
        # importancia, donde aparece como una fila más compitiendo por espacio.
        # Se anulan a proposito en vez de dejarlos entrar como constantes. La
        # señal de concentración para SISTEMA viaja por share_contraparte_lag1,
        # que es el FOCO sobre el total (el bbva_share_lag1 de v1).
        if clave_p == "sistema":
            df["share_propio_lag1"] = np.nan
            df["var_ccovn_propio_exceso_lag1"] = np.nan
        logger.info(f"  {banco}: ccovn propio='{clave_p}' contraparte="
                    f"'{clave_c or '—'}' ({_sub.shape[1]} cols traídas)")
    else:
        for _c in _ccovn_cols_finales:
            df[_c] = np.nan

    # ── Reducir a float32 antes de reordenar ────────────────────────────────
    # Mitad de memoria y evita el OOM en la consolidación interna de pandas
    float_cols = df.select_dtypes(include="float64").columns
    df[float_cols] = df[float_cols].astype("float32")

    # ── Excluir features definidos en FEATURES_EXCLUIR ──────────────────────
    excluir = [c for c in FEATURES_EXCLUIR
               if c not in ("fecha_t", "banco", "h", "target")]   # protege identidad
    if excluir:
        df = df.drop(columns=[c for c in excluir if c in df.columns], errors="ignore")
        logger.info(f"  {banco}: {len(excluir)} features excluidos → "
                    f"{df.shape[1]} columnas finales")

    # ── Reordenar columnas ───────────────────────────────────────────────────
    cols_id    = [c for c in ["fecha_t", "banco", "h", "log_h", "fecha_th"] if c in df.columns]
    cols_resto = [c for c in df.columns if c not in set(cols_id)]
    df = df[cols_id + cols_resto]

    # Reporte NaN
    total = len(df)
    nan_cols = {c: f"{df[c].isna().sum()/total:.1%}" for c in df.columns if df[c].isna().any()}
    logger.info(f"  {banco}: {total:,} filas | columnas con NaN: {len(nan_cols)}")

    return df


def build_full_matrix(
    params,
    datos_manuales,
    lista_bancos,
    macro_features,
    peru_bday,
    peru_holidays,
    fechas_elecciones,
    reporte_particion=None,   # v2: salida de aplicar_particion(); None = sin partición
):
    """
    Construye la matriz banco por banco y la escribe directamente al Parquet
    de salida usando PyArrow streaming. Nunca acumula más de un banco en RAM.
    Retorna el path del archivo generado (no el DataFrame completo).
    """
    import gc
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq_writer_lib
    except ImportError:
        logger.error("pyarrow no disponible. Instalar con: pip install pyarrow")
        return pd.DataFrame()

    logger.info("PARTE 6: Construyendo matriz completa de features...")

    df_bancarios = datos_manuales.get("bancarios", pd.DataFrame())
    ruta_output  = Path(params.get("ruta_output", "matriz_features.parquet"))
    ruta_output.parent.mkdir(parents=True, exist_ok=True)

    # ── Agregar SISTEMA: suma del sistema bancario completo ──────────────────
    # Las features de SISTEMA se calculan sobre la serie agregada (no suma de
    # features individuales) para que volatilidades y medias móviles sean correctas.
    NOMBRE_SISTEMA = "SISTEMA"
    # v2: los agregados de la partición viven en el mismo DataFrame, así que el
    # barrido por sufijo de v1 los sumaría al total y SISTEMA quedaría al doble
    # (FOCO + RESTO ya ES el sistema entero). Se excluyen explícitamente en vez
    # de confiar en que el sufijo alcance para distinguir un banco de un agregado.
    _derivadas = columnas_derivadas(reporte_particion, NOMBRE_SISTEMA)
    if not df_bancarios.empty:
        r_cols = [c for c in df_bancarios.columns
                  if c.endswith("_R") and c[:-2] not in _derivadas]
        d_cols = [c for c in df_bancarios.columns
                  if c.endswith("_D") and c[:-2] not in _derivadas]
        if r_cols and d_cols:
            df_bancarios = df_bancarios.copy()
            df_bancarios[f"{NOMBRE_SISTEMA}_R"] = df_bancarios[r_cols].sum(axis=1)
            df_bancarios[f"{NOMBRE_SISTEMA}_D"] = df_bancarios[d_cols].sum(axis=1)
            datos_manuales["bancarios"] = df_bancarios
            logger.info(
                f"  SISTEMA agregado: suma de {len(r_cols)} series "
                f"({', '.join(c.replace('_R','') for c in r_cols)})"
            )
            # Con partición activa, FOCO + RESTO tiene que reconstruir SISTEMA.
            # aplicar_particion() ya lo verificó sobre el pivot crudo, pero acá
            # el universo cambió: agrupar_bancos() colapsó los chicos en
            # Otros_bancos. Si el colapso hubiera perdido o duplicado algo, esta
            # es la única comprobación que lo ve.
            if reporte_particion and reporte_particion.get("activa"):
                _f, _r = reporte_particion["nombre_foco"], reporte_particion["nombre_resto"]
                for _s in ("_R", "_D"):
                    if f"{_f}{_s}" in df_bancarios.columns:
                        _peor = float((df_bancarios[f"{NOMBRE_SISTEMA}{_s}"]
                                       - df_bancarios[f"{_f}{_s}"]
                                       - df_bancarios[f"{_r}{_s}"]).abs().max())
                        if _peor > 1e-6:
                            raise AssertionError(
                                f"{_f} + {_r} != {NOMBRE_SISTEMA} en {_s} "
                                f"(peor desvío {_peor:.6g}) tras agrupar_bancos()")
                logger.info(f"  Partición verificada: {_f} + {_r} == {NOMBRE_SISTEMA}")

    # SISTEMA va primero en la lista para poder validarlo de forma aislada.
    # Los grupos de la partición van inmediatamente después: son los que se
    # comparan contra SISTEMA en la carrera de caballos.
    lista_bancos_full = [NOMBRE_SISTEMA]
    if reporte_particion and reporte_particion.get("activa"):
        lista_bancos_full += [reporte_particion["nombre_foco"],
                              reporte_particion["nombre_resto"]]
    lista_bancos_full += list(lista_bancos)

    # ── Auto-detección de días no hábiles adicionales ────────────────────────
    # Perú declara "puentes" por decreto supremo (DS) varios días al año: previos
    # a Fiestas Patrias, Navidad, Año Nuevo, APEC, etc. Estos días NO aparecen en
    # ningún calendario estándar (PE ni US), pero el sistema bancario reporta R=D=0.
    #
    # Criterio: si SISTEMA_R == 0 Y SISTEMA_D == 0, el sistema bancario completo
    # no operó ese día → debe tratarse igual que un feriado estándar.
    # Se complementa con la lista manual de PARAMS para casos donde los datos fuente
    # pudieran tener un valor residual no cero por error de captura.
    #
    # Impacto si no se corrige: lags (R_t0, D_t0, R_t-1, D_t-1) y rolling features
    # (ma_flujo_5d, sigma_flujo_ratio, flujo_neto_sum_5d, etc.) se contaminan con
    # ceros artificiales, enseñando al modelo un estado que no existe.
    _dias_extra_set = set()

    # 1. Auto-detección desde datos SISTEMA
    if not df_bancarios.empty:
        _sis_r_col = f"{NOMBRE_SISTEMA}_R"
        _sis_d_col = f"{NOMBRE_SISTEMA}_D"
        if _sis_r_col in df_bancarios.columns and _sis_d_col in df_bancarios.columns:
            _zero_mask = (df_bancarios[_sis_r_col] == 0) & (df_bancarios[_sis_d_col] == 0)
            _zero_dates = df_bancarios.index[_zero_mask].normalize()
            _hols_norm  = peru_holidays.normalize()
            _auto_extra = _zero_dates[~_zero_dates.isin(_hols_norm)]
            if len(_auto_extra):
                _dias_extra_set.update(_auto_extra.tolist())

    # 2. Lista manual de PARAMS (override / casos con valor residual no cero)
    _manual_extra = params.get("dias_no_habiles_adicionales", [])
    if _manual_extra:
        _manual_ts = pd.DatetimeIndex(pd.to_datetime(_manual_extra).normalize())
        _new_manual = [d for d in _manual_ts if d not in _dias_extra_set]
        if _new_manual:
            _dias_extra_set.update(_new_manual)

    # Imprimir tabla de días detectados
    if _dias_extra_set and not df_bancarios.empty:
        _sis_r_col = f"{NOMBRE_SISTEMA}_R"
        _sis_d_col = f"{NOMBRE_SISTEMA}_D"
        _dias_sorted = sorted(_dias_extra_set)
        _n_manual = len([
            d for d in _dias_extra_set
            if d in pd.DatetimeIndex(pd.to_datetime(params.get("dias_no_habiles_adicionales", [])).normalize())
            and d not in (
                _zero_dates[~_zero_dates.isin(_hols_norm)]
                if "_zero_dates" in dir() else []
            )
        ])
        print("\n" + "=" * 72)
        print(f"  DÍAS NO LABORABLES ADICIONALES DETECTADOS: {len(_dias_sorted)}")
        print(f"  (SISTEMA_R=SISTEMA_D=0, fuera del calendario PE+USA)")
        print("=" * 72)
        _dias_semana = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
        print(f"  {'Fecha':<14} {'Día':<5}  {'R_t-1 (MM USD)':>16}  {'D_t-1 (MM USD)':>16}  {'Fuente'}")
        print(f"  {'-'*14} {'-'*5}  {'-'*16}  {'-'*16}  {'-'*20}")
        for _f in _dias_sorted:
            _f_ts = pd.Timestamp(_f)
            _dia  = _dias_semana[_f_ts.weekday()]
            _prev = df_bancarios.index[df_bancarios.index < _f_ts]
            if len(_prev) and _sis_r_col in df_bancarios.columns:
                _r_ant = df_bancarios.loc[_prev[-1], _sis_r_col] / 1e6
                _d_ant = df_bancarios.loc[_prev[-1], _sis_d_col] / 1e6
                _r_str = f"{_r_ant:>14,.1f}"
                _d_str = f"{_d_ant:>14,.1f}"
            else:
                _r_str = f"{'N/D':>14}"
                _d_str = f"{'N/D':>14}"
            _manual_flag = (
                _f_ts.normalize() in pd.DatetimeIndex(
                    pd.to_datetime(params.get("dias_no_habiles_adicionales", [])).normalize()
                ) if params.get("dias_no_habiles_adicionales") else False
            )
            _fuente = "manual+auto" if _manual_flag else "auto"
            print(f"  {str(_f_ts.date()):<14} {_dia:<5}  {_r_str}  {_d_str}  {_fuente}")
        print("=" * 72 + "\n")
        logger.info(
            f"  {len(_dias_sorted)} días no hábiles adicionales detectados "
            f"({len(_manual_extra)} en PARAMS, resto auto-detectados)"
        )

    # Auto-whitelist: fechas clasificadas como "feriado" en el calendario (PE+USA
    # estándar o auto-detectadas) donde SISTEMA realmente operó (R > 0 ó D > 0).
    # Causas típicas:
    #   - Nochebuena (Dec 24) hardcodeada pero banco abierto varios años
    #   - US Federal holidays "observados" (viernes previo a sábado festivo):
    #     Veterans Day, Independence Day, Juneteenth, New Year's Eve observado, etc.
    #     La banca peruana no cierra por estos feriados observados de EEUU.
    #   - Feriados PE estándar (FiestasP2, Angamos) donde el banco operó ese año
    #
    # Se remueven de peru_holidays y _dias_extra_set ANTES de construir peru_bday_ext,
    # para que queden válidos tanto como fecha_t (filas de entrenamiento) como
    # como fecha_th (fechas objetivo con target real).
    _dias_si_habiles: set = set()
    if not df_bancarios.empty and f"{NOMBRE_SISTEMA}_R" in df_bancarios.columns:
        _sis_r_col_wl = f"{NOMBRE_SISTEMA}_R"
        _sis_d_col_wl = f"{NOMBRE_SISTEMA}_D"
        _sis_operated = (
            (df_bancarios[_sis_r_col_wl] > 0) | (df_bancarios[_sis_d_col_wl] > 0)
        )
        _operated_dates_wl = df_bancarios.index[_sis_operated].normalize()
        _hols_norm_wl   = set(peru_holidays.normalize())
        _extra_norm_wl  = {pd.Timestamp(d).normalize() for d in _dias_extra_set}
        _all_excl_wl    = _hols_norm_wl | _extra_norm_wl
        _false_excl     = set(_operated_dates_wl) & _all_excl_wl
        if _false_excl:
            _dias_si_habiles.update(_false_excl)

    # Whitelist manual desde PARAMS (override adicional)
    _manual_si = params.get("dias_si_habiles", [])
    if _manual_si:
        _dias_si_habiles.update(
            pd.DatetimeIndex(pd.to_datetime(_manual_si).normalize()).tolist()
        )

    if _dias_si_habiles:
        _si_idx = pd.DatetimeIndex(sorted(_dias_si_habiles))
        _si_dates_str = [pd.Timestamp(d).strftime("%Y-%m-%d") for d in sorted(_dias_si_habiles)]
        print("\n" + "=" * 72)
        print(f"  AUTO-WHITELIST: {len(_dias_si_habiles)} fechas 'feriado' con actividad SISTEMA")
        print("  → restauradas como días hábiles (excluidas del calendario)")
        print("=" * 72)
        for _d_wl in sorted(_dias_si_habiles):
            _t_wl = pd.Timestamp(_d_wl)
            _in_hols  = _t_wl.normalize() in set(peru_holidays.normalize())
            _in_extra = _t_wl.normalize() in {pd.Timestamp(x).normalize() for x in _dias_extra_set}
            _fuente_wl = ("feriado_PE_USA" if _in_hols else "") + \
                         (" dias_extra" if _in_extra else "")
            print(f"  {_t_wl.date()}  [{_fuente_wl.strip()}]")
        print("=" * 72 + "\n")
        logger.info(
            f"  Auto-whitelist: {len(_dias_si_habiles)} fechas restauradas como hábiles "
            f"({_si_dates_str})"
        )
        # Limpiar peru_holidays
        peru_holidays = peru_holidays[
            ~peru_holidays.normalize().isin(_si_idx.normalize())
        ]
        # Limpiar _dias_extra_set
        _si_norm_set = {pd.Timestamp(d).normalize() for d in _dias_si_habiles}
        _dias_extra_set -= _si_norm_set

    # 3. Calendario extendido: peru_bday + días no hábiles adicionales (puentes, APEC…).
    # Usado para _peru_bdays_idx (features shift/rolling) Y para fecha_th en build_feature_matrix.
    # Garantiza que fecha_th nunca caiga en un día no hábil adicional: si cayera, el target
    # sería 0 (R=D=0 real) pero la "distancia real" al próximo día operativo es mayor,
    # de modo que el target corresponde a otro horizonte efectivo, rompiendo la semántica de h.
    if _dias_extra_set:
        from pandas.tseries.offsets import CustomBusinessDay as _CBD
        _extra_hols = pd.DatetimeIndex(sorted(_dias_extra_set))
        _all_hols   = peru_holidays.normalize().union(_extra_hols.normalize())
        peru_bday_ext = _CBD(holidays=_all_hols)
    else:
        peru_bday_ext = peru_bday

    # Índice de días hábiles PE+USA sin días no hábiles adicionales.
    # peru_bday_ext ya los excluye → shift/rolling no se contamina con R=D=0 de esos días.
    _peru_bdays_idx = (
        pd.bdate_range(
            start=df_bancarios.index.min(),
            end=df_bancarios.index.max(),
            freq=peru_bday_ext,
        )
        if not df_bancarios.empty else pd.DatetimeIndex([])
    )
    # _extra_ts_drop: fallback explícito (por si algún día cae fuera del rango de _all_hols)
    _extra_ts_drop = (
        pd.DatetimeIndex(sorted(_dias_extra_set))
        if _dias_extra_set else pd.DatetimeIndex([])
    )

    # Pre-calcular features bancarias (una serie por banco, bajo consumo de RAM)
    bank_features_dict = {}
    for banco in lista_bancos_full:
        if not df_bancarios.empty and f"{banco}_R" in df_bancarios.columns:
            df_banco = df_bancarios[[f"{banco}_R", f"{banco}_D"]].rename(
                columns={f"{banco}_R": "R", f"{banco}_D": "D"}
            )
            # Reindex a peru_bday → excluye feriados PE+USA
            # Drop explícito de días no hábiles adicionales → shift/rolling sin contaminación
            df_banco_feat = df_banco.reindex(_peru_bdays_idx)
            if len(_extra_ts_drop):
                _to_drop = _extra_ts_drop.intersection(df_banco_feat.index)
                if len(_to_drop):
                    df_banco_feat = df_banco_feat.drop(index=_to_drop)
            df_banco_feat = df_banco_feat.dropna()
            bank_features_dict[banco] = build_bank_features(
                df_banco_feat,
                params["lags_cortos"],
                params["lag_semana"],
                params["lag_mes"],
                params["ventanas_vol"],
            )
        else:
            bank_features_dict[banco] = pd.DataFrame()

    # Pre-calcular features de encaje (solo para el banco configurado, ej. BBVA)
    banco_encaje = params.get("banco_encaje", "BBVA")
    # v2: el grupo FOCO también las recibe cuando el banco_encaje cae adentro.
    # Se resuelve contra la lista de bancos del reporte y no por coincidencia de
    # nombres, porque el grupo se llama FOCO_* y nunca va a coincidir por texto
    # con "BBVA". Sin esto, el modelo del foco corre sin encaje y la comparación
    # contra el modelo de BBVA mide la falta de features, no la partición.
    # Una sola regla para los dos bloques de encaje. La primera version usaba
    # "el grupo CONTIENE a banco_encaje", que le daba las features a
    # FOCO_GLOBALES: cinco bancos recibiendo la posicion de encaje de uno solo.
    # destinos_encaje_bbva() exige composicion EXACTA.
    _bancos_con_encaje = destinos_encaje_bbva(
        lista_bancos_full, banco_encaje, reporte_particion, NOMBRE_SISTEMA)
    _destinos_enc_bbva = _bancos_con_encaje
    logger.info(f"  Encaje de {banco_encaje} -> {sorted(_bancos_con_encaje)} "
                f"(politica {POLITICA_ENCAJE_BBVA})")
    df_encaje = load_encaje_data(params)
    encaje_feat = build_encaje_features(df_encaje, peru_bday)

    # Features de avance/exceso encaje desde bbva_encaje_features_modelo.xlsx
    # (generado por aux_encaje_2.py con días calendario — más preciso que EncajeD.xlsx)
    bbva_encaje_feat = load_bbva_encaje_features(params)

    # Pre-calcular features CC+OVN. flujo_neto_sistema = D_SISTEMA - R_SISTEMA,
    # necesario para residuo_ccovn_lag1.
    df_ccovn_raw = load_ccovn_data(params)
    flujo_neto_sistema = pd.Series(dtype=float)
    if not df_bancarios.empty:
        _r = f"{NOMBRE_SISTEMA}_R"
        _d = f"{NOMBRE_SISTEMA}_D"
        if _r in df_bancarios.columns and _d in df_bancarios.columns:
            # Reindex a peru_bday + drop días no hábiles adicionales + dropna
            _sis_clean = df_bancarios[[_r, _d]].reindex(_peru_bdays_idx)
            if len(_extra_ts_drop):
                _to_drop_sis = _extra_ts_drop.intersection(_sis_clean.index)
                if len(_to_drop_sis):
                    _sis_clean = _sis_clean.drop(index=_to_drop_sis)
            _sis_clean = _sis_clean.dropna()
            flujo_neto_sistema = _sis_clean[_d] - _sis_clean[_r]

    # v2: universo de bancos a emparejar contra Saldos_CCOVN.xlsx. Unión de los
    # bancos individuales que quedan modelados tras agrupar_bancos() (lista_bancos,
    # sin Otros_bancos) y los bancos de la partición activa (bancos_foco/resto,
    # que son PRE-agrupación — pueden incluir un banco que ya cayó dentro de
    # Otros_bancos, y aun así hace falta para sumar ccovn_foco/ccovn_resto).
    _bancos_canonicos_ccovn = set(lista_bancos) - {params.get("nombre_otros", "Otros_bancos")}
    if reporte_particion and reporte_particion.get("activa"):
        _bancos_canonicos_ccovn |= set(reporte_particion.get("bancos_foco", []))
        _bancos_canonicos_ccovn |= set(reporte_particion.get("bancos_resto", []))

    df_ccovn_ancho, _rep_match_ccovn = armar_ccovn_ancho(
        df_ccovn_raw, sorted(_bancos_canonicos_ccovn), reporte_particion)
    ccovn_feat = build_ccovn_features(df_ccovn_ancho, peru_bday,
                                      flujo_sistema=flujo_neto_sistema)

    # Pre-computar HMM expanding SOLO para SISTEMA (régimen sistémico, no por banco)
    # El mismo hmm_estado se aplica a todos los bancos individuales.
    hmm_sistema = pd.Series(dtype="Int8")
    if _HMM_DISPONIBLE:
        logger.info("  Pre-computando HMM expanding sobre SISTEMA...")
        bf_sis = bank_features_dict.get(NOMBRE_SISTEMA, pd.DataFrame())
        if not bf_sis.empty and "sigma_22d" in bf_sis.columns and \
                f"{NOMBRE_SISTEMA}_R" in df_bancarios.columns:
            flujo_sis = (df_bancarios[f"{NOMBRE_SISTEMA}_D"]
                         - df_bancarios[f"{NOMBRE_SISTEMA}_R"])
            hmm_sistema = _calcular_hmm_expanding(flujo_sis, bf_sis["sigma_22d"],
                                              primer_ventana_años=6)
            n_etiq = hmm_sistema.notna().sum()
            logger.info(f"    SISTEMA: hmm_estado calculado — {n_etiq:,} días etiquetados")
        else:
            logger.warning("  HMM: no se pudo calcular hmm_estado para SISTEMA (datos insuficientes)")

    # Escribir banco por banco al Parquet de salida — peak RAM = 1 banco a la vez
    pq_writer   = None
    total_filas = 0
    schema_ref  = None

    for banco in lista_bancos_full:
        df_banco_mat = build_feature_matrix(
            banco=banco,
            datos_manuales=datos_manuales,
            macro_features=macro_features,
            bank_features=bank_features_dict,
            peru_bday=peru_bday_ext,   # calendario extendido: excluye también días no hábiles adicionales
            peru_holidays=peru_holidays,
            fechas_elecciones=fechas_elecciones,
            h_min=params["h_min"],
            h_max=params["h_max"],
            hmm_features=hmm_sistema,       # mismo régimen sistémico para todos los bancos
            encaje_features=encaje_feat,    # features de encaje (solo aplican a banco_encaje)
            banco_encaje=banco_encaje,
            bancos_con_encaje=_bancos_con_encaje,
            recibe_encaje_bbva=(banco in _destinos_enc_bbva),
            bbva_encaje_feat=bbva_encaje_feat,  # avance/exceso desde aux_encaje_2 (días calendario)
            ccovn_features=ccovn_feat,      # saldos CC+OVN en BCR, resueltos a propio/contraparte por banco
            reporte_particion=reporte_particion,
            nombre_sistema=NOMBRE_SISTEMA,
            dias_no_habiles_adicionales=sorted(_dias_extra_set) if _dias_extra_set else None,
        )
        if df_banco_mat.empty:
            continue

        table = pa.Table.from_pandas(df_banco_mat, preserve_index=False)

        if pq_writer is None:
            schema_ref = table.schema
            pq_writer  = pq_writer_lib.ParquetWriter(
                ruta_output, schema_ref, compression="snappy"
            )

        # Alinear schema por si hay columnas con tipo diferente en algún banco
        table = table.cast(schema_ref)
        pq_writer.write_table(table)
        total_filas += len(df_banco_mat)

        del df_banco_mat, table
        gc.collect()
        logger.info(f"  {banco}: escrito al Parquet ({total_filas:,} filas acumuladas)")

    if pq_writer is not None:
        pq_writer.close()
    else:
        logger.warning("No se generaron matrices para ningún banco.")
        return pd.DataFrame()

    size_mb = ruta_output.stat().st_size / 1e6
    logger.info(f"\n{'='*60}")
    logger.info("  RESUMEN FINAL DE LA MATRIZ")
    logger.info(f"{'='*60}")
    logger.info(f"  Bancos modelados : {lista_bancos}")
    logger.info(f"  Total filas      : {total_filas:,}")
    logger.info(f"  Columnas         : {len(schema_ref.names)}")
    logger.info(f"  Archivo Parquet  : {ruta_output}  ({size_mb:.1f} MB)")
    logger.info(f"  Para cargar      : pd.read_parquet(r'{ruta_output}')")
    logger.info(f"{'='*60}\n")

    # Retornar DataFrame vacío con el schema correcto como señal de éxito
    return pd.DataFrame(columns=schema_ref.names)


###############################################################################
# Diccionario de variables
###############################################################################
def build_data_dictionary(params):
    """
    Retorna un DataFrame con todas las variables de la matriz,
    su fuente, descripción y rezago/horizonte.
    """
    lags_cortos  = params["lags_cortos"]
    lag_semana   = params["lag_semana"]
    lag_mes      = params["lag_mes"]
    ventanas_vol = params["ventanas_vol"]

    registros = []

    def add(variable, fuente, descripcion, rezago_dias=None, horizonte=None):
        registros.append({
            "variable"    : variable,
            "fuente"      : fuente,
            "descripcion" : descripcion,
            "rezago_dias" : rezago_dias,
            "horizonte"   : horizonte,
        })

    # ── Identificadores ──────────────────────────────────────────────────────
    add("fecha_t",  "Sistema", "Fecha de origen de la predicción",       None, None)
    add("banco",    "Sistema", "Identificador del banco",                 None, None)
    add("h",        "Sistema", "Horizonte de predicción en días hábiles", None, "t+h")
    add("log_h",    "Sistema", "Logaritmo natural del horizonte h",       None, "t+h")

    # ── Features bancarias ────────────────────────────────────────────────────
    # Estructura de información disponible en t:
    #   R (retiros, aviso 2 días hábiles antes):
    #     t-2 → R(t) disponible como R_t0
    #     t-1 → R(t+1) disponible como R_conf_t1
    #     t   → R(t+2) disponible como R_conf_t2
    #   D (depósitos, aviso 1 día hábil antes):
    #     t-1 → D(t) disponible como D_t0
    #     t   → D(t+1) disponible como D_conf_t1

    # Valores de hoy: conocidos por aviso anticipado
    add("R_t0", "Datos bancarios", "Retiro realizado en t — conocido hoy (avisado en t-2)", 0, None)
    add("D_t0", "Datos bancarios", "Depósito realizado en t — conocido hoy (avisado en t-1)", 0, None)

    # Rezagos históricos
    todos_lags = lags_cortos + [lag_semana, lag_mes]
    for l in todos_lags:
        add(f"R_t-{l}", "Datos bancarios", f"Retiro del banco en t-{l}", l, None)
        add(f"D_t-{l}", "Datos bancarios", f"Depósito del banco en t-{l}", l, None)

    for v in ventanas_vol:
        add(f"sigma_R_{v}d", "Datos bancarios", f"Desv. estándar rolling {v}d de retiros (incluye R_t0)",    None, None)
        add(f"sigma_D_{v}d", "Datos bancarios", f"Desv. estándar rolling {v}d de depósitos (incluye D_t0)",  None, None)
        add(f"ma_R_{v}d",    "Datos bancarios", f"Media móvil {v}d de retiros (incluye R_t0)",                None, None)
        add(f"ma_D_{v}d",    "Datos bancarios", f"Media móvil {v}d de depósitos (incluye D_t0)",              None, None)

    add("delta_R", "Datos bancarios", "Variación diaria de retiros: R_t0 - R(t-1)",   0, None)
    add("delta_D", "Datos bancarios", "Variación diaria de depósitos: D_t0 - D(t-1)", 0, None)

    # Volatilidad y nivel del flujo neto (D - R)
    for v in [5, 20]:
        add(f"sigma_flujo_{v}d", "Datos bancarios",
            f"Desv. estándar rolling {v}d del flujo neto D−R (régimen de volatilidad)", None, None)
        add(f"ma_flujo_{v}d",    "Datos bancarios",
            f"Media móvil {v}d del flujo neto D−R (nivel reciente del flujo)", None, None)
    add("sigma_22d", "Datos bancarios",
        "Desv. estándar rolling 22d del flujo neto D−R. Más reactiva que GARCH: "
        "cae rápido al salir de un episodio de stress (sin persistencia artificial).", None, None)
    add("hmm_estado", "Calculado (HMM expanding)",
        f"Estado de régimen oculto: 0=calma, 1=moderado, 2=severo. "
        f"Gaussian HMM {HMM_N_ESTADOS} estados sobre [flujo_neto, sigma_22d]. "
        f"Pre-computado con expanding window sin leakage (mín {HMM_MIN_AÑOS}a). "
        f"NaN en burn-in ({HMM_INICIO}→+{HMM_MIN_AÑOS}a) o si hmmlearn no instalado. "
        f"Si no aporta señal, agregar a FEATURES_EXCLUIR.", None, None)
    add("sigma_flujo_ratio", "Datos bancarios",
        "sigma_flujo_5d / sigma_flujo_20d — ratio de volatilidad corta/larga del flujo. "
        "> 1: régimen de alta vol reciente. NaN si sigma_flujo_20d = 0.", None, None)
    add("flujo_neto_acum_mes", "Datos bancarios",
        "Acumulado del flujo neto D−R desde el primer día hábil del mes hasta t. "
        "Captura la lógica de reversión intramensual: acumulación de depósitos netos "
        "anticipa mayores retiros en cierre de mes, y viceversa.", 0, None)
    add("flujo_neto_sum_5d",  "Datos bancarios",
        "Suma rolling 5dh del flujo neto D−R hasta t (semana). "
        "min_periods=5: NaN si historia < 5 días → imputado por mediana del fold.", 0, None)
    add("flujo_neto_sum_22d", "Datos bancarios",
        "Suma rolling 22dh del flujo neto D−R hasta t (mes). "
        "min_periods=22: NaN si historia < 22 días → imputado por mediana del fold.", 0, None)
    add("flujo_neto_sum_66d", "Datos bancarios",
        "Suma rolling 66dh del flujo neto D−R hasta t (trimestre). "
        "min_periods=66: NaN si historia < 66 días → imputado por mediana del fold.", 0, None)

    # ── Escala por posición del mes (rezagos alineados al cierre) ────────────
    add("esc_neto_min_pos", "Datos bancarios / Posición del mes",
        "PEOR flujo neto D−R (con signo, el más negativo) en los días de los meses "
        "previos con la MISMA posición en el mes que t+h (rezagos 1-4, solo los ya "
        "observables en t). Es el retiro neto extremo esperado en esa posición del "
        "ciclo: el predictor natural de la cola BAJA. Complementa a sigma_22d, que "
        "promedia sobre las 22 ruedas y borra la variación por posición.",
        None, "t+h")
    add("esc_neto_max_pos", "Datos bancarios / Posición del mes",
        "Ídem, el flujo neto MÁS POSITIVO: el depósito neto extremo esperado en esa "
        "posición. Predictor natural de la cola ALTA. Separar por signo evita que "
        "ambas colas reciban el mismo número, que es lo que ocurría con un único "
        "máximo de valores absolutos.", None, "t+h")
    add("esc_retiro_pos", "Datos bancarios / Posición del mes",
        "Ídem sobre R solo. Dirige la cola BAJA, la relevante para el portafolio "
        "de liquidez. En la permutación del modelo aparece por encima de la "
        "versión neta: descomponer el flujo en sus componentes aporta.", None, "t+h")
    add("acum_neto_min_pos", "Datos bancarios / Posición del mes",
        "Flujo neto ACUMULADO dentro del mes hasta el día con la misma posición "
        "que t+h, tomado el más negativo de los 4 meses previos. A diferencia de "
        "esc_neto_min_pos —un extremo de un solo día— es una cantidad de "
        "trayectoria: cuánto se había construido a esa altura del ciclo de "
        "encaje. Anclado en la fecha de origen esta señal no resultó relevante "
        "(ver flujo_neto_acum_mes, excluido).", None, "t+h")
    add("acum_neto_max_pos", "Datos bancarios / Posición del mes",
        "Ídem, el acumulado más positivo. Se separa por signo por la misma razón "
        "que el flujo diario: un máximo absoluto confundiría un ciclo de salida "
        "con uno de entrada.", None, "t+h")
    add("frec_flujo_pos", "Datos bancarios / Posición del mes",
        "Fracción de los últimos 12 meses en que el día con la MISMA posición en "
        "el mes que t+h superó su umbral de |flujo neto| (mediana móvil de 250 "
        "ruedas hasta t, point-in-time). Mide RECURRENCIA en vez de magnitud: una "
        "posición cargada en 8 de 12 meses es distinta de una cargada en 1 con un "
        "valor enorme. Con 12 rezagos cubre los 74 horizontes.", None, "t+h")
    add("capacidad_retiro_th", "aux_encaje_2 / Calendario",
        "MAX(0, encaje_diario_lag1 − min_por_dia_proyectado(t+h)) + overnight_lag1: "
        "cuánto puede retirar el banco en t+h sin incumplir el 100% del exigible "
        "al cierre del período. Reemplaza a presion_deadline_th/_t (eliminadas "
        "esta sesión), que medían la obligación de depositar en vez de la "
        "capacidad de retirar y cuyo nivel era casi constante dentro del mes "
        "por construcción algebraica. Proyecta EncajeAcumMes hacia t+h asumiendo "
        "que el banco mantiene el ritmo diario de encaje_diario_lag1 desde t-1 "
        "— único supuesto, explícito y acotado (el déficit proyectado nunca es "
        "negativo). Sin clip artificial de magnitud: cuando el banco va "
        "adelantado min_por_dia se satura en 0 y la capacidad reportada es el "
        "saldo completo. NaN si t+h cae en otro período de encaje —el cómputo "
        "se reinicia al cierre y no hay forma de proyectar el requerimiento del "
        "mes siguiente sin sus propios datos— o si falta el dato de origen "
        "(archivo desde 2016-07).",
        None, "t+h")
    add("esc_deposito_pos", "Datos bancarios / Posición del mes",
        "Ídem sobre D solo. Dirige la cola ALTA, donde el diagnóstico del oráculo "
        "encontró la peor calibración (factor de ensanchamiento necesario hasta "
        "3.1x). No es redundante con las otras dos: el estadístico es un máximo de "
        "valores absolutos y max|D-R| no se deduce de max|D| y max|R|.", None, "t+h")
    add("esc_neto_max_pos_ap", "Datos bancarios / Posición del mes",
        "Igual que esc_neto_max_pos pero con referencia='inicio': busca el día "
        "en meses previos a la MISMA distancia del PRIMER día hábil del mes, no "
        "del último. Apunta a la reversión observada al abrir el mes (depósito "
        "fuerte seguido de retiro fuerte días después) — el ancla de cierre es "
        "ciega a esto porque con meses de 19-23 días hábiles, 'días desde el "
        "inicio' y 'días hasta el cierre' no son la misma posición calendario. "
        "Solo el extremo positivo: el negativo cerca de la apertura ya lo cubre "
        "esc_neto_min_pos razonablemente.", None, "t+h")
    add("esc_deposito_pos_ap", "Datos bancarios / Posición del mes",
        "Igual que esc_deposito_pos pero con referencia='inicio'. Motivado por "
        "aux_importancia_calendario.py: dias_desde_cierre_mes (crudo, mismo "
        "principio de anclar a la apertura) domina la importancia de q95/q99, "
        "mientras dias_al_cierre_mes domina q01/q05 — la familia *_pos hasta "
        "ahora solo usaba el ancla de cierre para las 7 columnas activas.",
        None, "t+h")
    if GUARDAR_N_LAGS_POS:
        add("n_lags_pos", "Datos bancarios / Posición del mes",
            "Cuántos rezagos de posición estaban disponibles (0-4). Diagnóstico: "
            "es función de h, así que dentro de un modelo es casi constante y no "
            "se guarda por defecto (ver GUARDAR_N_LAGS_POS).", None, "t+h")

    # ── Confirmados futuros ───────────────────────────────────────────────────
    # Entrenamiento: proxy histórico = valor realizado en t+1/t+2 (shift negativo)
    # Producción:    valor real del correo de aviso (override desde confirmados.xlsx)
    add("R_conf_t1", "Datos bancarios / Confirmados operativos", "R(t+1) — proxy histórico o aviso real (2d anticipación)", -1, "t+1")
    add("R_conf_t2", "Datos bancarios / Confirmados operativos", "R(t+2) — proxy histórico o aviso real (2d anticipación)",  0, "t+2")
    add("D_conf_t1", "Datos bancarios / Confirmados operativos", "D(t+1) — proxy histórico o aviso real (1d anticipación)",  0, "t+1")

    # ── Features macroeconómicas (observadas en t) ───────────────────────────
    add("VIX",              "Yahoo Finance",  "Índice de volatilidad implícita S&P 500",          0, None)
    add("delta_VIX",        "Yahoo Finance",  "Variación diaria del VIX",                         1, None)
    add("VIX_ma22",         "Yahoo Finance",  "Media móvil 22d del VIX",                          None, None)
    add("TC_PEN_USD",       "BCRP Add-In",    "Tipo de cambio PEN/USD (promedio compra/venta)",   0, None)
    add("delta_TC",         "BCRP Add-In",    "Variación diaria del tipo de cambio",              1, None)
    add("tc_vol_5d",        "BCRP Add-In",    "Volatilidad rolling 5d de retornos del TC",        None, None)
    add("tc_vol_22d",       "BCRP Add-In",    "Volatilidad rolling 22d de retornos del TC",       None, None)
    add("tc_vol_ratio",     "Calculado",      "tc_vol_5d / tc_vol_22d — ratio de vol cambiaria corta/larga. "
        "> 1: estrés cambiario reciente. NaN si tc_vol_22d = 0.",               None, None)
    add("garch_vol_tc",     "BCRP Add-In",    "Volatilidad condicional GARCH(1,1) de retornos log del TC PEN/USD — detecta estrés cambiario", None, None)
    add("EMBI_PERU",        "BCRP Add-In",    "EMBI Perú (riesgo país)",                          0, None)
    add("delta_EMBI",       "BCRP Add-In",    "Variación diaria del EMBI Perú",                   1, None)
    add("garch_vol_embi",   "BCRP Add-In",    "Volatilidad condicional GARCH(1,1) de cambios diarios del EMBI Perú — detecta estrés político", None, None)
    add("TASA_REF_BCRP",    "BCRP Add-In",    "Tasa de referencia del BCRP",                      0, None)
    add("FED_FUNDS",        "FRED API",       "Tasa de política monetaria de la Fed",             0, None)
    add("diferencial_tasas","Calculado",      "TASA_REF_BCRP - FED_FUNDS",                        0, None)
    add("T10Y",             "Yahoo Finance",  "Rendimiento del bono del Tesoro EE.UU. a 10 años", 0, None)

    # ── Features FFD — diferenciación fraccional (López de Prado Cap. 5) ──────
    add("EMBI_PERU_frac",   "BCRP Add-In (FFD)",  "EMBI Perú diferenciado fraccionalmente (d mínimo que elimina raíz unitaria). Preserva memoria de largo plazo; más estacionario que el nivel.", None, None)
    add("T10Y_frac",        "Yahoo Finance (FFD)", "UST 10Y diferenciado fraccionalmente. Alternativa a delta_T10Y con menor pérdida de memoria.", None, None)
    add("CDS_PERU_5Y_frac", "Bloomberg (FFD)",     "CDS Perú 5Y diferenciado fraccionalmente (d_opt calibrado en TRAIN).", None, None)
    add("COPPER_frac",      "Bloomberg (FFD)",     "Cobre LME diferenciado fraccionalmente. Indicador de ciclo global.", None, None)
    add("VIX_frac",         "Yahoo Finance (FFD)", "VIX diferenciado fraccionalmente (d≈0 si ya es estacionario; preserva estructura).", None, None)

    # ── Features estacionales (en t+h — fecha futura, siempre conocidas) ─────
    add("dias_al_cierre_mes",    "Calendario", "Días hábiles restantes hasta fin de mes en t+h",          None, "t+h")
    add("dias_desde_cierre_mes", "Calendario", "Días hábiles transcurridos desde inicio de mes en t+h",   None, "t+h")
    add("dias_al_cierre_trim",   "Calendario", "Días hábiles restantes hasta fin de trimestre en t+h",     None, "t+h")
    add("dias_cal_al_cierre_mes","Calendario", "Días CALENDARIO (no hábiles) restantes hasta fin de mes en "
        "t+h — reloj regulatorio del encaje, distinto de dias_al_cierre_mes",  None, "t+h")
    add("es_mes_cierre_trim",    "Calendario", "1 si el mes de t+h es Mar/Jun/Sep/Dic — cierre de trimestre, "
        "donde se concentra el retiro de sobreencaje observado",              None, "t+h")
    add("pos_en_mes",            "Calendario", "Posición del día hábil dentro del mes (1=primero)",        None, "t+h")
    add("total_bdays_mes",       "Calendario", "Total de días hábiles del mes de t+h",                     None, "t+h")
    add("is_penult_bday_trim",   "Calendario", "1 si t+h es penúltimo día hábil del trimestre",            None, "t+h")
    add("is_ultimo_bday_trim",   "Calendario", "1 si t+h es último día hábil del trimestre",               None, "t+h")
    add("is_1er_bday_trim",      "Calendario", "1 si t+h es primer día hábil del trimestre",               None, "t+h")
    add("is_2do_bday_trim",      "Calendario", "1 si t+h es segundo día hábil del trimestre",              None, "t+h")
    add("is_3er_bday_trim",      "Calendario", "1 si t+h es tercer día hábil del trimestre",               None, "t+h")
    add("dia_semana",            "Calendario", "Día de la semana (0=lunes, 4=viernes)",                    None, "t+h")
    add("mes",                   "Calendario", "Mes del año (1–12)",                                       None, "t+h")
    add("is_quincena",           "Calendario", "1 si t+h es día 15 o último hábil del mes",                None, "t+h")
    add("is_cierre_encaje",      "Calendario", "1 si t+h está en los últimos 2 días hábiles del mes",      None, "t+h")
    add("is_fin_anio",           "Calendario", "1 si t+h es 28–31 de diciembre",                           None, "t+h")
    add("is_pre_feriado",        "Calendario / holidays.Peru + holidays.US", "1 si el día siguiente a t+h es feriado PE o US",  None, "t+h")
    add("is_post_feriado",       "Calendario / holidays.Peru + holidays.US", "1 si el día anterior a t+h es feriado PE o US",   None, "t+h")
    add("is_pre_eleccion",       "Calendario", "1 si t+h está dentro de los 7 días previos a elecciones presidenciales",    None, "t+h")
    add("is_post_eleccion",      "Calendario", "1 si t+h está dentro de los 7 días posteriores a elecciones presidenciales", None, "t+h")
    add("mes_sin",               "Calendario", "sin(2π·mes/12) — proyección cíclica del mes (P=12)",                         None, "t+h")
    add("mes_cos",               "Calendario", "cos(2π·mes/12) — proyección cíclica del mes (P=12)",                         None, "t+h")
    add("dias_sem_sin",          "Calendario", "sin(2π·dia_sem/5) — ciclo semanal en días hábiles (P=5)",                    None, "t+h")
    add("dias_sem_cos",          "Calendario", "cos(2π·dia_sem/5) — ciclo semanal en días hábiles (P=5)",                    None, "t+h")
    add("dias_al_cierre_mes_sin","Calendario", "sin(2π·pos_mes/P_mes) — ciclo mensual hábil dinámico",                       None, "t+h")
    add("dias_al_cierre_mes_cos","Calendario", "cos(2π·pos_mes/P_mes) — ciclo mensual hábil dinámico",                       None, "t+h")
    add("dias_al_cierre_trim_sin","Calendario","sin(2π·pos_trim/P_trim) — ciclo trimestral hábil dinámico",                  None, "t+h")
    add("dias_al_cierre_trim_cos","Calendario","cos(2π·pos_trim/P_trim) — ciclo trimestral hábil dinámico",                  None, "t+h")
    add("dias_al_cierre_anio_sin","Calendario","sin(2π·pos_anio/P_anio) — ciclo anual hábil dinámico",                       None, "t+h")
    add("dias_al_cierre_anio_cos","Calendario","cos(2π·pos_anio/P_anio) — ciclo anual hábil dinámico",                       None, "t+h")
    add("elec_sin",              "Calendario", "sin(2π·bd_desde_elec/1260) — ciclo electoral 5 años (solo 1ras vueltas)",    None, "t+h")
    add("elec_cos",              "Calendario", "cos(2π·bd_desde_elec/1260) — ciclo electoral 5 años (solo 1ras vueltas)",    None, "t+h")

    # ── Features de encaje BBVA (EncajeD.xlsx, rezago 1 día) ─────────────────
    _benc = params.get("banco_encaje", "BBVA")
    add("encaje_lag1",         f"EncajeD / {_benc}", "Caja + Cta Cte BCR en t-1 (M USD). Solo banco configurado en banco_encaje.", 1, "t")
    add("exceso_lag1",         f"EncajeD / {_benc}", "Encaje(t-1) − Exigible(t-1): exceso acumulado disponible (M USD).", 1, "t")
    add("faltante_lag1",       f"EncajeD / {_benc}", "Exigible_total_mes − encaje_acum(t-1): saldo aún pendiente de acumular en el mes (M USD).", 1, "t")
    add("techo_10h",           f"EncajeD / {_benc}", "CC+ON al día hábil 10 antes del cierre del mes, rezagado 1 día. Techo operativo de retiros.", 1, "t")
    add("techo_restante_lag1",   f"EncajeD / {_benc}", "techo_10h − retiro_acumulado_mes(t-1): presupuesto de retiro aún disponible (M USD).", 1, "t")
    add("proporcion_usada",      f"EncajeD / {_benc}", "retiro_acum_mes(t-1) / techo_10h: fracción del techo ya utilizada (0-1+).", 1, "t")
    add("avance_mes_lag1",       f"aux_encaje_2 / {_benc}", "EncajeAcumMes(t-1) / ExigibleTotalMes_est(t-1): fracción del req. mensual cubierta. Días calendario → ffill a días hábiles.", 1, "t")
    add("exceso_abs_lag1",       f"aux_encaje_2 / {_benc}", "MAX(0, encaje(t-1)−encaje_min_por_dia)+overnight(t-1): capacidad de retiro real sin incumplir 100% al cierre (M USD). Días calendario.", 1, "t")
    add("exceso_dia_lag1",       f"aux_encaje_2 / {_benc}", "exceso_abs(t-1)/dias_restantes(t-1): presión temporal — exceso ponderado por urgencia de fin de periodo. Q5 media: −290M.", 1, "t")
    add("encaje_ovn_lag1",       f"aux_encaje_2 / {_benc}", "overnight+cta_cte+caja en t-1 (M USD): posición total BCRP. Techo físico máximo de retiro posible.", 1, "t")
    add("ratio_ovn_total_lag1",  f"aux_encaje_2 / {_benc}", "overnight(t-1)/encaje_ovn(t-1): alto = banco aún no inició retiro (fondos en overnight sin mover).", 1, "t")

    # ── Features CC+OVN en BCR (Saldos_CCOVN.xlsx, rezago 1 día) ─────────────
    # v2: "propio"/"contraparte" se resuelven por entidad, no son columnas
    # fijas de BBVA. Ver resolver_ccovn_lados() — SISTEMA no tiene contraparte;
    # FOCO_x y RESTO_x son contraparte mutua; un banco individual solo hereda
    # contraparte si su composición coincide EXACTO con un lado de la partición
    # activa (ej. BBVA cuando particion_activa="bbva").
    add("ccovn_sistema_lag1",     "Saldos_CCOVN",
        "Saldo total CC+OVN del sistema en el BCR en t-1 (USD). Disponible desde 2010. "
        "Igual para todas las entidades — no depende de la partición.",
        1, "t")
    add("ccovn_propio_lag1",      "Saldos_CCOVN",
        "Saldo CC+OVN propio de la entidad modelada en t-1 (USD). Antes "
        "ccovn_bbva_lag1: era el saldo de BBVA para TODAS las entidades por "
        "igual. Ahora es el saldo de quien se está modelando — para SISTEMA "
        "coincide con ccovn_sistema_lag1, para FOCO_BBVA es BBVA, para "
        "RESTO_BBVA es el resto del sistema.",
        1, "t")
    add("var_ccovn_sistema_lag1", "Saldos_CCOVN",
        "Variación diaria del saldo CC+OVN del sistema en t-1 (USD).",
        1, "t")
    add("var_ccovn_propio_lag1",  "Saldos_CCOVN",
        "Variación diaria del saldo CC+OVN propio en t-1 (USD). Antes var_ccovn_bbva_lag1.",
        1, "t")
    add("ccovn_contraparte_lag1", "Saldos_CCOVN",
        "Saldo CC+OVN del OTRO lado de la partición activa en t-1 (USD). Nuevo "
        "en v2 — no existía en v1, que no tenía noción de 'el resto'. NaN para "
        "SISTEMA y para bancos individuales fuera de la partición activa.",
        1, "t")
    add("var_ccovn_contraparte_lag1", "Saldos_CCOVN",
        "Variación diaria del saldo de la contraparte en t-1 (USD). Nuevo en v2.",
        1, "t")
    add("share_propio_lag1",        "Saldos_CCOVN",
        "ccovn_propio(t-1) / ccovn_sistema(t-1): participación de la entidad "
        "modelada en el sistema en t-1. Antes bbva_share_lag1 — mismo estado "
        "(excluido), solo cambió el nombre.",
        1, "t")
    add("var_ccovn_propio_exceso_lag1", "Saldos_CCOVN",
        "Δsaldo_propio(t-1) − share_propio(t-1)×Δsaldo_sistema(t-1): componente "
        "idiosincrático de la entidad modelada en la variación diaria del "
        "saldo. Ortogonal a la variación sistémica. Antes "
        "var_ccovn_bbva_exceso_lag1 — mismo estado (excluido), solo cambió el nombre.",
        1, "t")
    add("ccovn_vs_dia_mes_lag1",  "Saldos_CCOVN",
        "saldo_sistema(t-1) − media_historica[rango_dia_habil_del_mes]: "
        "desviación del saldo respecto al nivel estacional esperado para ese "
        "puesto del mes hábil. Elimina la estacionalidad intramonth. "
        "Basada en el sistema, ajena a la partición.",
        1, "t")
    add("residuo_ccovn_lag1",     "Saldos_CCOVN + Datos bancarios",
        "Δsaldo_sistema(t-1) − flujo_neto_sistema(t-1): error de la identidad "
        "Δsaldo≈flujo. Señal de reversión para horizontes cortos. "
        "Solo disponible si flujo_neto_sistema está en bancarios.",
        1, "t")

    # ── Target ────────────────────────────────────────────────────────────────
    add("target", "Datos bancarios", "Flujo neto = D(b, t+h) - R(b, t+h). NaN si no hay datos.", None, "t+h")

    df_dict = pd.DataFrame(registros)

    # Marcar si la variable está activa en el parquet o fue excluida
    df_dict["en_modelo"] = df_dict["variable"].apply(
        lambda v: "No" if v in FEATURES_EXCLUIR else "Sí"
    )

    return df_dict


###############################################################################
# Datos demo — se usa cuando PARAMS["usar_datos_demo"] = True
###############################################################################
def generate_demo_data(params):
    """
    Genera datos bancarios y macroeconómicos sintéticos para pruebas.
    Simula 5 bancos de distintos tamaños para activar la agrupación automática.
    """
    logger.info("*** MODO DEMO: generando datos sintéticos ***")
    np.random.seed(42)

    fechas = pd.bdate_range(
        start=params["fecha_inicio_historico"],
        end="2024-12-31",
    )
    n = len(fechas)

    # Cinco bancos con escalas muy distintas → banco_4 y banco_5 quedarán en Otros
    escalas = {
        "banco_1": 500,
        "banco_2": 350,
        "banco_3": 200,
        "banco_4": 30,
        "banco_5": 20,
    }

    df_bancarios = pd.DataFrame(index=fechas)
    df_bancarios.index.name = "fecha"
    for banco, escala in escalas.items():
        tendencia = np.linspace(0, escala * 0.1, n)
        ruido_R = np.random.normal(0, escala * 0.15, n)
        ruido_D = np.random.normal(0, escala * 0.15, n)
        df_bancarios[f"{banco}_R"] = escala + tendencia + ruido_R
        df_bancarios[f"{banco}_D"] = escala + tendencia + ruido_D

    # Macro sintética con tendencias plausibles
    df_macro = pd.DataFrame(index=fechas)
    df_macro["VIX"]           = np.clip(20 + np.cumsum(np.random.normal(0, 0.3, n)), 10, 80)
    df_macro["TC_PEN_USD"]    = np.clip(3.3 + np.cumsum(np.random.normal(0, 0.005, n)), 3.0, 4.5)
    df_macro["T10Y"]          = np.clip(3.5 + np.cumsum(np.random.normal(0, 0.02, n)) * 0.05, 1.0, 6.0)
    df_macro["FED_FUNDS"]     = np.clip(1.0 + np.cumsum(np.random.normal(0, 0.01, n)) * 0.03, 0.0, 6.0)
    df_macro["EMBI_PERU"]     = np.clip(130 + np.cumsum(np.random.normal(0, 1.0, n)), 80, 400)
    df_macro["TASA_REF_BCRP"] = np.clip(4.0 + np.cumsum(np.random.normal(0, 0.01, n)) * 0.02, 1.5, 8.0)

    datos_manuales = {
        "bancarios"   : df_bancarios,
        "confirmados" : pd.DataFrame(columns=["banco", "R_conf_t1", "R_conf_t2", "D_conf_t1"]),
        "intervencion": pd.Series(dtype=float),
    }
    return datos_manuales, df_macro


###############################################################################
# Flujo principal
###############################################################################
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("  SISTEMA DE PREDICCIÓN DE LIQUIDEZ EN MONEDA EXTRANJERA")
    logger.info("  Construcción de Matriz de Features")
    logger.info(f"  Fecha de ejecución: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 60)

    # 1. Calendario
    peru_bday, peru_holidays, fechas_elecciones = build_peru_calendar(
        años=PARAMS["años_calendario"],
    )

    # 2. Datos manuales + macro
    if PARAMS.get("usar_datos_demo"):
        datos_manuales, macro_raw = generate_demo_data(PARAMS)
    else:
        datos_manuales = load_manual_data(PARAMS)
        macro_raw      = download_external_series(PARAMS)

    # 2b. Partición del sistema (v2) — ANTES de agrupar.
    # El orden no es cosmético: agrupar_bancos() elimina las columnas de los
    # bancos chicos después de sumarlas en Otros_bancos, y varios bancos globales
    # (Deutsche, ICBC, Bank of China, BCI) están por debajo del umbral del 1%.
    # Partir después dejaría a esos bancos del lado equivocado sin ningún aviso.
    datos_manuales["bancarios"], reporte_particion = aplicar_particion(
        datos_manuales["bancarios"],
        PARAMS.get("particion_activa"),
        nombre_otros=PARAMS["nombre_otros"],
    )

    # 3. Agrupación de bancos pequeños → Otros_bancos
    df_bancarios_agrupado, lista_bancos, reporte_agrupacion = agrupar_bancos(
        datos_manuales["bancarios"],
        PARAMS["umbral_banco_pequeño_pct"],
        PARAMS["bancos_otros"],
        PARAMS["nombre_otros"],
        excluir=columnas_derivadas(reporte_particion),
    )
    datos_manuales["bancarios"] = df_bancarios_agrupado

    if not lista_bancos:
        logger.warning(
            "No hay bancos en lista_bancos. "
            "Ejecutando con banco sintético para validar estructura."
        )
        lista_bancos = ["banco_demo"]

    # 4. Features macroeconómicas
    macro_features = build_macro_features(macro_raw)

    # 5. Matriz completa
    matriz = build_full_matrix(
        params=PARAMS,
        datos_manuales=datos_manuales,
        lista_bancos=lista_bancos,
        macro_features=macro_features,
        peru_bday=peru_bday,
        peru_holidays=peru_holidays,
        fechas_elecciones=fechas_elecciones,
        reporte_particion=reporte_particion,
    )

    ruta_out = Path(PARAMS.get("ruta_output", "matriz_features.parquet"))
    if ruta_out.exists() and ruta_out.stat().st_size > 0:
        print(f"\nMatriz generada correctamente → {ruta_out}")
    else:
        print("\nMatriz vacía. Verifique la configuración de PARAMS y los archivos de datos.")

    # Diccionario de variables
    data_dict = build_data_dictionary(PARAMS)
    df_activas  = data_dict[data_dict["en_modelo"] == "Sí"].drop(columns="en_modelo").reset_index(drop=True)
    df_excluidas = data_dict[data_dict["en_modelo"] == "No"].drop(columns="en_modelo").reset_index(drop=True)

    print(f"\n{'='*70}")
    print("  DICCIONARIO DE VARIABLES — en modelo")
    print(f"{'='*70}")
    print(df_activas.to_string(index=False))
    print(f"\nTotal en modelo: {len(df_activas)}  |  Excluidas: {len(df_excluidas)}")

    if len(df_excluidas):
        print(f"\n  Variables excluidas (FEATURES_EXCLUIR): "
              + ", ".join(df_excluidas["variable"].tolist()))

    # Exportar diccionario a Excel: dos pestañas
    ruta_dict = Path(PARAMS.get("ruta_diccionario", r"1. Data\Clean\diccionario_variables.xlsx"))
    try:
        ruta_dict.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(ruta_dict, engine="openpyxl") as writer:
            df_activas.to_excel(writer, index=False, sheet_name="Variables_modelo")
            df_excluidas.to_excel(writer, index=False, sheet_name="Variables_excluidas")

            # Resaltar pestaña de excluidas con fondo naranja en celdas de datos
            from openpyxl.styles import PatternFill
            fill_naranja = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
            ws = writer.sheets["Variables_excluidas"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.fill = fill_naranja

        logger.info(f"  Diccionario exportado a: {ruta_dict} "
                    f"(2 pestañas: Variables_modelo={len(df_activas)}, "
                    f"Variables_excluidas={len(df_excluidas)})")
    except Exception as e:
        logger.warning(f"  No se pudo exportar diccionario: {e}")
