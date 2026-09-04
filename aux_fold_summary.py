# -*- coding: utf-8 -*-
"""
aux_fold_summary.py
===================
Genera un Excel con tablas resumen de la estructura de folds para las 4
configuraciones del walk-forward CV (step005_walk_forward_cv_3.py):

    Expanding × {True, False}  ×  Val_años × {0.5, 1.0}

Por cada configuración se muestra una tabla con una fila por fold y columnas
que detallan fechas, duración y número de filas de cada etapa:
    Warm-up | Train efectivo | Embargo T→V | VAL | Embargo V→T | TEST
"""

import pathlib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Rutas ──────────────────────────────────────────────────────────────────────
BASE = pathlib.Path(
    r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente"
)
RUTA_MATRIZ = BASE / "1. Data" / "Clean" / "matriz_features.parquet"
RUTA_OUT    = BASE / "2. Output" / "aux_fold_summary.xlsx"

BANCO = "SISTEMA"
H_REF = 2        # horizonte de referencia para contar filas (evita duplicados por h)

# ── Parámetros fijos de step005 ────────────────────────────────────────────────
VENTANA_TRAIN_AÑOS = 5
VENTANA_TEST_AÑOS  = 1
PASO_AÑOS          = 1
PURGE_DIAS_HAB     = 75   # embargo train → val
PURGE_VAL_TEST     = 75   # embargo val   → test
BURN_IN_DIAS_HAB   = 22   # warm-up al inicio de train

# ── 4 configuraciones ──────────────────────────────────────────────────────────
CONFIGS = [
    {"expanding": True,  "val_años": 0.5, "label": "Expanding | Val 6m"},
    {"expanding": True,  "val_años": 1.0, "label": "Expanding | Val 12m"},
    {"expanding": False, "val_años": 0.5, "label": "Rolling   | Val 6m"},
    {"expanding": False, "val_años": 1.0, "label": "Rolling   | Val 12m"},
]

# ── Paleta de colores por etapa ────────────────────────────────────────────────
COLORES = {
    "warmup":     "FFF2CC",   # amarillo suave
    "train":      "DDEEFF",   # azul suave
    "embargo_tv": "F4CCCC",   # rojo suave
    "val":        "D9EAD3",   # verde suave
    "embargo_vt": "F4CCCC",   # rojo suave
    "test":       "E6D0F0",   # morado suave
    "header":     "1F4E79",   # azul oscuro para encabezados
    "subheader":  "2E75B6",   # azul medio para sub-encabezados
}


# ─────────────────────────────────────────────────────────────────────────────
# Lógica de generación de folds (replicada de step005 para autonomía)
# ─────────────────────────────────────────────────────────────────────────────

def generar_folds(fechas, expanding, val_años):
    folds    = []
    f_min    = fechas.min()
    f_max    = fechas.max()
    fold_idx = 0

    while True:
        if expanding:
            train_start = f_min
            train_end   = f_min + pd.DateOffset(
                months=int(round((VENTANA_TRAIN_AÑOS + fold_idx * PASO_AÑOS) * 12)))
        else:
            train_start = f_min + pd.DateOffset(
                months=int(round(fold_idx * PASO_AÑOS * 12)))
            train_end   = train_start + pd.DateOffset(
                months=int(round(VENTANA_TRAIN_AÑOS * 12)))

        val_start  = train_end  + pd.offsets.BusinessDay(PURGE_DIAS_HAB)
        val_end    = val_start  + pd.DateOffset(months=int(round(val_años * 12)))
        test_start = val_end    + pd.offsets.BusinessDay(PURGE_VAL_TEST)
        test_end   = test_start + pd.DateOffset(months=int(round(VENTANA_TEST_AÑOS * 12)))

        if test_end > f_max or train_end >= f_max or val_start >= f_max:
            break

        burn_cutoff = train_start + pd.offsets.BusinessDay(BURN_IN_DIAS_HAB)

        def _n(lo, hi):
            return int(((fechas >= lo) & (fechas <= hi)).sum())

        def _nbday(lo, hi):
            rng = pd.bdate_range(lo, hi)
            return len(rng)

        n_warmup    = _n(train_start, burn_cutoff - pd.Timedelta(days=1))
        n_train_eff = _n(burn_cutoff, train_end)
        n_emb_tv    = _nbday(train_end + pd.Timedelta(days=1),
                             val_start  - pd.Timedelta(days=1))
        n_val       = _n(val_start, val_end)
        n_emb_vt    = _nbday(val_end   + pd.Timedelta(days=1),
                             test_start - pd.Timedelta(days=1))
        n_test      = _n(test_start, test_end)

        if n_train_eff < 60 or n_val < 10 or n_test < 10:
            fold_idx += 1
            continue

        folds.append({
            "fold"         : fold_idx + 1,
            # Warm-up
            "wu_start"     : train_start,
            "wu_end"       : burn_cutoff - pd.Timedelta(days=1),
            "wu_dias"      : BURN_IN_DIAS_HAB,
            "wu_filas"     : n_warmup,
            # Train efectivo
            "tr_start"     : burn_cutoff,
            "tr_end"       : train_end,
            "tr_meses"     : round((train_end - burn_cutoff).days / 30.4, 1),
            "tr_filas"     : n_train_eff,
            # Embargo train → val
            "etv_dias_hab" : n_emb_tv,
            # VAL
            "val_start"    : val_start,
            "val_end"      : val_end,
            "val_meses"    : round(val_años * 12, 0),
            "val_filas"    : n_val,
            # Embargo val → test
            "evt_dias_hab" : n_emb_vt,
            # TEST
            "test_start"   : test_start,
            "test_end"     : test_end,
            "test_meses"   : round(VENTANA_TEST_AÑOS * 12, 0),
            "test_filas"   : n_test,
        })
        fold_idx += 1

    return folds


# ─────────────────────────────────────────────────────────────────────────────
# Construcción del DataFrame resumen por configuración
# ─────────────────────────────────────────────────────────────────────────────

def folds_a_df(folds):
    rows = []
    for f in folds:
        rows.append({
            "Fold"               : f["fold"],
            # Warm-up
            "WU inicio"          : f["wu_start"].date(),
            "WU fin"             : f["wu_end"].date(),
            "WU días háb."       : f["wu_dias"],
            "WU filas"           : f["wu_filas"],
            # Train efectivo
            "Train inicio"       : f["tr_start"].date(),
            "Train fin"          : f["tr_end"].date(),
            "Train meses"        : f["tr_meses"],
            "Train filas"        : f["tr_filas"],
            # Embargo Train→Val
            "Embargo T→V días h.": f["etv_dias_hab"],
            # VAL
            "Val inicio"         : f["val_start"].date(),
            "Val fin"            : f["val_end"].date(),
            "Val meses"          : int(f["val_meses"]),
            "Val filas"          : f["val_filas"],
            # Embargo Val→Test
            "Embargo V→T días h.": f["evt_dias_hab"],
            # TEST
            "Test inicio"        : f["test_start"].date(),
            "Test fin"           : f["test_end"].date(),
            "Test meses"         : int(f["test_meses"]),
            "Test filas"         : f["test_filas"],
        })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Grupos de columnas para colorear
# ─────────────────────────────────────────────────────────────────────────────

GRUPOS = [
    ("Warm-up",          ["WU inicio", "WU fin", "WU días háb.", "WU filas"],
                          "warmup"),
    ("Train efectivo",   ["Train inicio", "Train fin", "Train meses", "Train filas"],
                          "train"),
    ("Embargo T→V",      ["Embargo T→V días h."],
                          "embargo_tv"),
    ("VAL (Optuna)",     ["Val inicio", "Val fin", "Val meses", "Val filas"],
                          "val"),
    ("Embargo V→T",      ["Embargo V→T días h."],
                          "embargo_vt"),
    ("TEST (OOS)",       ["Test inicio", "Test fin", "Test meses", "Test filas"],
                          "test"),
]


# ─────────────────────────────────────────────────────────────────────────────
# Escritura a Excel con formato
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)

def _border_thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def _border_medium():
    s = Side(style="medium", color="555555")
    return Border(left=s, right=s, top=s, bottom=s)


def escribir_hoja(ws, df, config_label, expanding, val_años):
    all_cols = ["Fold"] + [c for g in GRUPOS for c in g[1]]

    thin   = _border_thin()
    medium = _border_medium()

    # ── Fila 1: título de la configuración ────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(all_cols))
    c = ws.cell(1, 1, config_label)
    c.font      = _font(bold=True, color="FFFFFF", size=12)
    c.fill      = _fill(COLORES["header"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Fila 2: parámetros clave ───────────────────────────────────────────────
    params_txt = (
        f"Expanding={'Sí' if expanding else 'No'}  |  "
        f"Train inicial={VENTANA_TRAIN_AÑOS} años  |  "
        f"Val={int(val_años*12)} meses  |  "
        f"Test={int(VENTANA_TEST_AÑOS*12)} meses  |  "
        f"Paso={PASO_AÑOS} año  |  "
        f"Purge T→V={PURGE_DIAS_HAB} dh  |  "
        f"Purge V→T={PURGE_VAL_TEST} dh  |  "
        f"Burn-in={BURN_IN_DIAS_HAB} dh"
    )
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=len(all_cols))
    c = ws.cell(2, 1, params_txt)
    c.font      = _font(bold=False, color="FFFFFF", size=9)
    c.fill      = _fill(COLORES["subheader"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Fila 3: encabezado de grupos ──────────────────────────────────────────
    col_cursor = 2   # col 1 = "Fold"
    # Celda Fold (abarca filas 3 y 4)
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    c = ws.cell(3, 1, "Fold")
    c.font      = _font(bold=True, color="FFFFFF", size=10)
    c.fill      = _fill(COLORES["header"])
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for grp_name, grp_cols, color_key in GRUPOS:
        n = len(grp_cols)
        ws.merge_cells(start_row=3, start_column=col_cursor,
                       end_row=3, end_column=col_cursor + n - 1)
        c = ws.cell(3, col_cursor, grp_name)
        c.font      = _font(bold=True, color="000000", size=10)
        c.fill      = _fill(COLORES[color_key])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = medium
        col_cursor += n

    # ── Fila 4: encabezado de columnas ────────────────────────────────────────
    col_cursor = 2
    for grp_name, grp_cols, color_key in GRUPOS:
        for col_name in grp_cols:
            c = ws.cell(4, col_cursor, col_name)
            c.font      = _font(bold=True, color="000000", size=9)
            c.fill      = _fill(COLORES[color_key])
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border    = thin
            col_cursor += 1

    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 28

    # ── Filas de datos (desde fila 5) ─────────────────────────────────────────
    for row_i, (_, row) in enumerate(df.iterrows(), start=5):
        # Col 1: número de fold
        c = ws.cell(row_i, 1, int(row["Fold"]))
        c.font      = _font(bold=True, size=10)
        c.fill      = _fill(COLORES["header"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = medium

        col_cursor = 2
        for grp_name, grp_cols, color_key in GRUPOS:
            for col_name in grp_cols:
                val = row[col_name]
                c   = ws.cell(row_i, col_cursor, val)
                c.fill      = _fill(COLORES[color_key])
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = thin
                c.font      = _font(size=9)
                col_cursor += 1

        ws.row_dimensions[row_i].height = 15

    # ── Fila de totales / resumen ──────────────────────────────────────────────
    total_row = 5 + len(df)
    ws.merge_cells(start_row=total_row, start_column=1,
                   end_row=total_row, end_column=3)
    c = ws.cell(total_row, 1, f"Total folds: {len(df)}")
    c.font      = _font(bold=True, size=10)
    c.fill      = _fill(COLORES["header"])
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Suma de filas train, val, test
    for col_name, col_idx in [
        ("Train filas", all_cols.index("Train filas") + 1),
        ("Val filas",   all_cols.index("Val filas")   + 1),
        ("Test filas",  all_cols.index("Test filas")  + 1),
    ]:
        total = df[col_name].sum()
        c = ws.cell(total_row, col_idx, f"Σ={total:,}")
        c.font      = _font(bold=True, size=9)
        c.fill      = _fill(COLORES["header"])
        c.alignment = Alignment(horizontal="center")

    ws.row_dimensions[total_row].height = 16

    # ── Ancho de columnas ─────────────────────────────────────────────────────
    anchos = {
        "Fold"               : 6,
        "WU inicio"          : 12, "WU fin"          : 12,
        "WU días háb."       : 10, "WU filas"        : 8,
        "Train inicio"       : 12, "Train fin"        : 12,
        "Train meses"        : 10, "Train filas"      : 9,
        "Embargo T→V días h.": 12,
        "Val inicio"         : 12, "Val fin"          : 12,
        "Val meses"          : 8,  "Val filas"        : 9,
        "Embargo V→T días h.": 12,
        "Test inicio"        : 12, "Test fin"         : 12,
        "Test meses"         : 8,  "Test filas"       : 9,
    }
    for i, col in enumerate(all_cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = anchos.get(col, 10)

    # Congelar paneles tras encabezados y columna fold
    ws.freeze_panes = "B5"


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("Cargando matriz de features...")
    df_raw = pd.read_parquet(
        RUTA_MATRIZ,
        columns=["fecha_t", "banco", "h"],
        filters=[("banco", "==", BANCO), ("h", "==", H_REF)],
    )
    df_raw["fecha_t"] = pd.to_datetime(df_raw["fecha_t"])
    fechas = df_raw["fecha_t"].drop_duplicates().sort_values()

    print(f"  Banco: {BANCO} | Fechas únicas: {len(fechas):,} "
          f"| {fechas.min().date()} → {fechas.max().date()}")

    writer = pd.ExcelWriter(RUTA_OUT, engine="openpyxl")
    # Crear libro vacío
    writer.book.create_sheet("dummy")

    for cfg in CONFIGS:
        label    = cfg["label"]
        exp      = cfg["expanding"]
        val_años = cfg["val_años"]

        print(f"\nGenerando folds — {label}...")
        folds = generar_folds(fechas, expanding=exp, val_años=val_años)
        df_folds = folds_a_df(folds)

        print(f"  {len(folds)} folds generados")
        for f in folds:
            print(f"  Fold {f['fold']:>2}: "
                  f"Train {f['tr_start'].date()} → {f['tr_end'].date()} "
                  f"({f['tr_filas']} filas) | "
                  f"Val {f['val_start'].date()} → {f['val_end'].date()} "
                  f"({f['val_filas']} filas) | "
                  f"Test {f['test_start'].date()} → {f['test_end'].date()} "
                  f"({f['test_filas']} filas)")

        sheet_name = label.replace("|", "").replace("  ", " ").strip()[:31]
        ws = writer.book.create_sheet(title=sheet_name)
        escribir_hoja(ws, df_folds, label, expanding=exp, val_años=val_años)

    # Eliminar hoja vacía inicial
    if "dummy" in writer.book.sheetnames:
        del writer.book["dummy"]

    writer.book.save(RUTA_OUT)
    print(f"\n✓ Exportado: {RUTA_OUT}")
    print(f"  Hojas: {[s for s in writer.book.sheetnames]}")


if __name__ == "__main__":
    main()
