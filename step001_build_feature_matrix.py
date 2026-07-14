# -*- coding: utf-8 -*-
"""
build_feature_matrix.py
Construye la matriz de features para el sistema de predicción de liquidez
en moneda extranjera del banco central peruano.

Modelo: Gradient Boosting con quantile regression (LightGBM).
h (horizonte) es un feature más → UN SOLO modelo por banco.
Los bancos pequeños se agrupan en "Otros_bancos" antes de construir features.
"""

import os
import time
import logging
import warnings
import getpass
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# Constantes para codificación cíclica de variables de calendario
# ─────────────────────────────────────────────────────────────────────────────
_P_ELEC = 1260  # 5 años × ~252 días hábiles

# Primeras vueltas de elecciones generales peruanas (cada 5 años, sin extraordinarias)
_ELECCIONES_GENERALES_NP = np.array([
    "2001-04-08", "2006-04-09", "2011-04-10",
    "2016-04-10", "2021-04-11", "2026-04-12",
    "2031-04-13", "2036-04-12", "2041-04-14",
], dtype="datetime64[D]")

# ---------------------------------------------------------------------------
# Proxy corporativo: lee credenciales de variable de entorno o las pide
# ---------------------------------------------------------------------------
if not os.environ.get("BCRP_PROXY"):
    print("=" * 60)
    print("  Configuración de proxy corporativo BCRP")
    print("=" * 60)
    _usuario = input("  Usuario de red (ej: 2579): ").strip()
    _password = getpass.getpass("  Contraseña: ")
    os.environ["BCRP_PROXY"] = f"http://{_usuario}:{_password}@bcrproxy:8080"
    print("  Proxy configurado correctamente.\n")

# ---------------------------------------------------------------------------
# Logging con timestamps
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


###############################################################################
# PARTE 0 — Parámetros globales
###############################################################################

# Fecha de corte dinámica: ayer (último día hábil confirmado).
# Lógica: D(t) se conoce con 1 día hábil de anticipación y R(t) con 2.
# Usando t-1 día hábil garantizamos que tanto R como D del último día
# estén completamente confirmados en Transacciones_BancaLocal.xlsx.
_hoy = pd.Timestamp.today().normalize()
_fin_historico = (_hoy - pd.offsets.BDay(1)).strftime("%Y-%m-%d")

# ─────────────────────────────────────────────────────────────────────────────
# FEATURES A EXCLUIR
# Columnas que se descartan del parquet final antes de guardarlo.
# Se calculan internamente (pueden ser necesarias para derivar otras) y luego
# se eliminan. Para agregar más tras revisar el heatmap de step005, añadir
# el nombre aquí. Las columnas identidad (fecha_t, banco, h, target) están
# protegidas y no pueden eliminarse desde aquí.
#
# Fuente: diccionario_variables_filtro.xlsx — columna "Color":
#   Rojo  → eliminar sin reemplazo (leakage, redundancia o escala bruta).
#   Azul  → reemplazadas por transformación sin/cos o ratio (ver abajo).
#   flujo_neto_acum_mes: re-activado (sin leakage; acumula desde inicio de mes
#                        usando D(t) conocido en t-1 y R(t) conocido en t-2).
#
# NOTA: sigma_flujo_5d/20d y tc_vol_5d/22d se calculan internamente para los
# ratios antes de que este listado los elimine del parquet final.
# ─────────────────────────────────────────────────────────────────────────────
FEATURES_EXCLUIR = [
    # ── Rojo: eliminar sin reemplazo ─────────────────────────────────────────
    "log_h",
    "R_t0", "D_t0",
    "R_t-1", "D_t-1", "R_t-2", "D_t-2", "R_t-3", "D_t-3",
    "R_t-5", "D_t-5", "R_t-22", "D_t-22",
    "sigma_R_5d", "sigma_D_5d", "ma_R_5d", "ma_D_5d",
    "sigma_R_22d", "sigma_D_22d", "ma_R_22d", "ma_D_22d",
    "delta_R", "delta_D",
    "VIX_ma22", "TC_PEN_USD", "delta_TC",
    "EMBI_PERU", "delta_EMBI", "garch_vol_embi",
    "diferencial_tasas",
    "EMBI_PERU_frac", "T10Y_frac", "VIX_frac",
    "dias_desde_cierre_mes", "pos_en_mes", "total_bdays_mes",
    "is_quincena", "is_cierre_encaje",
    # ── SADF: sin señal en episodios de stress (hit rate P95 < 5%) ───────────
    "sadf_vol_60d", "sadf_vol_120d", "sadf_vol_252d",
    # ── Azul: reemplazadas por transformación sin/cos o ratio ────────────────
    "sigma_flujo_5d", "sigma_flujo_20d",              # → sigma_flujo_ratio
    "tc_vol_5d", "tc_vol_22d",                        # → tc_vol_ratio
    "dias_al_cierre_mes",                             # → dias_al_cierre_mes_sin/cos
    "is_penult_bday_trim", "is_ultimo_bday_trim",     # → dias_al_cierre_trim_sin/cos
    "is_1er_bday_trim", "is_2do_bday_trim", "is_3er_bday_trim",
    "dia_semana",                                     # → dias_sem_sin/cos
    "mes",                                            # → mes_sin/cos
    "is_fin_anio",                                    # → dias_al_cierre_anio_sin/cos
    "is_pre_eleccion", "is_post_eleccion",            # → elec_sin/cos
]

# ─────────────────────────────────────────────────────────────────────────────
# DIFERENCIACIÓN FRACCIONAL (FFD) — López de Prado Cap. 5
# Calibrar d mínimo solo con datos hasta esta fecha → sin leakage futuro.
# Usar la mayor fecha disponible ANTES del período de test de producción.
# ─────────────────────────────────────────────────────────────────────────────
_FD_TRAIN_CUTOFF = "2010-12-31"   # calibrar d en ventana histórica inicial

# ─────────────────────────────────────────────────────────────────────────────
# HMM — Gaussian Hidden Markov Model (3 estados: calma / moderado / severo)
# Se pre-computa UNA SOLA VEZ por banco con ventana expanding (sin leakage).
# Para etiquetar el año Y se entrena en [HMM_INICIO, Y-1].
# Requiere hmmlearn y sklearn. Si no están instalados, la columna queda en NaN.
# Si no aporta señal, agregar "hmm_estado" a FEATURES_EXCLUIR.
# ─────────────────────────────────────────────────────────────────────────────
HMM_N_ESTADOS = 3
HMM_INICIO    = "2010-01-01"   # primer año de historia disponible
HMM_MIN_AÑOS  = 2              # mínimo de años antes de etiquetar
                               # → primeras etiquetas desde HMM_INICIO + 2 años

try:
    from hmmlearn import hmm as _hmmlearn
    from sklearn.preprocessing import StandardScaler as _StandardScaler
    _HMM_DISPONIBLE = True
except ImportError:
    _HMM_DISPONIBLE = False
    logger.warning("hmmlearn/sklearn no instalados — hmm_estado quedará en NaN. "
                   "Instalar: pip install hmmlearn scikit-learn")

PARAMS = {
    # Fechas
    "fecha_inicio_historico": "2010-01-01",
    "fecha_fin_historico": _fin_historico,   # dinámico: ayer (1 día hábil de rezago)

    # Modelo
    "h_min": 2,
    "h_max": 75,
    "quantiles": [0.01, 0.05, 0.50, 0.95, 0.99],

    # Modo demo (True mientras no lleguen los datos reales)
    "usar_datos_demo": False,

    # Alias de bancos: cambios de nombre históricos que deben unificarse
    # clave = nombre en los datos, valor = nombre canónico
    "alias_bancos": {
        "CONTINEN": "BBVA",   # Banco Continental → BBVA desde 2020
    },

    # Agrupación de bancos
    "umbral_banco_pequeño_pct": 0.01,   # bancos con < 1% del volumen → Otros
    "bancos_otros": [],                  # lista fija (tiene prioridad si no está vacía)
    "nombre_otros": "Otros_bancos",

    # Rezagos bancarios
    "lags_cortos": [1, 2, 3],
    "lag_semana": 5,
    "lag_mes": 22,
    "ventanas_vol": [5, 22],

    # Calendario
    "años_calendario": list(range(2009, 2042)),

    # Rutas archivos manuales
    "ruta_datos_bancarios": r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Raw\Transacciones_BancaLocal.xlsx",
    "ruta_encaje": r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Raw\EncajeD.xlsx",
    "ruta_bbva_encaje_features": r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\2. Output\encaje_bbva\bbva_encaje_features_modelo.xlsx",
    "banco_encaje": "BBVA",   # banco al que aplican los datos de encaje
    "ruta_ccovn":  r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Raw\Saldos_CCOVN.xlsx",
    "bbva_keyword": "bbva",   # subcadena (case-insensitive) que identifica la columna BBVA en Saldos_CCOVN
    "ruta_confirmados": r"RUTA\confirmados.xlsx",
    "ruta_intervencion": r"RUTA\intervencion.xlsx",
    # "ruta_igv" eliminado: pagos IGV en soles, no relevante para liquidez ME
    # ruta_elecciones: eliminado — fechas presidenciales hardcodeadas en build_peru_calendar()
    "ruta_aux_xgboost": r"RUTA\Aux_XGBoost.py",
    "ruta_output":      r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Clean\matriz_features.parquet",
    "ruta_diccionario": r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Clean\diccionario_variables.xlsx",

    # Series BCRP descargadas manualmente con Add-In BCRPData
    # Un solo archivo Excel, cada serie en su propia hoja
    "ruta_bcrp": r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Raw\series_bcrp.xlsx",
    "hoja_bcrp_embi":     "EMBIG",      # hoja con PD04709XD (Spread EMBIG Perú)
    "hoja_bcrp_tasa_ref": "TasaPM",     # hoja con PD12301MD (Tasa de Referencia)
    "hoja_bcrp_tc_compra": "TC Compra", # hoja con TC compra BCRP
    "hoja_bcrp_tc_venta":  "TC Venta",  # hoja con TC venta BCRP

    # Series Bloomberg — DataBBG.xlsx con valores estáticos (hojas: BVL, CDS, Cobre)
    "ruta_bloomberg": r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Raw\DataBBG.xlsx",

    # APIs externas
    "fred_api_key": "96fa168ee9a9a4c1fcf323983db5ba64",
    "proxy": os.environ.get("BCRP_PROXY", "http://bcrproxy:8080"),

    # Códigos BCRP (solo como referencia; la lectura es desde archivo)
    "bcrp_embi": "PD04709XD",      # Spread EMBIG Perú (pbs) — diaria
    "bcrp_tasa_ref": "PD12301MD",  # Tasa de Referencia Política Monetaria — diaria
    # TC se lee de BCRP (promedio compra/venta) — Yahoo Finance tenía outliers

    # Tickers Yahoo Finance
    "ticker_vix": "^VIX",
    "ticker_tc": "PEN=X",          # Tipo de cambio USD/PEN (venta)
    "ticker_t10y": "^TNX",

    # Código FRED
    "fred_fedfunds": "DFF",            # Daily Federal Funds Effective Rate (diaria)
}


###############################################################################
# PARTE 1 — Calendario hábil peruano
###############################################################################
def build_peru_calendar(años, ruta_elecciones=None):
    """
    Retorna (peru_bday, peru_holidays, fechas_elecciones).

    Calendario hábil conjunto PE + US: excluye feriados de ambos países.
    Elecciones presidenciales hardcodeadas 2000-2041 (no requieren archivo externo).
    """
    logger.info("PARTE 1: Construyendo calendario hábil peruano...")

    # ── Base garantizada: AbstractHolidayCalendar con Easter hardcoded ──────────
    # No depende de la librería 'holidays' para Jueves/Viernes Santo.
    # Incluye todos los feriados fijos PE + US más los variables de Semana Santa.
    from pandas.tseries.holiday import (
        AbstractHolidayCalendar, Holiday, GoodFriday, USFederalHolidayCalendar,
    )
    from pandas.tseries.offsets import CustomBusinessDay, Day as _Day
    from dateutil.relativedelta import Easter

    _f0 = str(min(años)) + "-01-01"
    _f1 = str(max(años)) + "-12-31"

    class _PeruCalendar(AbstractHolidayCalendar):
        rules = [
            Holiday("AnioNuevo",   month=1,  day=1),
            Holiday("JuevesSanto", month=1,  day=1, offset=[Easter(), _Day(-3)]),
            GoodFriday,                              # Viernes Santo
            Holiday("Trabajo",     month=5,  day=1),
            Holiday("SanPedro",    month=6,  day=29),
            Holiday("FiestasP1",   month=7,  day=28),
            Holiday("FiestasP2",   month=7,  day=29),
            Holiday("SantaRosa",   month=8,  day=30),
            Holiday("Angamos",     month=10, day=8),
            Holiday("TodosSantos", month=11, day=1),
            Holiday("Inmaculada",  month=12, day=8),
            Holiday("Nochebuena",  month=12, day=24),
            Holiday("Navidad",     month=12, day=25),
        ]

    _hols_pe = set(_PeruCalendar().holidays(_f0, _f1).normalize())
    _hols_us = set(USFederalHolidayCalendar().holidays(_f0, _f1).normalize())
    _hols_base = _hols_pe | _hols_us

    # ── Suplemento opcional: librería 'holidays' añade fechas adicionales ────────
    # (p.e. feriados municipales, reemplazos por lunes, etc.)
    # Si no está instalada se usa la base garantizada — Semana Santa siempre incluida.
    try:
        import holidays as _hlib
        _sup_pe = set(pd.to_datetime(list(_hlib.Peru(years=años).keys())).normalize())
        _sup_us = set(pd.to_datetime(list(_hlib.UnitedStates(years=años).keys())).normalize())
        _hols_all = _hols_base | _sup_pe | _sup_us
        logger.info(
            f"  Feriados: {len(_hols_base)} base garantizada + "
            f"{len(_hols_all) - len(_hols_base)} adicionales (librería 'holidays') "
            f"= {len(_hols_all)} únicos en {min(años)}-{max(años)}"
        )
    except ImportError:
        _hols_all = _hols_base
        logger.warning(
            f"  Librería 'holidays' no instalada — usando {len(_hols_base)} feriados "
            f"base (PE+US fijos + Semana Santa). Instalar con: pip install holidays"
        )

    peru_holidays = pd.to_datetime(sorted(_hols_all))
    peru_bday     = CustomBusinessDay(holidays=peru_holidays)

    # Elecciones presidenciales peruanas hardcodeadas 2000-2041
    # Fechas desde 2031 son estimadas (2do domingo de abril / 1er domingo de junio)
    ULTIMA_FECHA_CONFIRMADA = pd.Timestamp("2026-06-07")
    ULTIMO_ANIO_HORIZONTE   = 2041

    _elecciones_raw = [
        # ── Históricas confirmadas ──────────────────────────────
        ("2000-04-09", "1ra vuelta"),  ("2000-05-28", "2da vuelta"),
        ("2001-04-08", "1ra vuelta"),  ("2001-06-03", "2da vuelta"),
        ("2006-04-09", "1ra vuelta"),  ("2006-06-04", "2da vuelta"),
        ("2011-04-10", "1ra vuelta"),  ("2011-06-05", "2da vuelta"),
        ("2016-04-10", "1ra vuelta"),  ("2016-06-05", "2da vuelta"),
        ("2021-04-11", "1ra vuelta"),  ("2021-06-06", "2da vuelta"),
        # ── Programadas / confirmadas ───────────────────────────
        ("2026-04-12", "1ra vuelta"),  ("2026-06-07", "2da vuelta"),
        # ── Estimadas (cada 5 años, 2do domingo abril / ~7 sem después) ──
        ("2031-04-13", "1ra vuelta"),  ("2031-06-08", "2da vuelta"),
        ("2036-04-12", "1ra vuelta"),  ("2036-06-07", "2da vuelta"),
        ("2041-04-14", "1ra vuelta"),  ("2041-06-08", "2da vuelta"),
    ]

    fechas_elecciones = [pd.Timestamp(f) for f, _ in _elecciones_raw]
    logger.info(
        f"  Elecciones presidenciales cargadas: {len(fechas_elecciones)} fechas "
        f"(2000-{ULTIMO_ANIO_HORIZONTE}, estimadas desde 2031)"
    )

    # Avisar cuando el horizonte de predicción se acerca al límite estimado
    fecha_hoy = pd.Timestamp.today().normalize()
    dias_al_limite = (pd.Timestamp(f"{ULTIMO_ANIO_HORIZONTE}-12-31") - fecha_hoy).days
    if dias_al_limite < 365 * 3:
        logger.warning(
            f"  ⚠ HORIZONTE ELECTORAL: quedan ~{dias_al_limite // 365} años hasta el límite "
            f"estimado ({ULTIMO_ANIO_HORIZONTE}). Considera extender las fechas en el código."
        )
    primera_estimada = pd.Timestamp("2031-04-13")
    dias_a_estimada = (primera_estimada - fecha_hoy).days
    if dias_a_estimada < 365 * 2:
        logger.warning(
            f"  ⚠ ELECCIONES ESTIMADAS: en ~{dias_a_estimada // 30} meses se entra a fechas "
            f"estimadas (desde {primera_estimada.date()}). Verificar con JNE."
        )

    logger.info("  Calendario construido correctamente.")
    return peru_bday, peru_holidays, fechas_elecciones


def get_future_business_dates(t, h_max, peru_bday):
    """Retorna los h_max días hábiles siguientes desde t (excluye t)."""
    fechas = pd.date_range(start=t + peru_bday, periods=h_max, freq=peru_bday)
    return fechas


###############################################################################
# PARTE 2 — Carga de datos manuales
###############################################################################
def load_manual_data(params):
    """
    Carga datos manuales desde archivos Excel definidos en PARAMS.

    Si un archivo no está disponible, carga un DataFrame vacío y muestra advertencia.
    Retorna dict con claves: 'bancarios', 'confirmados', 'intervencion'.
    """
    logger.info("PARTE 2: Cargando datos manuales...")

    resultado = {
        "bancarios": pd.DataFrame(),
        "confirmados": pd.DataFrame(columns=["banco", "R_conf_t1", "R_conf_t2", "D_conf_t1"]),
        "intervencion": pd.Series(dtype=float),
    }

    # Datos bancarios — formato Transacciones_BancaLocal.xlsx
    # Columnas: Broker (banco), Fecha Valor (fecha), Delivery Principal Usd (monto)
    # Monto < 0 → Retiro (R) | Monto > 0 → Depósito (D)
    try:
        df_raw = pd.read_excel(params["ruta_datos_bancarios"])
        if df_raw.empty:
            logger.warning(f"  Archivo bancarios vacío: {params['ruta_datos_bancarios']}")
        else:
            df_raw["Fecha Valor"] = pd.to_datetime(df_raw["Fecha Valor"])
            df_raw["monto"] = pd.to_numeric(df_raw["Delivery Principal Usd"], errors="coerce")

            # Separar retiros y depósitos
            df_raw["R"] = df_raw["monto"].clip(upper=0).abs()  # negativos → positivos
            df_raw["D"] = df_raw["monto"].clip(lower=0)        # positivos

            # Agregar por banco y fecha
            df_banco = (
                df_raw.groupby(["Broker", "Fecha Valor"])[["R", "D"]]
                .sum()
                .reset_index()
                .rename(columns={"Broker": "banco", "Fecha Valor": "fecha"})
            )
            df_banco["fecha"] = pd.to_datetime(df_banco["fecha"])

            # Unificar alias históricos (ej. CONTINEN → BBVA)
            # Comparación case-insensitive para tolerar variaciones de capitalización
            alias = params.get("alias_bancos", {})
            if alias:
                alias_lower = {k.lower(): v for k, v in alias.items()}
                def _aplicar_alias(nombre):
                    return alias_lower.get(nombre.lower(), nombre)
                antes = set(df_banco["banco"].unique())
                df_banco["banco"] = df_banco["banco"].apply(_aplicar_alias)
                despues = set(df_banco["banco"].unique())
                aplicados = antes - despues
                for viejo in aplicados:
                    nuevo = _aplicar_alias(viejo)
                    logger.info(f"  Alias aplicado: '{viejo}' → '{nuevo}'")
                # Re-agregar por si el mismo banco tiene filas con ambos nombres en la misma fecha
                df_banco = (
                    df_banco.groupby(["banco", "fecha"])[["R", "D"]]
                    .sum()
                    .reset_index()
                )

            df_banco = df_banco.sort_values(["banco", "fecha"])
            n_bancos = df_banco["banco"].nunique()
            f_min = df_banco["fecha"].min().date()
            f_max = df_banco["fecha"].max().date()
            logger.info(
                f"  Datos bancarios cargados: {len(df_banco):,} filas | "
                f"{n_bancos} bancos | {f_min} → {f_max}"
            )

            # Convertir formato LARGO → ANCHO: índice=fecha, columnas={banco}_R / {banco}_D
            # Necesario para compatibilidad con agrupar_bancos y build_feature_matrix
            df_wide = df_banco.pivot_table(
                index="fecha", columns="banco", values=["R", "D"], aggfunc="sum"
            )
            # MultiIndex ('R', 'BBVA') → 'BBVA_R'
            df_wide.columns = [f"{col[1]}_{col[0]}" for col in df_wide.columns]
            df_wide = df_wide.sort_index()
            df_wide.index.name = "fecha"

            # Rellenar calendario completo con 0: días sin transacción = flujo cero,
            # no dato desconocido. Evita NaN masivos en lags, sigmas y medias móviles.
            # reindex() añade fechas faltantes (fill_value=0).
            # fillna(0.0) cubre además los NaN internos del pivot (banco sin transacción
            # en una fecha donde otro banco sí la tuvo — pivot_table los deja como NaN).
            idx_completo = pd.bdate_range(start=df_wide.index.min(), end=df_wide.index.max())
            df_wide = df_wide.reindex(idx_completo, fill_value=0.0).fillna(0.0)
            df_wide.index.name = "fecha"

            resultado["bancarios"] = df_wide
            logger.info(
                f"  Matriz bancaria pivotada: {df_wide.shape[0]:,} fechas × "
                f"{df_wide.shape[1]} columnas ({n_bancos} bancos × 2) "
                f"[calendario completo, ceros en días sin transacción]"
            )
    except Exception as e:
        logger.warning(f"  No se pudo cargar datos bancarios: {params['ruta_datos_bancarios']} | {e}")

    # Confirmados (proxy histórico: valores realizados en t+1/t+2)
    ruta_conf = params.get("ruta_confirmados", "")
    if ruta_conf and ruta_conf != r"RUTA\confirmados.xlsx":
        try:
            df = pd.read_excel(ruta_conf)
            if df.empty:
                logger.warning(f"  Archivo confirmados vacío: {ruta_conf}")
            else:
                resultado["confirmados"] = df
                logger.info(f"  Datos confirmados cargados: {df.shape}")
        except Exception as e:
            logger.warning(f"  No se pudo cargar confirmados: {ruta_conf} | {e}")
    else:
        logger.info("  Confirmados: usando proxy histórico (D_t+1, R_t+2 realizados).")

    logger.info("  Carga de datos manuales completada.")
    return resultado


def load_encaje_data(params):
    """
    Carga EncajeD.xlsx con datos de encaje de BBVA.

    Columnas esperadas (nombres normalizados a minúsculas y sin tildes):
      fecha, caja, cta_cte_bcr, overnight_bcr, activos, encaje_exigible, retiro_neto

    Retorna DataFrame indexado por fecha con las columnas normalizadas,
    o DataFrame vacío si el archivo no está disponible.
    """
    ruta = params.get("ruta_encaje", "")
    if not ruta or ruta == r"RUTA\EncajeD.xlsx":
        logger.info("  Encaje: ruta no configurada — features de encaje omitidas.")
        return pd.DataFrame()

    try:
        df = pd.read_excel(ruta)
        if df.empty:
            logger.warning(f"  EncajeD.xlsx vacío: {ruta}")
            return pd.DataFrame()

        # Normalizar nombres de columna: minúsculas, sin tildes, sin espacios
        import unicodedata
        def _norm(s):
            s = str(s).lower().strip()
            s = "".join(c for c in unicodedata.normalize("NFD", s)
                        if unicodedata.category(c) != "Mn")
            return s.replace(" ", "_").replace(".", "_")

        df.columns = [_norm(c) for c in df.columns]

        # Mapear variantes de nombres al estándar interno
        _alias = {
            "cta__cte__bcr": "cta_cte_bcr",
            "cta_cte": "cta_cte_bcr",
            "overnight": "overnight_bcr",
            "exigible": "encaje_exigible",
            "retiro": "retiro_neto",
        }
        df = df.rename(columns=_alias)

        # Detectar columna de fecha
        fecha_col = next((c for c in df.columns if "fecha" in c), None)
        if fecha_col is None:
            logger.warning("  EncajeD.xlsx: no se encontró columna de fecha.")
            return pd.DataFrame()

        df["fecha"] = pd.to_datetime(df[fecha_col])
        df = df.set_index("fecha").sort_index()

        # Conservar solo columnas numéricas relevantes
        cols_requeridas = ["caja", "cta_cte_bcr", "overnight_bcr", "activos",
                           "encaje_exigible", "retiro_neto"]
        cols_presentes = [c for c in cols_requeridas if c in df.columns]
        df = df[cols_presentes].apply(pd.to_numeric, errors="coerce")

        logger.info(
            f"  EncajeD cargado: {len(df):,} filas | "
            f"{df.index.min().date()} → {df.index.max().date()} | "
            f"cols: {cols_presentes}"
        )
        return df

    except Exception as e:
        logger.warning(f"  No se pudo cargar EncajeD.xlsx: {ruta} | {e}")
        return pd.DataFrame()


def load_ccovn_data(params):
    """
    Carga Saldos_CCOVN.xlsx: saldos de cierre CC + OVN por banco en el BCR.

    Formato esperado:
      - Col A: Fecha (d-m-yyyy)
      - Cols B…: un banco por columna (valores numéricos en USD)

    Retorna DataFrame con columnas:
      ccovn_sistema : suma de todos los bancos (USD)
      ccovn_bbva    : columna que contiene el keyword "bbva" (case-insensitive)

    Indexado por fecha, o DataFrame vacío si el archivo no está disponible.
    """
    ruta = params.get("ruta_ccovn", "")
    if not ruta:
        logger.info("  CCOVN: ruta no configurada — features ccovn omitidas.")
        return pd.DataFrame()

    try:
        raw = pd.read_excel(ruta, header=0)
        if raw.empty:
            logger.warning(f"  Saldos_CCOVN.xlsx vacío: {ruta}")
            return pd.DataFrame()

        col_fecha   = raw.columns[0]
        cols_bancos = raw.columns[1:].tolist()

        raw[col_fecha] = pd.to_datetime(raw[col_fecha], dayfirst=True, errors="coerce")
        raw = raw.dropna(subset=[col_fecha]).sort_values(col_fecha)

        for c in cols_bancos:
            raw[c] = pd.to_numeric(raw[c], errors="coerce")

        keyword = params.get("bbva_keyword", "bbva")
        cols_bbva = [c for c in cols_bancos if keyword in str(c).lower()]
        col_bbva  = cols_bbva[0] if cols_bbva else None

        df_out = pd.DataFrame()
        df_out["ccovn_sistema"] = raw[cols_bancos].sum(axis=1, skipna=True).values
        df_out["ccovn_bbva"]    = raw[col_bbva].values if col_bbva else np.nan
        df_out.index = pd.DatetimeIndex(raw[col_fecha].values)
        df_out.index.name = "fecha"
        df_out = df_out.sort_index()

        logger.info(
            f"  Saldos_CCOVN cargado: {len(df_out):,} filas | "
            f"{df_out.index.min().date()} → {df_out.index.max().date()} | "
            f"col_bbva='{col_bbva}'"
        )
        return df_out

    except Exception as e:
        logger.warning(f"  No se pudo cargar Saldos_CCOVN.xlsx: {ruta} | {e}")
        return pd.DataFrame()


###############################################################################
# PARTE 3 — Agrupación de bancos
###############################################################################
def agrupar_bancos(df_bancarios, umbral_pct, bancos_otros, nombre_otros):
    """
    Agrupa bancos pequeños en una categoría única antes de modelar.

    Retorna (df_agrupado, lista_bancos, reporte).
    """
    logger.info("PARTE 3: Agrupando bancos pequeños...")

    if df_bancarios.empty:
        logger.warning("  DataFrame bancarios vacío. No hay bancos que agrupar.")
        return df_bancarios, [], {}

    # Detectar bancos disponibles (columnas tipo banco_X_R y banco_X_D)
    cols_R = [c for c in df_bancarios.columns if c.endswith("_R")]
    cols_D = [c for c in df_bancarios.columns if c.endswith("_D")]
    bancos_disponibles = [c[:-2] for c in cols_R if c[:-2] + "_D" in cols_D]

    if not bancos_disponibles:
        logger.warning("  No se detectaron columnas con formato {banco}_R / {banco}_D.")
        return df_bancarios, [], {}

    # Determinar bancos pequeños
    if bancos_otros:
        # Opción B: lista fija
        pequeños = [b for b in bancos_otros if b in bancos_disponibles]
        logger.info(f"  Usando lista fija de bancos pequeños: {pequeños}")
    else:
        # Opción A: umbral automático por volumen
        volumen = {}
        for b in bancos_disponibles:
            r = df_bancarios.get(f"{b}_R", pd.Series(dtype=float))
            d = df_bancarios.get(f"{b}_D", pd.Series(dtype=float))
            volumen[b] = (r.abs() + d.abs()).mean()

        total_vol = sum(volumen.values())
        participacion = {b: v / total_vol for b, v in volumen.items()} if total_vol > 0 else {b: 0 for b in bancos_disponibles}
        pequeños = [b for b, p in participacion.items() if p < umbral_pct]
        logger.info(f"  Bancos detectados automáticamente como pequeños (< {umbral_pct:.0%}): {pequeños}")

    grandes = [b for b in bancos_disponibles if b not in pequeños]

    df_agrupado = df_bancarios.copy()

    if pequeños:
        # Sumar R y D de bancos pequeños
        cols_R_pequeños = [f"{b}_R" for b in pequeños if f"{b}_R" in df_agrupado.columns]
        cols_D_pequeños = [f"{b}_D" for b in pequeños if f"{b}_D" in df_agrupado.columns]

        df_agrupado[f"{nombre_otros}_R"] = df_agrupado[cols_R_pequeños].sum(axis=1)
        df_agrupado[f"{nombre_otros}_D"] = df_agrupado[cols_D_pequeños].sum(axis=1)

        cols_a_eliminar = cols_R_pequeños + cols_D_pequeños
        df_agrupado = df_agrupado.drop(columns=cols_a_eliminar)

        lista_bancos = grandes + [nombre_otros]
    else:
        lista_bancos = grandes

    # Reporte
    reporte = {
        "bancos_individuales": grandes,
        "bancos_agrupados": pequeños,
        "nombre_grupo": nombre_otros,
    }

    if not df_bancarios.empty and not bancos_otros:
        reporte["participaciones"] = participacion if "participacion" in dir() else {}

    # Imprimir tabla resumen
    logger.info(f"\n{'='*55}")
    logger.info(f"  RESUMEN DE AGRUPACIÓN DE BANCOS")
    logger.info(f"{'='*55}")
    logger.info(f"  Bancos individuales ({len(grandes)}): {grandes}")
    logger.info(f"  Bancos agrupados en '{nombre_otros}' ({len(pequeños)}): {pequeños}")
    logger.info(f"  Total entidades a modelar: {len(lista_bancos)}")
    logger.info(f"{'='*55}\n")

    return df_agrupado, lista_bancos, reporte


###############################################################################
# PARTE 4 — Descarga automática de series externas
###############################################################################
def _retry_download(func, max_intentos=3, espera=2):
    """Ejecuta func hasta max_intentos veces con espera exponencial."""
    for intento in range(1, max_intentos + 1):
        try:
            resultado = func()
            if resultado is not None:
                return resultado
        except Exception as e:
            logger.warning(f"  Intento {intento}/{max_intentos} fallido: {e}")
            if intento < max_intentos:
                time.sleep(espera * intento)
    return None


def _descargar_yahoo(ticker, inicio, fin, proxies, nombre):
    """
    Descarga una serie desde Yahoo Finance via requests directo con headers de navegador.
    Funciona detrás de proxies corporativos que interceptan SSL.
    """
    import urllib3
    urllib3.disable_warnings()

    ticker_encoded = ticker.replace("^", "%5E")
    inicio_ts = int(pd.Timestamp(inicio).timestamp())
    fin_ts    = int(pd.Timestamp(fin).timestamp())
    url = (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker_encoded}"
        f"?interval=1d&period1={inicio_ts}&period2={fin_ts}"
    )
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://finance.yahoo.com",
    }

    def _descarga():
        import requests as req
        r = req.get(
            url,
            headers=headers,
            proxies=proxies if proxies else None,
            verify=False,
            timeout=30,
        )
        if r.status_code != 200:
            raise ValueError(f"HTTP {r.status_code} para {ticker}")
        data   = r.json()
        result = data["chart"]["result"]
        if not result:
            raise ValueError(f"Sin datos en respuesta para {ticker}")
        timestamps = result[0]["timestamp"]
        closes     = result[0]["indicators"]["quote"][0]["close"]
        s = pd.Series(
            closes,
            index=pd.to_datetime(timestamps, unit="s").tz_localize(None),
            name=nombre,
            dtype=float,
        )
        s.index = s.index.normalize()
        if s.dropna().empty:
            raise ValueError(f"Serie vacía para {ticker}")
        return s

    resultado = _retry_download(_descarga)
    if resultado is None:
        logger.warning(f"  No se pudo descargar {nombre} ({ticker}) desde Yahoo Finance.")
        return pd.Series(dtype=float, name=nombre)
    logger.info(f"  {nombre} descargado: {len(resultado)} observaciones.")
    return resultado


def _descargar_fred(serie_id, inicio, fin, api_key, proxies, nombre):
    """Descarga una serie desde la API REST de FRED (sin fredapi, solo requests)."""
    import urllib3
    urllib3.disable_warnings()

    url = (
        f"https://api.stlouisfed.org/fred/series/observations"
        f"?series_id={serie_id}"
        f"&api_key={api_key}"
        f"&file_type=json"
        f"&observation_start={pd.Timestamp(inicio).strftime('%Y-%m-%d')}"
        f"&observation_end={pd.Timestamp(fin).strftime('%Y-%m-%d')}"
    )

    def _descarga():
        import requests
        r = requests.get(
            url,
            proxies=proxies if proxies else None,
            verify=False,
            timeout=30,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        if r.status_code != 200:
            raise ValueError(f"FRED HTTP {r.status_code}: {r.text[:200]}")
        obs = r.json().get("observations", [])
        if not obs:
            raise ValueError(f"Sin observaciones para {serie_id}")
        registros = []
        for o in obs:
            try:
                val = float(o["value"])
                registros.append({"fecha": pd.Timestamp(o["date"]), nombre: val})
            except (ValueError, TypeError):
                pass  # "." = dato faltante en FRED
        s = pd.DataFrame(registros).set_index("fecha")[nombre]
        return s

    resultado = _retry_download(_descarga)
    if resultado is None:
        logger.warning(f"  No se pudo descargar {nombre} ({serie_id}) desde FRED.")
        return pd.Series(dtype=float, name=nombre)
    logger.info(f"  {nombre} descargado: {len(resultado)} observaciones.")
    return resultado


def _leer_bcrp_excel(ruta, nombre, hoja=0):
    """
    Lee una serie BCRP desde una hoja de Excel descargado con el Add-In BCRPData.

    Formato esperado del Add-In:
        Fila 1: código de serie
        Fila 2: Descripción
        Fila 3: Frecuencia
        Fila 4: Periodo
        Fila 5: vacía
        Fila 6: encabezados  Date | Valores
        Fila 7+: datos       fecha | valor

    hoja: nombre o índice de la hoja dentro del Excel (default=0, primera hoja).
    Si el archivo no existe devuelve Series vacía con advertencia.
    """
    if not ruta or not os.path.exists(ruta):
        logger.warning(
            f"  Archivo BCRP '{nombre}' no encontrado en: {ruta}\n"
            f"  → Descargarlo con Add-In BCRPData y guardarlo en esa ruta."
        )
        return pd.Series(dtype=float, name=nombre)

    try:
        # header=None para leer desde fila 1 sin asumir encabezados
        raw = pd.read_excel(ruta, sheet_name=hoja, header=None)

        # Buscar la fila de encabezado "Date" / "Valores"
        header_row = None
        for i, row in raw.iterrows():
            vals = [str(v).strip().lower() for v in row if pd.notna(v)]
            if any(v in ("date", "fecha") for v in vals):
                header_row = i
                break

        if header_row is None:
            header_row = 5  # fallback: fila 6

        df = pd.read_excel(ruta, sheet_name=hoja, header=header_row)
        df.columns = [str(c).strip() for c in df.columns]

        # Identificar columnas de fecha y valor
        col_fecha = next((c for c in df.columns if c.lower() in ("date", "fecha")), df.columns[0])
        col_valor = next((c for c in df.columns if c.lower() in ("valores", "value", "values")), df.columns[1])

        df = df[[col_fecha, col_valor]].copy()
        df.columns = ["fecha", nombre]

        # El Add-In BCRPData usa meses en español: Set=Sep, Ene=Jan, etc.
        _meses_es = {
            "Ene": "Jan", "Feb": "Feb", "Mar": "Mar", "Abr": "Apr",
            "May": "May", "Jun": "Jun", "Jul": "Jul", "Ago": "Aug",
            "Set": "Sep", "Oct": "Oct", "Nov": "Nov", "Dic": "Dec",
        }
        def _parse_fecha_bcrp(s):
            s = str(s).strip()
            for es, en in _meses_es.items():
                s = s.replace(es, en)
            return pd.to_datetime(s, errors="coerce", dayfirst=True)

        df["fecha"] = df["fecha"].apply(_parse_fecha_bcrp)
        df[nombre] = pd.to_numeric(
            df[nombre].astype(str).str.replace(",", "."), errors="coerce"
        )
        df = df.dropna(subset=["fecha"]).set_index("fecha").sort_index()

        # Advertir si los datos tienen más de 7 días de antigüedad
        ultima = df.index.max()
        dias_atras = (pd.Timestamp.today() - ultima).days
        if dias_atras > 7:
            logger.warning(
                f"  {nombre}: última fecha {ultima.date()} ({dias_atras} días atrás). "
                f"Considera actualizar el archivo."
            )

        logger.info(f"  {nombre} leído desde archivo: {len(df)} obs, hasta {ultima.date()}")
        return df[nombre]

    except Exception as e:
        logger.warning(f"  Error leyendo {nombre} desde {ruta}: {e}")
        return pd.Series(dtype=float, name=nombre)


_BBG_HOJAS     = {"CDS_PERU_5Y": "CDS", "COPPER": "Cobre"}
_BBG_DATE_FMTS = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y%m%d"]


def _leer_bloomberg_excel(ruta, nombre):
    """
    Lee directamente desde DataBBG.xlsx (hojas BVL / CDS / Cobre).
    Maneja el formato Bloomberg estándar: filas de metadatos + datos.
    """
    if not ruta or not os.path.exists(ruta):
        logger.warning(f"  Bloomberg '{nombre}' no encontrado en: {ruta}")
        return pd.Series(dtype=float, name=nombre)
    hoja = _BBG_HOJAS.get(nombre, nombre)
    try:
        raw = pd.read_excel(ruta, sheet_name=hoja, header=None, dtype=str)
        etiquetas = {"security", "start date", "end date", "period", "currency",
                     "pricing source", "date", "dates", "nan", ""}
        fila_inicio = None
        for i, row in raw.iterrows():
            val = str(row.iloc[0]).strip().lower()
            if val in etiquetas:
                continue
            for fmt in _BBG_DATE_FMTS:
                try:
                    ts = pd.to_datetime(val.split(" ")[0], format=fmt)
                    if pd.Timestamp("1990-01-01") <= ts <= pd.Timestamp("2100-01-01"):
                        fila_inicio = i
                        break
                except Exception:
                    continue
            if fila_inicio is not None:
                break
        if fila_inicio is None:
            logger.warning(f"  {nombre}: no se encontraron filas de datos en hoja '{hoja}'")
            return pd.Series(dtype=float, name=nombre)

        datos = raw.iloc[fila_inicio:, :2].copy()
        datos.columns = ["fecha", nombre]
        datos = datos.dropna(subset=["fecha"])
        datos = datos[datos["fecha"].str.strip() != ""]
        for fmt in _BBG_DATE_FMTS:
            try:
                datos["fecha"] = pd.to_datetime(datos["fecha"], format=fmt, errors="raise")
                break
            except Exception:
                continue
        else:
            datos["fecha"] = pd.to_datetime(datos["fecha"], dayfirst=True, errors="coerce")
        datos[nombre] = (datos[nombre].astype(str)
                         .str.replace(",", ".", regex=False)
                         .pipe(pd.to_numeric, errors="coerce"))
        datos = datos.dropna(subset=["fecha", nombre]).set_index("fecha").sort_index()
        datos = datos[~datos.index.duplicated(keep="last")]
        s = datos[nombre]
        logger.info(f"  {nombre} [{hoja}]: {len(s):,} obs, "
                    f"{s.index.min().date()} → {s.index.max().date()}")
        return s
    except Exception as e:
        logger.warning(f"  Error leyendo Bloomberg '{nombre}' (hoja '{hoja}'): {e}")
        return pd.Series(dtype=float, name=nombre)


def download_external_series(params):
    """
    Descarga todas las series externas y las retorna en un DataFrame
    alineado al calendario hábil peruano.

    Forward fill máximo 3 días para series de menor frecuencia.
    """
    logger.info("PARTE 4: Descargando series externas...")

    inicio = params["fecha_inicio_historico"]
    fin = params["fecha_fin_historico"]
    proxy = params.get("proxy", "")
    proxies = {"http": proxy, "https": proxy} if proxy and "USUARIO" not in proxy else None

    series = {}

    # 4a. Yahoo Finance
    series["VIX"]  = _descargar_yahoo(params["ticker_vix"],  inicio, fin, proxies, "VIX")
    series["T10Y"] = _descargar_yahoo(params["ticker_t10y"], inicio, fin, proxies, "T10Y")

    # 4b. FRED
    api_key = params.get("fred_api_key", "")
    if api_key and api_key != "TU_API_KEY":
        series["FED_FUNDS"] = _descargar_fred(
            params["fred_fedfunds"], inicio, fin, api_key, proxies, "FED_FUNDS"
        )
    else:
        logger.warning("  FRED API key no configurada. FED_FUNDS quedará como NaN.")
        series["FED_FUNDS"] = pd.Series(dtype=float, name="FED_FUNDS")

    # 4c. Series BCRP — leídas desde Excel descargado manualmente con Add-In BCRPData
    ruta_bcrp = params["ruta_bcrp"]
    series["EMBI_PERU"]     = _leer_bcrp_excel(ruta_bcrp, "EMBI_PERU",     hoja=params["hoja_bcrp_embi"])
    series["TASA_REF_BCRP"] = _leer_bcrp_excel(ruta_bcrp, "TASA_REF_BCRP", hoja=params["hoja_bcrp_tasa_ref"])

    # TC_PEN_USD = promedio(TC Compra, TC Venta) del BCRP — reemplaza Yahoo Finance (tenía outliers)
    tc_compra = _leer_bcrp_excel(ruta_bcrp, "TC_COMPRA", hoja=params["hoja_bcrp_tc_compra"])
    tc_venta  = _leer_bcrp_excel(ruta_bcrp, "TC_VENTA",  hoja=params["hoja_bcrp_tc_venta"])
    if not tc_compra.empty and not tc_venta.empty:
        tc_df = pd.concat([tc_compra.rename("compra"), tc_venta.rename("venta")], axis=1)
        series["TC_PEN_USD"] = tc_df.mean(axis=1).rename("TC_PEN_USD")
        logger.info("  TC_PEN_USD: promedio BCRP compra/venta cargado.")
    elif not tc_venta.empty:
        series["TC_PEN_USD"] = tc_venta.rename("TC_PEN_USD")
        logger.warning("  TC_PEN_USD: solo TC Venta disponible (sin TC Compra).")
    elif not tc_compra.empty:
        series["TC_PEN_USD"] = tc_compra.rename("TC_PEN_USD")
        logger.warning("  TC_PEN_USD: solo TC Compra disponible (sin TC Venta).")
    else:
        logger.warning("  TC_PEN_USD: no se encontraron hojas TC Compra/Venta en series_bcrp.xlsx.")
        series["TC_PEN_USD"] = pd.Series(dtype=float, name="TC_PEN_USD")

    # 4d. Bloomberg — CDS Perú 5Y y Cobre LME (DataBBG.xlsx, hojas CDS / Cobre)
    ruta_bbg = params.get("ruta_bloomberg", "")
    for _nombre_bbg in ("CDS_PERU_5Y", "COPPER"):
        series[_nombre_bbg] = _leer_bloomberg_excel(ruta_bbg, _nombre_bbg)

    # Alinear al índice de fechas del rango
    idx = pd.bdate_range(start=inicio, end=fin)
    df = pd.DataFrame(index=idx)
    for nombre, s in series.items():
        if not s.empty:
            s_alineada = s.reindex(idx)
            s_alineada = s_alineada.ffill()   # sin límite: series mensuales/trimestrales se propagan correctamente
            df[nombre] = s_alineada
        else:
            df[nombre] = np.nan

    cols_esperadas = ["VIX", "TC_PEN_USD", "T10Y", "FED_FUNDS", "EMBI_PERU", "TASA_REF_BCRP",
                      "CDS_PERU_5Y", "COPPER"]
    for c in cols_esperadas:
        if c not in df.columns:
            df[c] = np.nan

    logger.info(f"  DataFrame externo construido: {df.shape}")
    return df


###############################################################################
# PARTE 5 — Construcción de features
###############################################################################

# 5a. Features bancarias
def _garch_vol(flujo: pd.Series) -> pd.Series:
    """
    GARCH(1,1) implementado en NumPy + scipy.optimize — sin dependencia del
    paquete 'arch' ni de statsmodels, evitando incompatibilidades con scipy 1.9+.

    Modelo:
        σ²_t = ω + α·ε²_{t-1} + β·σ²_{t-1}

    donde ε_t = x_t (flujo neto centrado). Los parámetros ω, α, β se estiman
    por Máxima Verosimilitud Gaussiana usando L-BFGS-B.

    σ_t usa solo información hasta t-1 → sin look-ahead bias.
    Fallback a rolling std 20d si la serie es constante o muy corta.
    """
    if flujo.dropna().std() < 1e-9 or len(flujo.dropna()) < 60:
        return flujo.rolling(20).std()

    from scipy.optimize import minimize

    # Escalar a varianza unitaria para estabilidad numérica del optimizador
    escala  = flujo.std()
    x       = (flujo / escala).fillna(0.0).values.astype(float)
    n       = len(x)
    var_unc = float(np.var(x))   # varianza incondicional → σ²_0 inicial

    def _sigma2(params):
        """Calcula la varianza condicional para un vector de parámetros."""
        omega, alpha, beta = params
        s2 = np.empty(n)
        s2[0] = var_unc
        for t in range(1, n):
            s2[t] = omega + alpha * x[t-1]**2 + beta * s2[t-1]
        return s2

    def _neg_loglik(params):
        """Log-verosimilitud gaussiana negativa (a minimizar)."""
        omega, alpha, beta = params
        # Restricciones de estacionariedad: ω>0, α>0, β>0, α+β<1
        if omega <= 0 or alpha <= 0 or beta <= 0 or alpha + beta >= 0.9999:
            return 1e10
        s2 = _sigma2(params)
        if np.any(s2 <= 0):
            return 1e10
        return 0.5 * float(np.sum(np.log(s2) + x**2 / s2))

    # Punto de partida: parámetros típicos de series financieras diarias
    p0     = [0.01, 0.08, 0.88]
    bounds = [(1e-7, 0.5), (1e-7, 0.5), (1e-7, 0.9999)]

    try:
        res = minimize(_neg_loglik, p0, method="L-BFGS-B", bounds=bounds,
                       options={"maxiter": 500, "ftol": 1e-10, "gtol": 1e-7})
        if not res.success and res.fun > 1e9:
            raise ValueError(res.message)

        sigma2     = _sigma2(res.x)
        sigma_cond = pd.Series(np.sqrt(np.maximum(sigma2, 0)) * escala,
                               index=flujo.index)
        logger.info(
            f"    GARCH(1,1) ajustado: ω={res.x[0]:.5f}  "
            f"α={res.x[1]:.4f}  β={res.x[2]:.4f}  "
            f"persistencia={res.x[1]+res.x[2]:.4f}"
        )
        return sigma_cond

    except Exception as e:
        logger.warning(f"  GARCH no convergió ({e}) — usando rolling std 20d")
        return flujo.rolling(20).std()


###############################################################################
# FFD — Diferenciación Fraccional de Ancho Fijo (López de Prado, Cap. 5)
###############################################################################

def _ffd_weights(d: float, thresh: float = 1e-5) -> np.ndarray:
    """Calcula los pesos de FFD por expansión binomial: w_k = Π_{j=1}^{k} (d-j+1)/j * (-1)^k."""
    w, k = [1.0], 1
    while True:
        w_k = -w[-1] * (d - k + 1) / k
        if abs(w_k) < thresh:
            break
        w.append(w_k)
        k += 1
    return np.array(w[::-1])   # más reciente al final → conveniente para dot()


def _fracdiff_fixed_width(series: pd.Series, d: float, thresh: float = 1e-5) -> pd.Series:
    """Aplica FFD con ventana de ancho fijo a una Serie de pandas."""
    w     = _ffd_weights(d, thresh)
    width = len(w)
    vals  = series.values.astype(float)
    n     = len(vals)
    out   = np.full(n, np.nan)
    for i in range(width - 1, n):
        chunk = vals[i - width + 1: i + 1]
        if not np.any(np.isnan(chunk)):
            out[i] = float(np.dot(w, chunk))
    return pd.Series(out, index=series.index, name=series.name)


def _find_min_d(
    series: pd.Series,
    thresh: float = 1e-5,
    max_d: float  = 1.0,
    target_pval: float = 0.05,
    n_steps: int  = 20,
) -> float:
    """
    Encuentra el d mínimo en (0, max_d] tal que la serie diferenciada
    sea estacionaria (ADF p-value ≤ target_pval).
    Requiere statsmodels; si no está disponible retorna 0.4 como fallback.
    """
    try:
        from statsmodels.tsa.stattools import adfuller
    except ImportError:
        logger.warning("  statsmodels no disponible → d_opt fallback = 0.4")
        return 0.4

    for d in np.linspace(0.05, max_d, n_steps):
        fd = _fracdiff_fixed_width(series.dropna(), round(float(d), 4), thresh)
        fd_clean = fd.dropna()
        if len(fd_clean) < 30:
            continue
        try:
            pval = adfuller(fd_clean, maxlag=1, regression="c", autolag=None)[1]
            if pval <= target_pval:
                return round(float(d), 4)
        except Exception:
            continue
    return round(max_d, 4)


def _calcular_hmm_expanding(flujo: pd.Series, sigma_22d: pd.Series,
                            primer_ventana_años: int = 6) -> pd.Series:
    """
    Pre-computa hmm_estado[t] para todo t con expanding window, sin leakage.

    Alineado con la estructura walk-forward (igual que GARCH):
      1. Primer modelo: entrena en [HMM_INICIO, HMM_INICIO + primer_ventana_años - 1]
         → etiqueta ESE mismo periodo (in-sample, como GARCH) + el siguiente año
      2. Años siguientes: expanding, entrena en [HMM_INICIO, Y-1] → etiqueta Y

    Con primer_ventana_años=6 e HMM_INICIO=2010:
      - Modelo 1: 2010-2015 → etiqueta 2010-2015 (in-sample) + 2016
      - Modelo 2: 2010-2016 → etiqueta 2017
      - Modelo 3: 2010-2017 → etiqueta 2018
      - ...

    Estados ordenados por det(Σ_i) ascendente: 0=calma, 1=moderado, 2=severo.
    Devuelve pd.Series de Int8 sin NaN (toda la historia desde HMM_INICIO etiquetada).
    """
    if not _HMM_DISPONIBLE:
        return pd.Series(pd.NA, index=flujo.index, dtype="Int8", name="hmm_estado")

    df_base = pd.concat([flujo.rename("flujo"), sigma_22d.rename("sigma_22d")],
                        axis=1).dropna()
    df_base = df_base[df_base.index >= HMM_INICIO]
    if df_base.empty:
        return pd.Series(pd.NA, index=flujo.index, dtype="Int8", name="hmm_estado")

    año_inicio    = df_base.index.year.min()
    año_fin_prim  = año_inicio + primer_ventana_años - 1   # último año del primer bloque
    años          = sorted(df_base.index.year.unique())
    resultado     = pd.Series(pd.NA, index=df_base.index, dtype="Int8", name="hmm_estado")

    def _fit_and_label(X_train, df_target):
        """Ajusta HMM en X_train y predice estados para df_target."""
        scaler = _StandardScaler()
        Xs     = scaler.fit_transform(X_train)
        modelo = _hmmlearn.GaussianHMM(
            n_components=HMM_N_ESTADOS, covariance_type="full",
            n_iter=1000, random_state=42,
        )
        modelo.fit(Xs)
        varianzas     = [np.linalg.det(modelo.covars_[s]) for s in range(HMM_N_ESTADOS)]
        sorted_states = np.argsort(varianzas)
        mapa          = {sorted_states[i]: i for i in range(HMM_N_ESTADOS)}
        Xs_target     = scaler.transform(df_target.values)
        estados_raw   = modelo.predict(Xs_target)
        return np.array([mapa[e] for e in estados_raw], dtype="int8")

    # ── Paso 1: primer bloque (in-sample, alineado con primer fold de train) ──
    df_prim = df_base[df_base.index.year <= año_fin_prim]
    if len(df_prim) >= 50:
        try:
            estados_prim = _fit_and_label(df_prim.values, df_prim)
            resultado.loc[df_prim.index] = pd.array(estados_prim, dtype="Int8")
        except Exception:
            pass

    # ── Paso 2: expanding año a año desde año_fin_prim + 1 ───────────────────
    for año in años:
        if año <= año_fin_prim:
            continue
        X_train = df_base[df_base.index.year < año].values
        df_año  = df_base[df_base.index.year == año]
        if df_año.empty or len(X_train) < 50:
            continue
        try:
            estados_año = _fit_and_label(X_train, df_año)
            resultado.loc[df_año.index] = pd.array(estados_año, dtype="Int8")
        except Exception:
            continue

    return resultado


def build_bank_features(df_banco, lags_cortos, lag_semana, lag_mes, ventanas_vol):
    """
    Recibe serie temporal de un banco con columnas R y D.
    Genera features de rezagos, volatilidades, medias móviles y cambios diarios.

    Incluye features de régimen de volatilidad del flujo neto (D−R):
      sigma_flujo_5d / 20d : std rolling del flujo neto realizado.
      ma_flujo_5d / 20d    : media rolling del flujo neto (nivel reciente).
      sigma_22d            : desviación estándar rolling 22d del flujo neto.
                             Más reactiva que GARCH al salir de periodos de estrés.
    """
    VENTANAS_FLUJO = [5, 20]

    if df_banco.empty or "R" not in df_banco.columns or "D" not in df_banco.columns:
        cols = (
            ["R_t0", "D_t0"]
            + [f"R_t-{l}" for l in lags_cortos + [lag_semana, lag_mes]]
            + [f"D_t-{l}" for l in lags_cortos + [lag_semana, lag_mes]]
            + [f"sigma_R_{v}d" for v in ventanas_vol]
            + [f"sigma_D_{v}d" for v in ventanas_vol]
            + [f"ma_R_{v}d" for v in ventanas_vol]
            + [f"ma_D_{v}d" for v in ventanas_vol]
            + ["delta_R", "delta_D"]
            + ["R_conf_t1", "R_conf_t2", "D_conf_t1"]
            + [f"sigma_flujo_{v}d" for v in VENTANAS_FLUJO]
            + [f"ma_flujo_{v}d"    for v in VENTANAS_FLUJO]
            + ["sigma_22d", "flujo_neto_acum_mes"]
            + ["flujo_neto_sum_5d", "flujo_neto_sum_22d", "flujo_neto_sum_66d"]
        )
        return pd.DataFrame(columns=cols)

    df = df_banco[["R", "D"]].copy()
    resultado = pd.DataFrame(index=df.index)

    # Valores de hoy: disponibles en t por aviso anticipado
    resultado["R_t0"] = df["R"]
    resultado["D_t0"] = df["D"]

    todos_lags = lags_cortos + [lag_semana, lag_mes]
    for l in todos_lags:
        resultado[f"R_t-{l}"] = df["R"].shift(l)
        resultado[f"D_t-{l}"] = df["D"].shift(l)

    for v in ventanas_vol:
        resultado[f"sigma_R_{v}d"] = df["R"].rolling(v).std()
        resultado[f"sigma_D_{v}d"] = df["D"].rolling(v).std()
        resultado[f"ma_R_{v}d"]    = df["R"].rolling(v).mean()
        resultado[f"ma_D_{v}d"]    = df["D"].rolling(v).mean()

    resultado["delta_R"] = df["R"].diff(1)
    resultado["delta_D"] = df["D"].diff(1)

    resultado["R_conf_t1"] = df["R"].shift(-1)
    resultado["R_conf_t2"] = df["R"].shift(-2)
    resultado["D_conf_t1"] = df["D"].shift(-1)

    # Features de régimen: rolling std/media + GARCH(1,1)
    flujo = df["D"] - df["R"]
    for v in VENTANAS_FLUJO:
        resultado[f"sigma_flujo_{v}d"] = flujo.rolling(v).std()
        resultado[f"ma_flujo_{v}d"]    = flujo.rolling(v).mean()

    resultado["sigma_22d"] = flujo.rolling(22, min_periods=5).std()

    # Ratio de volatilidad corta/larga del flujo neto: > 1 → régimen de alta vol reciente
    resultado["sigma_flujo_ratio"] = (
        resultado["sigma_flujo_5d"] / resultado["sigma_flujo_20d"].replace(0, np.nan)
    )

    # Acumulado del flujo neto (D−R) desde el primer día hábil del mes hasta t.
    # Lógica: una acumulación de depósitos netos a lo largo del mes anticipa
    # mayores retiros en los últimos días, y viceversa.
    resultado["flujo_neto_acum_mes"] = (
        flujo.groupby(flujo.index.to_period("M")).cumsum()
    )

    # Agregados rolling del flujo neto — ventana siempre relativa a t
    # min_periods=N garantiza NaN para filas con historia insuficiente (→ imputadas por mediana del fold)
    resultado["flujo_neto_sum_5d"]  = flujo.rolling( 5, min_periods= 5).sum()
    resultado["flujo_neto_sum_22d"] = flujo.rolling(22, min_periods=22).sum()
    resultado["flujo_neto_sum_66d"] = flujo.rolling(66, min_periods=66).sum()

    return resultado


# 5b. Features macroeconómicas
def build_macro_features(macro_df):
    """
    Genera features macroeconómicas a partir del DataFrame de series externas.
    """
    if macro_df.empty:
        cols = [
            "VIX", "delta_VIX", "VIX_ma22",
            "TC_PEN_USD", "delta_TC", "tc_vol_5d", "tc_vol_22d", "garch_vol_tc",
            "EMBI_PERU", "delta_EMBI", "garch_vol_embi",
            "TASA_REF_BCRP", "FED_FUNDS", "diferencial_tasas", "T10Y",
            "EMBI_PERU_frac", "T10Y_frac", "CDS_PERU_5Y_frac", "COPPER_frac", "VIX_frac",
        ]
        return pd.DataFrame(columns=cols)

    df = macro_df.copy()
    resultado = pd.DataFrame(index=df.index)

    resultado["VIX"] = df.get("VIX", np.nan)
    resultado["delta_VIX"] = df.get("VIX", pd.Series(dtype=float)).diff(1)
    resultado["VIX_ma22"] = df.get("VIX", pd.Series(dtype=float)).rolling(22).mean()

    resultado["TC_PEN_USD"] = df.get("TC_PEN_USD", np.nan)
    tc = df.get("TC_PEN_USD", pd.Series(dtype=float))
    resultado["delta_TC"] = tc.diff(1)
    retornos_tc = tc.pct_change()
    resultado["tc_vol_5d"] = retornos_tc.rolling(5).std()
    resultado["tc_vol_22d"] = retornos_tc.rolling(22).std()

    # Ratio de volatilidad cambiaria corta/larga: > 1 → estrés cambiario reciente
    resultado["tc_vol_ratio"] = (
        resultado["tc_vol_5d"] / resultado["tc_vol_22d"].replace(0, np.nan)
    )

    retornos_log_tc = np.log(tc / tc.shift(1))
    resultado["garch_vol_tc"] = _garch_vol(retornos_log_tc)

    resultado["EMBI_PERU"] = df.get("EMBI_PERU", np.nan)
    delta_embi = df.get("EMBI_PERU", pd.Series(dtype=float)).diff(1)
    resultado["delta_EMBI"] = delta_embi
    resultado["garch_vol_embi"] = _garch_vol(delta_embi)

    resultado["TASA_REF_BCRP"] = df.get("TASA_REF_BCRP", np.nan)
    resultado["FED_FUNDS"] = df.get("FED_FUNDS", np.nan)

    tasa_ref = df.get("TASA_REF_BCRP", pd.Series(dtype=float))
    fed_funds = df.get("FED_FUNDS", pd.Series(dtype=float))
    if not tasa_ref.empty and not fed_funds.empty:
        resultado["diferencial_tasas"] = tasa_ref.values - fed_funds.reindex(tasa_ref.index).values
    else:
        resultado["diferencial_tasas"] = np.nan

    resultado["T10Y"] = df.get("T10Y", np.nan)

    # ── FFD: diferenciación fraccional para series I(1) ────────────────────────
    # d_opt calibrado en datos hasta _FD_TRAIN_CUTOFF (sin leakage).
    # Si statsmodels no está disponible usa d_opt=0.4 como fallback conservador.
    _fd_cutoff = pd.Timestamp(_FD_TRAIN_CUTOFF)
    _ffd_targets = [
        ("EMBI_PERU",   "EMBI_PERU_frac"),
        ("T10Y",        "T10Y_frac"),
        ("CDS_PERU_5Y", "CDS_PERU_5Y_frac"),
        ("COPPER",      "COPPER_frac"),
        ("VIX",         "VIX_frac"),
    ]
    for col_raw, col_frac in _ffd_targets:
        serie = df.get(col_raw, pd.Series(dtype=float))
        if isinstance(serie, pd.Series) and serie.dropna().shape[0] >= 60:
            s_train = serie[serie.index <= _fd_cutoff].dropna()
            d_opt   = _find_min_d(s_train) if len(s_train) >= 60 else 0.4
            resultado[col_frac] = _fracdiff_fixed_width(serie, d_opt)
            logger.info(f"  FFD {col_raw} → {col_frac}  d_opt={d_opt:.4f}")
        else:
            resultado[col_frac] = np.nan

    return resultado


# 5c. Features de encaje (BBVA-específicas, con rezago 1 día)
def build_encaje_features(df_encaje, peru_bday):
    """
    Deriva features de encaje a partir de EncajeD.xlsx.
    Todas las variables usan información de t-1 (sin leakage).

    Features generadas:
      encaje_lag1         : Caja + Cta Cte BCR del día t-1 (M USD)
      exceso_lag1         : encaje_lag1 - encaje_exigible_lag1
      faltante_lag1       : exigible_total_mes - encaje_acum(t-1)
                            = lo que aún necesita acumular para cerrar el mes
      techo_10h           : CC + ON al día hábil que queda a 10 días hábiles
                            del cierre del mes. Disponible solo cuando t está
                            dentro de los últimos 10 días hábiles del mes.
      techo_restante_lag1 : techo_10h - retiro_acum_mes(t-1)
                            = presupuesto de retiro aún disponible
      proporcion_usada    : retiro_acum_mes(t-1) / techo_10h (0 si techo=0)

    Retorna DataFrame indexado por fecha, alineado con los datos de encaje.
    """
    if df_encaje.empty:
        return pd.DataFrame()

    df = df_encaje.copy()
    resultado = pd.DataFrame(index=df.index)

    # Encaje disponible (solo Caja + Cta Cte BCR; Overnight NO cuenta)
    if "caja" in df.columns and "cta_cte_bcr" in df.columns:
        encaje = df["caja"] + df["cta_cte_bcr"]
    else:
        logger.warning("  build_encaje_features: faltan columnas caja / cta_cte_bcr")
        return pd.DataFrame()

    overnight = df.get("overnight_bcr", pd.Series(0.0, index=df.index))
    exigible  = df.get("encaje_exigible", pd.Series(np.nan, index=df.index))
    retiro    = df.get("retiro_neto", pd.Series(np.nan, index=df.index))

    # ── 1. Encaje y exceso con rezago 1 día ─────────────────────────────────
    resultado["encaje_lag1"]   = encaje.shift(1)
    resultado["exceso_lag1"]   = (encaje - exigible).shift(1)

    # ── 2. Faltante = lo que el banco aún necesita acumular este mes ─────────
    # exigible_total_mes = encaje_exigible × días_calendario_en_mes
    # encaje_acum(t-1)   = suma de encaje diario desde inicio del mes hasta t-1
    #
    # Sólo días donde encaje_exigible no es NaN participan en el cálculo.
    exigible_mensual = exigible.groupby(exigible.index.to_period("M")).transform("first")
    dias_mes = pd.Series(
        df.index.to_series().dt.daysinmonth.values, index=df.index
    )
    exigible_total = exigible_mensual * dias_mes

    encaje_acum = encaje.groupby(encaje.index.to_period("M")).cumsum()
    faltante_raw = exigible_total - encaje_acum          # positivo = necesita más
    resultado["faltante_lag1"] = faltante_raw.shift(1)

    # ── 3. Techo (CC + ON a 10 días hábiles del cierre del mes) ─────────────
    # Para cada fecha, identificar el día que queda a 10 días hábiles del
    # cierre del mes en su respectivo mes. El techo se define solo una vez
    # por mes (en ese día puntual) y se propaga hacia adelante dentro del mes.
    cc_on = encaje + overnight   # CC + ON es el instrumento observado antes de ON

    techo_por_fecha = pd.Series(np.nan, index=df.index)
    for periodo, grupo in cc_on.groupby(cc_on.index.to_period("M")):
        inicio = periodo.start_time
        if inicio.month == 12:
            fin = pd.Timestamp(year=inicio.year + 1, month=1, day=1) - pd.Timedelta(days=1)
        else:
            fin = pd.Timestamp(year=inicio.year, month=inicio.month + 1, day=1) - pd.Timedelta(days=1)

        bdays_mes = pd.bdate_range(start=inicio, end=fin, freq=peru_bday)
        n_bd = len(bdays_mes)
        if n_bd < 10:
            continue

        fecha_techo = bdays_mes[n_bd - 10]   # día hábil que está a 10 días del cierre
        if fecha_techo in cc_on.index:
            val_techo = cc_on.loc[fecha_techo]
            fechas_grupo = grupo.index
            fechas_post = fechas_grupo[fechas_grupo >= fecha_techo]
            techo_por_fecha.loc[fechas_post] = val_techo

    resultado["techo_10h"] = techo_por_fecha.shift(1)   # rezago 1 día (conocido en t)

    # ── 4. Retiro acumulado del mes hasta t-1 ────────────────────────────────
    retiro_acum_mes = retiro.groupby(retiro.index.to_period("M")).cumsum().shift(1)

    # ── 5. Techo restante y proporción usada ─────────────────────────────────
    resultado["techo_restante_lag1"] = resultado["techo_10h"] - retiro_acum_mes.abs()
    resultado["proporcion_usada"] = (
        retiro_acum_mes.abs() / resultado["techo_10h"].replace(0, np.nan)
    ).clip(0, None)

    n_validos = resultado["faltante_lag1"].notna().sum()
    logger.info(
        f"  build_encaje_features: {n_validos:,} filas con faltante_lag1 válido | "
        f"techo_10h disponible: {resultado['techo_10h'].notna().sum():,} filas"
    )
    return resultado


# 5c-bis. Features de encaje BBVA desde Excel generado por aux_encaje_2.py
def load_bbva_encaje_features(params):
    """
    Lee bbva_encaje_features_modelo.xlsx (generado por aux_encaje_2.py).
    Retorna un DataFrame indexado en días CALENDARIO con las 4 features
    lag1 ya calculadas correctamente sobre ese índice.

    La adaptación al calendario hábil de step001 se hace con merge_asof
    en build_feature_matrix (busca la fecha calendario más reciente ≤ fecha_t).
    Así se evita el problema de construir un índice hábil pre-computado que
    no coincide exactamente con los fecha_t del dataset real.
    """
    _COLS = ["avance_mes_lag1", "exceso_abs_lag1", "exceso_dia_lag1",
             "encaje_ovn_lag1", "ratio_ovn_total_lag1"]

    ruta = params.get("ruta_bbva_encaje_features", "")
    if not ruta or not Path(ruta).exists():
        logger.info("  bbva_encaje_features: archivo no encontrado — features de avance/exceso omitidas.")
        return pd.DataFrame()

    try:
        df_raw = pd.read_excel(ruta, sheet_name="Datos")
        # Seleccionar solo columnas disponibles — si el Excel no tiene alguna
        # (p.ej. versión anterior sin exceso_dia_lag1) se omite sin fallar.
        cols_ok = ["fecha"] + [c for c in _COLS if c in df_raw.columns]
        cols_missing = [c for c in _COLS if c not in df_raw.columns]
        if cols_missing:
            logger.warning(f"  bbva_encaje_features: columnas ausentes en Excel — {cols_missing}")
        df = df_raw[cols_ok].copy()
        df["fecha"] = pd.to_datetime(df["fecha"]).dt.normalize()
        df = df.sort_values("fecha").reset_index(drop=True)

        n_val = df["avance_mes_lag1"].notna().sum() if "avance_mes_lag1" in df.columns else 0
        logger.info(
            f"  bbva_encaje_features: {len(df):,} filas calendario cargadas "
            f"({df['fecha'].min().date()} → {df['fecha'].max().date()}) | "
            f"avance_mes_lag1 válido: {n_val:,} | cols: {cols_ok[1:]}"
        )
        return df

    except Exception as e:
        logger.warning(f"  bbva_encaje_features: error al cargar — {e}")
        return pd.DataFrame()


# 5d. Features CC+OVN en BCR (Saldos_CCOVN.xlsx, rezago 1 día)
def build_ccovn_features(df_ccovn, peru_bday, flujo_sistema=None):
    """
    Deriva features de saldos CC+OVN en el BCR a partir de Saldos_CCOVN.xlsx.
    Todas las variables usan información de t-1 (shift(1)) → sin leakage.

    Reindexación a freq="B" (lunes-viernes) antes de diff():
      - Los feriados peruanos (sin datos) se forward-fill desde el día previo.
      - Los feriados de EE.UU. que NO son feriados en Perú (ej. MLK Day,
        Presidents Day) quedan incluidos como días hábiles normales, lo que
        es correcto porque los bancos peruanos operan esos días y el BCRP
        reporta el saldo. No se usa peru_bday (joint PE+US) para este
        reindex porque ese calendario los excluiría, generando NaN en la
        matriz donde sí existen filas de fecha_t.

    Parámetros:
      flujo_sistema : pd.Series opcional (D-R del sistema, indexada por fecha).
                      Si se provee, se calcula residuo_ccovn_lag1.

    Features generadas (todas con sufijo _lag1):
      ccovn_sistema_lag1        : saldo total del sistema en el BCR en t-1 (USD)
      ccovn_bbva_lag1           : saldo BBVA en el BCR en t-1 (USD)
      var_ccovn_sistema_lag1    : variación entre días hábiles consecutivos del sistema en t-1
      var_ccovn_bbva_lag1       : variación entre días hábiles consecutivos de BBVA en t-1
      bbva_share_lag1           : ccovn_bbva / ccovn_sistema en t-1
      var_ccovn_bbva_exceso_lag1: Δsaldo_bbva - bbva_share×Δsaldo_sistema en t-1
                                  (componente idiosincrático de BBVA, ortogonal al sistema)
      ccovn_vs_dia_mes_lag1     : saldo_sistema(t-1) - media_historica[rango_dia_habil_mes]
                                  (desviación respecto al nivel estacional esperado para ese
                                  puesto del mes: elimina estacionalidad intramonth)
      residuo_ccovn_lag1        : Δsaldo_sistema(t-1) - flujo_neto_sistema(t-1)
                                  (error de la identidad Δsaldo≈flujo; solo si flujo_sistema
                                  no es None)

    Retorna DataFrame indexado por días lun-vie (freq="B").
    """
    if df_ccovn.empty:
        return pd.DataFrame()

    # Reindexar a lun-vie: cubre feriados PE (ffill) y feriados US (incluidos
    # como hábiles en Perú). Esto alinea con todas las fechas posibles de la
    # matriz aunque el calendario sea PE+US.
    idx_bd = pd.bdate_range(
        start=df_ccovn.index.min(),
        end=df_ccovn.index.max(),
        freq="B",
    )
    df_bd = df_ccovn.reindex(idx_bd).ffill()

    sis = df_bd["ccovn_sistema"]
    bbv = df_bd.get("ccovn_bbva", pd.Series(np.nan, index=df_bd.index))

    resultado = pd.DataFrame(index=df_bd.index)
    resultado["ccovn_sistema_lag1"]     = sis.shift(1)
    resultado["ccovn_bbva_lag1"]        = bbv.shift(1)
    # diff() sobre idx_bd → variación entre días lun-vie consecutivos (correcto)
    resultado["var_ccovn_sistema_lag1"] = sis.diff().shift(1)
    resultado["var_ccovn_bbva_lag1"]    = bbv.diff().shift(1)
    share = bbv / sis.replace(0, np.nan)
    resultado["bbva_share_lag1"]        = share.shift(1)

    # ── var_ccovn_bbva_exceso_lag1 ───────────────────────────────────────────
    # Componente idiosincrático de BBVA en la variación diaria del saldo:
    #   exceso = Δsaldo_bbva - bbva_share × Δsaldo_sistema
    # Mide cuánto se apartó BBVA del comportamiento proporcional al sistema.
    # Ortogonal a la variación sistémica → señal específica del banco.
    resultado["var_ccovn_bbva_exceso_lag1"] = (
        bbv.diff() - share * sis.diff()
    ).shift(1)

    # ── ccovn_vs_dia_mes_lag1 ────────────────────────────────────────────────
    # Desviación del saldo del sistema respecto a su media histórica para ese
    # mismo puesto (rango) dentro del mes hábil:
    #   rango = 1 si es el 1er día hábil del mes, 2 si es el 2do, etc.
    #   media_por_rango = media expandida (solo datos hasta t-1) del saldo para
    #                     cada rango, evitando leakage de información futura.
    #   ccovn_vs_dia_mes = saldo_sistema - media_por_rango
    # Elimina el patrón estacional intramonth (ingresos inicio / retiros fin).
    # XGBoost ya tiene dia_habil_del_mes como feature; esta variable aporta
    # la desviación INESPERADA respecto a ese patrón estacional.
    rango_dia = (
        pd.Series(np.arange(len(df_bd)), index=df_bd.index)
        .groupby(df_bd.index.to_period("M"))
        .transform(lambda x: np.arange(1, len(x) + 1))
    )
    media_por_rango = (
        sis.groupby(rango_dia)
           .transform(lambda x: x.expanding().mean().shift(1))
    )
    resultado["ccovn_vs_dia_mes_lag1"] = (sis - media_por_rango).shift(1)

    # ── residuo_ccovn_lag1 ───────────────────────────────────────────────────
    # Error de la identidad Δsaldo_sistema ≈ flujo_neto_sistema:
    #   residuo = Δsaldo_sistema - flujo_neto_sistema
    # Captura variación en el saldo del BCR no explicada por los flujos D-R
    # reportados (diferencias de timing, componentes no observados, etc.).
    # Para horizontes cortos actúa como señal de reversión; si residuo > 0,
    # el saldo cayó más de lo que justifican los flujos → corrección esperada.
    if flujo_sistema is not None and not flujo_sistema.empty:
        flujo_bd = flujo_sistema.reindex(idx_bd).ffill()
        resultado["residuo_ccovn_lag1"] = (sis.diff() - flujo_bd).shift(1)
        n_res = resultado["residuo_ccovn_lag1"].notna().sum()
        logger.info(f"  build_ccovn_features: residuo_ccovn_lag1 calculado "
                    f"({n_res:,} filas válidas)")
    else:
        logger.info("  build_ccovn_features: flujo_sistema no disponible — "
                    "residuo_ccovn_lag1 omitido")

    n_val = resultado["ccovn_sistema_lag1"].notna().sum()
    logger.info(f"  build_ccovn_features: {n_val:,} filas válidas en ccovn_sistema_lag1")
    return resultado


# 5e. Features estacionales en t+h
def _get_bdays_en_mes(fecha, peru_bday):
    """Retorna todos los días hábiles del mes de la fecha dada."""
    inicio_mes = fecha.replace(day=1)
    if fecha.month == 12:
        fin_mes = fecha.replace(year=fecha.year + 1, month=1, day=1) - pd.Timedelta(days=1)
    else:
        fin_mes = fecha.replace(month=fecha.month + 1, day=1) - pd.Timedelta(days=1)
    return pd.bdate_range(start=inicio_mes, end=fin_mes, freq=peru_bday)


def _get_bdays_en_trim(fecha, peru_bday):
    """Retorna todos los días hábiles del trimestre de la fecha dada."""
    trim = (fecha.month - 1) // 3
    inicio_trim = fecha.replace(month=trim * 3 + 1, day=1)
    if trim == 3:
        fin_trim = fecha.replace(year=fecha.year + 1, month=1, day=1) - pd.Timedelta(days=1)
    else:
        fin_trim = fecha.replace(month=(trim + 1) * 3 + 1, day=1) - pd.Timedelta(days=1)
    return pd.bdate_range(start=inicio_trim, end=fin_trim, freq=peru_bday)


###############################################################################
# PARTE 6 — Construcción de la matriz de features completa
###############################################################################
def _build_seasonal_table(fechas_unicas, peru_holidays, fechas_elecciones, peru_bday):
    """
    Calcula features estacionales para un conjunto de fechas únicas (t+h).
    Retorna DataFrame indexado por fecha.
    """
    # Pre-calcular días hábiles por mes y trimestre una sola vez
    meses_unicos = pd.PeriodIndex(fechas_unicas, freq="M").unique()
    trims_unicos = pd.PeriodIndex(fechas_unicas, freq="Q").unique()

    # Mapa: fecha → días hábiles de su mes
    bdays_por_mes = {}
    for m in meses_unicos:
        inicio = m.start_time
        fin    = m.end_time
        bdays_por_mes[m] = pd.bdate_range(start=inicio, end=fin, freq=peru_bday)

    # Mapa: fecha → días hábiles de su trimestre
    bdays_por_trim = {}
    for q in trims_unicos:
        inicio = q.start_time
        fin    = q.end_time
        bdays_por_trim[q] = pd.bdate_range(start=inicio, end=fin, freq=peru_bday)

    # Mapa: año → días hábiles + dict de posición O(1)
    anios_unicos = pd.PeriodIndex(fechas_unicas, freq="Y").unique()
    bdays_por_anio     = {}
    bdays_pos_por_anio = {}
    for _a in anios_unicos:
        _bd = pd.bdate_range(start=_a.start_time, end=_a.end_time, freq=peru_bday)
        bdays_por_anio[_a]     = _bd
        bdays_pos_por_anio[_a] = {d: i + 1 for i, d in enumerate(_bd)}

    # Sets para lookups rápidos
    hols_set  = set(peru_holidays)
    elec_set  = set(pd.DatetimeIndex(fechas_elecciones)) if fechas_elecciones else set()
    elec_arr  = np.array(sorted(elec_set), dtype="datetime64[D]") if elec_set else np.array([], dtype="datetime64[D]")

    # Feriados en formato numpy para np.busday_count (ciclo electoral)
    hols_np = (peru_holidays.values.astype("datetime64[D]")
               if len(peru_holidays) > 0
               else np.array([], dtype="datetime64[D]"))

    TWO_PI = 2.0 * np.pi

    registros = []
    for fecha in fechas_unicas:
        ts = pd.Timestamp(fecha)
        m  = ts.to_period("M")
        q  = ts.to_period("Q")

        bd_mes  = bdays_por_mes.get(m,  pd.DatetimeIndex([]))
        bd_trim = bdays_por_trim.get(q, pd.DatetimeIndex([]))
        bd_mes_list  = list(bd_mes)
        bd_trim_list = list(bd_trim)

        total_bdays_mes  = len(bd_mes_list)
        total_bdays_trim = len(bd_trim_list)

        try:
            pos_en_mes  = bd_mes_list.index(ts) + 1
        except ValueError:
            pos_en_mes  = 1
        try:
            pos_en_trim = bd_trim_list.index(ts) + 1
        except ValueError:
            pos_en_trim = 1

        dias_al_cierre_mes    = total_bdays_mes - pos_en_mes
        dias_desde_cierre_mes = pos_en_mes - 1

        is_penult_bday_trim = int(pos_en_trim == total_bdays_trim - 1)
        is_ultimo_bday_trim = int(pos_en_trim == total_bdays_trim)
        is_1er_bday_trim    = int(pos_en_trim == 1)
        is_2do_bday_trim    = int(pos_en_trim == 2)
        is_3er_bday_trim    = int(pos_en_trim == 3)

        # Feriados (busca en ventana ±2 días calendario)
        is_pre_feriado  = int(any((ts + pd.Timedelta(days=d)) in hols_set for d in [1, 2]))
        is_post_feriado = int(any((ts - pd.Timedelta(days=d)) in hols_set for d in [1, 2]))

        # Elecciones
        if len(elec_arr) > 0:
            ts_d  = np.datetime64(ts, "D")
            diffs = (elec_arr - ts_d).astype(int)
            futuros_e        = diffs[diffs > 0]
            pasados_e        = (-diffs)[diffs < 0]
            is_pre_eleccion  = int(len(futuros_e) > 0 and futuros_e.min() <= 7)
            is_post_eleccion = int(len(pasados_e) > 0 and pasados_e.min() <= 7)
        else:
            is_pre_eleccion = is_post_eleccion = 0

        mes = ts.month

        # ── Posición en el año (días hábiles) ────────────────────────────────
        _a               = ts.to_period("Y")
        total_bdays_anio = len(bdays_por_anio.get(_a, []))
        pos_en_anio      = bdays_pos_por_anio.get(_a, {}).get(ts, 1)
        dias_al_cierre_anio = total_bdays_anio - pos_en_anio

        # ── Días hábiles desde última elección general ────────────────────────
        _ts64     = np.datetime64(ts, "D")
        _pasados  = _ELECCIONES_GENERALES_NP[_ELECCIONES_GENERALES_NP <= _ts64]
        n_bd_elec = (int(np.busday_count(_pasados[-1], _ts64, holidays=hols_np))
                     if len(_pasados) > 0 else 0)

        # ── Codificaciones cíclicas (sin/cos) ─────────────────────────────────
        # mes (P=12, calendario)
        mes_sin = float(np.sin(TWO_PI * mes / 12))
        mes_cos = float(np.cos(TWO_PI * mes / 12))

        # día de semana (P=5 hábiles; dayofweek 0=lun→4=vie, usamos 1-5)
        _dsem        = ts.dayofweek + 1
        dias_sem_sin = float(np.sin(TWO_PI * _dsem / 5))
        dias_sem_cos = float(np.cos(TWO_PI * _dsem / 5))

        # cierre de mes (P=total_bdays_mes hábiles)
        _p_mes = total_bdays_mes or 1
        dias_al_cierre_mes_sin = float(np.sin(TWO_PI * pos_en_mes / _p_mes))
        dias_al_cierre_mes_cos = float(np.cos(TWO_PI * pos_en_mes / _p_mes))

        # cierre de trimestre (P=total_bdays_trim hábiles)
        _p_trim = total_bdays_trim or 1
        dias_al_cierre_trim_sin = float(np.sin(TWO_PI * pos_en_trim / _p_trim))
        dias_al_cierre_trim_cos = float(np.cos(TWO_PI * pos_en_trim / _p_trim))

        # cierre de año (P=total_bdays_anio hábiles)
        _p_anio = total_bdays_anio or 1
        dias_al_cierre_anio_sin = float(np.sin(TWO_PI * pos_en_anio / _p_anio))
        dias_al_cierre_anio_cos = float(np.cos(TWO_PI * pos_en_anio / _p_anio))

        # ciclo electoral (P=1260 hábiles ≈ 5 años, solo elecciones generales)
        elec_sin = float(np.sin(TWO_PI * n_bd_elec / _P_ELEC))
        elec_cos = float(np.cos(TWO_PI * n_bd_elec / _P_ELEC))

        registros.append({
            "fecha_th"             : ts,
            "dias_al_cierre_mes"   : dias_al_cierre_mes,
            "dias_desde_cierre_mes": dias_desde_cierre_mes,
            "pos_en_mes"           : pos_en_mes,
            "total_bdays_mes"      : total_bdays_mes,
            "is_penult_bday_trim"  : is_penult_bday_trim,
            "is_ultimo_bday_trim"  : is_ultimo_bday_trim,
            "is_1er_bday_trim"     : is_1er_bday_trim,
            "is_2do_bday_trim"     : is_2do_bday_trim,
            "is_3er_bday_trim"     : is_3er_bday_trim,
            "dia_semana"           : ts.dayofweek,
            "mes"                  : mes,
            "is_quincena"          : int(pos_en_mes == 15 or dias_al_cierre_mes == 0),
            "is_cierre_encaje"     : int(dias_al_cierre_mes <= 1),
            "is_fin_anio"          : int(mes == 12 and ts.day >= 28),
            "is_pre_feriado"       : is_pre_feriado,
            "is_post_feriado"      : is_post_feriado,
            "is_pre_eleccion"      : is_pre_eleccion,
            "is_post_eleccion"     : is_post_eleccion,
            # ── Cíclicas (sin/cos) ────────────────────────────────────────────
            "mes_sin"                : mes_sin,
            "mes_cos"                : mes_cos,
            "dias_sem_sin"           : dias_sem_sin,
            "dias_sem_cos"           : dias_sem_cos,
            "dias_al_cierre_mes_sin" : dias_al_cierre_mes_sin,
            "dias_al_cierre_mes_cos" : dias_al_cierre_mes_cos,
            "dias_al_cierre_trim_sin": dias_al_cierre_trim_sin,
            "dias_al_cierre_trim_cos": dias_al_cierre_trim_cos,
            "dias_al_cierre_anio_sin": dias_al_cierre_anio_sin,
            "dias_al_cierre_anio_cos": dias_al_cierre_anio_cos,
            "elec_sin"               : elec_sin,
            "elec_cos"               : elec_cos,
        })

    return pd.DataFrame(registros).set_index("fecha_th")


def build_feature_matrix(
    banco,
    datos_manuales,
    macro_features,
    bank_features,
    peru_bday,
    peru_holidays,
    fechas_elecciones,
    h_min,
    h_max,
    hmm_features=None,     # pd.Series(hmm_estado, index=fecha_t) pre-computada
    encaje_features=None,        # pd.DataFrame(index=fecha_t) con features de encaje (BBVA)
    banco_encaje="BBVA",         # banco al que aplican esas features
    bbva_encaje_feat=None,       # pd.DataFrame(index=fecha_t) avance/exceso desde aux_encaje_2
    ccovn_features=None,         # pd.DataFrame(index=fecha_t) con saldos CC+OVN para todos los bancos
):
    """
    Construye el dataset de entrenamiento para un banco de forma vectorizada.
    """
    logger.info(f"  Construyendo matriz para banco: {banco}")

    df_bancarios  = datos_manuales.get("bancarios",  pd.DataFrame())
    df_confirmados = datos_manuales.get("confirmados", pd.DataFrame())

    # Índice temporal de referencia
    if not df_bancarios.empty and f"{banco}_R" in df_bancarios.columns:
        fechas_t = df_bancarios.index
    elif not macro_features.empty:
        fechas_t = macro_features.index
    else:
        logger.warning(f"  Sin fechas de referencia para {banco}.")
        return pd.DataFrame()

    fechas_t = pd.DatetimeIndex(fechas_t)

    # Eliminar feriados de fecha_t.
    # El reindex de bancarios usa pd.bdate_range (Lun-Vie sin exclusión de feriados)
    # para preservar los ceros de días festivos en los lags — correcto, el banco
    # realmente tuvo 0 transacciones ese día. Pero esos días NO deben ser puntos
    # de predicción: (a) target duplica al día hábil anterior porque peru_bday los
    # salta al calcular fecha_th, y (b) todas sus features son cero o forward-fill.
    if len(peru_holidays) > 0:
        _n_antes = len(fechas_t)
        fechas_t = fechas_t[~fechas_t.normalize().isin(peru_holidays.normalize())]
        _n_feriados = _n_antes - len(fechas_t)
        if _n_feriados:
            logger.info(
                f"    {banco}: {_n_feriados} fechas de feriados eliminadas de fecha_t "
                f"({_n_antes} → {len(fechas_t)})"
            )

    # ── 1. Grid (fecha_t × h) ────────────────────────────────────────────────
    hs = np.arange(h_min, h_max + 1)
    grid = pd.MultiIndex.from_product([fechas_t, hs], names=["fecha_t", "h"])
    df = pd.DataFrame(index=grid).reset_index()
    df["log_h"] = np.log(df["h"])
    df["banco"] = banco

    # Calcular fecha_th vectorialmente: aplicar el offset h veces
    # Construimos un mapa fecha_t → array de fechas futuras
    logger.info(f"    Calculando fechas futuras ({len(fechas_t)} × {len(hs)})...")
    th_map = {}
    for t in fechas_t:
        th_map[t] = list(get_future_business_dates(t, h_max, peru_bday))

    def get_th(row):
        fechas = th_map.get(row["fecha_t"], [])
        idx = row["h"] - 1
        return fechas[idx] if idx < len(fechas) else pd.NaT

    df["fecha_th"] = df.apply(get_th, axis=1)
    df = df.dropna(subset=["fecha_th"])

    # ── 1b. Features de gap de calendario ───────────────────────────────────
    # dias_desde_ultimo_habil: días calendarios desde el BDay anterior.
    #   Lun normal = 3 | Lun post-Semana Santa = 5 | cualquier Mar-Vie = 1
    # es_post_feriado: 1 si el gap es mayor a un fin de semana estándar (>3 días).
    #   Señala al árbol que los flujos entrantes pueden incluir acumulación de feriados.
    # Nota: al calcular sobre fechas_t ya filtradas de feriados, shift(1) apunta
    # siempre al día hábil inmediato anterior, con el gap correcto.
    _ft_sorted = fechas_t.sort_values()
    _gap = pd.Series(
        (_ft_sorted.to_series().diff().dt.days).fillna(1).astype(int).values,
        index=_ft_sorted,
        name="dias_desde_ultimo_habil",
    )
    _post_feriado = (_gap > 3).astype(int).rename("es_post_feriado")
    _cal_gap_df = pd.DataFrame({"dias_desde_ultimo_habil": _gap,
                                 "es_post_feriado": _post_feriado})

    # ── 2. Features bancarias (merge por fecha_t) ────────────────────────────
    bf = bank_features.get(banco, pd.DataFrame())
    if not bf.empty:
        df = df.merge(bf.add_prefix(""), left_on="fecha_t", right_index=True, how="left")

    df = df.merge(_cal_gap_df, left_on="fecha_t", right_index=True, how="left")

    # ── 3. Datos bancarios → target ──────────────────────────────────────────
    tiene_bancarios = (
        not df_bancarios.empty
        and f"{banco}_R" in df_bancarios.columns
        and f"{banco}_D" in df_bancarios.columns
    )
    if tiene_bancarios:
        df_banco = df_bancarios[[f"{banco}_R", f"{banco}_D"]].copy()
        df_banco.columns = ["R_th", "D_th"]
        df_banco["target"] = df_banco["D_th"] - df_banco["R_th"]
        df = df.merge(df_banco[["target"]], left_on="fecha_th", right_index=True, how="left")
    else:
        df["target"] = np.nan

    # ── 4. Confirmados ───────────────────────────────────────────────────────
    # Base: valores calculados en build_bank_features (shift -1/-2 sobre realizados)
    # Override: si existen confirmados operativos reales (correos), reemplazan la fila exacta
    if not df_confirmados.empty and "banco" in df_confirmados.columns and "fecha" in df_confirmados.columns:
        conf = df_confirmados[df_confirmados["banco"] == banco].set_index("fecha")
        for col in ["R_conf_t1", "R_conf_t2", "D_conf_t1"]:
            if col in conf.columns:
                override = df["fecha_t"].map(conf[col])
                df[col] = override.combine_first(df[col])

    # ── 5. Macro (merge por fecha_t) ─────────────────────────────────────────
    if not macro_features.empty:
        df = df.merge(macro_features, left_on="fecha_t", right_index=True, how="left")
    else:
        for col in ["VIX","delta_VIX","VIX_ma22","TC_PEN_USD","delta_TC",
                    "tc_vol_5d","tc_vol_22d","garch_vol_tc",
                    "EMBI_PERU","delta_EMBI","garch_vol_embi",
                    "TASA_REF_BCRP","FED_FUNDS","diferencial_tasas","T10Y",
                    "EMBI_PERU_frac","T10Y_frac","CDS_PERU_5Y_frac","COPPER_frac","VIX_frac"]:
            df[col] = np.nan

    # ── 6. Estacionales (calculadas una vez por fecha_th única) ──────────────
    fechas_th_unicas = pd.DatetimeIndex(df["fecha_th"].unique())
    logger.info(f"    Calculando estacionales para {len(fechas_th_unicas):,} fechas únicas...")
    df_seasonal = _build_seasonal_table(
        fechas_th_unicas, peru_holidays, fechas_elecciones, peru_bday
    )
    df = df.merge(df_seasonal, left_on="fecha_th", right_index=True, how="left")

    # ── 7. HMM estado de régimen (pre-computado, sin leakage) ────────────────
    # hmm_estado es una Serie indexada por fecha_t calculada una sola vez en
    # build_full_matrix con expanding window. El merge por fecha_t es seguro:
    # hmm_estado[Y] fue calculado con datos hasta Y-1, exactamente lo que el
    # fold vería como train. Años de burn-in quedan en NaN.
    if hmm_features is not None and not hmm_features.empty:
        df = df.merge(
            hmm_features.rename("hmm_estado").to_frame(),
            left_on="fecha_t", right_index=True, how="left",
        )
    else:
        df["hmm_estado"] = pd.NA

    # ── 8. Features de encaje (solo para el banco configurado, ej. BBVA) ─────
    # encaje_features está indexado por fecha_t y contiene valores del día t-1.
    # Columnas: encaje_lag1, exceso_lag1, faltante_lag1, techo_10h,
    #           techo_restante_lag1, proporcion_usada.
    # Para bancos distintos al banco_encaje → columnas quedan en NaN.
    if encaje_features is not None and not encaje_features.empty:
        if banco == banco_encaje:
            df = df.merge(encaje_features, left_on="fecha_t", right_index=True, how="left")
            logger.info(f"  {banco}: features de encaje incorporadas ({len(encaje_features.columns)} cols)")
        else:
            for col in encaje_features.columns:
                df[col] = np.nan

    # ── 8b. Features de avance/exceso encaje BBVA (todos los bancos) ───────────
    # BBVA es el principal driver del movimiento sistémico en ME; sus features
    # de encaje son señales sistémicas relevantes para SISTEMA y cualquier otro
    # banco modelado. Se aplican a todas las entidades sin restricción.
    # bbva_encaje_feat está en días calendario con lag1 ya computado en aux_encaje_2.
    # merge_asof alinea al calendario hábil de cada entidad.
    _bbva_feat_cols = ["avance_mes_lag1", "exceso_abs_lag1", "exceso_dia_lag1",
                       "encaje_ovn_lag1", "ratio_ovn_total_lag1"]
    if bbva_encaje_feat is not None and not bbva_encaje_feat.empty:
        # Construir lookup: fecha_t_única → valor del Excel más reciente ≤ fecha_t.
        # merge_asof sobre fechas únicas evita ambigüedades con múltiples h por fecha.
        _ft_norm = pd.to_datetime(df["fecha_t"]).dt.normalize()
        _dates_unicas = pd.DataFrame(
            {"fecha_t": _ft_norm.unique()}
        ).sort_values("fecha_t")

        _feat_r = bbva_encaje_feat[["fecha"] + _bbva_feat_cols].copy()
        _feat_r["fecha"] = pd.to_datetime(_feat_r["fecha"]).dt.normalize()
        _feat_r = _feat_r.rename(columns={"fecha": "fecha_t"}) \
                          .sort_values("fecha_t")

        _lookup = pd.merge_asof(
            _dates_unicas,
            _feat_r,
            on="fecha_t",
            direction="backward",
        ).set_index("fecha_t")

        for col in _bbva_feat_cols:
            df[col] = _ft_norm.map(_lookup[col]).values

        n_ok = df["avance_mes_lag1"].notna().sum()
        logger.info(f"  {banco}: bbva_encaje_feat incorporadas — "
                    f"{n_ok:,}/{len(df):,} filas con valores")

    # ── 9. Features CC+OVN en BCR (Saldos_CCOVN.xlsx, todos los bancos) ──────
    # ccovn_features está indexado por fecha y contiene valores del día t-1.
    # ccovn_sistema_lag1 aplica a todos los bancos.
    # ccovn_bbva_lag1 aplica a todos (es una variable sistémica disponible para
    # cualquier banco que quiera observar el comportamiento de BBVA).
    if ccovn_features is not None and not ccovn_features.empty:
        df = df.merge(ccovn_features, left_on="fecha_t", right_index=True, how="left")
        logger.info(f"  {banco}: features ccovn incorporadas ({len(ccovn_features.columns)} cols)")

    # ── Reducir a float32 antes de reordenar ────────────────────────────────
    # Mitad de memoria y evita el OOM en la consolidación interna de pandas
    float_cols = df.select_dtypes(include="float64").columns
    df[float_cols] = df[float_cols].astype("float32")

    # ── Excluir features definidos en FEATURES_EXCLUIR ──────────────────────
    excluir = [c for c in FEATURES_EXCLUIR
               if c not in ("fecha_t", "banco", "h", "target")]   # protege identidad
    if excluir:
        df = df.drop(columns=[c for c in excluir if c in df.columns], errors="ignore")
        logger.info(f"  {banco}: {len(excluir)} features excluidos → "
                    f"{df.shape[1]} columnas finales")

    # ── Reordenar columnas ───────────────────────────────────────────────────
    cols_id    = [c for c in ["fecha_t", "banco", "h", "log_h"] if c in df.columns]
    cols_resto = [c for c in df.columns if c not in cols_id + ["fecha_th"]]
    df = df[cols_id + cols_resto].drop(columns=["fecha_th"], errors="ignore")

    # Reporte NaN
    total = len(df)
    nan_cols = {c: f"{df[c].isna().sum()/total:.1%}" for c in df.columns if df[c].isna().any()}
    logger.info(f"  {banco}: {total:,} filas | columnas con NaN: {len(nan_cols)}")

    return df


def build_full_matrix(
    params,
    datos_manuales,
    lista_bancos,
    macro_features,
    peru_bday,
    peru_holidays,
    fechas_elecciones,
):
    """
    Construye la matriz banco por banco y la escribe directamente al Parquet
    de salida usando PyArrow streaming. Nunca acumula más de un banco en RAM.
    Retorna el path del archivo generado (no el DataFrame completo).
    """
    import gc
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq_writer_lib
    except ImportError:
        logger.error("pyarrow no disponible. Instalar con: pip install pyarrow")
        return pd.DataFrame()

    logger.info("PARTE 6: Construyendo matriz completa de features...")

    df_bancarios = datos_manuales.get("bancarios", pd.DataFrame())
    ruta_output  = Path(params.get("ruta_output", "matriz_features.parquet"))
    ruta_output.parent.mkdir(parents=True, exist_ok=True)

    # ── Agregar SISTEMA: suma del sistema bancario completo ──────────────────
    # Las features de SISTEMA se calculan sobre la serie agregada (no suma de
    # features individuales) para que volatilidades y medias móviles sean correctas.
    NOMBRE_SISTEMA = "SISTEMA"
    if not df_bancarios.empty:
        r_cols = [c for c in df_bancarios.columns if c.endswith("_R")]
        d_cols = [c for c in df_bancarios.columns if c.endswith("_D")]
        if r_cols and d_cols:
            df_bancarios = df_bancarios.copy()
            df_bancarios[f"{NOMBRE_SISTEMA}_R"] = df_bancarios[r_cols].sum(axis=1)
            df_bancarios[f"{NOMBRE_SISTEMA}_D"] = df_bancarios[d_cols].sum(axis=1)
            datos_manuales["bancarios"] = df_bancarios
            logger.info(
                f"  SISTEMA agregado: suma de {len(r_cols)} series "
                f"({', '.join(c.replace('_R','') for c in r_cols)})"
            )

    # SISTEMA va primero en la lista para poder validarlo de forma aislada
    lista_bancos_full = [NOMBRE_SISTEMA] + list(lista_bancos)

    # Pre-calcular features bancarias (una serie por banco, bajo consumo de RAM)
    bank_features_dict = {}
    for banco in lista_bancos_full:
        if not df_bancarios.empty and f"{banco}_R" in df_bancarios.columns:
            df_banco = df_bancarios[[f"{banco}_R", f"{banco}_D"]].rename(
                columns={f"{banco}_R": "R", f"{banco}_D": "D"}
            )
            bank_features_dict[banco] = build_bank_features(
                df_banco,
                params["lags_cortos"],
                params["lag_semana"],
                params["lag_mes"],
                params["ventanas_vol"],
            )
        else:
            bank_features_dict[banco] = pd.DataFrame()

    # Pre-calcular features de encaje (solo para el banco configurado, ej. BBVA)
    banco_encaje = params.get("banco_encaje", "BBVA")
    df_encaje = load_encaje_data(params)
    encaje_feat = build_encaje_features(df_encaje, peru_bday)

    # Features de avance/exceso encaje desde bbva_encaje_features_modelo.xlsx
    # (generado por aux_encaje_2.py con días calendario — más preciso que EncajeD.xlsx)
    bbva_encaje_feat = load_bbva_encaje_features(params)

    # Pre-calcular features CC+OVN (todos los bancos; sistémico + BBVA).
    # flujo_neto_sistema = D_SISTEMA - R_SISTEMA, necesario para residuo_ccovn_lag1.
    df_ccovn   = load_ccovn_data(params)
    flujo_neto_sistema = pd.Series(dtype=float)
    if not df_bancarios.empty:
        _r = f"{NOMBRE_SISTEMA}_R"
        _d = f"{NOMBRE_SISTEMA}_D"
        if _r in df_bancarios.columns and _d in df_bancarios.columns:
            flujo_neto_sistema = df_bancarios[_d] - df_bancarios[_r]
    ccovn_feat = build_ccovn_features(df_ccovn, peru_bday,
                                      flujo_sistema=flujo_neto_sistema)

    # Pre-computar HMM expanding SOLO para SISTEMA (régimen sistémico, no por banco)
    # El mismo hmm_estado se aplica a todos los bancos individuales.
    hmm_sistema = pd.Series(dtype="Int8")
    if _HMM_DISPONIBLE:
        logger.info("  Pre-computando HMM expanding sobre SISTEMA...")
        bf_sis = bank_features_dict.get(NOMBRE_SISTEMA, pd.DataFrame())
        if not bf_sis.empty and "sigma_22d" in bf_sis.columns and \
                f"{NOMBRE_SISTEMA}_R" in df_bancarios.columns:
            flujo_sis = (df_bancarios[f"{NOMBRE_SISTEMA}_D"]
                         - df_bancarios[f"{NOMBRE_SISTEMA}_R"])
            hmm_sistema = _calcular_hmm_expanding(flujo_sis, bf_sis["sigma_22d"],
                                              primer_ventana_años=6)
            n_etiq = hmm_sistema.notna().sum()
            logger.info(f"    SISTEMA: hmm_estado calculado — {n_etiq:,} días etiquetados")
        else:
            logger.warning("  HMM: no se pudo calcular hmm_estado para SISTEMA (datos insuficientes)")

    # Escribir banco por banco al Parquet de salida — peak RAM = 1 banco a la vez
    pq_writer   = None
    total_filas = 0
    schema_ref  = None

    for banco in lista_bancos_full:
        df_banco_mat = build_feature_matrix(
            banco=banco,
            datos_manuales=datos_manuales,
            macro_features=macro_features,
            bank_features=bank_features_dict,
            peru_bday=peru_bday,
            peru_holidays=peru_holidays,
            fechas_elecciones=fechas_elecciones,
            h_min=params["h_min"],
            h_max=params["h_max"],
            hmm_features=hmm_sistema,       # mismo régimen sistémico para todos los bancos
            encaje_features=encaje_feat,    # features de encaje (solo aplican a banco_encaje)
            banco_encaje=banco_encaje,
            bbva_encaje_feat=bbva_encaje_feat,  # avance/exceso desde aux_encaje_2 (días calendario)
            ccovn_features=ccovn_feat,      # saldos CC+OVN en BCR (aplican a todos los bancos)
        )
        if df_banco_mat.empty:
            continue

        table = pa.Table.from_pandas(df_banco_mat, preserve_index=False)

        if pq_writer is None:
            schema_ref = table.schema
            pq_writer  = pq_writer_lib.ParquetWriter(
                ruta_output, schema_ref, compression="snappy"
            )

        # Alinear schema por si hay columnas con tipo diferente en algún banco
        table = table.cast(schema_ref)
        pq_writer.write_table(table)
        total_filas += len(df_banco_mat)

        del df_banco_mat, table
        gc.collect()
        logger.info(f"  {banco}: escrito al Parquet ({total_filas:,} filas acumuladas)")

    if pq_writer is not None:
        pq_writer.close()
    else:
        logger.warning("No se generaron matrices para ningún banco.")
        return pd.DataFrame()

    size_mb = ruta_output.stat().st_size / 1e6
    logger.info(f"\n{'='*60}")
    logger.info("  RESUMEN FINAL DE LA MATRIZ")
    logger.info(f"{'='*60}")
    logger.info(f"  Bancos modelados : {lista_bancos}")
    logger.info(f"  Total filas      : {total_filas:,}")
    logger.info(f"  Columnas         : {len(schema_ref.names)}")
    logger.info(f"  Archivo Parquet  : {ruta_output}  ({size_mb:.1f} MB)")
    logger.info(f"  Para cargar      : pd.read_parquet(r'{ruta_output}')")
    logger.info(f"{'='*60}\n")

    # Retornar DataFrame vacío con el schema correcto como señal de éxito
    return pd.DataFrame(columns=schema_ref.names)


###############################################################################
# Diccionario de variables
###############################################################################
def build_data_dictionary(params):
    """
    Retorna un DataFrame con todas las variables de la matriz,
    su fuente, descripción y rezago/horizonte.
    """
    lags_cortos  = params["lags_cortos"]
    lag_semana   = params["lag_semana"]
    lag_mes      = params["lag_mes"]
    ventanas_vol = params["ventanas_vol"]

    registros = []

    def add(variable, fuente, descripcion, rezago_dias=None, horizonte=None):
        registros.append({
            "variable"    : variable,
            "fuente"      : fuente,
            "descripcion" : descripcion,
            "rezago_dias" : rezago_dias,
            "horizonte"   : horizonte,
        })

    # ── Identificadores ──────────────────────────────────────────────────────
    add("fecha_t",  "Sistema", "Fecha de origen de la predicción",       None, None)
    add("banco",    "Sistema", "Identificador del banco",                 None, None)
    add("h",        "Sistema", "Horizonte de predicción en días hábiles", None, "t+h")
    add("log_h",    "Sistema", "Logaritmo natural del horizonte h",       None, "t+h")

    # ── Features bancarias ────────────────────────────────────────────────────
    # Estructura de información disponible en t:
    #   R (retiros, aviso 2 días hábiles antes):
    #     t-2 → R(t) disponible como R_t0
    #     t-1 → R(t+1) disponible como R_conf_t1
    #     t   → R(t+2) disponible como R_conf_t2
    #   D (depósitos, aviso 1 día hábil antes):
    #     t-1 → D(t) disponible como D_t0
    #     t   → D(t+1) disponible como D_conf_t1

    # Valores de hoy: conocidos por aviso anticipado
    add("R_t0", "Datos bancarios", "Retiro realizado en t — conocido hoy (avisado en t-2)", 0, None)
    add("D_t0", "Datos bancarios", "Depósito realizado en t — conocido hoy (avisado en t-1)", 0, None)

    # Rezagos históricos
    todos_lags = lags_cortos + [lag_semana, lag_mes]
    for l in todos_lags:
        add(f"R_t-{l}", "Datos bancarios", f"Retiro del banco en t-{l}", l, None)
        add(f"D_t-{l}", "Datos bancarios", f"Depósito del banco en t-{l}", l, None)

    for v in ventanas_vol:
        add(f"sigma_R_{v}d", "Datos bancarios", f"Desv. estándar rolling {v}d de retiros (incluye R_t0)",    None, None)
        add(f"sigma_D_{v}d", "Datos bancarios", f"Desv. estándar rolling {v}d de depósitos (incluye D_t0)",  None, None)
        add(f"ma_R_{v}d",    "Datos bancarios", f"Media móvil {v}d de retiros (incluye R_t0)",                None, None)
        add(f"ma_D_{v}d",    "Datos bancarios", f"Media móvil {v}d de depósitos (incluye D_t0)",              None, None)

    add("delta_R", "Datos bancarios", "Variación diaria de retiros: R_t0 - R(t-1)",   0, None)
    add("delta_D", "Datos bancarios", "Variación diaria de depósitos: D_t0 - D(t-1)", 0, None)

    # Volatilidad y nivel del flujo neto (D - R)
    for v in [5, 20]:
        add(f"sigma_flujo_{v}d", "Datos bancarios",
            f"Desv. estándar rolling {v}d del flujo neto D−R (régimen de volatilidad)", None, None)
        add(f"ma_flujo_{v}d",    "Datos bancarios",
            f"Media móvil {v}d del flujo neto D−R (nivel reciente del flujo)", None, None)
    add("sigma_22d", "Datos bancarios",
        "Desv. estándar rolling 22d del flujo neto D−R. Más reactiva que GARCH: "
        "cae rápido al salir de un episodio de stress (sin persistencia artificial).", None, None)
    add("hmm_estado", "Calculado (HMM expanding)",
        f"Estado de régimen oculto: 0=calma, 1=moderado, 2=severo. "
        f"Gaussian HMM {HMM_N_ESTADOS} estados sobre [flujo_neto, sigma_22d]. "
        f"Pre-computado con expanding window sin leakage (mín {HMM_MIN_AÑOS}a). "
        f"NaN en burn-in ({HMM_INICIO}→+{HMM_MIN_AÑOS}a) o si hmmlearn no instalado. "
        f"Si no aporta señal, agregar a FEATURES_EXCLUIR.", None, None)
    add("sigma_flujo_ratio", "Datos bancarios",
        "sigma_flujo_5d / sigma_flujo_20d — ratio de volatilidad corta/larga del flujo. "
        "> 1: régimen de alta vol reciente. NaN si sigma_flujo_20d = 0.", None, None)
    add("flujo_neto_acum_mes", "Datos bancarios",
        "Acumulado del flujo neto D−R desde el primer día hábil del mes hasta t. "
        "Captura la lógica de reversión intramensual: acumulación de depósitos netos "
        "anticipa mayores retiros en cierre de mes, y viceversa.", 0, None)
    add("flujo_neto_sum_5d",  "Datos bancarios",
        "Suma rolling 5dh del flujo neto D−R hasta t (semana). "
        "min_periods=5: NaN si historia < 5 días → imputado por mediana del fold.", 0, None)
    add("flujo_neto_sum_22d", "Datos bancarios",
        "Suma rolling 22dh del flujo neto D−R hasta t (mes). "
        "min_periods=22: NaN si historia < 22 días → imputado por mediana del fold.", 0, None)
    add("flujo_neto_sum_66d", "Datos bancarios",
        "Suma rolling 66dh del flujo neto D−R hasta t (trimestre). "
        "min_periods=66: NaN si historia < 66 días → imputado por mediana del fold.", 0, None)

    # ── Confirmados futuros ───────────────────────────────────────────────────
    # Entrenamiento: proxy histórico = valor realizado en t+1/t+2 (shift negativo)
    # Producción:    valor real del correo de aviso (override desde confirmados.xlsx)
    add("R_conf_t1", "Datos bancarios / Confirmados operativos", "R(t+1) — proxy histórico o aviso real (2d anticipación)", -1, "t+1")
    add("R_conf_t2", "Datos bancarios / Confirmados operativos", "R(t+2) — proxy histórico o aviso real (2d anticipación)",  0, "t+2")
    add("D_conf_t1", "Datos bancarios / Confirmados operativos", "D(t+1) — proxy histórico o aviso real (1d anticipación)",  0, "t+1")

    # ── Features macroeconómicas (observadas en t) ───────────────────────────
    add("VIX",              "Yahoo Finance",  "Índice de volatilidad implícita S&P 500",          0, None)
    add("delta_VIX",        "Yahoo Finance",  "Variación diaria del VIX",                         1, None)
    add("VIX_ma22",         "Yahoo Finance",  "Media móvil 22d del VIX",                          None, None)
    add("TC_PEN_USD",       "BCRP Add-In",    "Tipo de cambio PEN/USD (promedio compra/venta)",   0, None)
    add("delta_TC",         "BCRP Add-In",    "Variación diaria del tipo de cambio",              1, None)
    add("tc_vol_5d",        "BCRP Add-In",    "Volatilidad rolling 5d de retornos del TC",        None, None)
    add("tc_vol_22d",       "BCRP Add-In",    "Volatilidad rolling 22d de retornos del TC",       None, None)
    add("tc_vol_ratio",     "Calculado",      "tc_vol_5d / tc_vol_22d — ratio de vol cambiaria corta/larga. "
        "> 1: estrés cambiario reciente. NaN si tc_vol_22d = 0.",               None, None)
    add("garch_vol_tc",     "BCRP Add-In",    "Volatilidad condicional GARCH(1,1) de retornos log del TC PEN/USD — detecta estrés cambiario", None, None)
    add("EMBI_PERU",        "BCRP Add-In",    "EMBI Perú (riesgo país)",                          0, None)
    add("delta_EMBI",       "BCRP Add-In",    "Variación diaria del EMBI Perú",                   1, None)
    add("garch_vol_embi",   "BCRP Add-In",    "Volatilidad condicional GARCH(1,1) de cambios diarios del EMBI Perú — detecta estrés político", None, None)
    add("TASA_REF_BCRP",    "BCRP Add-In",    "Tasa de referencia del BCRP",                      0, None)
    add("FED_FUNDS",        "FRED API",       "Tasa de política monetaria de la Fed",             0, None)
    add("diferencial_tasas","Calculado",      "TASA_REF_BCRP - FED_FUNDS",                        0, None)
    add("T10Y",             "Yahoo Finance",  "Rendimiento del bono del Tesoro EE.UU. a 10 años", 0, None)

    # ── Features FFD — diferenciación fraccional (López de Prado Cap. 5) ──────
    add("EMBI_PERU_frac",   "BCRP Add-In (FFD)",  "EMBI Perú diferenciado fraccionalmente (d mínimo que elimina raíz unitaria). Preserva memoria de largo plazo; más estacionario que el nivel.", None, None)
    add("T10Y_frac",        "Yahoo Finance (FFD)", "UST 10Y diferenciado fraccionalmente. Alternativa a delta_T10Y con menor pérdida de memoria.", None, None)
    add("CDS_PERU_5Y_frac", "Bloomberg (FFD)",     "CDS Perú 5Y diferenciado fraccionalmente (d_opt calibrado en TRAIN).", None, None)
    add("COPPER_frac",      "Bloomberg (FFD)",     "Cobre LME diferenciado fraccionalmente. Indicador de ciclo global.", None, None)
    add("VIX_frac",         "Yahoo Finance (FFD)", "VIX diferenciado fraccionalmente (d≈0 si ya es estacionario; preserva estructura).", None, None)

    # ── Features estacionales (en t+h — fecha futura, siempre conocidas) ─────
    add("dias_al_cierre_mes",    "Calendario", "Días hábiles restantes hasta fin de mes en t+h",          None, "t+h")
    add("dias_desde_cierre_mes", "Calendario", "Días hábiles transcurridos desde inicio de mes en t+h",   None, "t+h")
    add("pos_en_mes",            "Calendario", "Posición del día hábil dentro del mes (1=primero)",        None, "t+h")
    add("total_bdays_mes",       "Calendario", "Total de días hábiles del mes de t+h",                     None, "t+h")
    add("is_penult_bday_trim",   "Calendario", "1 si t+h es penúltimo día hábil del trimestre",            None, "t+h")
    add("is_ultimo_bday_trim",   "Calendario", "1 si t+h es último día hábil del trimestre",               None, "t+h")
    add("is_1er_bday_trim",      "Calendario", "1 si t+h es primer día hábil del trimestre",               None, "t+h")
    add("is_2do_bday_trim",      "Calendario", "1 si t+h es segundo día hábil del trimestre",              None, "t+h")
    add("is_3er_bday_trim",      "Calendario", "1 si t+h es tercer día hábil del trimestre",               None, "t+h")
    add("dia_semana",            "Calendario", "Día de la semana (0=lunes, 4=viernes)",                    None, "t+h")
    add("mes",                   "Calendario", "Mes del año (1–12)",                                       None, "t+h")
    add("is_quincena",           "Calendario", "1 si t+h es día 15 o último hábil del mes",                None, "t+h")
    add("is_cierre_encaje",      "Calendario", "1 si t+h está en los últimos 2 días hábiles del mes",      None, "t+h")
    add("is_fin_anio",           "Calendario", "1 si t+h es 28–31 de diciembre",                           None, "t+h")
    add("is_pre_feriado",        "Calendario / holidays.Peru + holidays.US", "1 si el día siguiente a t+h es feriado PE o US",  None, "t+h")
    add("is_post_feriado",       "Calendario / holidays.Peru + holidays.US", "1 si el día anterior a t+h es feriado PE o US",   None, "t+h")
    add("is_pre_eleccion",       "Calendario", "1 si t+h está dentro de los 7 días previos a elecciones presidenciales",    None, "t+h")
    add("is_post_eleccion",      "Calendario", "1 si t+h está dentro de los 7 días posteriores a elecciones presidenciales", None, "t+h")
    add("mes_sin",               "Calendario", "sin(2π·mes/12) — proyección cíclica del mes (P=12)",                         None, "t+h")
    add("mes_cos",               "Calendario", "cos(2π·mes/12) — proyección cíclica del mes (P=12)",                         None, "t+h")
    add("dias_sem_sin",          "Calendario", "sin(2π·dia_sem/5) — ciclo semanal en días hábiles (P=5)",                    None, "t+h")
    add("dias_sem_cos",          "Calendario", "cos(2π·dia_sem/5) — ciclo semanal en días hábiles (P=5)",                    None, "t+h")
    add("dias_al_cierre_mes_sin","Calendario", "sin(2π·pos_mes/P_mes) — ciclo mensual hábil dinámico",                       None, "t+h")
    add("dias_al_cierre_mes_cos","Calendario", "cos(2π·pos_mes/P_mes) — ciclo mensual hábil dinámico",                       None, "t+h")
    add("dias_al_cierre_trim_sin","Calendario","sin(2π·pos_trim/P_trim) — ciclo trimestral hábil dinámico",                  None, "t+h")
    add("dias_al_cierre_trim_cos","Calendario","cos(2π·pos_trim/P_trim) — ciclo trimestral hábil dinámico",                  None, "t+h")
    add("dias_al_cierre_anio_sin","Calendario","sin(2π·pos_anio/P_anio) — ciclo anual hábil dinámico",                       None, "t+h")
    add("dias_al_cierre_anio_cos","Calendario","cos(2π·pos_anio/P_anio) — ciclo anual hábil dinámico",                       None, "t+h")
    add("elec_sin",              "Calendario", "sin(2π·bd_desde_elec/1260) — ciclo electoral 5 años (solo 1ras vueltas)",    None, "t+h")
    add("elec_cos",              "Calendario", "cos(2π·bd_desde_elec/1260) — ciclo electoral 5 años (solo 1ras vueltas)",    None, "t+h")

    # ── Features de encaje BBVA (EncajeD.xlsx, rezago 1 día) ─────────────────
    _benc = params.get("banco_encaje", "BBVA")
    add("encaje_lag1",         f"EncajeD / {_benc}", "Caja + Cta Cte BCR en t-1 (M USD). Solo banco configurado en banco_encaje.", 1, "t")
    add("exceso_lag1",         f"EncajeD / {_benc}", "Encaje(t-1) − Exigible(t-1): exceso acumulado disponible (M USD).", 1, "t")
    add("faltante_lag1",       f"EncajeD / {_benc}", "Exigible_total_mes − encaje_acum(t-1): saldo aún pendiente de acumular en el mes (M USD).", 1, "t")
    add("techo_10h",           f"EncajeD / {_benc}", "CC+ON al día hábil 10 antes del cierre del mes, rezagado 1 día. Techo operativo de retiros.", 1, "t")
    add("techo_restante_lag1",   f"EncajeD / {_benc}", "techo_10h − retiro_acumulado_mes(t-1): presupuesto de retiro aún disponible (M USD).", 1, "t")
    add("proporcion_usada",      f"EncajeD / {_benc}", "retiro_acum_mes(t-1) / techo_10h: fracción del techo ya utilizada (0-1+).", 1, "t")
    add("avance_mes_lag1",       f"aux_encaje_2 / {_benc}", "EncajeAcumMes(t-1) / ExigibleTotalMes_est(t-1): fracción del req. mensual cubierta. Días calendario → ffill a días hábiles.", 1, "t")
    add("exceso_abs_lag1",       f"aux_encaje_2 / {_benc}", "MAX(0, encaje(t-1)−encaje_min_por_dia)+overnight(t-1): capacidad de retiro real sin incumplir 100% al cierre (M USD). Días calendario.", 1, "t")
    add("exceso_dia_lag1",       f"aux_encaje_2 / {_benc}", "exceso_abs(t-1)/dias_restantes(t-1): presión temporal — exceso ponderado por urgencia de fin de periodo. Q5 media: −290M.", 1, "t")
    add("encaje_ovn_lag1",       f"aux_encaje_2 / {_benc}", "overnight+cta_cte+caja en t-1 (M USD): posición total BCRP. Techo físico máximo de retiro posible.", 1, "t")
    add("ratio_ovn_total_lag1",  f"aux_encaje_2 / {_benc}", "overnight(t-1)/encaje_ovn(t-1): alto = banco aún no inició retiro (fondos en overnight sin mover).", 1, "t")

    # ── Features CC+OVN en BCR (Saldos_CCOVN.xlsx, rezago 1 día) ─────────────
    add("ccovn_sistema_lag1",     "Saldos_CCOVN",
        "Saldo total CC+OVN del sistema en el BCR en t-1 (USD). Disponible desde 2010.",
        1, "t")
    add("ccovn_bbva_lag1",        "Saldos_CCOVN",
        "Saldo CC+OVN de BBVA en el BCR en t-1 (USD). Disponible desde 2010.",
        1, "t")
    add("var_ccovn_sistema_lag1", "Saldos_CCOVN",
        "Variación diaria del saldo CC+OVN del sistema en t-1 (USD).",
        1, "t")
    add("var_ccovn_bbva_lag1",    "Saldos_CCOVN",
        "Variación diaria del saldo CC+OVN de BBVA en t-1 (USD).",
        1, "t")
    add("bbva_share_lag1",        "Saldos_CCOVN",
        "ccovn_bbva(t-1) / ccovn_sistema(t-1): participación de BBVA en el sistema en t-1.",
        1, "t")
    add("var_ccovn_bbva_exceso_lag1", "Saldos_CCOVN",
        "Δsaldo_bbva(t-1) − bbva_share(t-1)×Δsaldo_sistema(t-1): componente "
        "idiosincrático de BBVA en la variación diaria del saldo. "
        "Ortogonal a la variación sistémica.",
        1, "t")
    add("ccovn_vs_dia_mes_lag1",  "Saldos_CCOVN",
        "saldo_sistema(t-1) − media_historica[rango_dia_habil_del_mes]: "
        "desviación del saldo respecto al nivel estacional esperado para ese "
        "puesto del mes hábil. Elimina la estacionalidad intramonth.",
        1, "t")
    add("residuo_ccovn_lag1",     "Saldos_CCOVN + Datos bancarios",
        "Δsaldo_sistema(t-1) − flujo_neto_sistema(t-1): error de la identidad "
        "Δsaldo≈flujo. Señal de reversión para horizontes cortos. "
        "Solo disponible si flujo_neto_sistema está en bancarios.",
        1, "t")

    # ── Target ────────────────────────────────────────────────────────────────
    add("target", "Datos bancarios", "Flujo neto = D(b, t+h) - R(b, t+h). NaN si no hay datos.", None, "t+h")

    df_dict = pd.DataFrame(registros)

    # Marcar si la variable está activa en el parquet o fue excluida
    df_dict["en_modelo"] = df_dict["variable"].apply(
        lambda v: "No" if v in FEATURES_EXCLUIR else "Sí"
    )

    return df_dict


###############################################################################
# Datos demo — se usa cuando PARAMS["usar_datos_demo"] = True
###############################################################################
def generate_demo_data(params):
    """
    Genera datos bancarios y macroeconómicos sintéticos para pruebas.
    Simula 5 bancos de distintos tamaños para activar la agrupación automática.
    """
    logger.info("*** MODO DEMO: generando datos sintéticos ***")
    np.random.seed(42)

    fechas = pd.bdate_range(
        start=params["fecha_inicio_historico"],
        end="2024-12-31",
    )
    n = len(fechas)

    # Cinco bancos con escalas muy distintas → banco_4 y banco_5 quedarán en Otros
    escalas = {
        "banco_1": 500,
        "banco_2": 350,
        "banco_3": 200,
        "banco_4": 30,
        "banco_5": 20,
    }

    df_bancarios = pd.DataFrame(index=fechas)
    df_bancarios.index.name = "fecha"
    for banco, escala in escalas.items():
        tendencia = np.linspace(0, escala * 0.1, n)
        ruido_R = np.random.normal(0, escala * 0.15, n)
        ruido_D = np.random.normal(0, escala * 0.15, n)
        df_bancarios[f"{banco}_R"] = escala + tendencia + ruido_R
        df_bancarios[f"{banco}_D"] = escala + tendencia + ruido_D

    # Macro sintética con tendencias plausibles
    df_macro = pd.DataFrame(index=fechas)
    df_macro["VIX"]           = np.clip(20 + np.cumsum(np.random.normal(0, 0.3, n)), 10, 80)
    df_macro["TC_PEN_USD"]    = np.clip(3.3 + np.cumsum(np.random.normal(0, 0.005, n)), 3.0, 4.5)
    df_macro["T10Y"]          = np.clip(3.5 + np.cumsum(np.random.normal(0, 0.02, n)) * 0.05, 1.0, 6.0)
    df_macro["FED_FUNDS"]     = np.clip(1.0 + np.cumsum(np.random.normal(0, 0.01, n)) * 0.03, 0.0, 6.0)
    df_macro["EMBI_PERU"]     = np.clip(130 + np.cumsum(np.random.normal(0, 1.0, n)), 80, 400)
    df_macro["TASA_REF_BCRP"] = np.clip(4.0 + np.cumsum(np.random.normal(0, 0.01, n)) * 0.02, 1.5, 8.0)

    datos_manuales = {
        "bancarios"   : df_bancarios,
        "confirmados" : pd.DataFrame(columns=["banco", "R_conf_t1", "R_conf_t2", "D_conf_t1"]),
        "intervencion": pd.Series(dtype=float),
    }
    return datos_manuales, df_macro


###############################################################################
# Flujo principal
###############################################################################
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("  SISTEMA DE PREDICCIÓN DE LIQUIDEZ EN MONEDA EXTRANJERA")
    logger.info("  Construcción de Matriz de Features")
    logger.info(f"  Fecha de ejecución: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 60)

    # 1. Calendario
    peru_bday, peru_holidays, fechas_elecciones = build_peru_calendar(
        años=PARAMS["años_calendario"],
    )

    # 2. Datos manuales + macro
    if PARAMS.get("usar_datos_demo"):
        datos_manuales, macro_raw = generate_demo_data(PARAMS)
    else:
        datos_manuales = load_manual_data(PARAMS)
        macro_raw      = download_external_series(PARAMS)

    # 3. Agrupación de bancos pequeños → Otros_bancos
    df_bancarios_agrupado, lista_bancos, reporte_agrupacion = agrupar_bancos(
        datos_manuales["bancarios"],
        PARAMS["umbral_banco_pequeño_pct"],
        PARAMS["bancos_otros"],
        PARAMS["nombre_otros"],
    )
    datos_manuales["bancarios"] = df_bancarios_agrupado

    if not lista_bancos:
        logger.warning(
            "No hay bancos en lista_bancos. "
            "Ejecutando con banco sintético para validar estructura."
        )
        lista_bancos = ["banco_demo"]

    # 4. Features macroeconómicas
    macro_features = build_macro_features(macro_raw)

    # 5. Matriz completa
    matriz = build_full_matrix(
        params=PARAMS,
        datos_manuales=datos_manuales,
        lista_bancos=lista_bancos,
        macro_features=macro_features,
        peru_bday=peru_bday,
        peru_holidays=peru_holidays,
        fechas_elecciones=fechas_elecciones,
    )

    ruta_out = Path(PARAMS.get("ruta_output", "matriz_features.parquet"))
    if ruta_out.exists() and ruta_out.stat().st_size > 0:
        print(f"\nMatriz generada correctamente → {ruta_out}")
    else:
        print("\nMatriz vacía. Verifique la configuración de PARAMS y los archivos de datos.")

    # Diccionario de variables
    data_dict = build_data_dictionary(PARAMS)
    df_activas  = data_dict[data_dict["en_modelo"] == "Sí"].drop(columns="en_modelo").reset_index(drop=True)
    df_excluidas = data_dict[data_dict["en_modelo"] == "No"].drop(columns="en_modelo").reset_index(drop=True)

    print(f"\n{'='*70}")
    print("  DICCIONARIO DE VARIABLES — en modelo")
    print(f"{'='*70}")
    print(df_activas.to_string(index=False))
    print(f"\nTotal en modelo: {len(df_activas)}  |  Excluidas: {len(df_excluidas)}")

    if len(df_excluidas):
        print(f"\n  Variables excluidas (FEATURES_EXCLUIR): "
              + ", ".join(df_excluidas["variable"].tolist()))

    # Exportar diccionario a Excel: dos pestañas
    ruta_dict = Path(PARAMS.get("ruta_diccionario", r"1. Data\Clean\diccionario_variables.xlsx"))
    try:
        ruta_dict.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(ruta_dict, engine="openpyxl") as writer:
            df_activas.to_excel(writer, index=False, sheet_name="Variables_modelo")
            df_excluidas.to_excel(writer, index=False, sheet_name="Variables_excluidas")

            # Resaltar pestaña de excluidas con fondo naranja en celdas de datos
            from openpyxl.styles import PatternFill
            fill_naranja = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
            ws = writer.sheets["Variables_excluidas"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.fill = fill_naranja

        logger.info(f"  Diccionario exportado a: {ruta_dict} "
                    f"(2 pestañas: Variables_modelo={len(df_activas)}, "
                    f"Variables_excluidas={len(df_excluidas)})")
    except Exception as e:
        logger.warning(f"  No se pudo exportar diccionario: {e}")
