# -*- coding: utf-8 -*-
"""
aux_depuracion_features_v2.py
Diagnóstico cuantitativo de features en 4 capas independientes.

Capas:
  1. Estadística pura   — varianza, nulos, correlación inter-feature, VIF
  2. Modelo (step005)   — importancia por fold y por cuantil (si existe el CSV)
  3. Horizonte          — correlación feature → target por h ∈ H_EVALUAR
  4. Recomendación      — semáforo consolidado: Mantener / Evaluar / Eliminar

Output:
  aux_depuracion_features_v2.xlsx  →  5 hojas
  Consola: lista lista de candidatos a agregar a FEATURES_EXCLUIR

Uso:
  1. Correr step001 y step005 al menos una vez.
  2. Ejecutar este script (sin cambiar el modelo).
  3. Revisar Excel + consola.
  4. Agregar candidatos confirmados a FEATURES_EXCLUIR en step001.
  5. Regenerar parquet y re-correr step005.
"""

import warnings
warnings.filterwarnings("ignore")

import os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── VIF (opcional) ────────────────────────────────────────────────────────────
try:
    from statsmodels.stats.outliers_influence import variance_inflation_factor
    _VIF_OK = True
except ImportError:
    _VIF_OK = False
    print("  [INFO] statsmodels no instalado — VIF omitido (pip install statsmodels)")

###############################################################################
# CONFIGURACIÓN
###############################################################################

BASE = Path(r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente")

RUTA_PARQUET   = BASE / "1. Data" / "Clean" / "matriz_features.parquet"
DIR_STEP005    = BASE / "2. Output" / "step005_wfcv_v3"
DIR_OUTPUT     = BASE / "2. Output" / "aux_depuracion_features_v2"
RUTA_EXCEL_OUT = DIR_OUTPUT / "aux_depuracion_features_v2.xlsx"
RUTA_PLOT_OUT  = DIR_OUTPUT / "correlacion_inter_features.png"

# ── Selector de CSV de importancias ───────────────────────────────────────────
# None  → busca automáticamente el más reciente en DIR_STEP005
# Ruta  → usa ese CSV exacto (recomendado cuando hay varios modelos)
# Ejemplos:
#   RUTA_CSV_IMPORTANCIAS = DIR_STEP005 / "xgb_qt_expanding_5051" / "wfcv_v3_importancias_SISTEMA_20250530.csv"
#   RUTA_CSV_IMPORTANCIAS = DIR_STEP005 / "xgb_expanding_5051"    / "wfcv_v3_importancias_SISTEMA_20250530.csv"
RUTA_CSV_IMPORTANCIAS = DIR_STEP005 / "final"   # busca el CSV dentro de esta carpeta

# Banco de referencia para los análisis estadísticos (el de mayor volumen)
BANCO_REF      = "SISTEMA"

# Usar solo observaciones de TRAIN (hasta este corte) para evitar contaminación
TRAIN_CUTOFF   = "2022-12-31"

# Horizontes a evaluar para la correlación feature → target
H_EVALUAR      = [3, 5, 10, 22, 45, 90]

# Umbrales estadísticos
UMBRAL_NULOS       = 0.20   # > 20% NaN → señal de alerta
UMBRAL_CORR_FEAT   = 0.85   # |corr(fi,fj)| > 0.85 → par redundante
UMBRAL_VIF         = 10.0   # VIF > 10 → inflado
UMBRAL_GAIN_BAJO   = 0.50   # feature con gain < mediana en >UMBRAL_GAIN_BAJO de folds → "bajo"
UMBRAL_CORR_TARGET = 0.05   # |corr(feat, target)| < 0.05 en todos los h → "sin señal"

# Columnas que nunca son features
COLS_EXCLUIR_SIEMPRE = {"fecha_t", "banco", "h", "target", "fecha_th"}

# Features ya excluidos en step001 (para marcarlos como "Ya excluido")
FEATURES_YA_EXCLUIDOS = [
    "log_h", "ma_R_5d", "ma_D_5d", "delta_R", "delta_D",
    "VIX_ma22", "diferencial_tasas", "dias_desde_cierre_mes", "pos_en_mes",
]

DIR_OUTPUT.mkdir(parents=True, exist_ok=True)

###############################################################################
# HELPERS DE FORMATO EXCEL
###############################################################################

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill():
    return PatternFill("solid", fgColor="1F4E79")


def _write_header_row(ws, headers, row=1, col_start=1):
    hf = _header_fill()
    for j, h in enumerate(headers, col_start):
        c = ws.cell(row, j, h)
        c.fill = hf
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def _color_semaforo(valor):
    """Devuelve fgColor según semáforo estadístico."""
    mapping = {
        "OK":           "C6EFCE",   # verde
        "Alerta":       "FFEB9C",   # amarillo
        "Crítico":      "FFC7CE",   # rojo
        "Ya excluido":  "D9D9D9",   # gris
        "—":            "F2F2F2",
    }
    return mapping.get(str(valor), "FFFFFF")


def _color_recomendacion(valor):
    mapping = {
        "Mantener":     "C6EFCE",
        "Evaluar":      "FFEB9C",
        "Eliminar":     "FFC7CE",
        "Ya excluido":  "D9D9D9",
        "—":            "F2F2F2",
    }
    return mapping.get(str(valor), "FFFFFF")


def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


###############################################################################
# CARGA DE DATOS
###############################################################################

print("=" * 65)
print("  DIAGNÓSTICO DE FEATURES — aux_depuracion_features_v2")
print("=" * 65)

print(f"\n[1] Cargando parquet ({BANCO_REF}, hasta {TRAIN_CUTOFF})...")
try:
    df_full = pd.read_parquet(RUTA_PARQUET, filters=[("banco", "==", BANCO_REF)])
    df_full["fecha_t"] = pd.to_datetime(df_full["fecha_t"])
except FileNotFoundError:
    raise FileNotFoundError(
        f"No se encontró la matriz: {RUTA_PARQUET}\n"
        "Ejecuta step001_build_feature_matrix.py primero."
    )

# Restricción a TRAIN para evitar contaminación
df = df_full[df_full["fecha_t"] <= TRAIN_CUTOFF].copy()
print(f"    Total parquet: {len(df_full):,} filas  |  TRAIN (≤{TRAIN_CUTOFF}): {len(df):,} filas")

# Feature columns
feat_cols = [c for c in df.columns
             if c not in COLS_EXCLUIR_SIEMPRE
             and df[c].dtype.kind in ("f", "i", "u", "b")]
print(f"    Features detectados en parquet: {len(feat_cols)}")


###############################################################################
# CAPA 1 — ESTADÍSTICA PURA
###############################################################################

print("\n[2] Capa 1 — estadística pura...")

# Una sola observación por fecha (usar solo h=10 como referencia para no duplicar)
# Una sola observación por fecha de origen (no por horizonte):
# El parquet tiene una fila por (banco, fecha_t, h=2..90), lo que
# duplica ~89× las observaciones. Para estadísticas de features
# (varianza, correlación inter-feature, VIF) se fija h=10 como
# referencia → evita inflar n_obs y sesgos de multicolinealidad.
df_stat = df[df["h"] == 10][feat_cols].copy() if 10 in df["h"].values else df[feat_cols].copy()

stat_rows = []
for feat in feat_cols:
    serie = df_stat[feat]
    n_total = len(serie)
    n_nan   = serie.isna().sum()
    pct_nan = n_nan / n_total
    serie_v = serie.dropna()

    std_v   = float(serie_v.std()) if len(serie_v) > 1 else 0.0
    rango   = float(serie_v.max() - serie_v.min()) if len(serie_v) > 1 else 0.0
    cv      = std_v / abs(float(serie_v.mean())) if abs(float(serie_v.mean())) > 1e-9 else np.nan
    n_uniq  = int(serie_v.nunique())
    pct_cero= float((serie_v == 0).mean()) if len(serie_v) > 0 else np.nan

    # Alerta de varianza casi cero
    if feat in FEATURES_YA_EXCLUIDOS:
        estado_var = "Ya excluido"
    elif n_uniq <= 1:
        estado_var = "Crítico"
    elif n_uniq <= 5 or (std_v < 1e-6 and rango < 1e-6):
        estado_var = "Alerta"
    else:
        estado_var = "OK"

    # Alerta de nulos
    if feat in FEATURES_YA_EXCLUIDOS:
        estado_nan = "Ya excluido"
    elif pct_nan > UMBRAL_NULOS:
        estado_nan = "Crítico" if pct_nan > 0.50 else "Alerta"
    else:
        estado_nan = "OK"

    stat_rows.append({
        "feature":       feat,
        "n_obs":         n_total,
        "n_nan":         int(n_nan),
        "pct_nan":       round(pct_nan, 4),
        "std":           round(std_v, 4),
        "rango":         round(rango, 4),
        "cv":            round(cv, 4) if not np.isnan(cv) else np.nan,
        "n_unicos":      n_uniq,
        "pct_cero":      round(pct_cero, 4),
        "estado_varianza": estado_var,
        "estado_nulos":    estado_nan,
        "ya_excluido":   feat in FEATURES_YA_EXCLUIDOS,
    })

df_stat_res = pd.DataFrame(stat_rows)

# ── Correlación inter-feature ─────────────────────────────────────────────────
print("    Calculando matriz de correlación inter-feature...")
feat_activos = [f for f in feat_cols if f not in FEATURES_YA_EXCLUIDOS]
corr_matrix  = df_stat[feat_activos].corr(method="pearson")

# Pares con |corr| > umbral
pares_altos = []
for i, fi in enumerate(feat_activos):
    for j, fj in enumerate(feat_activos):
        if j <= i:
            continue
        r = corr_matrix.loc[fi, fj]
        if abs(r) >= UMBRAL_CORR_FEAT:
            pares_altos.append({
                "feature_A": fi, "feature_B": fj,
                "correlacion": round(r, 4),
                "nivel": "Crítico" if abs(r) > 0.95 else "Alerta",
            })

df_pares = pd.DataFrame(pares_altos).sort_values("correlacion", ascending=False, key=abs)

# Features involucrados en al menos un par alto
feats_multicolm = set(df_pares["feature_A"]) | set(df_pares["feature_B"])

# Agregar columna a df_stat_res
df_stat_res["n_pares_altos"] = df_stat_res["feature"].apply(
    lambda f: int(((df_pares["feature_A"] == f) | (df_pares["feature_B"] == f)).sum())
)
df_stat_res["estado_multicol"] = df_stat_res.apply(
    lambda r: "Ya excluido" if r["ya_excluido"] else
              ("Crítico" if r["n_pares_altos"] >= 3 else
               ("Alerta"  if r["n_pares_altos"] >= 1 else "OK")),
    axis=1
)

# ── VIF ───────────────────────────────────────────────────────────────────────
df_vif = pd.DataFrame(columns=["feature", "VIF", "estado_vif"])
if _VIF_OK:
    print("    Calculando VIF...")
    feat_vif = [f for f in feat_activos
                if df_stat_res.loc[df_stat_res["feature"] == f, "pct_nan"].values[0] < 0.05]
    df_vif_data = df_stat[feat_vif].dropna().astype(float)
    if len(df_vif_data) > len(feat_vif) + 10:
        vif_vals = []
        for i, f in enumerate(feat_vif):
            try:
                v = float(variance_inflation_factor(df_vif_data.values, i))
            except Exception:
                v = np.nan
            vif_vals.append(v)
        df_vif = pd.DataFrame({
            "feature": feat_vif,
            "VIF": [round(v, 2) for v in vif_vals],
            "estado_vif": [
                "Crítico" if (not np.isnan(v) and v > UMBRAL_VIF) else
                "Alerta"  if (not np.isnan(v) and v > 5) else "OK"
                for v in vif_vals
            ],
        })

        # Para VIF > 10: todos los features con |corr| >= 0.70 (pueden ser 2, 3 o más)
        UMBRAL_CORR_VIF = 0.70

        def _corr_partners_significativos(feat):
            if feat not in corr_matrix.columns:
                return ""
            serie = corr_matrix[feat].drop(feat)
            signif = serie[serie.abs() >= UMBRAL_CORR_VIF].abs().sort_values(ascending=False)
            if signif.empty:
                # Si ninguno supera 0.70, mostrar el top 3 igual
                signif = serie.abs().sort_values(ascending=False).head(3)
            return "  |  ".join(
                f"{f} ({corr_matrix.loc[feat, f]:+.2f})" for f in signif.index
            )

        df_vif["top_corr"] = df_vif.apply(
            lambda r: _corr_partners_significativos(r["feature"]) if r["VIF"] > UMBRAL_VIF else "",
            axis=1
        )
    else:
        print("    [VIF] Datos insuficientes para calcular VIF")

print(f"    Pares con |corr| ≥ {UMBRAL_CORR_FEAT}: {len(df_pares)}")
print(f"    Features involucrados en pares altos: {len(feats_multicolm)}")

# ── Plot heatmap de correlación ───────────────────────────────────────────────
# Limitar a 25 features y dpi bajo para evitar MemoryError en carpetas de red
print("    Generando heatmap de correlación...")
top_feat = df_stat_res.nlargest(25, "n_pares_altos")["feature"].tolist()
if len(top_feat) < 5:
    top_feat = feat_activos[:25]
corr_sub = corr_matrix.loc[top_feat, top_feat]
n_f = len(top_feat)

fig, ax = plt.subplots(figsize=(min(n_f * 0.55, 13), min(n_f * 0.50, 12)))
cmap = plt.cm.RdYlGn   # colormap estándar — evita allocación extra de LinearSegmented
im = ax.imshow(corr_sub.values, vmin=-1, vmax=1, cmap=cmap, aspect="auto",
               interpolation="none")
plt.colorbar(im, ax=ax, shrink=0.7, label="Correlación de Pearson")
ax.set_xticks(range(n_f))
ax.set_yticks(range(n_f))
ax.set_xticklabels(top_feat, rotation=90, fontsize=7)
ax.set_yticklabels(top_feat, fontsize=7)
ax.set_title(
    f"Correlación inter-feature (Top {n_f} por n_pares_altos) — {BANCO_REF}\n"
    f"Datos TRAIN ≤ {TRAIN_CUTOFF}  |  Umbral: |r| ≥ {UMBRAL_CORR_FEAT}",
    fontweight="bold", fontsize=10,
)
for i in range(n_f):
    for j in range(n_f):
        v = corr_sub.values[i, j]
        if abs(v) >= UMBRAL_CORR_FEAT and i != j:
            ax.text(j, i, f"{v:.2f}", ha="center", va="center",
                    fontsize=6, color="black")
plt.tight_layout()
plt.savefig(RUTA_PLOT_OUT, dpi=90, bbox_inches="tight")
plt.close()
print(f"    Heatmap guardado: {RUTA_PLOT_OUT.name}")


###############################################################################
# CAPA 2 — IMPORTANCIAS DEL MODELO (step005)
###############################################################################

print("\n[3] Capa 2 — importancias del modelo (step005)...")

# Listar siempre todos los CSV disponibles para que el usuario sepa qué hay
csv_todos = sorted(DIR_STEP005.rglob(f"*importancias*{BANCO_REF}*.csv"), reverse=True)
if not csv_todos:
    csv_todos = sorted(DIR_STEP005.rglob("*importancias*.csv"), reverse=True)

if csv_todos:
    print(f"    CSVs de importancias disponibles ({len(csv_todos)}):")
    for i, p in enumerate(csv_todos):
        marca = " ← seleccionado" if i == 0 and RUTA_CSV_IMPORTANCIAS is None else ""
        print(f"      [{i+1}] {p.relative_to(DIR_STEP005)}{marca}")
else:
    print(f"    No se encontraron CSVs en {DIR_STEP005}")

# Seleccionar CSV: explícito, carpeta, o el más reciente
if RUTA_CSV_IMPORTANCIAS is not None:
    ruta_imp_usar = Path(RUTA_CSV_IMPORTANCIAS)
    if ruta_imp_usar.is_dir():
        # Es una carpeta → buscar el CSV más reciente dentro de ella
        candidatos_dir = sorted(ruta_imp_usar.glob(f"*importancias*{BANCO_REF}*.csv"), reverse=True)
        if not candidatos_dir:
            candidatos_dir = sorted(ruta_imp_usar.glob("*importancias*.csv"), reverse=True)
        if candidatos_dir:
            ruta_imp_usar = candidatos_dir[0]
            print(f"    Carpeta indicada → usando: {ruta_imp_usar.name}")
        else:
            print(f"    [ERROR] No se encontró CSV de importancias en: {ruta_imp_usar}")
            ruta_imp_usar = None
    elif not ruta_imp_usar.exists():
        print(f"    [ERROR] RUTA_CSV_IMPORTANCIAS no existe: {ruta_imp_usar}")
        ruta_imp_usar = None
    else:
        print(f"    Usando CSV explícito: {ruta_imp_usar.relative_to(DIR_STEP005)}")
elif csv_todos:
    ruta_imp_usar = csv_todos[0]
    print(f"    Usando el más reciente. Para cambiar, edita RUTA_CSV_IMPORTANCIAS en la config.")
else:
    ruta_imp_usar = None

df_imp_res = pd.DataFrame()
tiene_importancias = False

if ruta_imp_usar is not None:
    ruta_imp = ruta_imp_usar
    print(f"    CSV seleccionado: {ruta_imp.name}")
    df_imp_raw = pd.read_csv(ruta_imp)
    # Estructura esperada: columnas = feature + fold_1, fold_2, ...
    fold_cols = [c for c in df_imp_raw.columns if c not in ("feature",) and str(c).isdigit()]
    if not fold_cols and "feature" in df_imp_raw.columns:
        fold_cols = [c for c in df_imp_raw.columns if c != "feature"]

    if "feature" in df_imp_raw.columns and fold_cols:
        df_imp_raw = df_imp_raw.set_index("feature")[fold_cols]
        n_folds    = len(fold_cols)

        # Normalizar por fold (suma = 1 en cada fold)
        df_norm = df_imp_raw.div(df_imp_raw.sum(axis=0).replace(0, 1), axis=1)

        imp_rows = []
        for feat in df_imp_raw.index:
            gains  = df_imp_raw.loc[feat].values.astype(float)
            norms  = df_norm.loc[feat].values.astype(float)
            mediana_gain  = float(np.nanmedian(gains))
            media_norm    = float(np.nanmean(norms))
            # Porcentaje de folds con gain < mediana global del fold
            med_global_by_fold = np.nanmedian(df_imp_raw.values, axis=0)
            pct_bajo   = float(np.mean(gains < med_global_by_fold))

            if feat in FEATURES_YA_EXCLUIDOS:
                estado_imp = "Ya excluido"
            elif pct_bajo > UMBRAL_GAIN_BAJO:
                estado_imp = "Crítico" if pct_bajo > 0.75 else "Alerta"
            else:
                estado_imp = "OK"

            imp_rows.append({
                "feature":        feat,
                "gain_mediana":   round(mediana_gain, 4),
                "gain_norm_media":round(media_norm, 6),
                "pct_folds_bajo": round(pct_bajo, 3),
                "n_folds":        n_folds,
                "estado_imp":     estado_imp,
                "rank_global":    0,   # se rellena después
            })

        df_imp_res = pd.DataFrame(imp_rows)
        df_imp_res["rank_global"] = df_imp_res["gain_mediana"].rank(ascending=False, method="min").astype(int)
        df_imp_res = df_imp_res.sort_values("rank_global")
        tiene_importancias = True
        print(f"    {len(df_imp_res)} features con importancia | {n_folds} folds")
    else:
        print(f"    [Aviso] Formato CSV inesperado: {df_imp_raw.columns.tolist()[:6]}")
else:
    print(f"    [Aviso] No se encontró CSV de importancias en {DIR_STEP005}")
    print(f"    → Ejecuta step005 con GUARDAR_MODELOS_TODOS_FOLDS=True para generarlo")


###############################################################################
# CAPA 3 — CORRELACIÓN FEATURE → TARGET POR HORIZONTE h
###############################################################################

print("\n[4] Capa 3 — correlación feature → target por horizonte h...")

corr_rows = []
df_train_full = df_full[df_full["fecha_t"] <= TRAIN_CUTOFF].copy()

# Para correlación feature → target sí se usa cada h por separado:
# se filtra df[df["h"] == h] para cada horizonte → una observación
# por (fecha_t) para ese h específico, sin repetición.
for feat in feat_cols:
    row_h = {"feature": feat}
    max_abs_corr = 0.0
    for h in H_EVALUAR:
        sub = df_train_full[df_train_full["h"] == h][[feat, "target"]].dropna()
        if len(sub) < 30:
            row_h[f"corr_h{h}"] = np.nan
            continue
        r = float(sub[feat].corr(sub["target"]))
        row_h[f"corr_h{h}"] = round(r, 4)
        max_abs_corr = max(max_abs_corr, abs(r))

    row_h["max_abs_corr"] = round(max_abs_corr, 4)

    if feat in FEATURES_YA_EXCLUIDOS:
        row_h["estado_corr_target"] = "Ya excluido"
    elif max_abs_corr < UMBRAL_CORR_TARGET:
        row_h["estado_corr_target"] = "Crítico"
    elif max_abs_corr < UMBRAL_CORR_TARGET * 2:
        row_h["estado_corr_target"] = "Alerta"
    else:
        row_h["estado_corr_target"] = "OK"

    corr_rows.append(row_h)

df_corr_h = pd.DataFrame(corr_rows).sort_values("max_abs_corr", ascending=False, key=abs)
print(f"    Features con max|corr_target| < {UMBRAL_CORR_TARGET}: "
      f"{(df_corr_h['estado_corr_target'] == 'Crítico').sum()}")
print(f"    Features con max|corr_target| en [{UMBRAL_CORR_TARGET},{UMBRAL_CORR_TARGET*2}): "
      f"{(df_corr_h['estado_corr_target'] == 'Alerta').sum()}")


###############################################################################
# CAPA 4 — RECOMENDACIÓN CONSOLIDADA
###############################################################################

print("\n[5] Capa 4 — recomendación consolidada...")

# Unir las tres capas
df_rec = df_stat_res[["feature", "pct_nan", "estado_varianza",
                        "estado_nulos", "estado_multicol", "n_pares_altos",
                        "ya_excluido"]].copy()

df_rec = df_rec.merge(
    df_corr_h[["feature", "max_abs_corr", "estado_corr_target"]],
    on="feature", how="left"
)

if tiene_importancias:
    df_rec = df_rec.merge(
        df_imp_res[["feature", "gain_mediana", "gain_norm_media",
                    "pct_folds_bajo", "rank_global", "estado_imp"]],
        on="feature", how="left"
    )
else:
    df_rec["gain_mediana"]   = np.nan
    df_rec["rank_global"]    = np.nan
    df_rec["pct_folds_bajo"] = np.nan
    df_rec["estado_imp"]     = "Sin datos"

if not df_vif.empty:
    df_rec = df_rec.merge(df_vif[["feature", "VIF", "estado_vif"]], on="feature", how="left")
else:
    df_rec["VIF"] = np.nan
    df_rec["estado_vif"] = "—"


def _consolidar_recomendacion(row):
    if row["ya_excluido"]:
        return "Ya excluido", "Ya en FEATURES_EXCLUIR de step001"

    criticos = []
    alertas  = []

    # Capa 1
    if row["estado_varianza"]  == "Crítico": criticos.append("varianza casi cero")
    elif row["estado_varianza"] == "Alerta": alertas.append("varianza baja")
    if row["estado_nulos"]     == "Crítico": criticos.append(f"nulos={row['pct_nan']:.0%}")
    elif row["estado_nulos"]   == "Alerta":  alertas.append(f"nulos={row['pct_nan']:.0%}")
    if row.get("estado_vif")   == "Crítico": criticos.append(f"VIF={row.get('VIF','?'):.1f}")
    elif row.get("estado_vif") == "Alerta":  alertas.append(f"VIF>{5}")
    if row["estado_multicol"]  == "Crítico": criticos.append(f"{row['n_pares_altos']} pares corr≥{UMBRAL_CORR_FEAT}")
    elif row["estado_multicol"]== "Alerta":  alertas.append(f"corr≥{UMBRAL_CORR_FEAT} con {row['n_pares_altos']} feature(s)")

    # Capa 2 (importancias)
    if row.get("estado_imp") == "Crítico":   criticos.append(f"gain bajo en {row.get('pct_folds_bajo',0):.0%} folds")
    elif row.get("estado_imp") == "Alerta":  alertas.append(f"gain bajo en {row.get('pct_folds_bajo',0):.0%} folds")

    # Capa 3 (corr target)
    if row["estado_corr_target"] == "Crítico":  criticos.append(f"sin señal target (max|r|={row['max_abs_corr']:.3f})")
    elif row["estado_corr_target"] == "Alerta": alertas.append(f"señal débil target (max|r|={row['max_abs_corr']:.3f})")

    if len(criticos) >= 2:
        return "Eliminar", "; ".join(criticos)
    elif len(criticos) == 1 and len(alertas) >= 1:
        return "Eliminar", f"{criticos[0]} + {alertas[0]}"
    elif len(criticos) == 1:
        return "Evaluar", criticos[0]
    elif len(alertas) >= 2:
        return "Evaluar", "; ".join(alertas)
    elif len(alertas) == 1:
        return "Evaluar", alertas[0]
    else:
        return "Mantener", "Sin señales de alerta en ninguna capa"


rec_result = df_rec.apply(_consolidar_recomendacion, axis=1, result_type="expand")
df_rec["recomendacion"] = rec_result[0]
df_rec["razon"]         = rec_result[1]

# Ordenar: Eliminar primero, luego Evaluar, luego Mantener
orden_rec = {"Eliminar": 0, "Evaluar": 1, "Mantener": 2, "Ya excluido": 3}
df_rec["_orden"] = df_rec["recomendacion"].map(orden_rec)
df_rec = df_rec.sort_values(["_orden", "max_abs_corr"], ascending=[True, True]).drop(columns=["_orden"])

resumen_rec = df_rec["recomendacion"].value_counts()
print(f"\n    Recomendación consolidada:")
for k in ["Eliminar", "Evaluar", "Mantener", "Ya excluido"]:
    print(f"      {k:15s}: {resumen_rec.get(k, 0):3d} features")

# Candidatos a agregar a FEATURES_EXCLUIR
candidatos_eliminar = df_rec[df_rec["recomendacion"] == "Eliminar"]["feature"].tolist()
print(f"\n    Candidatos a agregar a FEATURES_EXCLUIR:")
for f in candidatos_eliminar:
    razon = df_rec.loc[df_rec["feature"] == f, "razon"].values[0]
    print(f"      · {f:<30s}  ← {razon}")


###############################################################################
# EXPORTAR EXCEL — 5 HOJAS
###############################################################################

print("\n[6] Exportando Excel...")

# ── Preparar datos para exportar ──────────────────────────────────────────────
# Hoja 1: Estadística
df_h1 = df_stat_res[[
    "feature", "n_obs", "pct_nan", "n_unicos", "std", "rango", "cv",
    "pct_cero", "estado_varianza", "estado_nulos",
    "n_pares_altos", "estado_multicol",
]].copy()
if not df_vif.empty:
    vif_cols = ["feature", "VIF", "estado_vif"]
    if "top_corr" in df_vif.columns:
        vif_cols.append("top_corr")
    df_h1 = df_h1.merge(df_vif[vif_cols], on="feature", how="left")
else:
    df_h1["VIF"] = "—"; df_h1["estado_vif"] = "—"; df_h1["top_corr"] = ""

df_h1["ya_excluido"] = df_h1["feature"].apply(
    lambda f: "Sí" if f in FEATURES_YA_EXCLUIDOS else "No"
)

# Hoja 2: Importancias
if tiene_importancias:
    df_h2 = df_imp_res.copy()
else:
    df_h2 = pd.DataFrame({"info": ["No se encontró CSV de importancias de step005"]})

# Hoja 3: Correlación horizonte
df_h3 = df_corr_h.copy()

# Hoja 4: Recomendación
df_h4 = df_rec[[
    "feature", "pct_nan", "n_pares_altos", "max_abs_corr",
    "gain_mediana", "rank_global", "pct_folds_bajo",
    "VIF" if "VIF" in df_rec.columns else "estado_multicol",
    "recomendacion", "razon",
]].copy()

# Hoja 5: Pares multicolinealidad
df_h5 = df_pares.copy()

# ── Escribir con openpyxl ─────────────────────────────────────────────────────

# Primero guardar con pandas para que cree el archivo base
with pd.ExcelWriter(RUTA_EXCEL_OUT, engine="openpyxl") as writer:
    df_h1.to_excel(writer, index=False, sheet_name="1_Estadistica")
    df_h2.to_excel(writer, index=False, sheet_name="2_Importancias_modelo")
    df_h3.to_excel(writer, index=False, sheet_name="3_Corr_horizonte")
    df_h4.to_excel(writer, index=False, sheet_name="4_Recomendacion")
    df_h5.to_excel(writer, index=False, sheet_name="5_Pares_multicolm")

# Reabrir para formatear
wb = load_workbook(RUTA_EXCEL_OUT)

# ── Hoja 1 — Estadística ──────────────────────────────────────────────────────
ws1 = wb["1_Estadistica"]
_write_header_row(ws1, [c.value for c in ws1[1]])
_set_col_widths(ws1, {"A": 28, "B": 8, "C": 10, "D": 9, "E": 12, "F": 10,
                       "G": 8, "H": 9, "I": 14, "J": 13, "K": 12, "L": 14,
                       "M": 8, "N": 12, "O": 10, "P": 60})
# Índice de columnas por nombre
h1_cols = {c.value: get_column_letter(c.column) for c in ws1[1]}

for row_idx in range(2, ws1.max_row + 1):
    feat = ws1.cell(row_idx, 1).value
    ya_exc = feat in FEATURES_YA_EXCLUIDOS

    for cell in ws1[row_idx]:
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.font = Font(size=9, color="808080" if ya_exc else "000000")

    # Colorear columnas de estado
    for col_name, col_letter in h1_cols.items():
        if "estado" in col_name.lower() or col_name == "ya_excluido":
            cell = ws1[f"{col_letter}{row_idx}"]
            val  = str(cell.value)
            cell.fill = PatternFill("solid", fgColor=_color_semaforo(val))
            cell.font = Font(bold=True, size=9, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_name == "top_corr":
            cell = ws1[f"{col_letter}{row_idx}"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(size=8, color="000000", italic=True)
            # Fondo naranja claro si tiene contenido (VIF > 10)
            if cell.value:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
    ws1.row_dimensions[row_idx].height = 28

ws1.freeze_panes = "A2"

# ── Hoja 2 — Importancias ─────────────────────────────────────────────────────
ws2 = wb["2_Importancias_modelo"]
_write_header_row(ws2, [c.value for c in ws2[1]])
_set_col_widths(ws2, {"A": 28, "B": 13, "C": 15, "D": 14, "E": 9, "F": 14, "G": 10})
h2_cols = {c.value: get_column_letter(c.column) for c in ws2[1]}

for row_idx in range(2, ws2.max_row + 1):
    feat   = ws2.cell(row_idx, 1).value
    ya_exc = feat in FEATURES_YA_EXCLUIDOS
    for cell in ws2[row_idx]:
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="center")
        cell.font = Font(size=9, color="808080" if ya_exc else "000000")
    if "estado_imp" in h2_cols:
        c = ws2[f"{h2_cols['estado_imp']}{row_idx}"]
        c.fill = PatternFill("solid", fgColor=_color_semaforo(str(c.value)))
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")

ws2.freeze_panes = "A2"

# ── Hoja 3 — Correlación por horizonte ───────────────────────────────────────
ws3 = wb["3_Corr_horizonte"]
_write_header_row(ws3, [c.value for c in ws3[1]])
_set_col_widths(ws3, {"A": 28, "B": 9, "C": 9, "D": 9, "E": 9, "F": 9, "G": 9, "H": 13, "I": 16})
h3_cols = {c.value: get_column_letter(c.column) for c in ws3[1]}

# Escala de color para correlación: positivo=verde, negativo=rojo, cero=blanco
def _corr_color(v):
    if pd.isna(v) or not isinstance(v, (int, float)):
        return "F2F2F2"
    v = float(v)
    intensity = min(abs(v) / 0.30, 1.0)
    if v > 0:
        r = int(198 - intensity * (198 - 0))
        g = int(239 - intensity * (239 - 128))
        b = int(206 - intensity * (206 - 0))
    else:
        r = int(255 - intensity * (255 - 139))
        g = int(235 - intensity * 235)
        b = int(156 - intensity * 156)
    return f"{r:02X}{g:02X}{b:02X}"

for row_idx in range(2, ws3.max_row + 1):
    feat   = ws3.cell(row_idx, 1).value
    ya_exc = feat in FEATURES_YA_EXCLUIDOS
    for cell in ws3[row_idx]:
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.font = Font(size=9, color="808080" if ya_exc else "000000")

    for h in H_EVALUAR:
        col_h = f"corr_h{h}"
        if col_h in h3_cols:
            c   = ws3[f"{h3_cols[col_h]}{row_idx}"]
            val = c.value
            if val is not None:
                c.fill = PatternFill("solid", fgColor=_corr_color(val))

    if "estado_corr_target" in h3_cols:
        c = ws3[f"{h3_cols['estado_corr_target']}{row_idx}"]
        c.fill = PatternFill("solid", fgColor=_color_semaforo(str(c.value)))
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")

ws3.freeze_panes = "A2"

# ── Hoja 4 — Recomendación ────────────────────────────────────────────────────
ws4 = wb["4_Recomendacion"]
_write_header_row(ws4, [c.value for c in ws4[1]])
_set_col_widths(ws4, {"A": 28, "B": 9, "C": 10, "D": 12, "E": 13, "F": 10,
                       "G": 13, "H": 10, "I": 14, "J": 65})
h4_cols = {c.value: get_column_letter(c.column) for c in ws4[1]}

for row_idx in range(2, ws4.max_row + 1):
    feat = ws4.cell(row_idx, 1).value
    ya_exc = feat in FEATURES_YA_EXCLUIDOS

    for cell in ws4[row_idx]:
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.font = Font(size=9, color="808080" if ya_exc else "000000")

    if "recomendacion" in h4_cols:
        c   = ws4[f"{h4_cols['recomendacion']}{row_idx}"]
        val = str(c.value)
        c.fill = PatternFill("solid", fgColor=_color_recomendacion(val))
        c.font = Font(bold=True, size=10, color="000000")
        c.alignment = Alignment(horizontal="center", vertical="center")

    if "razon" in h4_cols:
        c = ws4[f"{h4_cols['razon']}{row_idx}"]
        c.alignment = Alignment(wrap_text=True, vertical="top")

    ws4.row_dimensions[row_idx].height = 28

ws4.freeze_panes = "A2"

# ── Hoja 5 — Pares multicolinealidad ─────────────────────────────────────────
ws5 = wb["5_Pares_multicolm"]
_write_header_row(ws5, [c.value for c in ws5[1]])
_set_col_widths(ws5, {"A": 28, "B": 28, "C": 14, "D": 12})
h5_cols = {c.value: get_column_letter(c.column) for c in ws5[1]}

for row_idx in range(2, ws5.max_row + 1):
    for cell in ws5[row_idx]:
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.font = Font(size=9)
    if "nivel" in h5_cols:
        c = ws5[f"{h5_cols['nivel']}{row_idx}"]
        c.fill = PatternFill("solid", fgColor=_color_semaforo(str(c.value)))
        c.font = Font(bold=True, size=9)
    if "correlacion" in h5_cols:
        c   = ws5[f"{h5_cols['correlacion']}{row_idx}"]
        val = c.value
        if val is not None:
            c.fill = PatternFill("solid", fgColor=_corr_color(val))

ws5.freeze_panes = "A2"

wb.save(RUTA_EXCEL_OUT)
print(f"    Excel guardado: {RUTA_EXCEL_OUT}")

###############################################################################
# RESUMEN FINAL EN CONSOLA
###############################################################################

print(f"\n{'='*65}")
print("  RESUMEN FINAL")
print(f"{'='*65}")
print(f"  Banco ref.:  {BANCO_REF}  (TRAIN ≤ {TRAIN_CUTOFF})")
print(f"  Features analizados: {len(feat_cols)}")
print(f"  Ya excluidos (FEATURES_EXCLUIR): {len(FEATURES_YA_EXCLUIDOS)}")
print(f"\n  Recomendación consolidada (features activos):")
for k in ["Eliminar", "Evaluar", "Mantener"]:
    n = (df_rec["recomendacion"] == k).sum()
    print(f"    {k:12s}: {n:3d}")

if candidatos_eliminar:
    print(f"\n  ─── Candidatos a añadir a FEATURES_EXCLUIR en step001 ───────────")
    for f in candidatos_eliminar:
        razon = df_rec.loc[df_rec["feature"] == f, "razon"].values[0]
        print(f"    \"{f}\",  # {razon}")

print(f"\n  Archivos generados en: {DIR_OUTPUT}")
print(f"  · {RUTA_EXCEL_OUT.name}  — diagnóstico completo (5 hojas)")
print(f"  · {RUTA_PLOT_OUT.name}   — heatmap correlación inter-feature")
print()
