# -*- coding: utf-8 -*-
"""
aux_visualizar_series.py
Visualiza las series base (no transformaciones) y features de calendario
del pipeline de predicción de liquidez en ME.
"""

import sys
from pathlib import Path
sys.path.insert(0, r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente")

import importlib
import step001_build_feature_matrix as bfm
importlib.reload(bfm)

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
from pandas.tseries.offsets import CustomBusinessDay

PARAMS = bfm.PARAMS.copy()

BASE_SISTEMA = Path(r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente")
DIR_OUTPUT   = BASE_SISTEMA / "2. Output" / "aux_visualizar_series"
DIR_OUTPUT.mkdir(parents=True, exist_ok=True)

# ─────────────────────────────────────────────────────────────────
# Cargar datos reales
# ─────────────────────────────────────────────────────────────────
peru_bday, peru_holidays, fechas_elecciones = bfm.build_peru_calendar(
    años=PARAMS["años_calendario"],
)
datos    = bfm.load_manual_data(PARAMS)
df_macro = bfm.download_external_series(PARAMS)
df_banc  = datos["bancarios"]

# step001 devuelve formato ANCHO ({banco}_R / {banco}_D).
# Reconvertir a formato LARGO (fecha, banco, R, D) que usa este script.
if not df_banc.empty:
    _bancos = [c[:-2] for c in df_banc.columns if c.endswith("_R")]
    _partes = []
    for _b in _bancos:
        _tmp = df_banc[[f"{_b}_R", f"{_b}_D"]].rename(
            columns={f"{_b}_R": "R", f"{_b}_D": "D"}
        )
        _tmp["banco"] = _b
        _partes.append(_tmp.reset_index())
    df_banc = pd.concat(_partes, ignore_index=True)

# ─────────────────────────────────────────────────────────────────
# BLOQUE 1: Series externas base
# ─────────────────────────────────────────────────────────────────
series_macro = {
    "VIX":           "Índice VIX",
    "TC_PEN_USD":    "Tipo de Cambio USD/PEN",
    "EMBI_PERU":     "EMBI Perú (pbs)",
    "TASA_REF_BCRP": "Tasa Ref. BCRP (%)",
    "FED_FUNDS":     "Fed Funds (%)",
    "T10Y":          "Bono Tesoro 10Y (%)",
}

fig1, axes = plt.subplots(3, 2, figsize=(14, 11))
axes = axes.flatten()
for i, (col, titulo) in enumerate(series_macro.items()):
    ax = axes[i]
    if col in df_macro.columns and df_macro[col].notna().any():
        s = df_macro[col].dropna()
        ax.plot(s.index, s.values, lw=0.8, color="steelblue")
        ax.set_title(titulo, fontweight="bold", fontsize=10)
        ax.xaxis.set_major_locator(mdates.YearLocator(2))
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
        ax.tick_params(axis="x", rotation=45, labelsize=8)
        ax.grid(True, alpha=0.3)
        ultima = s.index.max()
        ax.annotate(f"Última: {ultima.date()}", xy=(0.98, 0.05),
                    xycoords="axes fraction", ha="right", fontsize=7, color="red")
    else:
        ax.text(0.5, 0.5, f"{col}\n(sin datos)", ha="center", va="center",
                transform=ax.transAxes, color="gray")
        ax.set_title(titulo, fontsize=10)

plt.suptitle("Series Externas Base — Datos Reales", fontsize=13, fontweight="bold")
plt.tight_layout()
plt.savefig(DIR_OUTPUT / "01_series_base_macro.png", dpi=150, bbox_inches="tight")
plt.show()

# ─────────────────────────────────────────────────────────────────
# BLOQUE 2a: Participación de volumen por banco
# ─────────────────────────────────────────────────────────────────
UMBRAL_PCT = 0.01

if not df_banc.empty:
    vol = df_banc.groupby("banco")[["R", "D"]].sum()
    vol["total"] = vol["R"] + vol["D"]
    vol["pct"]   = vol["total"] / vol["total"].sum()
    vol = vol.sort_values("pct", ascending=False)

    grandes  = vol[vol["pct"] >= UMBRAL_PCT].index.tolist()
    pequeños = vol[vol["pct"] <  UMBRAL_PCT].index.tolist()

    print(f"\nParticipación de volumen por banco (umbral = {UMBRAL_PCT:.0%}):")
    print(f"{'Banco':<25} {'Retiros (MM)':>14} {'Depósitos (MM)':>16} {'Total (MM)':>12} {'Part.':>7} {'Grupo':>10}")
    print("-" * 90)
    for banco, row in vol.iterrows():
        grupo = "GRANDE" if row["pct"] >= UMBRAL_PCT else "pequeño"
        print(f"{banco:<25} {row['R']/1e6:>14,.1f} {row['D']/1e6:>16,.1f} "
              f"{row['total']/1e6:>12,.1f} {row['pct']:>7.1%} {grupo:>10}")
    print(f"\n  Grandes ({len(grandes)}): {', '.join(grandes)}")
    print(f"  Pequeños → Otros_bancos ({len(pequeños)}): {', '.join(pequeños)}")

    # Exportar tabla a Excel para revisión manual
    tabla_export = pd.DataFrame({
        "retiros_musd"   : vol["R"]     / 1e6,
        "depositos_musd" : vol["D"]     / 1e6,
        "total_musd"     : vol["total"] / 1e6,
        "participacion"  : vol["pct"],
        "grupo_auto"     : ["GRANDE" if p >= UMBRAL_PCT else "Otros_bancos" for p in vol["pct"]],
        "incluir_en_otros": ["" if p >= UMBRAL_PCT else "X" for p in vol["pct"]],
    })
    tabla_export.index.name = "banco"

    ruta_tabla = DIR_OUTPUT / "tabla_clasificacion_bancos.xlsx"
    with pd.ExcelWriter(ruta_tabla, engine="openpyxl") as writer:
        tabla_export.to_excel(writer, sheet_name="Clasificacion", float_format="%.4f")
        ws = writer.sheets["Clasificacion"]

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        azul   = PatternFill("solid", fgColor="BDD7EE")
        rojo   = PatternFill("solid", fgColor="FFCCCC")
        header = PatternFill("solid", fgColor="2F5496")
        thin   = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

        # Encabezado
        for cell in ws[1]:
            cell.fill      = header
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.border    = thin

        # Filas de datos
        for i, (banco, row) in enumerate(tabla_export.iterrows(), start=2):
            fill = azul if row["grupo_auto"] == "GRANDE" else rojo
            for cell in ws[i]:
                cell.fill   = fill
                cell.border = thin
            # Porcentaje en columna participacion (col 5 = E)
            ws[f"E{i}"].number_format = "0.00%"
            ws[f"E{i}"].alignment = Alignment(horizontal="center")
            ws[f"F{i}"].alignment = Alignment(horizontal="center")
            ws[f"G{i}"].alignment = Alignment(horizontal="center")

        # Ancho de columnas
        anchos = [28, 16, 16, 14, 14, 18, 20]
        for col_idx, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho

        # Nota instructiva
        fila_nota = len(tabla_export) + 3
        ws.cell(row=fila_nota, column=1,
                value="INSTRUCCIONES: Revise la columna 'incluir_en_otros'. "
                      "Marque con 'X' los bancos que quiera agrupar manualmente, "
                      "independientemente del umbral automático. "
                      f"Umbral automático actual: {UMBRAL_PCT:.0%} del volumen total.")
        ws.cell(row=fila_nota, column=1).font = Font(italic=True, color="666666")

    print(f"\nTabla de clasificación exportada: {ruta_tabla}")

    # Gráfico de barras horizontales
    colores = ["steelblue" if p >= UMBRAL_PCT else "tomato" for p in vol["pct"]]
    fig2a, ax2a = plt.subplots(figsize=(14, max(5, len(vol) * 0.45)))
    bars = ax2a.barh(vol.index, vol["pct"] * 100, color=colores, alpha=0.8, edgecolor="white")
    ax2a.axvline(UMBRAL_PCT * 100, color="black", lw=1.2, ls="--")
    for bar, (_, row) in zip(bars, vol.iterrows()):
        ax2a.text(bar.get_width() + 0.3, bar.get_y() + bar.get_height() / 2,
                  f"{row['pct']:.1%}", va="center", fontsize=8)
    ax2a.set_xlabel("Participación en volumen total (%)")
    ax2a.set_title("Participación por Banco — Volumen Total (R + D)", fontweight="bold")
    ax2a.invert_yaxis()
    ax2a.grid(True, axis="x", alpha=0.3)
    from matplotlib.patches import Patch
    ax2a.legend(handles=[
        Patch(fc="steelblue", alpha=0.8, label=f"Grande (≥ {UMBRAL_PCT:.0%}) — modelo individual"),
        Patch(fc="tomato",    alpha=0.8, label=f"Pequeño (< {UMBRAL_PCT:.0%}) → Otros_bancos"),
        plt.Line2D([0], [0], color="black", lw=1.2, ls="--", label=f"Umbral {UMBRAL_PCT:.0%}"),
    ], fontsize=8, loc="lower right")
    plt.tight_layout()
    plt.savefig(DIR_OUTPUT / "02a_participacion_bancos.png", dpi=150, bbox_inches="tight")
    plt.show()

# ─────────────────────────────────────────────────────────────────
# BLOQUE 2: Series por banco — ANTES del filtro (todos los bancos)
# ─────────────────────────────────────────────────────────────────
def _plot_bancos(df_banc, bancos, titulo, fname):
    n = len(bancos)
    fig, axes = plt.subplots(n, 2, figsize=(14, 4 * n))
    if n == 1:
        axes = axes.reshape(1, -1)
    for i, banco in enumerate(bancos):
        sub = df_banc[df_banc["banco"] == banco].set_index("fecha")
        ax_r, ax_d = axes[i, 0], axes[i, 1]
        ax_r.bar(sub.index, sub["R"], color="tomato",    width=1, alpha=0.7)
        ax_d.bar(sub.index, sub["D"], color="steelblue", width=1, alpha=0.7)
        for ax, etiqueta in [(ax_r, "Retiros (USD)"), (ax_d, "Depósitos (USD)")]:
            ax.set_title(f"{banco} — {etiqueta}", fontsize=9, fontweight="bold")
            ax.xaxis.set_major_locator(mdates.YearLocator(2))
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
            ax.tick_params(axis="x", rotation=45, labelsize=7)
            ax.grid(True, alpha=0.3)
    fig.suptitle(titulo, fontsize=13, fontweight="bold", y=1.0)
    fig.tight_layout(rect=[0, 0, 1, 0.99])
    fig.savefig(DIR_OUTPUT / fname, dpi=150, bbox_inches="tight")
    plt.show()

if not df_banc.empty:
    todos_bancos = sorted(df_banc["banco"].unique())
    _plot_bancos(
        df_banc, todos_bancos,
        f"Retiros y Depósitos por Banco — Antes del filtro ({UMBRAL_PCT:.0%})",
        "02_series_bancarias_todos.png",
    )

    print("\nResumen por banco (USD):")
    resumen = (
        df_banc.groupby("banco")[["R", "D"]]
        .sum()
        .assign(neto=lambda x: x["D"] - x["R"])
    )
    print(resumen.to_string())
else:
    print("Sin datos bancarios cargados.")

# ─────────────────────────────────────────────────────────────────
# BLOQUE 2b: Series por banco — DESPUÉS del filtro (grandes + Otros)
# ─────────────────────────────────────────────────────────────────
if not df_banc.empty:
    # Construir Otros_bancos agregado
    nombre_otros = PARAMS.get("nombre_otros", "Otros_bancos")
    bancos_fijos = PARAMS.get("bancos_otros", [])

    if bancos_fijos:
        pequeños_2b = [b for b in bancos_fijos if b in df_banc["banco"].unique()]
    else:
        pequeños_2b = vol[vol["pct"] < UMBRAL_PCT].index.tolist()

    grandes_2b = [b for b in df_banc["banco"].unique() if b not in pequeños_2b]

    if pequeños_2b:
        df_otros = (
            df_banc[df_banc["banco"].isin(pequeños_2b)]
            .groupby("fecha")[["R", "D"]]
            .sum()
            .reset_index()
            .assign(banco=nombre_otros)
        )
        df_filtrado = pd.concat(
            [df_banc[df_banc["banco"].isin(grandes_2b)], df_otros],
            ignore_index=True,
        )
    else:
        df_filtrado = df_banc[df_banc["banco"].isin(grandes_2b)].copy()

    bancos_filtrados = sorted(grandes_2b) + ([nombre_otros] if pequeños_2b else [])
    _plot_bancos(
        df_filtrado, bancos_filtrados,
        f"Retiros y Depósitos por Banco — Después del filtro ({UMBRAL_PCT:.0%})",
        "02b_series_bancarias_filtrado.png",
    )

# ─────────────────────────────────────────────────────────────────
# BLOQUE 3: Flujos del sistema completo (D, R, neto)
# ─────────────────────────────────────────────────────────────────
if not df_banc.empty:
    fechas_con_data = set(df_banc.groupby("fecha").groups.keys())
    sistema = df_banc.groupby("fecha")[["R", "D"]].sum()

    cal_completo = pd.bdate_range(
        start=sistema.index.min(), end=sistema.index.max(), freq=peru_bday
    )
    sistema = sistema.reindex(cal_completo, fill_value=0)
    sistema.index.name = "fecha"
    sistema["neto"] = sistema["D"] - sistema["R"]

    d_mm    =  sistema["D"] / 1e6
    r_mm    = -sistema["R"] / 1e6
    neto_mm =  sistema["neto"] / 1e6

    fig3, ax = plt.subplots(figsize=(14, 6))

    ax.bar(sistema.index, d_mm,    color="steelblue", width=1, alpha=0.8, zorder=1, label="Depósitos")
    ax.bar(sistema.index, r_mm,    color="tomato",    width=1, alpha=0.8, zorder=1, label="Retiros")
    ax.plot(sistema.index, neto_mm, color="black",    lw=0.8, alpha=0.9,  zorder=2, label="Neto")
    ax.axhline(0, color="black", lw=0.6, zorder=3)

    ax.set_ylabel("MM USD (neto diario)")
    ax.set_title("Flujos Sistema Financiero — Depósitos / Retiros / Neto (millones USD)",
                 fontweight="bold")
    ax.xaxis.set_major_locator(mdates.YearLocator(2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    ax.tick_params(axis="x", rotation=0)
    ax.legend(fontsize=8, loc="upper left",
              handles=[
                  plt.Line2D([0],[0], color="black", lw=1.5, label="Neto"),
                  plt.Rectangle((0,0),1,1, fc="steelblue", alpha=0.8, label="Depósitos"),
                  plt.Rectangle((0,0),1,1, fc="tomato",    alpha=0.8, label="Retiros"),
              ])
    ax.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(DIR_OUTPUT / "03_flujos_sistema.png", dpi=150, bbox_inches="tight")
    plt.show()

    # Exportar flujos diarios a Excel
    ruta_flujos_clean = DIR_OUTPUT / "flujos_diarios.xlsx"
    flujos_export = pd.DataFrame({
        "depositos_musd":  sistema["D"] / 1e6,
        "retiros_musd":    sistema["R"] / 1e6,
        "neto_musd":       sistema["neto"] / 1e6,
    })
    flujos_export.index.name = "fecha"
    flujos_export.to_excel(ruta_flujos_clean)
    print(f"\nFlujos diarios exportados: {ruta_flujos_clean}")
    print(f"  Filas  : {len(flujos_export):,}")
    print(f"  Periodo: {flujos_export.index.min().date()} → {flujos_export.index.max().date()}")

# ─────────────────────────────────────────────────────────────────
# BLOQUE 3b: Flujos sistema — zoom últimos 2 años
# ─────────────────────────────────────────────────────────────────
if not df_banc.empty:
    fecha_fin_3b   = sistema.index.max()
    fecha_ini_3b   = fecha_fin_3b - pd.DateOffset(years=2)
    mask_2y        = (sistema.index >= fecha_ini_3b) & (sistema.index <= fecha_fin_3b)
    sis_2y         = sistema[mask_2y]

    d_2y    =  sis_2y["D"] / 1e6
    r_2y    = -sis_2y["R"] / 1e6
    neto_2y =  sis_2y["neto"] / 1e6

    fig3b, ax3b = plt.subplots(figsize=(14, 6))

    ax3b.bar(sis_2y.index, d_2y,    color="steelblue", width=1, alpha=0.8, zorder=1)
    ax3b.bar(sis_2y.index, r_2y,    color="tomato",    width=1, alpha=0.8, zorder=1)
    ax3b.plot(sis_2y.index, neto_2y, color="black",    lw=0.9, alpha=0.9, zorder=2)
    ax3b.axhline(0, color="black", lw=0.6, zorder=3)

    ax3b.set_ylabel("MM USD (neto diario)")
    ax3b.set_title(
        f"Flujos Sistema Financiero — Zoom Últimos 2 Años "
        f"({fecha_ini_3b.strftime('%b %Y')} – {fecha_fin_3b.strftime('%b %Y')})",
        fontweight="bold"
    )
    ax3b.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
    ax3b.xaxis.set_major_formatter(mdates.DateFormatter("%b\n%Y"))
    ax3b.tick_params(axis="x", rotation=0)
    ax3b.legend(
        fontsize=8, loc="upper left",
        handles=[
            plt.Line2D([0], [0], color="black", lw=1.5, label="Neto"),
            plt.Rectangle((0, 0), 1, 1, fc="steelblue", alpha=0.8, label="Depositos"),
            plt.Rectangle((0, 0), 1, 1, fc="tomato",    alpha=0.8, label="Retiros"),
        ]
    )
    ax3b.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(DIR_OUTPUT / "03b_flujos_sistema_2y.png", dpi=150, bbox_inches="tight")
    plt.show()

# ─────────────────────────────────────────────────────────────────
# BLOQUE 3c: Flujo neto por banco — contribución de cada entidad
# ─────────────────────────────────────────────────────────────────
if not df_banc.empty:
    import matplotlib.cm as cm
    from matplotlib.patches import Patch as _Patch

    SMOOTH_DIAS_3C = 22

    # Top 5 fijos para este plot; el resto se agrupa como residual
    _TOP5         = ["BBVA", "CREDITO", "CITIBANK", "SCOTIABANK", "INTERBANK"]
    _nombre_otros = "Otros_bancos"
    _bancos_disp  = df_banc["banco"].unique()
    _top5_disp    = [b for b in _TOP5 if b in _bancos_disp]
    _resto        = [b for b in _bancos_disp if b not in _top5_disp]

    # Pivot base desde df_banc (todos los bancos, sin filtro previo)
    _df_nc       = df_banc.copy()
    _df_nc["neto"] = _df_nc["D"] - _df_nc["R"]
    _pivot_base  = (
        _df_nc.groupby(["fecha", "banco"])["neto"]
        .sum()
        .unstack("banco")
        .reindex(sistema.index, fill_value=0)
    )

    # Construir pivot_nc: top5 individuales + Otros_bancos agregado
    pivot_nc = _pivot_base[_top5_disp].copy()
    if _resto:
        pivot_nc[_nombre_otros] = _pivot_base[_resto].sum(axis=1)

    _all_cols = _top5_disp + ([_nombre_otros] if _nombre_otros in pivot_nc.columns else [])
    pivot_nc  = pivot_nc[_all_cols]

    # Paleta: color fijo por banco, gris neutro para Otros_bancos
    _pal  = [cm.tab10(i) for i in range(len(_top5_disp))]
    _pal += [(0.60, 0.60, 0.60, 1.0)] * int(_nombre_otros in pivot_nc.columns)

    # ── Gráfico 03c: historia completa suavizada + participación relativa ──
    _piv_sm = pivot_nc.rolling(SMOOTH_DIAS_3C, min_periods=1).mean() / 1e6  # MM USD
    _net_sm = (sistema["neto"] / 1e6).rolling(SMOOTH_DIAS_3C, min_periods=1).mean()
    _idx_sm = _piv_sm.index

    fig3c, (ax_hist, ax_share) = plt.subplots(
        2, 1, figsize=(14, 10),
        gridspec_kw={"height_ratios": [3, 1]},
        sharex=True,
    )

    _pos_acc = np.zeros(len(_piv_sm))
    _neg_acc = np.zeros(len(_piv_sm))
    for col, clr in zip(_all_cols, _pal):
        v  = _piv_sm[col].values
        vp = np.where(v > 0, v, 0.0)
        vn = np.where(v < 0, v, 0.0)
        ax_hist.fill_between(_idx_sm, _pos_acc, _pos_acc + vp, color=clr, alpha=0.80)
        ax_hist.fill_between(_idx_sm, _neg_acc, _neg_acc + vn, color=clr, alpha=0.80)
        _pos_acc += vp
        _neg_acc += vn

    ax_hist.plot(_net_sm.index, _net_sm.values, color="black", lw=1.3, zorder=5)
    ax_hist.axhline(0, color="black", lw=0.5, zorder=4)
    ax_hist.set_ylabel(f"MM USD (media {SMOOTH_DIAS_3C}d)")
    ax_hist.set_title(
        f"Flujo Neto (D − R) por Banco — Historia Completa  [media móvil {SMOOTH_DIAS_3C} días]\n"
        f"Top 5 individuales  ·  {_nombre_otros}: resto del sistema",
        fontweight="bold",
    )
    _ley_hist = [_Patch(fc=clr, alpha=0.8, label=col) for col, clr in zip(_all_cols, _pal)]
    _ley_hist.append(plt.Line2D([0], [0], color="black", lw=1.5, label="Neto total"))
    ax_hist.legend(handles=_ley_hist, fontsize=7.5, loc="upper left", ncol=3, framealpha=0.9)
    ax_hist.grid(True, alpha=0.3)

    # Panel inferior: participación relativa (% del flujo absoluto)
    _abs_tot = _piv_sm.abs().sum(axis=1).replace(0, np.nan)
    _shares  = _piv_sm.abs().div(_abs_tot, axis=0).fillna(0) * 100
    _sh_acc  = np.zeros(len(_piv_sm))
    for col, clr in zip(_all_cols, _pal):
        _s = _shares[col].values
        ax_share.fill_between(_idx_sm, _sh_acc, _sh_acc + _s, color=clr, alpha=0.80)
        _sh_acc += _s

    ax_share.set_ylim(0, 100)
    ax_share.set_ylabel("Part. (%)")
    ax_share.set_title("Participación en el flujo neto (valor absoluto)", fontsize=9)
    ax_share.xaxis.set_major_locator(mdates.YearLocator(2))
    ax_share.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    ax_share.tick_params(axis="x", rotation=0)
    ax_share.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(DIR_OUTPUT / "03c_neto_por_banco.png", dpi=150, bbox_inches="tight")
    plt.show()

    # ── Gráfico 03d: zoom últimos 2 años, barras diarias apiladas ──
    _f_fin  = sistema.index.max()
    _f_ini  = _f_fin - pd.DateOffset(years=2)
    _mask2y = (pivot_nc.index >= _f_ini) & (pivot_nc.index <= _f_fin)
    _piv2y  = pivot_nc[_mask2y] / 1e6
    _idx2y  = _piv2y.index

    fig3d, ax3d = plt.subplots(figsize=(14, 6))
    _pb2 = np.zeros(len(_idx2y))
    _nb2 = np.zeros(len(_idx2y))
    for col, clr in zip(_all_cols, _pal):
        v2  = _piv2y[col].values
        vp2 = np.where(v2 > 0, v2, 0.0)
        vn2 = np.where(v2 < 0, v2, 0.0)
        ax3d.bar(_idx2y, vp2, bottom=_pb2, color=clr, alpha=0.85, width=1)
        ax3d.bar(_idx2y, vn2, bottom=_nb2, color=clr, alpha=0.85, width=1)
        _pb2 += vp2
        _nb2 += vn2

    _neto2y = (sistema["neto"] / 1e6)[_mask2y]
    ax3d.plot(_neto2y.index, _neto2y.values, color="black", lw=0.9, zorder=5)
    ax3d.axhline(0, color="black", lw=0.5)
    ax3d.set_ylabel("MM USD")
    ax3d.set_title(
        f"Flujo Neto (D − R) por Banco — Zoom Últimos 2 Años "
        f"({_f_ini.strftime('%b %Y')} – {_f_fin.strftime('%b %Y')})\n"
        f"Top 5 individuales  ·  {_nombre_otros}: resto del sistema",
        fontweight="bold",
    )
    ax3d.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
    ax3d.xaxis.set_major_formatter(mdates.DateFormatter("%b\n%Y"))
    _ley3d = [_Patch(fc=clr, alpha=0.85, label=col) for col, clr in zip(_all_cols, _pal)]
    _ley3d.append(plt.Line2D([0], [0], color="black", lw=1.5, label="Neto total"))
    ax3d.legend(handles=_ley3d, fontsize=7.5, loc="upper left", ncol=3, framealpha=0.9)
    ax3d.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(DIR_OUTPUT / "03d_neto_por_banco_2y.png", dpi=150, bbox_inches="tight")
    plt.show()

# ─────────────────────────────────────────────────────────────────
# BLOQUE 4: Features de calendario — año 2021
# ─────────────────────────────────────────────────────────────────
peru_bday2 = CustomBusinessDay(holidays=peru_holidays)
fechas_cal = pd.date_range("2021-01-01", "2021-12-31", freq=peru_bday2)
df_cal = bfm._build_seasonal_table(
    fechas_cal, peru_holidays, fechas_elecciones, peru_bday2
)

# Diagnóstico: verificar si las fechas electorales de 2021 son feriados
fechas_elec_2021 = [pd.Timestamp("2021-04-11"), pd.Timestamp("2021-06-06")]
for fe in fechas_elec_2021:
    en_feriados = fe in peru_holidays
    en_cal      = fe in fechas_cal
    print(f"  {fe.date()} — en peru_holidays: {en_feriados} | en fechas_cal: {en_cal} "
          f"| is_eleccion en df_cal: "
          f"{int(df_cal.loc[fe, 'is_eleccion']) if fe in df_cal.index else 'NO está en índice'}")

cols_cal = [
    "dia_semana", "mes", "pos_en_mes", "dias_al_cierre_mes",
    "dias_desde_cierre_mes", "total_bdays_mes",
    "is_quincena", "is_cierre_encaje",
    "is_penult_bday_trim", "is_ultimo_bday_trim",
    "is_1er_bday_trim", "is_2do_bday_trim", "is_3er_bday_trim",
    "is_fin_anio",
    "is_pre_feriado", "is_post_feriado",
    "is_pre_eleccion", "is_post_eleccion",
]

n_cols_cal = len(cols_cal)
n_rows_cal = (n_cols_cal + 2) // 3
fig4, axes4 = plt.subplots(n_rows_cal, 3, figsize=(16, 4 * n_rows_cal))
axes4 = axes4.flatten()

for i, col in enumerate(cols_cal):
    ax = axes4[i]
    if col in df_cal.columns:
        ax.plot(df_cal.index, df_cal[col], lw=0.8, color="steelblue")
        ax.set_title(col, fontsize=9, fontweight="bold")
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b"))
        ax.tick_params(axis="x", rotation=45, labelsize=7)
        ax.grid(True, alpha=0.3)
        if df_cal[col].max() <= 1 and df_cal[col].min() >= 0:
            ax.set_ylim(-0.1, 1.3)
    else:
        ax.text(0.5, 0.5, f"{col}\n(no disponible)", ha="center", va="center",
                transform=ax.transAxes, color="gray", fontsize=9)
        ax.set_title(col, fontsize=9)

for j in range(n_cols_cal, len(axes4)):
    axes4[j].set_visible(False)

plt.suptitle("Features de Calendario — Año 2021", fontsize=13, fontweight="bold")
plt.tight_layout()
plt.savefig(DIR_OUTPUT / "04_features_calendario_2021.png", dpi=150, bbox_inches="tight")
plt.show()

# ─────────────────────────────────────────────────────────────────
# BLOQUE 5: Flujos sistema 2021 con marcadores electorales
# ─────────────────────────────────────────────────────────────────
if not df_banc.empty:
    elec_1ra = pd.Timestamp("2021-04-11")
    elec_2da = pd.Timestamp("2021-06-06")

    mask21 = (sistema.index >= "2021-01-01") & (sistema.index <= "2021-12-31")
    sis21  = sistema[mask21]

    d21    =  sis21["D"] / 1e6
    r21    = -sis21["R"] / 1e6
    neto21 =  sis21["neto"] / 1e6

    fig5, ax5 = plt.subplots(figsize=(14, 6))

    ax5.bar(sis21.index, d21,    color="steelblue", width=1, alpha=0.8, label="Depositos")
    ax5.bar(sis21.index, r21,    color="tomato",    width=1, alpha=0.8, label="Retiros")
    ax5.plot(sis21.index, neto21, color="black",    lw=1.0, alpha=0.9,  label="Neto")
    ax5.axhline(0, color="black", lw=0.6)

    ax5.axvspan(elec_1ra, elec_2da, color="crimson", alpha=0.07)
    ax5.axvline(elec_1ra, color="crimson", lw=1.8, ls="--",
                label="1ra Vuelta (11-abr)")
    ax5.axvline(elec_2da, color="purple",  lw=1.8, ls="--",
                label="2da Vuelta (6-jun)")

    ymin = min(r21.min(), neto21.min()) * 1.15
    ax5.text(elec_1ra, ymin, "1ra Vuelta\n(11-abr)",
             rotation=90, va="bottom", ha="right", fontsize=8, color="crimson")
    ax5.text(elec_2da, ymin, "2da Vuelta\n(6-jun)",
             rotation=90, va="bottom", ha="right", fontsize=8, color="purple")

    ax5.set_ylabel("MM USD (neto diario)")
    ax5.set_title("Transferencias Exterior - 2021", fontweight="bold", fontsize=12)
    ax5.xaxis.set_major_locator(mdates.MonthLocator())
    ax5.xaxis.set_major_formatter(mdates.DateFormatter("%b"))
    ax5.tick_params(axis="x", rotation=0)
    ax5.legend(fontsize=8, loc="upper right")
    ax5.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(DIR_OUTPUT / "05_flujos_2021_electoral.png", dpi=150, bbox_inches="tight")
    plt.show()
