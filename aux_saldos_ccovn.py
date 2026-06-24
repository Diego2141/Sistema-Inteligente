# -*- coding: utf-8 -*-
"""
aux_saldos_ccovn.py
===================
Lee Saldos_CCOVN.xlsx (saldos de cierre CC + OVN por banco),
construye dos series diarias y exporta para validación:

  - ccovn_sistema : suma de todos los bancos
  - ccovn_bbva    : solo BBVA

Uso: python aux_saldos_ccovn.py
     → revisa el output en consola y el Excel generado antes de incorporar a step001
"""

import pathlib
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

RUTA_IN  = pathlib.Path(
    r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Raw\Saldos_CCOVN.xlsx"
)
RUTA_OUT = RUTA_IN.parent.parent / "Clean" / "saldos_ccovn_features.xlsx"
RUTA_PLOT = RUTA_IN.parent.parent.parent / "2. Output" / "saldos_ccovn.png"

# Término de búsqueda para identificar la columna BBVA (case-insensitive)
BBVA_KEYWORD = "bbva"

# ── Carga ──────────────────────────────────────────────────────────────────────
print("Leyendo Saldos_CCOVN.xlsx...")
raw = pd.read_excel(RUTA_IN, header=0)

# Primera columna = fechas
col_fecha = raw.columns[0]
cols_bancos = raw.columns[1:].tolist()

raw[col_fecha] = pd.to_datetime(raw[col_fecha])
raw = raw.sort_values(col_fecha).reset_index(drop=True)

print(f"\n  Período   : {raw[col_fecha].min().date()} → {raw[col_fecha].max().date()}")
print(f"  Filas     : {len(raw):,}")
print(f"  Columnas  : {len(cols_bancos)} bancos")
print(f"\n  Columnas encontradas:")
for i, c in enumerate(cols_bancos, 1):
    n_validos = raw[c].notna().sum()
    print(f"    {i:>2}. {str(c):<35}  ({n_validos:,} obs no-nulos)")

# ── Identificar columna BBVA ───────────────────────────────────────────────────
cols_bbva = [c for c in cols_bancos if BBVA_KEYWORD in str(c).lower()]

if len(cols_bbva) == 0:
    print(f"\n  AVISO: no se encontró columna con '{BBVA_KEYWORD}' en el nombre.")
    print("  Ajusta BBVA_KEYWORD en el script con el nombre exacto de la columna.")
    col_bbva = None
elif len(cols_bbva) > 1:
    print(f"\n  AVISO: múltiples columnas con '{BBVA_KEYWORD}': {cols_bbva}")
    print("  Usando la primera. Ajusta BBVA_KEYWORD si es incorrecto.")
    col_bbva = cols_bbva[0]
else:
    col_bbva = cols_bbva[0]
    print(f"\n  Columna BBVA identificada: '{col_bbva}'")

# ── Construir features ────────────────────────────────────────────────────────
df = pd.DataFrame()
df["fecha"] = raw[col_fecha]

# Convertir columnas a numérico
for c in cols_bancos:
    raw[c] = pd.to_numeric(raw[c], errors="coerce")

# SISTEMA = suma de todos los bancos
df["ccovn_sistema"] = raw[cols_bancos].sum(axis=1, skipna=True)

# BBVA
if col_bbva:
    df["ccovn_bbva"] = raw[col_bbva]
else:
    df["ccovn_bbva"] = np.nan

# Variaciones diarias
df["var_ccovn_sistema"] = df["ccovn_sistema"].diff()
df["var_ccovn_bbva"]    = df["ccovn_bbva"].diff()

# BBVA share del sistema
df["bbva_share_sistema"] = df["ccovn_bbva"] / df["ccovn_sistema"].replace(0, np.nan)

# Medias móviles (5d y 22d) — útiles como features
for ventana in [5, 22]:
    df[f"ccovn_sistema_ma{ventana}"] = df["ccovn_sistema"].rolling(ventana, min_periods=ventana).mean()
    df[f"ccovn_bbva_ma{ventana}"]    = df["ccovn_bbva"].rolling(ventana, min_periods=ventana).mean()

# ── Resumen estadístico ───────────────────────────────────────────────────────
print(f"\n{'─'*60}")
print(f"  RESUMEN — features calculados")
print(f"{'─'*60}")
for col in ["ccovn_sistema", "ccovn_bbva", "var_ccovn_sistema", "var_ccovn_bbva",
            "bbva_share_sistema"]:
    s = df[col].dropna()
    if len(s) == 0:
        continue
    print(f"\n  {col}")
    print(f"    Obs válidos : {len(s):,}")
    print(f"    Media       : {s.mean()/1e6:+.1f}M")
    print(f"    Mediana     : {s.median()/1e6:+.1f}M")
    print(f"    Std         : {s.std()/1e6:.1f}M")
    print(f"    Min / Max   : {s.min()/1e6:.1f}M  /  {s.max()/1e6:.1f}M")

# ── Gráfico de validación ─────────────────────────────────────────────────────
fig, axes = plt.subplots(3, 1, figsize=(15, 12), sharex=True)
fig.suptitle("Saldos CC + OVN — validación de features\n"
             "ccovn_sistema y ccovn_bbva",
             fontsize=12, fontweight="bold")

# Panel 1: nivel sistema y BBVA
ax = axes[0]
ax.plot(df["fecha"], df["ccovn_sistema"] / 1e9,
        color="#1565C0", lw=1.2, label="Sistema")
if col_bbva:
    ax.plot(df["fecha"], df["ccovn_bbva"] / 1e9,
            color="#F44336", lw=1.2, label="BBVA")
ax.set_ylabel("Saldo (B USD)")
ax.set_title("Nivel de saldo CC + OVN")
ax.legend(fontsize=9)
ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.1f}B"))

# Panel 2: variación diaria
ax = axes[1]
cols_var  = np.where(df["var_ccovn_sistema"] >= 0, "#4CAF50", "#F44336")
ax.bar(df["fecha"], df["var_ccovn_sistema"] / 1e9,
       color=cols_var, alpha=0.6, width=1, label="Var. Sistema")
if col_bbva:
    ax.plot(df["fecha"], df["var_ccovn_bbva"] / 1e9,
            color="navy", lw=0.8, alpha=0.7, label="Var. BBVA")
ax.axhline(0, color="k", lw=0.7)
ax.set_ylabel("Variación (B USD)")
ax.set_title("Variación diaria")
ax.legend(fontsize=9)
ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.2f}B"))

# Panel 3: BBVA share del sistema
ax = axes[2]
if col_bbva:
    ax.plot(df["fecha"], df["bbva_share_sistema"] * 100,
            color="#7B1FA2", lw=1.0)
    ax.axhline(df["bbva_share_sistema"].median() * 100, color="red",
               lw=1, ls="--",
               label=f"Mediana = {df['bbva_share_sistema'].median()*100:.1f}%")
    ax.set_ylabel("BBVA / Sistema (%)")
    ax.set_title("Participación BBVA en el sistema")
    ax.legend(fontsize=9)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.0f}%"))

plt.tight_layout()
RUTA_PLOT.parent.mkdir(parents=True, exist_ok=True)
fig.savefig(RUTA_PLOT, dpi=130, bbox_inches="tight")
plt.close()
print(f"\n  Gráfico guardado: {RUTA_PLOT.name}")

# ── Exportar ──────────────────────────────────────────────────────────────────
RUTA_OUT.parent.mkdir(parents=True, exist_ok=True)

with pd.ExcelWriter(RUTA_OUT, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Features", index=False)

    # Hoja de columnas originales para auditoría
    df_orig = raw[[col_fecha] + cols_bancos].copy()
    df_orig.columns = ["fecha"] + [str(c) for c in cols_bancos]
    df_orig.to_excel(writer, sheet_name="Raw_bancos", index=False)

    # Hoja de descripción
    desc = pd.DataFrame([
        {"feature": "ccovn_sistema",        "descripcion": "Suma CC+OVN de todos los bancos (USD)",        "rezago": 0, "desde": df['fecha'].min().date()},
        {"feature": "ccovn_bbva",           "descripcion": f"CC+OVN BBVA — col '{col_bbva}' (USD)",         "rezago": 0, "desde": df['fecha'].min().date()},
        {"feature": "var_ccovn_sistema",    "descripcion": "Variación diaria ccovn_sistema",                "rezago": 0, "desde": df['fecha'].min().date()},
        {"feature": "var_ccovn_bbva",       "descripcion": "Variación diaria ccovn_bbva",                   "rezago": 0, "desde": df['fecha'].min().date()},
        {"feature": "bbva_share_sistema",   "descripcion": "ccovn_bbva / ccovn_sistema — participación",    "rezago": 0, "desde": df['fecha'].min().date()},
        {"feature": "ccovn_sistema_ma5",    "descripcion": "Media móvil 5d de ccovn_sistema",               "rezago": 0, "desde": df['fecha'].min().date()},
        {"feature": "ccovn_sistema_ma22",   "descripcion": "Media móvil 22d de ccovn_sistema",              "rezago": 0, "desde": df['fecha'].min().date()},
        {"feature": "ccovn_bbva_ma5",       "descripcion": "Media móvil 5d de ccovn_bbva",                  "rezago": 0, "desde": df['fecha'].min().date()},
        {"feature": "ccovn_bbva_ma22",      "descripcion": "Media móvil 22d de ccovn_bbva",                 "rezago": 0, "desde": df['fecha'].min().date()},
    ])
    desc.to_excel(writer, sheet_name="Descripcion_features", index=False)

print(f"\n  Exportado: {RUTA_OUT}")
print(f"    Hoja 'Features'              : {len(df):,} filas × {len(df.columns)} cols")
print(f"    Hoja 'Raw_bancos'            : datos originales por banco")
print(f"    Hoja 'Descripcion_features'  : diccionario de features")
print(f"\n  Features para step001:")
print(f"    ccovn_sistema  (desde {df['fecha'].min().date()})")
print(f"    ccovn_bbva     (desde {df['fecha'].min().date()})")
