# -*- coding: utf-8 -*-
"""
aux_graficos_hallazgos.py
==========================
Genera los 3 gráficos que respaldan los hallazgos mostrados a Operaciones
Monetarias, a partir de Transacciones_BancaLocal.xlsx.

Replica la carga/agregación de step001_build_feature_matrix.py (líneas
541-561) para no reinventar reglas de negocio:
  - Broker (banco), Fecha Valor (fecha), Delivery Principal Usd (monto)
  - monto < 0 → Retiro (R), en valor absoluto | monto > 0 → Depósito (D)
  - Alias histórico: CONTINEN → BBVA (cambio de nombre 2020)

La definición de "período de encaje" (día hábil dentro del mes, cierre =
último día hábil del mes) y la bandera de mes-cierre-trimestre replican
step001_build_feature_matrix.py líneas 2249-2278.

Hallazgo 1 → 01_ciclo_encaje_slider.html   (con barra deslizante por año)
Hallazgo 2 → 02_retiro_mas_temprano.html
Hallazgo 3 → 03_bbva_arrastra_agregado.html

Uso:
    python aux_graficos_hallazgos.py [ruta_al_excel]
Si no se pasa ruta, usa la ruta por defecto de step001 (params["ruta_datos_bancarios"]).
"""

import sys
import json
import pathlib

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ── Configuración ────────────────────────────────────────────────────────────

RUTA_DEFAULT = pathlib.Path(
    r"H:\DPINV\CARPETAS PERSONALES\DIEGO\3. Sistema Inteligente\1. Data\Raw\Transacciones_BancaLocal.xlsx"
)
RUTA_SALIDA = pathlib.Path(__file__).parent / "2. Output" / "graficos_hallazgos"

ALIAS_BANCOS = {"CONTINEN": "BBVA"}
BANCO_FOCO = "BBVA"

# Cómo se incluye plotly.js en los HTML generados:
#   "directory" → escribe plotly.min.js una vez en la carpeta de salida y los tres
#                 HTML lo referencian. Funciona sin internet (red corporativa con
#                 CDN bloqueado) y no repite la librería en cada archivo.
#   "cdn"       → los HTML piden plotly.js a internet. Archivos de ~45 KB, pero se
#                 ven en blanco si el CDN está bloqueado.
#   True        → embebe la librería en cada HTML (~4.8 MB c/u). Úsalo si necesitas
#                 mandar UN archivo suelto por correo, sin la carpeta.
PLOTLYJS_MODE = "directory"

# Ventana de cierre: los últimos N días hábiles del mes (dias_al_cierre_mes < N,
# es decir 0..N-1). UNA sola definición para todos los gráficos — antes convivían
# dos ventanas distintas (10 días para intensidad/anticipación, 3 para magnitud),
# lo que obligaba a explicar en la reunión por qué "cierre" significaba dos cosas.
VENTANA_CIERRE_BDAYS = 5
# Umbral de masa acumulada de retiro para fijar el "día de inicio del retiro"
UMBRAL_INICIO_RETIRO = 0.5

pd.options.mode.chained_assignment = None


# ── 1. Carga + validación ────────────────────────────────────────────────────

def cargar_transacciones(ruta):
    print(f"Leyendo {ruta} ...")
    df_raw = pd.read_excel(ruta)
    if df_raw.empty:
        raise ValueError("El archivo está vacío.")

    faltantes = [c for c in ("Broker", "Fecha Valor", "Delivery Principal Usd") if c not in df_raw.columns]
    if faltantes:
        raise ValueError(f"Faltan columnas esperadas: {faltantes}")

    n_fecha_bad = pd.to_datetime(df_raw["Fecha Valor"], errors="coerce").isna().sum()
    n_monto_bad = pd.to_numeric(df_raw["Delivery Principal Usd"], errors="coerce").isna().sum()
    if n_fecha_bad or n_monto_bad:
        print(f"  ⚠ {n_fecha_bad:,} 'Fecha Valor' no parseables, {n_monto_bad:,} montos no numéricos (se descartan).")

    for col in ("Tipo Instrumento", "Product Type"):
        if col in df_raw.columns:
            vals = df_raw[col].dropna().unique()
            if len(vals) > 1:
                print(f"  ⚠ '{col}' trae {len(vals)} valores distintos, no solo el uniforme esperado: {list(vals)[:10]}")

    df_raw["Fecha Valor"] = pd.to_datetime(df_raw["Fecha Valor"], errors="coerce")
    df_raw["monto"] = pd.to_numeric(df_raw["Delivery Principal Usd"], errors="coerce")
    df_raw = df_raw.dropna(subset=["Fecha Valor", "monto", "Broker"])

    df_raw["R"] = df_raw["monto"].clip(upper=0).abs()
    df_raw["D"] = df_raw["monto"].clip(lower=0)

    df = (
        df_raw.groupby(["Broker", "Fecha Valor"])[["R", "D"]]
        .sum()
        .reset_index()
        .rename(columns={"Broker": "banco", "Fecha Valor": "fecha"})
    )

    alias_lower = {k.lower(): v for k, v in ALIAS_BANCOS.items()}
    df["banco"] = df["banco"].astype(str).str.strip()
    df["banco"] = df["banco"].apply(lambda b: alias_lower.get(b.lower(), b))
    df = df.groupby(["banco", "fecha"])[["R", "D"]].sum().reset_index()
    df = df.sort_values(["banco", "fecha"])

    n_bancos = df["banco"].nunique()
    f_min, f_max = df["fecha"].min().date(), df["fecha"].max().date()
    volumen = (df.groupby("banco")[["R", "D"]].sum().sum(axis=1).sort_values(ascending=False))
    top5 = volumen.head(5)

    print(f"  {len(df):,} filas agregadas | {n_bancos} bancos | {f_min} → {f_max}")
    print("  Top 5 bancos por volumen (R+D):")
    for b, v in top5.items():
        print(f"    {b:<20} {v:,.0f}")
    if BANCO_FOCO not in volumen.index:
        print(f"  ⚠ '{BANCO_FOCO}' no aparece en los datos — revisar alias.")

    return df


# ── 2. Calendario de período de encaje (replica step001) ────────────────────

def calendario_habil_peru(años):
    """Días hábiles PE+US, replicando build_peru_calendar() de
    step001_build_feature_matrix.py (líneas 406-468).

    Se replica en vez de importarse porque step001 es un módulo pesado con
    efectos al importar; los feriados deben quedar IDÉNTICOS a los de allá."""
    from pandas.tseries.holiday import (
        AbstractHolidayCalendar, Holiday, GoodFriday, USFederalHolidayCalendar, Easter,
    )
    from pandas.tseries.offsets import Day as _Day

    f0, f1 = f"{min(años)}-01-01", f"{max(años)}-12-31"

    class _PeruCalendar(AbstractHolidayCalendar):
        rules = [
            Holiday("AnioNuevo",   month=1,  day=1),
            Holiday("JuevesSanto", month=1,  day=1, offset=[Easter(), _Day(-3)]),
            GoodFriday,
            Holiday("Trabajo",     month=5,  day=1),
            Holiday("SanPedro",    month=6,  day=29),
            Holiday("FiestasP1",   month=7,  day=28),
            Holiday("FiestasP2",   month=7,  day=29),
            Holiday("SantaRosa",   month=8,  day=30),
            Holiday("Angamos",     month=10, day=8),
            Holiday("TodosSantos", month=11, day=1),
            Holiday("Inmaculada",  month=12, day=8),
            Holiday("Nochebuena",  month=12, day=24),
            Holiday("Navidad",     month=12, day=25),
        ]

    hols = set(_PeruCalendar().holidays(f0, f1).normalize())
    hols |= set(USFederalHolidayCalendar().holidays(f0, f1).normalize())
    try:
        import holidays as _hlib
        hols |= set(pd.to_datetime(list(_hlib.Peru(years=list(años)).keys())).normalize())
        hols |= set(pd.to_datetime(list(_hlib.UnitedStates(years=list(años)).keys())).normalize())
    except ImportError:
        print("  ⚠ librería 'holidays' no instalada: se usan solo los feriados fijos "
              "PE+US y Semana Santa. Instalar con: pip install holidays")

    feriados = pd.to_datetime(sorted(hols))
    return pd.DatetimeIndex(
        [d for d in pd.bdate_range(f0, f1) if d not in set(feriados)]
    ), feriados


def anotar_calendario_encaje(df):
    """Agrega dias_al_cierre_mes (0 = último día hábil del mes) y es_mes_cierre_trim.

    EL CALENDARIO SALE DE LOS DATOS: días hábiles = fechas que aparecen en
    Transacciones_BancaLocal. Si el sistema operó ese día, hay movimiento; si no
    aparece, no fue hábil. Es auto-consistente y no depende de mantener una lista
    de feriados al día.

    Antes se usaba pd.bdate_range, que solo excluye sábados y domingos. Eso metía
    Navidad en la ventana de cierre de diciembre y la corría dos días: tomaba el
    25 y 26 en vez del 23 y 24, dejando fuera los de volumen grande.

    Los meses que la serie no cubre enteros se marcan con mes_completo = 0 y
    quedan fuera de los cortes por ventana de cierre: si los datos terminan a
    mitad de mes, ese último día no es el cierre del mes.

    Se cruza contra el calendario PE+US de step001 solo para avisar de rarezas
    (fechas en fin de semana, o días hábiles sin ningún movimiento)."""
    f_min, f_max = df["fecha"].min(), df["fecha"].max()

    habiles = pd.DatetimeIndex(sorted(df["fecha"].unique()))
    cal = pd.DataFrame({"fecha": habiles})
    cal["mes"] = cal["fecha"].dt.to_period("M")
    cal["pos_en_mes"] = cal.groupby("mes").cumcount() + 1
    cal["total_bdays_mes"] = cal.groupby("mes")["fecha"].transform("count")
    cal["dias_al_cierre_mes"] = cal["total_bdays_mes"] - cal["pos_en_mes"]
    cal["es_mes_cierre_trim"] = cal["fecha"].dt.month.isin([3, 6, 9, 12]).astype(int)
    cal["anio"] = cal["fecha"].dt.year
    cal["trimestre"] = cal["fecha"].dt.to_period("Q")

    # Un mes está completo si la serie abarca el mes entero de punta a punta.
    meses = cal["mes"].unique()
    completos = {m for m in meses if m.start_time >= f_min and m.end_time <= f_max}
    cal["mes_completo"] = cal["mes"].isin(completos).astype(int)

    incompletos = sorted(set(meses) - completos)
    print(f"  Calendario tomado de los datos: {len(habiles):,} días hábiles "
          f"({f_min.date()} → {f_max.date()})")
    if incompletos:
        print(f"  {len(incompletos)} mes(es) que la serie no cubre entero, fuera de la "
              f"ventana de cierre: {', '.join(str(m) for m in incompletos)}")

    # ── Cruce con el calendario PE+US, solo informativo ──────────────────────
    finde = cal[cal["fecha"].dt.dayofweek >= 5]
    if len(finde):
        print(f"  ⚠ {len(finde)} fechas del archivo caen en sábado o domingo "
              f"(p.ej. {sorted(finde['fecha'].dt.date)[:3]}): se cuentan como hábiles "
              f"porque hay movimiento, pero conviene revisarlas.")
    try:
        pe_bdays, _ = calendario_habil_peru(range(f_min.year, f_max.year + 1))
        pe_bdays = pe_bdays[(pe_bdays >= f_min) & (pe_bdays <= f_max)]
        faltan = pe_bdays.difference(habiles)
        if len(faltan):
            print(f"  · {len(faltan)} días que el calendario PE+US considera hábiles no "
                  f"tienen ningún movimiento (p.ej. {[d.date() for d in faltan[:3]]}): "
                  f"no cuentan como día hábil aquí.")
    except Exception as e:
        print(f"  · no se pudo cruzar con el calendario PE+US: {e}")

    # pos_en_mes (1 = primer día hábil) se exporta junto a dias_al_cierre_mes
    # (0 = último): los meses tienen entre ~19 y ~23 días hábiles, así que el
    # inicio y el cierre necesitan anclajes distintos. Contar el inicio desde el
    # cierre haría que las posiciones más lejanas solo existan en los meses
    # largos, y la barra de ese día promediaría un subconjunto sesgado.
    cols = ["fecha", "pos_en_mes", "dias_al_cierre_mes", "es_mes_cierre_trim",
            "anio", "trimestre", "mes_completo"]
    return df.merge(cal[cols], on="fecha", how="left")


# ── 3. Gráfico 1 — ciclo de encaje, barra deslizante por año ────────────────

def graf_ciclo_encaje(df):
    """R y D del sistema como % del volumen mensual, por día-al-cierre,
    separando meses cierre-trimestre vs. resto. Slider recorre los años."""
    sistema = df.groupby(["fecha", "pos_en_mes", "dias_al_cierre_mes", "es_mes_cierre_trim", "anio"])[["R", "D"]].sum().reset_index()

    total_mes = (
        df.assign(mes=df["fecha"].dt.to_period("M"))
        .groupby("mes")[["R", "D"]].sum()
    )
    sistema["mes"] = sistema["fecha"].dt.to_period("M")
    sistema = sistema.merge(total_mes, on="mes", suffixes=("", "_total_mes"))
    sistema["R_pct"] = np.where(sistema["R_total_mes"] > 0, sistema["R"] / sistema["R_total_mes"] * 100, 0)
    sistema["D_pct"] = np.where(sistema["D_total_mes"] > 0, sistema["D"] / sistema["D_total_mes"] * 100, 0)

    # Dos agregados con anclajes distintos: uno contando desde el primer día
    # hábil del mes y otro desde el último. Un mes de 19 días hábiles y uno de 23
    # aportan ambos a "día 1" y a "día 0 al cierre", pero solo el largo llega a
    # "22 días al cierre" — por eso el inicio se ancla al inicio.
    agg = (
        sistema.groupby(["anio", "es_mes_cierre_trim", "dias_al_cierre_mes"])[["R_pct", "D_pct"]]
        .mean()
        .reset_index()
    )
    agg_ini = (
        sistema.groupby(["anio", "es_mes_cierre_trim", "pos_en_mes"])[["R_pct", "D_pct"]]
        .mean()
        .reset_index()
    )
    # Cuántos meses aportan a cada posición: si una barra promedia pocos meses,
    # conviene saberlo antes de leerla como representativa.
    n_ini = (sistema.groupby(["anio", "es_mes_cierre_trim", "pos_en_mes"])["fecha"]
             .nunique().rename("n_meses").reset_index())
    n_fin = (sistema.groupby(["anio", "es_mes_cierre_trim", "dias_al_cierre_mes"])["fecha"]
             .nunique().rename("n_meses").reset_index())
    agg_ini = agg_ini.merge(n_ini, on=["anio", "es_mes_cierre_trim", "pos_en_mes"])
    agg = agg.merge(n_fin, on=["anio", "es_mes_cierre_trim", "dias_al_cierre_mes"])

    anios = sorted(agg["anio"].unique())
    fig = go.Figure()
    trace_meta = []
    for anio in anios:
        for es_trim, nombre, color_r, color_d in [(1, "cierre de trimestre", "#a85a2b", "#1f4d5c"),
                                                     (0, "resto de meses", "#c99a7a", "#7fa8b0")]:
            sub = agg[(agg["anio"] == anio) & (agg["es_mes_cierre_trim"] == es_trim)].sort_values("dias_al_cierre_mes", ascending=False)
            visible = (anio == anios[-1])
            fig.add_trace(go.Scatter(
                x=-sub["dias_al_cierre_mes"], y=sub["R_pct"], mode="lines",
                name=f"Retiro — {nombre}", line=dict(color=color_r, width=2.4),
                visible=visible,
            ))
            fig.add_trace(go.Scatter(
                x=-sub["dias_al_cierre_mes"], y=sub["D_pct"], mode="lines",
                name=f"Depósito — {nombre}", line=dict(color=color_d, width=2, dash="dot"),
                visible=visible,
            ))
            trace_meta.append(anio)
            trace_meta.append(anio)

    steps = []
    for anio in anios:
        vis = [m == anio for m in trace_meta]
        steps.append(dict(method="update", args=[{"visible": vis}, {"title": f"Ciclo de encaje — {anio}"}], label=str(anio)))

    fig.update_layout(
        title=f"Ciclo de encaje — {anios[-1]}",
        xaxis_title="días hábiles al cierre de mes (0 = último día hábil)",
        yaxis_title="% del volumen mensual concentrado ese día",
        sliders=[dict(active=len(anios) - 1, steps=steps, currentvalue={"prefix": "Año: "})],
        template="plotly_white",
        legend=dict(orientation="h", y=-0.2),
    )
    ren = {"anio": "año", "es_mes_cierre_trim": "es_mes_cierre_trimestre",
           "R_pct": "retiro_pct_del_mes", "D_pct": "deposito_pct_del_mes"}
    datos = agg.rename(columns=ren).sort_values(
        ["año", "es_mes_cierre_trimestre", "dias_al_cierre_mes"], ascending=[True, False, False])
    datos_ini = agg_ini.rename(columns=ren).sort_values(
        ["año", "es_mes_cierre_trimestre", "pos_en_mes"], ascending=[True, False, True])
    return fig, datos, datos_ini


# ── 4. Gráfico 2 — el retiro llega cada vez más temprano ────────────────────

def graf_retiro_mas_temprano(df):
    """Por cada mes cierre-trimestre, día (dias_al_cierre_mes) en que el
    retiro acumulado de la ventana de cierre alcanza UMBRAL_INICIO_RETIRO
    del retiro total de esa ventana. Serie a lo largo del tiempo + tendencia."""
    trim = df[(df["es_mes_cierre_trim"] == 1) & (df["mes_completo"] == 1) & (df["dias_al_cierre_mes"] < VENTANA_CIERRE_BDAYS)]
    sistema = trim.groupby(["trimestre", "dias_al_cierre_mes"])["R"].sum().reset_index()

    puntos = []
    for t, sub in sistema.groupby("trimestre"):
        sub = sub.sort_values("dias_al_cierre_mes", ascending=False)  # de más lejos a más cerca del cierre
        total = sub["R"].sum()
        if total <= 0:
            continue
        sub["acum_pct"] = sub["R"].cumsum() / total
        primero = sub[sub["acum_pct"] >= UMBRAL_INICIO_RETIRO].head(1)
        if primero.empty:
            continue
        puntos.append({"trimestre": t.to_timestamp(), "dia_inicio": int(primero["dias_al_cierre_mes"].iloc[0])})

    serie = pd.DataFrame(puntos).sort_values("trimestre")
    if serie.empty:
        raise ValueError("No se pudo calcular 'día de inicio del retiro' — revisar VENTANA_CIERRE_BDAYS.")

    x_num = np.arange(len(serie))
    coef = np.polyfit(x_num, serie["dia_inicio"], 1) if len(serie) > 1 else (0, serie["dia_inicio"].mean())
    tendencia = np.polyval(coef, x_num)

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=serie["trimestre"], y=serie["dia_inicio"], mode="markers+lines",
                              name="día de inicio del retiro", line=dict(color="#a85a2b")))
    fig.add_trace(go.Scatter(x=serie["trimestre"], y=tendencia, mode="lines", name="tendencia",
                              line=dict(color="#1f4d5c", dash="dash")))
    fig.update_layout(
        title=f"El retiro empieza cada vez más temprano (umbral {UMBRAL_INICIO_RETIRO:.0%} del tramo de cierre)",
        xaxis_title="cierre de trimestre", yaxis_title="días hábiles antes del cierre en que empieza el retiro",
        template="plotly_white",
    )
    datos = serie.copy()
    datos["tendencia_ajustada"] = tendencia
    datos["trimestre"] = datos["trimestre"].dt.to_period("Q").astype(str)
    datos = datos.rename(columns={"dia_inicio": "dias_antes_del_cierre_inicio_retiro"})
    # Detalle día a día que alimenta el cálculo, para poder auditar el umbral
    detalle = sistema.copy()
    detalle["trimestre"] = detalle["trimestre"].astype(str)
    detalle = detalle.rename(columns={"R": "retiro_sistema"}).sort_values(
        ["trimestre", "dias_al_cierre_mes"], ascending=[True, False]
    )
    return fig, datos, detalle


# ── 5. Gráfico 3 — BBVA arrastra el agregado (solo con datos de flujo) ──────

def graf_bbva_arrastra(df):
    """Panel A: cuota de BBVA en el DRENAJE NETO del sistema en la ventana de cierre.
    Panel B: intensidad — salida neta diaria en la ventana por unidad de flujo
    bruto diario del banco — de BBVA vs. el resto de bancos grandes.

    Usa neto (R - D), no retiro bruto: los depósitos son entradas de capital que
    compensan la salida, así que el bruto sobreestima cuánto se drena de verdad.
    A cambio, el neto puede acercarse a cero o cambiar de signo, y ahí tanto el
    porcentaje como el ratio dejan de tener sentido — por eso cada métrica se
    calcula bajo guarda y lo descartado se reporta, no se esconde."""
    ventana = df[(df["es_mes_cierre_trim"] == 1) & (df["mes_completo"] == 1) & (df["dias_al_cierre_mes"] < VENTANA_CIERRE_BDAYS)]
    resto = df[(df["es_mes_cierre_trim"] == 1) & (df["mes_completo"] == 1) & (df["dias_al_cierre_mes"] >= VENTANA_CIERRE_BDAYS)]

    neto = lambda g: g["R"].sum() - g["D"].sum()   # salida neta: + = sale plata
    N_tramo_banco = ventana.groupby(["trimestre", "banco"]).apply(neto)

    # ── Participación: cuota del DRENAJE, no del neto agregado ────────────────
    # El neto del sistema no sirve de denominador: los bancos que están entrando
    # capital lo achican, y la cuota de BBVA se dispara por encima de 100% o
    # cambia de signo sin que su comportamiento haya cambiado. Se usa la suma de
    # los netos positivos — cuánto se drena en total —, que siempre da 0-100%.
    drenaje_sistema = N_tramo_banco.clip(lower=0).groupby("trimestre").sum()
    part_bbva = (N_tramo_banco.xs(BANCO_FOCO, level="banco").clip(lower=0)
                 / drenaje_sistema.where(drenaje_sistema > 0) * 100)
    part_bbva = part_bbva.replace([np.inf, -np.inf], np.nan).dropna()

    # ── Montos absolutos, para la tabla del one-pager ─────────────────────────
    # El neto del SISTEMA aquí es el agregado real (suma de todos los bancos,
    # con signo), no la suma de netos positivos: es lo que se compara contra el
    # neto de BBVA. Puede ser <= 0 — trimestres en que el sistema recibió más de
    # lo que sacó —, y en esos la proporción no significa nada y se deja vacía.
    N_bbva = N_tramo_banco.xs(BANCO_FOCO, level="banco")
    N_sistema = ventana.groupby("trimestre").apply(neto)
    prop_bbva = (N_bbva / N_sistema.where(N_sistema > 0) * 100).replace([np.inf, -np.inf], np.nan)
    MUSD = 1e6
    n_sin_prop = int((N_sistema <= 0).sum())
    if n_sin_prop:
        print(f"  · {n_sin_prop} trimestres sin proporción: el sistema fue receptor neto "
              f"en la ventana (proporción vacía, no cero)")

    # ── Tamaño de BBVA, para contrastarlo con su cuota del drenaje ────────────
    # Una cuota alta del drenaje no prueba nada por sí sola: un banco grande
    # mueve más en términos absolutos. La referencia honesta es su peso en el
    # flujo BRUTO del sistema en la misma ventana — cuánto del movimiento total
    # es suyo. La brecha entre "cuota del drenaje" y "cuota del tamaño" es lo
    # que realmente mide si drena más de lo que le correspondería.
    bruto_tramo_banco = ventana.groupby(["trimestre", "banco"])[["R", "D"]].sum().sum(axis=1)
    bruto_tramo_sistema = bruto_tramo_banco.groupby("trimestre").sum()
    tam_bbva = (bruto_tramo_banco.xs(BANCO_FOCO, level="banco")
                / bruto_tramo_sistema.where(bruto_tramo_sistema > 0) * 100)
    tam_bbva = tam_bbva.replace([np.inf, -np.inf], np.nan).dropna()

    # ── Intensidad: neto del cierre contra la ESCALA del banco ────────────────
    # El denominador natural sería el neto diario fuera de la ventana, pero ese
    # neto es negativo casi siempre: los bancos acumulan durante el mes y drenan
    # al cierre — justamente el patrón que describe el one-pager. Dividir por él
    # da un ratio negativo y sin sentido. Se divide por el flujo BRUTO diario,
    # que mide el tamaño operativo del banco y nunca es ~0: la métrica pasa a ser
    # "cuánto drena por día en el cierre, por unidad de su actividad diaria".
    dias_resto = resto.groupby(["trimestre", "banco"])["fecha"].nunique()
    N_diario_resto = resto.groupby(["trimestre", "banco"]).apply(neto) / dias_resto
    N_diario_tramo = N_tramo_banco / VENTANA_CIERRE_BDAYS
    bruto_diario_resto = (resto.groupby(["trimestre", "banco"])[["R", "D"]].sum().sum(axis=1) / dias_resto)
    escala = bruto_diario_resto.where(bruto_diario_resto > 0)
    intensidad = (N_diario_tramo / escala).replace([np.inf, -np.inf], np.nan)

    n_sin_escala = int(bruto_diario_resto.notna().sum() - escala.notna().sum())
    n_sin_drenaje = int((drenaje_sistema <= 0).sum())
    print(f"  Neto: {n_sin_drenaje} trimestres sin drenaje agregado, "
          f"{n_sin_escala} pares banco-trimestre sin escala (banco inactivo ese mes)")
    n_neg = int((N_tramo_banco.xs(BANCO_FOCO, level="banco") < 0).sum())
    if n_neg:
        print(f"  · {n_neg} trimestres en que {BANCO_FOCO} fue receptor neto en la ventana "
              f"(participación 0% esos trimestres, no negativa)")

    volumen_banco = df.groupby("banco")[["R", "D"]].sum().sum(axis=1).sort_values(ascending=False)  # tamaño = flujo bruto
    grandes = [b for b in volumen_banco.index[:6] if b != BANCO_FOCO]

    int_bbva = intensidad.xs(BANCO_FOCO, level="banco").dropna()
    int_pares = intensidad[intensidad.index.get_level_values("banco").isin(grandes)]
    int_pares_prom = int_pares.groupby("trimestre").mean().dropna()

    fig = make_subplots(rows=2, cols=1, shared_xaxes=True,
                         subplot_titles=("% de la salida NETA del sistema (ventana de cierre) que es de BBVA",
                                          "Salida neta diaria en la ventana, por unidad de flujo bruto diario"))
    fig.add_trace(go.Scatter(x=part_bbva.index.to_timestamp(), y=part_bbva.values, mode="lines+markers",
                              name="% BBVA del drenaje del sistema", line=dict(color="#a85a2b")), row=1, col=1)
    fig.add_trace(go.Scatter(x=tam_bbva.index.to_timestamp(), y=tam_bbva.values, mode="lines",
                              name="% BBVA del flujo bruto (su tamaño)",
                              line=dict(color="#5c6673", dash="dot")), row=1, col=1)
    fig.add_trace(go.Scatter(x=int_bbva.index.to_timestamp(), y=int_bbva.values, mode="lines+markers",
                              name="Intensidad BBVA", line=dict(color="#a85a2b")), row=2, col=1)
    fig.add_trace(go.Scatter(x=int_pares_prom.index.to_timestamp(), y=int_pares_prom.values, mode="lines+markers",
                              name="Intensidad — otros bancos grandes (prom.)", line=dict(color="#5c6673", dash="dot")), row=2, col=1)

    fig.update_yaxes(title_text="% del drenaje del sistema", row=1, col=1)
    fig.update_yaxes(title_text="neto de cierre / actividad diaria", row=2, col=1)
    fig.update_layout(title="BBVA arrastra el agregado del sistema (salida neta)", template="plotly_white", legend=dict(orientation="h", y=-0.15))

    datos = pd.DataFrame({
        "participacion_bbva_pct_retiro_sistema": part_bbva,   # cuota del DRENAJE neto
        "tamano_bbva_pct_bruto_sistema": tam_bbva,            # cuota del FLUJO BRUTO = su tamaño
        "neto_bbva_musd": N_bbva / MUSD,                      # millones USD, + = salio plata
        "neto_sistema_musd": N_sistema / MUSD,                # idem, agregado real con signo
        "proporcion_bbva_pct": prop_bbva,                     # vacia si el sistema no drenó
        "intensidad_bbva": int_bbva,
        "intensidad_otros_grandes_prom": int_pares_prom,
    })
    datos.index = datos.index.astype(str)
    datos = datos.reset_index().rename(columns={"index": "trimestre"})

    # Detalle por banco: insumo de ambos paneles, para auditar banco por banco
    detalle = pd.DataFrame({
        "neto_ventana_cierre": N_tramo_banco,
        "neto_diario_ventana_cierre": N_diario_tramo,
        "neto_diario_resto_del_mes": N_diario_resto,
        "bruto_diario_resto_del_mes": bruto_diario_resto,   # denominador de la intensidad
        "intensidad": intensidad,
    }).reset_index()
    detalle["trimestre"] = detalle["trimestre"].astype(str)
    detalle["es_banco_grande"] = detalle["banco"].isin(grandes + [BANCO_FOCO]).astype(int)
    detalle = detalle.sort_values(["trimestre", "banco"])

    return fig, datos, detalle



# ── 5b. Cuota anual de cada banco en la salida neta ─────────────────────────

# Bancos con matriz en el exterior. La hipótesis a contrastar es que el patrón
# de salida al cierre lo siguen los bancos globales y no los de capital local.
# Se listan TODOS los que califican, no solo los que sostienen la hipótesis: si
# Scotiabank, ICBC o Bank of China no muestran el patrón, eso tiene que verse.
# Match por subcadena sobre el nombre normalizado, porque el archivo escribe
# "J.P. MORGAN", "BANK OF CHINA (PERÚ)" y variantes con puntos y paréntesis.
CLAVES_GLOBALES = ("BBVA", "CONTINEN", "CITI", "JPMORGAN", "DEUTSCHE",
                   "SCOTIABANK", "BANKOFCHINA", "ICBC", "BCI", "SANTANDER", "HSBC")


def _es_global(nombre):
    n = "".join(ch for ch in str(nombre).upper() if ch.isalnum())
    return int(any(k in n for k in CLAVES_GLOBALES))


def datos_cuota_por_banco(df):
    """Cuota de cada banco en la salida neta del sistema, por AÑO.

    Es un agregado anual de verdad: suma los netos del año y recién ahí divide.
    No es el promedio de las cuotas trimestrales, que le daría el mismo peso a
    un trimestre chico que a uno grande.

    Denominador = suma de los netos positivos del año, es decir cuánta salida
    hubo en total. Los bancos que fueron receptores netos no lo achican."""
    ventana = df[(df["es_mes_cierre_trim"] == 1) & (df["mes_completo"] == 1)
                 & (df["dias_al_cierre_mes"] < VENTANA_CIERRE_BDAYS)]
    if ventana.empty:
        return pd.DataFrame(), pd.DataFrame()

    neto_banco = (ventana.groupby(["anio", "banco"])
                  .apply(lambda g: g["R"].sum() - g["D"].sum())
                  .rename("neto"))
    salida_total = neto_banco.clip(lower=0).groupby("anio").sum().rename("salida_sistema")

    d = neto_banco.reset_index().merge(salida_total.reset_index(), on="anio")
    d["cuota_pct"] = np.where(d["salida_sistema"] > 0,
                              d["neto"].clip(lower=0) / d["salida_sistema"] * 100, np.nan)
    d["neto_musd"] = d["neto"] / 1e6
    d["es_global"] = d["banco"].apply(_es_global)

    # Trimestres cerrados por año: para marcar el año en curso como parcial
    trim = (ventana.groupby("anio")["trimestre"].nunique().rename("trimestres_cerrados"))
    d = d.merge(trim.reset_index(), on="anio")
    d = d.rename(columns={"anio": "año"}).sort_values(["año", "cuota_pct"],
                                                       ascending=[True, False])

    # Serie ancha: una fila por año, una columna por banco
    ancha = d.pivot(index="año", columns="banco", values="cuota_pct").reset_index()
    ancha = ancha.merge(trim.reset_index().rename(columns={"anio": "año"}), on="año")
    return d, ancha


# ── 6. Magnitud absoluta del retiro, indexada ───────────────────────────────

def datos_magnitud_retiro(df):
    """Retiro TOTAL (no % del mes) del sistema en la ventana de cierre
    (VENTANA_CIERRE_BDAYS días hábiles) de cada cierre de trimestre, sumado por
    año e indexado al primer año = 100.

    Los otros gráficos normalizan a % del mes o usan proporciones/ratios — por
    diseño no pueden mostrar que el monto absoluto creció. Esta serie es el
    complemento: mide el crecimiento real, no la forma. Se exporta como índice,
    no como monto en soles/dólares, para no imprimir cifras de flujo absolutas
    del sistema en un artifact que puede terminar compartido.

    Mide SALIDA NETA (R - D), no retiro bruto: en la ventana de cierre también
    entran depósitos, y contar solo los retiros sobreestima cuánto se drena."""
    tramo = df[(df["es_mes_cierre_trim"] == 1) & (df["mes_completo"] == 1) & (df["dias_al_cierre_mes"] < VENTANA_CIERRE_BDAYS)]
    por_anio = tramo.groupby("anio")[["R", "D"]].sum().reset_index().rename(columns={"anio": "año"})
    por_anio["salida_neta"] = por_anio["R"] - por_anio["D"]
    por_anio["retiro_bruto"] = por_anio["R"]          # se conserva para comparar en el Excel
    por_anio = por_anio.sort_values("año").drop(columns=["R", "D"])

    base = por_anio["salida_neta"].iloc[0] if len(por_anio) else 0
    if base > 0:
        por_anio["indice_retiro"] = (por_anio["salida_neta"] / base * 100).round(1)
        n_neg = int((por_anio["salida_neta"] < 0).sum())
        if n_neg:
            print(f"  ⚠ {n_neg} años con salida neta NEGATIVA en la ventana de cierre: "
                  f"esos años el sistema fue receptor neto, y el índice sale bajo cero.")
    else:
        # Indexar contra una base <= 0 da un índice sin sentido (signo invertido o
        # división por ~0). Mejor no publicar la serie que publicarla mal.
        por_anio["indice_retiro"] = np.nan
        print(f"  ⚠ Año base con salida neta <= 0 ({base:,.0f}): no se puede indexar. "
              f"Índice omitido — revisar el año base o usar retiro bruto.")
    return por_anio


# ── 7. Export para el one-pager (artifact) ────────────────────────────────

def exportar_json_artifact(datos1, datos1_ini, datos2, datos3, datos4, datos5, ruta):
    """Escribe un JSON compacto con SOLO las series que se dibujan.

    Deliberadamente NO incluye el detalle por banco ni el agregado diario:
    el one-pager solo necesita las curvas, y así lo que sale de la máquina
    es el mínimo indispensable. Los valores van redondeados."""
    # NaN → None. json.dump escribe los NaN de pandas como el literal `NaN`,
    # que NO es JSON valido: un parser estricto lo rechaza y, peor, si el JSON
    # se incrusta como literal JS el `NaN` se cuela silenciosamente y propaga
    # NaN a todo el calculo del grafico. Los NaN aqui son reales (bancos sin
    # retiro fuera del tramo de cierre → division sin definir), asi que van
    # como null y el grafico los dibuja como hueco, no como cero.
    def _registros(d, cols=None):
        d = d[cols].round(3) if cols else d.round(3)
        return d.astype(object).where(pd.notna(d), None).to_dict(orient="records")

    payload = {
        "meta": {
            "ventana_cierre_bdays": VENTANA_CIERRE_BDAYS,
            "umbral_inicio_retiro": UMBRAL_INICIO_RETIRO,
            "banco_foco": BANCO_FOCO,
        },
        # Dos series con anclajes distintos: g1 cuenta días hacia atrás desde el
        # cierre, g1_inicio cuenta hacia adelante desde el primer día hábil.
        "g1_ciclo_encaje": _registros(datos1),
        "g1_inicio_mes": _registros(datos1_ini),
        "g2_inicio_retiro": _registros(datos2),
        "g3_bbva": _registros(datos3),
        # Del índice sale solo el número indexado, no el monto que lo genera.
        # Los montos de g3 (neto_bbva_musd / neto_sistema_musd) SÍ viajan: la
        # tabla del one-pager los muestra, a pedido — quedan legibles en el
        # código fuente de la página publicada.
        "g4_magnitud": _registros(datos4, cols=["año", "indice_retiro"]),
        # Cuota anual de cada banco en la salida neta. Va el porcentaje y el
        # monto: el monto permite ver si una cuota alta viene de un año chico.
        "g5_cuota_bancos": _registros(
            datos5, cols=["año", "banco", "cuota_pct", "neto_musd", "es_global",
                          "trimestres_cerrados"]),
    }
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"), allow_nan=False)


# ── 8. Export a Excel ─────────────────────────────────────────────────────

def _hoja(writer, nombre, datos, nota):
    """Escribe una hoja con la nota en la fila 1 y la tabla desde la fila 3."""
    pd.DataFrame({"": [nota]}).to_excel(writer, sheet_name=nombre, index=False, header=False, startrow=0)
    datos.to_excel(writer, sheet_name=nombre, index=False, startrow=2)

    ws = writer.sheets[nombre]
    from openpyxl.styles import Font, PatternFill
    ws.cell(row=1, column=1).font = Font(italic=True, color="5C6673")
    for cell in ws[3]:
        if cell.value is not None:
            cell.fill = PatternFill("solid", fgColor="1F4D5C")
            cell.font = Font(color="FFFFFF", bold=True)
    for col in ws.columns:
        largo = max((len(str(c.value)) for c in col[2:] if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = max(14, min(largo + 2, 34))
    ws.freeze_panes = "A4"



# ── Verificación línea por línea de un trimestre ────────────────────────────

def verificar_trimestre(df_raw, df, trimestre):
    """Imprime el rastro completo de cómo se llega a las cifras de un trimestre,
    para poder contrastarlo contra un cálculo hecho a mano en Excel."""
    q = pd.Period(trimestre, freq="Q")
    print("\n" + "=" * 74)
    print(f"  VERIFICACIÓN — {q}")
    print("=" * 74)

    print(f"\n[1] FILAS DEL ARCHIVO")
    print(f"    filas leídas ................. {len(df_raw):,}")
    usadas = df_raw.dropna(subset=["Fecha Valor", "monto", "Broker"])
    print(f"    descartadas por dato faltante  {len(df_raw) - len(usadas):,}")
    print(f"    filas usadas ................. {len(usadas):,}")
    print(f"    columnas: Broker | Fecha Valor | Delivery Principal Usd")
    print(f"    NO se filtra por Tipo Instrumento ni Product Type: entran todas.")

    print(f"\n[2] SIGNO")
    print(f"    monto < 0 -> retiro (R, en valor absoluto) | monto > 0 -> depósito (D)")
    print(f"    salida neta = R - D  (positivo = salió plata)")

    print(f"\n[3] ALIAS DE BANCOS")
    for k, v in ALIAS_BANCOS.items():
        n = (df_raw["Broker"].astype(str).str.strip().str.lower() == k.lower()).sum()
        print(f"    {k} -> {v}  ({n:,} filas en todo el archivo)")

    sub = df[df["trimestre"] == q]
    if sub.empty:
        print(f"\n    ⚠ sin datos para {q}")
        return

    mes_cierre = sub[sub["es_mes_cierre_trim"] == 1]
    print(f"\n[4] CALENDARIO")
    completo = int(mes_cierre["mes_completo"].max()) if len(mes_cierre) else 0
    print(f"    mes de cierre del trimestre completo en la serie: {'sí' if completo else 'NO -> excluido'}")

    ventana = mes_cierre[(mes_cierre["mes_completo"] == 1)
                         & (mes_cierre["dias_al_cierre_mes"] < VENTANA_CIERRE_BDAYS)]
    dias = sorted(ventana["fecha"].dt.date.unique())
    print(f"\n[5] VENTANA DE CIERRE — últimos {VENTANA_CIERRE_BDAYS} días hábiles")
    print(f"    fechas incluidas ({len(dias)}):")
    for d in dias:
        dd = ventana[ventana["fecha"].dt.date == d]["dias_al_cierre_mes"].iloc[0]
        print(f"      {d}  ({pd.Timestamp(d).strftime('%a')})  dias_al_cierre = {int(dd)}")
    if not dias:
        print("      (ninguna)")
        return

    print(f"\n[6] POR BANCO EN ESA VENTANA (millones USD)")
    print(f"    {'banco':<16}{'retiros':>14}{'depósitos':>14}{'neto':>14}")
    agg = ventana.groupby("banco")[["R", "D"]].sum()
    agg["neto"] = agg["R"] - agg["D"]
    for b, r in agg.sort_values("neto", ascending=False).iterrows():
        marca = "  <<<" if b == BANCO_FOCO else ""
        print(f"    {b:<16}{r['R']/1e6:>14,.0f}{r['D']/1e6:>14,.0f}{r['neto']/1e6:>14,.0f}{marca}")

    nb = agg.loc[BANCO_FOCO, "neto"] if BANCO_FOCO in agg.index else float("nan")
    ns = agg["neto"].sum()
    print(f"\n[7] AGREGADOS (millones USD)")
    print(f"    neto {BANCO_FOCO:<11} {nb/1e6:>14,.0f}")
    print(f"    neto sistema      {ns/1e6:>14,.0f}")
    print(f"    proporción        {(nb/ns*100 if ns > 0 else float('nan')):>13,.1f}%"
          + ("" if ns > 0 else "   (vacía: el sistema no drenó)"))
    print("=" * 74 + "\n")


# ── main ──────────────────────────────────────────────────────────────────

def main():
    # --verificar 2025Q4  imprime el rastro completo de ese trimestre
    args = [a for a in sys.argv[1:]]
    global VERIFICAR
    VERIFICAR = None
    if "--verificar" in args:
        i = args.index("--verificar")
        VERIFICAR = args[i + 1] if i + 1 < len(args) else "2025Q4"
        del args[i:i + 2]
    ruta = pathlib.Path(args[0]) if args else RUTA_DEFAULT
    if not ruta.exists():
        print(f"✗ No se encuentra el archivo: {ruta}")
        print("  Pásalo como argumento: python aux_graficos_hallazgos.py <ruta.xlsx>")
        sys.exit(1)

    df_raw = pd.read_excel(ruta)
    df_raw["Fecha Valor"] = pd.to_datetime(df_raw["Fecha Valor"], errors="coerce")
    df_raw["monto"] = pd.to_numeric(df_raw["Delivery Principal Usd"], errors="coerce")

    df = cargar_transacciones(ruta)
    df = anotar_calendario_encaje(df)

    if VERIFICAR:
        verificar_trimestre(df_raw, df, VERIFICAR)

    RUTA_SALIDA.mkdir(parents=True, exist_ok=True)

    print("\nGenerando gráfico 1 — ciclo de encaje (slider por año)...")
    fig1, datos1, datos1_ini = graf_ciclo_encaje(df)
    fig1.write_html(RUTA_SALIDA / "01_ciclo_encaje_slider.html", include_plotlyjs=PLOTLYJS_MODE)

    print("Generando gráfico 2 — retiro cada vez más temprano...")
    fig2, datos2, detalle2 = graf_retiro_mas_temprano(df)
    fig2.write_html(RUTA_SALIDA / "02_retiro_mas_temprano.html", include_plotlyjs=PLOTLYJS_MODE)

    print("Generando gráfico 3 — BBVA arrastra el agregado...")
    fig3, datos3, detalle3 = graf_bbva_arrastra(df)
    fig3.write_html(RUTA_SALIDA / "03_bbva_arrastra_agregado.html", include_plotlyjs=PLOTLYJS_MODE)

    print("Calculando magnitud absoluta del retiro (índice por año)...")
    datos4 = datos_magnitud_retiro(df)

    print("Calculando cuota anual por banco...")
    datos5, cuota_ancha = datos_cuota_por_banco(df)
    if not datos5.empty:
        ult = datos5[datos5["año"] == datos5["año"].max()]
        print(f"  Cuota de la salida neta en {int(datos5['año'].max())} "
              f"({int(ult['trimestres_cerrados'].iloc[0])} trimestres):")
        for _, r in ult.head(6).iterrows():
            if pd.notna(r["cuota_pct"]):
                g = " [global]" if r["es_global"] else ""
                print(f"    {r['banco']:<14}{r['cuota_pct']:>6.1f}%{g}")

    # Excel de validación: una hoja por serie graficada + el agregado base
    print("Exportando Excel de validación...")
    base = df.copy()
    base["trimestre"] = base["trimestre"].astype(str)
    ruta_xlsx = RUTA_SALIDA / "datos_graficos_hallazgos.xlsx"
    with pd.ExcelWriter(ruta_xlsx, engine="openpyxl") as writer:
        _hoja(writer, "G1_inicio_mes", datos1_ini,
              "Grafico 1 (inicio) — mismo calculo anclado al PRIMER dia habil del mes. "
              "n_meses indica cuantos meses aportan a cada posicion.")
        _hoja(writer, "G1_ciclo_encaje", datos1,
              "Grafico 1 — promedio por año/día-al-cierre del % del volumen mensual (una fila por línea del slider)")
        _hoja(writer, "G2_inicio_retiro", datos2,
              f"Grafico 2 — día en que el retiro acumulado del tramo de cierre supera {UMBRAL_INICIO_RETIRO:.0%}")
        _hoja(writer, "G2_detalle_diario", detalle2,
              "Grafico 2 (insumo) — retiro del sistema por trimestre y día-al-cierre")
        _hoja(writer, "G3_series", datos3,
              "Grafico 3 — participación de BBVA e intensidad de retiro, por trimestre")
        _hoja(writer, "G3_detalle_banco", detalle3,
              "Grafico 3 (insumo) — retiro e intensidad banco por banco, por trimestre")
        _hoja(writer, "G4_magnitud", datos4,
              f"Grafico 4 — SALIDA NETA (retiros - depositos) del sistema en los últimos "
              f"{VENTANA_CIERRE_BDAYS} días hábiles de cierre-trimestre, por año, indexada al primer año=100. "
              f"'retiro_bruto' se incluye para comparar cuanto sobreestima el bruto. "
              f"Al JSON del one-pager solo va 'indice_retiro'.")
        _hoja(writer, "G5_cuota_por_banco", datos5,
              "Grafico 5 — cuota de cada banco en la salida neta del sistema, por AÑO. "
              "Agregado anual real (suma los netos del año y despues divide), no promedio "
              "de cuotas trimestrales. es_global marca los bancos con matriz en el exterior.")
        _hoja(writer, "G5_cuota_ancha", cuota_ancha,
              "Grafico 5 (vista ancha) — una fila por año, una columna por banco.")
        _hoja(writer, "Base_banco_fecha", base,
              "Agregado base: R y D por banco y fecha, tras alias y limpieza (insumo de los 3 gráficos)")

    # JSON compacto para incrustar las curvas reales en el one-pager
    print("Exportando JSON para el one-pager...")
    ruta_json = RUTA_SALIDA / "datos_para_onepager.json"
    exportar_json_artifact(datos1, datos1_ini, datos2, datos3, datos4, datos5, ruta_json)

    print(f"\nListo. Archivos en: {RUTA_SALIDA}")
    print(f"  Excel de validación : {ruta_xlsx.name}")
    print(f"  JSON one-pager      : {ruta_json.name} ({ruta_json.stat().st_size / 1024:.0f} KB)")


if __name__ == "__main__":
    main()
